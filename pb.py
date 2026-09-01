# -*- coding: utf-8 -*-
import sys
import json
import math
import copy
import os
import hashlib
from xml.sax.saxutils import escape as xml_escape
from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog

# --- Configuration & Constants ---
DEFAULT_FONT = "SimSun"
DEFAULT_FONT_SIZE = 24
COLUMN_SPACING = 10
LINE_HEIGHT_RATIO = 1.2
CORELDRAW_EXPORT_DPI = 600  # Scene pixels to millimeters for CorelDRAW SVG export.
APP_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(APP_DIR, "assets")  # 素材库目录
DEFAULT_LINE_WIDTH = 3  # 默认连接线粗细（像素）
CONFIG_FILE = os.path.join(APP_DIR, "config.json")  # 配置文件
DEFAULT_LICENSE_PASSWORD = "12345678"


def hash_license_password(password):
    """生成授权密码哈希，避免在配置文件中直接保存明文密码。"""
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def license_is_valid(config):
    """检查授权日期；未设置到期日时视为长期有效。"""
    expiry_text = str(config.get("license_expiry_date", "") or "").strip()
    if not expiry_text:
        return True, ""

    expiry_date = QDate.fromString(expiry_text, "yyyy-MM-dd")
    if not expiry_date.isValid():
        return False, f"授权到期日期无效：{expiry_text}"
    if QDate.currentDate() > expiry_date:
        return False, f"软件授权已于 {expiry_text} 到期"
    return True, ""


def recover_expired_license(config_manager):
    """到期后允许管理员验证密码并延长授权日期。"""
    password, ok = QInputDialog.getText(
        None,
        "软件已到期",
        "请输入授权密码以修改到期日期：",
        QLineEdit.EchoMode.Password
    )
    if not ok:
        return False

    stored_hash = config_manager.get(
        "license_password_hash",
        hash_license_password(DEFAULT_LICENSE_PASSWORD)
    )
    if hash_license_password(password) != stored_hash:
        QMessageBox.critical(None, "授权验证失败", "密码不正确，软件无法使用。")
        return False

    dialog = QDialog()
    dialog.setWindowTitle("延长软件授权")
    layout = QFormLayout(dialog)
    layout.addRow(QLabel("请输入新的到期日期，保存后软件才能继续使用。"))
    expiry_edit = QDateEdit(QDate.currentDate(), dialog)
    expiry_edit.setCalendarPopup(True)
    expiry_edit.setDisplayFormat("yyyy-MM-dd")
    expiry_edit.setMinimumDate(QDate.currentDate())
    layout.addRow("新的到期日期：", expiry_edit)

    buttons = QDialogButtonBox(
        QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
    )
    buttons.button(QDialogButtonBox.StandardButton.Ok).setText("保存并继续")
    buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("退出")
    buttons.accepted.connect(dialog.accept)
    buttons.rejected.connect(dialog.reject)
    layout.addRow(buttons)

    if dialog.exec() != QDialog.DialogCode.Accepted:
        return False

    config_manager.set(
        "license_expiry_date",
        expiry_edit.date().toString("yyyy-MM-dd")
    )
    return True

# Vertically sensitive characters (Simple Heuristic for demo)
ROTATE_CHARS = {'—', '…', '(', ')', '[', ']', '{', '}', '《', '》', '-', '_'}
OFFSET_CHARS = {'，', '。', '、', '：', '；', '！', '？', ',', '.', '!', '?'}

class ConfigManager:
    """配置管理器"""
    def __init__(self):
        self.config_file = CONFIG_FILE
        self.config = self.load_config()
    
    def load_config(self):
        """加载配置"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                print(f"配置已加载: {self.config_file}")
                return config
            except Exception as e:
                print(f"加载配置失败: {e}")
                return self.get_default_config()
        else:
            print("配置文件不存在，使用默认配置")
            return self.get_default_config()
    
    def get_default_config(self):
        """获取默认配置"""
        return {
            'default_background_image': '',
            'background_opacity': 1.0,
            'background_scale_mode': 'fit',
            'default_font_family': DEFAULT_FONT,
            'default_font_size': DEFAULT_FONT_SIZE,
            'default_text_format': {
                'font_family': DEFAULT_FONT,
                'font_size': DEFAULT_FONT_SIZE,
                'text_color': '#000000',
                'chars_per_column': 15,
                'column_spacing': COLUMN_SPACING,
                'character_spacing': 0,
                'manual_line_break': True,
            },
            'text_format_presets': [],
            'default_line_width': DEFAULT_LINE_WIDTH,  # 默认连线粗细
            'bg_above_connectors': False,  # 背景图片是否在连线之上
            'marquee_only_images': False,  # 框选时仅选择图片（兼容旧版）
            'marquee_only_connected': False,  # 框选时仅选择有连接点的元素（兼容旧版）
            'marquee_mode': 'all',  # 框选模式：all / images / connected
            'insert_image_to_bottom': False,  # 插入图片时置于底层
            'nudge_large_step': 10,  # Shift+方向键大步长（像素）
            'horizontal_move_only': False,  # 拖动/方向键移动时只允许水平移动
            'image_right_edge_snap_enabled': False,  # 水平移动图片时吸附右边缘X
            'image_top_edge_snap_enabled': False,  # 移动图片时吸附顶部Y
            'startup_horizontal_guides_enabled': False,  # 启动/新建画布时添加默认横向辅助线
            'startup_horizontal_guides': [],  # 启动时横向辅助线Y坐标（场景像素）
            'startup_horizontal_guides_unit': 'px',  # 启动横向辅助线输入单位：px / mm
            'display_unit': 'mm',  # 全局显示/输入单位：px / mm
            'smart_brush_radius': 18,  # 智能笔刷默认大小
            'default_save_dir': '',  # 默认保存目录
            'insert_image_max_width_ratio': 0.3,  # 插入图片最大宽度（画布宽度的比例）
            'insert_image_default_width': 0,  # 插入图片默认宽度（px），0表示按比例自动
            'insert_image_default_height': 0,  # 插入图片默认高度（px），0表示按宽度比例自动
            'insert_image_use_custom_size': False,  # 是否使用自定义插入尺寸
            'insert_image_fit_canvas': False,  # 插入图片时自动适应画布大小
            'hide_connection_points_when_image_selected': False,
            'show_text_hover_tooltip': False,
            'favorite_fonts': ['SimSun', 'Microsoft YaHei', '黑体', '楷体', 'Arial'],
            'favorite_sizes': [10, 12, 14, 16, 18, 20, 24, 30, 36, 48, 72],
            'license_password_hash': hash_license_password(DEFAULT_LICENSE_PASSWORD),
            'license_expiry_date': '',
        }
    
    def save_config(self):
        """保存配置"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
            print(f"配置已保存: {self.config_file}")
            return True
        except Exception as e:
            print(f"保存配置失败: {e}")
            return False
    
    def get(self, key, default=None):
        """获取配置项"""
        return self.config.get(key, default)
    
    def set(self, key, value):
        """设置配置项"""
        self.config[key] = value
        self.save_config()

class AssetManager:
    """素材管理器"""
    def __init__(self):
        self.assets_file = os.path.join(ASSETS_DIR, "assets.json")
        self.ensure_assets_dir()
        self.load_assets()
    
    def ensure_assets_dir(self):
        """确保素材目录存在"""
        if not os.path.exists(ASSETS_DIR):
            os.makedirs(ASSETS_DIR)
            print(f"创建素材目录: {ASSETS_DIR}")
        else:
            print(f"素材目录已存在 {ASSETS_DIR}")
    
    def load_assets(self):
        """加载素材库"""
        print(f"加载素材库 {self.assets_file}")
        if os.path.exists(self.assets_file):
            try:
                with open(self.assets_file, 'r', encoding='utf-8') as f:
                    self.assets = json.load(f)
                print(f"成功加载素材库 文字{len(self.assets.get('texts', []))}条 图片{len(self.assets.get('images', []))}条 组合{len(self.assets.get('groups', []))}条")
            except Exception as e:
                print(f"加载素材库失败 {e}")
                self.assets = {"texts": [], "images": [], "groups": []}
        else:
            print("素材库文件不存在，创建新的")
            self.assets = {"texts": [], "images": [], "groups": []}
    
    def save_assets(self):
        """保存素材库"""
        print(f"保存素材库到: {self.assets_file}")
        print(f"素材数据: 文字{len(self.assets.get('texts', []))}条 图片{len(self.assets.get('images', []))}条 组合{len(self.assets.get('groups', []))}条")
        with open(self.assets_file, 'w', encoding='utf-8') as f:
            json.dump(self.assets, f, indent=2, ensure_ascii=False)
        print("素材库保存完成")
    
    def add_text_asset(self, text_item):
        """添加文字素材"""
        print(f"开始保存文字素材 {text_item.full_text[:20]}")
        asset_data = {
            'id': len(self.assets['texts']),
            'name': text_item.full_text[:20] + ('...' if len(text_item.full_text) > 20 else ''),
            'text': text_item.full_text,
            'font_size': text_item.font_size,
            'box_height': text_item.box_height,
            'font_family': text_item.font_family,
            'text_color': text_item.text_color.name(),
            'chars_per_column': text_item.chars_per_column,
            'column_spacing': text_item.column_spacing,
            'character_spacing': text_item.character_spacing,
            'auto_height': text_item.auto_height,
            'manual_line_break': text_item.manual_line_break,
            'created_time': QDateTime.currentDateTime().toString()
        }
        self.assets['texts'].append(asset_data)
        print(f"文字素材数据: {asset_data}")
        self.save_assets()
        print(f"当前文字素材总数: {len(self.assets['texts'])}")
        return asset_data
    
    def add_image_asset(self, image_item):
        """添加图片素材"""
        print(f"开始保存图片素材 {image_item.file_path}")
        # 复制图片到素材目录
        original_path = image_item.file_path
        filename = os.path.basename(original_path)
        asset_path = os.path.join(ASSETS_DIR, f"img_{len(self.assets['images'])}_{filename}")
        
        try:
            # 复制文件
            import shutil
            shutil.copy2(original_path, asset_path)
            print(f"图片已复制到: {asset_path}")
            
            asset_data = {
                'id': len(self.assets['images']),
                'name': filename,
                'path': asset_path,
                'original_path': original_path,
                'width': image_item.target_width,
                'created_time': QDateTime.currentDateTime().toString()
            }
            self.assets['images'].append(asset_data)
            print(f"图片素材数据: {asset_data}")
            self.save_assets()
            print(f"当前图片素材总数: {len(self.assets['images'])}")
            return asset_data
        except Exception as e:
            print(f"复制图片失败: {e}")
            return None
    
    def get_text_assets(self):
        """获取所有文字素材"""
        return self.assets.get('texts', [])
    
    def get_image_assets(self):
        """获取所有图片素材"""
        return self.assets.get('images', [])
    
    def add_group_asset(self, items, scene):
        """添加组合素材(支持父子关系和图文连接)"""
        if not items:
            return None
        
        # 创建项目到索引的映射
        item_to_index = {item: idx for idx, item in enumerate(items)}
        
        # 保存所有项目的数据
        items_data = []
        for idx, item in enumerate(items):
            if isinstance(item, VTextItem):
                # 保存连接点状态
                connection_point_visible = item.connection_point.isVisible() if item.connection_point else True
                connection_point_deleted = item.connection_point is None

                item_data = {
                    'type': 'VTextItem',
                    'text': item.full_text,
                    'font_size': item.font_size,
                    'box_height': item.box_height,
                    'font_family': item.font_family,
                    'text_color': item.text_color.name(),
                    'chars_per_column': item.chars_per_column,
                    'column_spacing': item.column_spacing,
                    'character_spacing': item.character_spacing,
                    'auto_height': item.auto_height,
                    'manual_line_break': item.manual_line_break,
                    'layer_eye_color': getattr(item, 'layer_eye_color', None),
                    'connection_point_visible': connection_point_visible,
                    'connection_point_deleted': connection_point_deleted,
                    'scene_pos': (item.scenePos().x(), item.scenePos().y()),
                    'local_pos': (item.x(), item.y()),
                    'parent_index': item_to_index.get(item.parentItem(), -1) if isinstance(item.parentItem(), BaseElement) else -1
                }
                items_data.append(item_data)
            elif isinstance(item, VImageItem):
                # 复制图片文件
                original_path = item.file_path
                filename = os.path.basename(original_path)
                
                # 清理文件名：移除已有的 group_ 前缀，避免重复叠加
                import re
                # 匹配 group_数字_数字_ 开头的模式并移除
                clean_filename = re.sub(r'^(group_\d+_\d+_)+', '', filename)
                
                # 生成新的唯一文件名
                import time
                timestamp = int(time.time() * 1000)  # 毫秒级时间戳
                asset_path = os.path.join(ASSETS_DIR, f"group_{timestamp}_{idx}_{clean_filename}")

                try:
                    import shutil
                    shutil.copy2(original_path, asset_path)

                    # 保存连接点状态
                    connection_point_visible = item.connection_point.isVisible() if item.connection_point else True
                    connection_point_deleted = item.connection_point is None

                    item_data = {
                        'type': 'VImageItem',
                        'path': asset_path,
                        'original_path': original_path,
                        'width': item.target_width,
                        'connection_point_visible': connection_point_visible,
                        'connection_point_deleted': connection_point_deleted,
                        'scene_pos': (item.scenePos().x(), item.scenePos().y()),
                        'local_pos': (item.x(), item.y()),
                        'parent_index': item_to_index.get(item.parentItem(), -1) if isinstance(item.parentItem(), BaseElement) else -1
                    }
                    items_data.append(item_data)
                except Exception as e:
                    print(f"复制图片失败: {e}")
                    return None
        
        # 保存图文连接关系
        image_text_connections = []
        for conn in scene.image_text_connectors:
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                img_idx = item_to_index.get(conn.image_item, -1)
                text_idx = item_to_index.get(conn.text_item, -1)
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                img_idx = item_to_index.get(conn.item1, -1)
                text_idx = item_to_index.get(conn.item2, -1)
            else:
                continue
            if img_idx != -1 and text_idx != -1:
                image_text_connections.append((img_idx, text_idx))
        
        # 生成组合名称
        text_count = sum(1 for item in items if isinstance(item, VTextItem))
        image_count = sum(1 for item in items if isinstance(item, VImageItem))
        group_name = f"组合_{text_count}文字_{image_count}图片"

        # 生成缩略图：把选中元素渲染到 QPixmap
        thumb_path = ''
        try:
            if scene and scene.views():
                # 计算所有元素的包围盒（场景坐标）
                combined = QRectF()
                for item in items:
                    r = QRectF(item.scenePos(), item.boundingRect().size())
                    combined = combined.united(r)
                if not combined.isEmpty():
                    margin = 10
                    combined = combined.adjusted(-margin, -margin, margin, margin)
                    # 渲染到 QImage，最大 300x400
                    max_w, max_h = 300, 400
                    scale = min(max_w / combined.width(), max_h / combined.height(), 1.0)
                    img_w = max(1, int(combined.width() * scale))
                    img_h = max(1, int(combined.height() * scale))
                    img = QImage(img_w, img_h, QImage.Format.Format_ARGB32)
                    if img.isNull():
                        raise ValueError("QImage 创建失败，尺寸无效")
                    img.fill(QColor(250, 250, 245))
                    painter = QPainter(img)
                    if not painter.isActive():
                        raise ValueError("QPainter 启动失败")
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    # 隐藏连接点和辅助线，只渲染元素本身
                    scene._rendering_thumb = True
                    scene.render(painter, QRectF(0, 0, img_w, img_h), combined)
                    scene._rendering_thumb = False
                    painter.end()
                    # 使用时间戳生成唯一的缩略图文件名
                    import time
                    timestamp = int(time.time() * 1000)
                    thumb_path = os.path.join(ASSETS_DIR, f"group_{timestamp}_thumb.png")
                    img.save(thumb_path)
                    print(f"缩略图已保存: {thumb_path}")
        except Exception as e:
            print(f"生成缩略图失败: {e}")
            thumb_path = ''

        # 创建组合素材数据 - 使用唯一 ID 避免重复
        existing_ids = {asset.get('id', -1) for asset in self.assets.get('groups', [])}
        new_id = 0
        while new_id in existing_ids:
            new_id += 1
        
        group_asset = {
            'id': new_id,
            'name': group_name,
            'items': items_data,
            'image_text_connections': image_text_connections,
            'item_count': len(items),
            'thumb_path': thumb_path,
            'created_time': QDateTime.currentDateTime().toString()
        }
        
        self.assets['groups'].append(group_asset)
        self.save_assets()
        return group_asset
    
    def get_group_assets(self):
        """获取所有组合素材"""
        return self.assets.get('groups', [])
    
    def remove_group_asset(self, asset_id):
        """删除组合素材"""
        asset_to_remove = None
        for asset in self.assets['groups']:
            if asset['id'] == asset_id:
                asset_to_remove = asset
                break

        if asset_to_remove:
            for item_data in asset_to_remove['items']:
                if item_data['type'] == 'VImageItem':
                    try:
                        if os.path.exists(item_data['path']):
                            os.remove(item_data['path'])
                    except:
                        pass
            thumb = asset_to_remove.get('thumb_path', '')
            if thumb and os.path.exists(thumb):
                try:
                    os.remove(thumb)
                except:
                    pass
            self.assets['groups'] = [a for a in self.assets['groups'] if a['id'] != asset_id]
            self.save_assets()

    def _delete_group_asset_files(self, asset, keep_paths=None):
        """删除组合素材关联文件，可指定保留路径避免误删新素材。"""
        keep_paths = {os.path.abspath(p) for p in (keep_paths or []) if p}

        def safe_remove(path):
            if not path:
                return
            abs_path = os.path.abspath(path)
            if abs_path in keep_paths:
                return
            try:
                if os.path.exists(abs_path):
                    os.remove(abs_path)
            except Exception as e:
                print(f"删除素材文件失败 {abs_path}: {e}")

        for item_data in asset.get('items', []):
            if item_data.get('type') == 'VImageItem':
                safe_remove(item_data.get('path', ''))
        safe_remove(asset.get('thumb_path', ''))

    def update_group_asset(self, asset_id, items, scene):
        """用当前画布上的元素更新已有组合素材"""
        # 找到原素材位置
        idx = next((i for i, a in enumerate(self.assets['groups']) if a['id'] == asset_id), None)
        if idx is None:
            return None
        old_asset = self.assets['groups'][idx]
        old_name = old_asset.get('name', '')

        # 先重建新素材，确认成功后再替换旧素材；避免先删旧图片导致复制源文件丢失
        try:
            new_asset = self.add_group_asset(items, scene)
        except Exception as e:
            print(f"更新组合素材失败: {e}")
            return None

        if new_asset:
            # 恢复原有 id 和 name
            new_asset['id'] = asset_id
            new_asset['name'] = old_name
            # 替换回原位置
            if new_asset in self.assets['groups']:
                self.assets['groups'].remove(new_asset)
            self.assets['groups'][idx] = new_asset
            self.save_assets()
            keep_paths = [
                item_data.get('path', '')
                for item_data in new_asset.get('items', [])
                if item_data.get('type') == 'VImageItem'
            ]
            keep_paths.append(new_asset.get('thumb_path', ''))
            self._delete_group_asset_files(old_asset, keep_paths)
        return new_asset
    
    def remove_text_asset(self, asset_id):
        """删除文字素材"""
        self.assets['texts'] = [a for a in self.assets['texts'] if a['id'] != asset_id]
        self.save_assets()
    
    def remove_image_asset(self, asset_id):
        """删除图片素材"""
        # 找到要删除的素材
        asset_to_remove = None
        for asset in self.assets['images']:
            if asset['id'] == asset_id:
                asset_to_remove = asset
                break
        
        if asset_to_remove:
            # 删除文件
            try:
                if os.path.exists(asset_to_remove['path']):
                    os.remove(asset_to_remove['path'])
            except:
                pass
            
            # 从列表中删除
            self.assets['images'] = [a for a in self.assets['images'] if a['id'] != asset_id]
            self.save_assets()

class BatchCopyDialog(QDialog):
    """步长和重复 - 仿 CDR 风格"""
    SPINBOX_STYLE = """
        QSpinBox, QDoubleSpinBox {
            background: white; border: 1px solid #aaa; border-radius: 4px;
            padding: 3px 22px 3px 6px; min-height: 22px; color: #323130;
        }
        QSpinBox::up-button, QDoubleSpinBox::up-button {
            subcontrol-origin: border; subcontrol-position: top right; width: 18px;
        }
        QSpinBox::down-button, QDoubleSpinBox::down-button {
            subcontrol-origin: border; subcontrol-position: bottom right; width: 18px;
        }
        QGroupBox { font-weight: bold; margin-top: 6px; padding-top: 10px; }
        QGroupBox::title { subcontrol-origin: margin; left: 8px; }
        QRadioButton { spacing: 5px; }
        QLabel { color: #323130; }
    """

    def __init__(self, item_w=0, item_h=0, parent=None):
        super().__init__(parent)
        self.setWindowTitle("步长和重复")
        self.setFixedWidth(340)
        self.setStyleSheet(self.SPINBOX_STYLE)
        self.item_w = item_w  # 选中元素宽度，用于"按对象尺寸"预设
        self.item_h = item_h

        root = QVBoxLayout(self)
        root.setSpacing(8)

        # ── 副本数量 ──────────────────────────────
        count_group = QGroupBox("副本数量")
        cg = QFormLayout(count_group)
        self.spin_count = QSpinBox()
        self.spin_count.setRange(1, 999)
        self.spin_count.setValue(3)
        cg.addRow("份数:", self.spin_count)
        root.addWidget(count_group)

        # ── 水平偏移 ──────────────────────────────
        h_group = QGroupBox("水平设置")
        hg = QVBoxLayout(h_group)

        h_mode_layout = QHBoxLayout()
        self.rb_h_none   = QRadioButton("不偏移")
        self.rb_h_offset = QRadioButton("指定偏移")
        self.rb_h_space  = QRadioButton("按间距")
        self.rb_h_obj    = QRadioButton("按对象宽度")
        self.rb_h_offset.setChecked(True)
        for rb in (self.rb_h_none, self.rb_h_offset, self.rb_h_space, self.rb_h_obj):
            h_mode_layout.addWidget(rb)
        hg.addLayout(h_mode_layout)

        h_val_layout = QFormLayout()
        self.dspin_h_offset = QDoubleSpinBox()
        self.dspin_h_offset.setRange(-9999, 9999); self.dspin_h_offset.setValue(100); self.dspin_h_offset.setSuffix(" px")
        self.dspin_h_space  = QDoubleSpinBox()
        self.dspin_h_space.setRange(-9999, 9999);  self.dspin_h_space.setValue(20);  self.dspin_h_space.setSuffix(" px")
        h_val_layout.addRow("X 偏移量:", self.dspin_h_offset)
        h_val_layout.addRow("间距:", self.dspin_h_space)
        hg.addLayout(h_val_layout)
        root.addWidget(h_group)

        # ── 垂直偏移 ──────────────────────────────
        v_group = QGroupBox("垂直设置")
        vg = QVBoxLayout(v_group)

        v_mode_layout = QHBoxLayout()
        self.rb_v_none   = QRadioButton("不偏移")
        self.rb_v_offset = QRadioButton("指定偏移")
        self.rb_v_space  = QRadioButton("按间距")
        self.rb_v_obj    = QRadioButton("按对象高度")
        self.rb_v_none.setChecked(True)
        for rb in (self.rb_v_none, self.rb_v_offset, self.rb_v_space, self.rb_v_obj):
            v_mode_layout.addWidget(rb)
        vg.addLayout(v_mode_layout)

        v_val_layout = QFormLayout()
        self.dspin_v_offset = QDoubleSpinBox()
        self.dspin_v_offset.setRange(-9999, 9999); self.dspin_v_offset.setValue(0); self.dspin_v_offset.setSuffix(" px")
        self.dspin_v_space  = QDoubleSpinBox()
        self.dspin_v_space.setRange(-9999, 9999);  self.dspin_v_space.setValue(20);  self.dspin_v_space.setSuffix(" px")
        v_val_layout.addRow("Y 偏移量:", self.dspin_v_offset)
        v_val_layout.addRow("间距:", self.dspin_v_space)
        vg.addLayout(v_val_layout)
        root.addWidget(v_group)

        # ── 按钮 ──────────────────────────────────
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        # 联动
        for rb in (self.rb_h_none, self.rb_h_offset, self.rb_h_space, self.rb_h_obj):
            rb.toggled.connect(self._update_controls)
        for rb in (self.rb_v_none, self.rb_v_offset, self.rb_v_space, self.rb_v_obj):
            rb.toggled.connect(self._update_controls)
        self._update_controls()

    def _update_controls(self):
        h_off = self.rb_h_offset.isChecked()
        h_sp  = self.rb_h_space.isChecked()
        self.dspin_h_offset.setEnabled(h_off)
        self.dspin_h_space.setEnabled(h_sp)

        v_off = self.rb_v_offset.isChecked()
        v_sp  = self.rb_v_space.isChecked()
        self.dspin_v_offset.setEnabled(v_off)
        self.dspin_v_space.setEnabled(v_sp)

    def _h_step(self):
        if self.rb_h_none.isChecked():   return 0
        if self.rb_h_offset.isChecked(): return self.dspin_h_offset.value()
        if self.rb_h_space.isChecked():  return self.item_w + self.dspin_h_space.value()
        if self.rb_h_obj.isChecked():    return self.item_w
        return 0

    def _v_step(self):
        if self.rb_v_none.isChecked():   return 0
        if self.rb_v_offset.isChecked(): return self.dspin_v_offset.value()
        if self.rb_v_space.isChecked():  return self.item_h + self.dspin_v_space.value()
        if self.rb_v_obj.isChecked():    return self.item_h
        return 0

    def get_params(self):
        return {
            'count': self.spin_count.value(),
            'step_x': self._h_step(),
            'step_y': self._v_step(),
        }


class FontPickerDialog(QDialog):
    """自定义字体选择器 - 支持常用字体/字号快捷按钮，可由用户自定义"""

    def __init__(self, current_font: QFont, config_manager, parent=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.setWindowTitle("字体选择器")
        self.setMinimumSize(820, 560)
        self.resize(900, 600)

        self._selected_family = current_font.family()
        self._selected_size = current_font.pointSize() if current_font.pointSize() > 0 else 24

        self._all_families = QFontDatabase.families()
        self._build_ui()
        self._refresh_fav_fonts()
        self._refresh_fav_sizes()
        self._update_preview()

    # ── UI 构建 ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        root = QHBoxLayout(self)
        root.setSpacing(12)
        root.setContentsMargins(12, 12, 12, 12)

        # ── 左侧：字体列表 ────────────────────────────────────────────────
        left = QVBoxLayout()
        left.setSpacing(6)

        search_row = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索字体...")
        self.search_edit.textChanged.connect(self._filter_fonts)
        search_row.addWidget(self.search_edit)
        left.addLayout(search_row)

        self.font_list = QListWidget()
        self.font_list.setMinimumWidth(280)
        for f in self._all_families:
            item = QListWidgetItem(f)
            item.setFont(QFont(f, 13))
            self.font_list.addItem(item)
        self.font_list.currentTextChanged.connect(self._on_family_changed)
        # 定位到当前字体
        matches = self.font_list.findItems(self._selected_family, Qt.MatchFlag.MatchExactly)
        if matches:
            self.font_list.setCurrentItem(matches[0])
            self.font_list.scrollToItem(matches[0])
        left.addWidget(self.font_list)

        root.addLayout(left, 3)

        # ── 右侧 ──────────────────────────────────────────────────────────
        right = QVBoxLayout()
        right.setSpacing(8)

        # 常用字体 + 常用字号 并排
        fav_row = QHBoxLayout()
        fav_row.setSpacing(12)

        # 常用字体区
        fav_font_box = QGroupBox("常用字体")
        fav_font_layout = QVBoxLayout(fav_font_box)
        fav_font_layout.setSpacing(4)
        self.fav_font_grid = QGridLayout()
        self.fav_font_grid.setSpacing(4)
        fav_font_layout.addLayout(self.fav_font_grid)
        btn_edit_fav_fonts = QPushButton("编辑常用字体...")
        btn_edit_fav_fonts.setFixedHeight(24)
        btn_edit_fav_fonts.clicked.connect(self._edit_fav_fonts)
        fav_font_layout.addWidget(btn_edit_fav_fonts)
        fav_row.addWidget(fav_font_box, 3)

        # 常用字号区
        fav_size_box = QGroupBox("常用字号")
        fav_size_layout = QVBoxLayout(fav_size_box)
        fav_size_layout.setSpacing(4)

        # 顶部：spinbox + 滑块
        size_ctrl_row = QHBoxLayout()
        self.size_spin = QSpinBox()
        self.size_spin.setRange(4, 400)
        self.size_spin.setValue(self._selected_size)
        self.size_spin.setSuffix(" pt")
        self.size_spin.setFixedWidth(80)
        self.size_spin.valueChanged.connect(self._on_size_spin_changed)
        size_ctrl_row.addWidget(self.size_spin)
        fav_size_layout.addLayout(size_ctrl_row)

        self.size_slider = QSlider(Qt.Orientation.Horizontal)
        self.size_slider.setRange(4, 200)
        self.size_slider.setValue(min(self._selected_size, 200))
        self.size_slider.valueChanged.connect(self._on_slider_changed)
        fav_size_layout.addWidget(self.size_slider)

        self.fav_size_grid = QGridLayout()
        self.fav_size_grid.setSpacing(4)
        fav_size_layout.addLayout(self.fav_size_grid)

        btn_edit_fav_sizes = QPushButton("编辑常用字号...")
        btn_edit_fav_sizes.setFixedHeight(24)
        btn_edit_fav_sizes.clicked.connect(self._edit_fav_sizes)
        fav_size_layout.addWidget(btn_edit_fav_sizes)
        fav_row.addWidget(fav_size_box, 2)

        right.addLayout(fav_row)

        # 预览区
        preview_box = QGroupBox("示例")
        preview_layout = QVBoxLayout(preview_box)
        self.preview_label = QLabel("YyZz 你好世界 AaBbCc")
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_label.setMinimumHeight(100)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)

        self.custom_preview_edit = QLineEdit()
        self.custom_preview_edit.setPlaceholderText("自定义文本...")
        self.custom_preview_edit.textChanged.connect(self._update_preview)
        preview_layout.addWidget(self.custom_preview_edit)
        right.addWidget(preview_box, 1)

        # 确定/取消
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("确定")
        btn_ok.setDefault(True)
        btn_ok.setFixedWidth(80)
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.setFixedWidth(80)
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        right.addLayout(btn_row)

        root.addLayout(right, 4)

    # ── 刷新常用字体按钮 ─────────────────────────────────────────────────────
    def _refresh_fav_fonts(self):
        # 清空旧按钮
        while self.fav_font_grid.count():
            w = self.fav_font_grid.takeAt(0).widget()
            if w:
                w.deleteLater()

        default_favs = ['SimSun', 'Microsoft YaHei', '黑体', '楷体', 'Arial']
        favs = self.config_manager.get('favorite_fonts', default_favs) if self.config_manager else default_favs
        cols = 2
        for idx, fname in enumerate(favs):
            btn = QPushButton(fname)
            btn.setFont(QFont(fname, 10))
            btn.setFixedHeight(28)
            btn.setToolTip(fname)
            btn.clicked.connect(lambda _, f=fname: self._select_family(f))
            self.fav_font_grid.addWidget(btn, idx // cols, idx % cols)

    # ── 刷新常用字号按钮 ─────────────────────────────────────────────────────
    def _refresh_fav_sizes(self):
        while self.fav_size_grid.count():
            w = self.fav_size_grid.takeAt(0).widget()
            if w:
                w.deleteLater()

        default_sizes = [10, 12, 14, 16, 18, 20, 24, 30, 36, 48, 72]
        sizes = self.config_manager.get('favorite_sizes', default_sizes) if self.config_manager else default_sizes
        cols = 4
        for idx, sz in enumerate(sizes):
            btn = QPushButton(str(sz))
            btn.setFixedSize(44, 28)
            btn.clicked.connect(lambda _, s=sz: self._select_size(s))
            self.fav_size_grid.addWidget(btn, idx // cols, idx % cols)

    # ── 事件处理 ─────────────────────────────────────────────────────────────
    def _filter_fonts(self, text):
        for i in range(self.font_list.count()):
            item = self.font_list.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def _on_family_changed(self, family):
        if family:
            self._selected_family = family
            self._update_preview()

    def _select_family(self, family):
        self._selected_family = family
        matches = self.font_list.findItems(family, Qt.MatchFlag.MatchExactly)
        if matches:
            self.font_list.setCurrentItem(matches[0])
            self.font_list.scrollToItem(matches[0])
        self._update_preview()

    def _on_size_spin_changed(self, val):
        self._selected_size = val
        self.size_slider.blockSignals(True)
        self.size_slider.setValue(min(val, 200))
        self.size_slider.blockSignals(False)
        self._update_preview()

    def _on_slider_changed(self, val):
        self._selected_size = val
        self.size_spin.blockSignals(True)
        self.size_spin.setValue(val)
        self.size_spin.blockSignals(False)
        self._update_preview()

    def _select_size(self, size):
        self._selected_size = size
        self.size_spin.blockSignals(True)
        self.size_spin.setValue(size)
        self.size_spin.blockSignals(False)
        self.size_slider.blockSignals(True)
        self.size_slider.setValue(min(size, 200))
        self.size_slider.blockSignals(False)
        self._update_preview()

    def _update_preview(self):
        if not hasattr(self, 'preview_label'):
            return
        text = self.custom_preview_edit.text() if hasattr(self, 'custom_preview_edit') and self.custom_preview_edit.text() else "YyZz 你好世界 AaBbCc"
        preview_size = min(self._selected_size, 72)
        font = QFont(self._selected_family, preview_size)
        self.preview_label.setFont(font)
        self.preview_label.setText(text)

    # ── 编辑常用字体 ──────────────────────────────────────────────────────────
    def _edit_fav_fonts(self):
        if not self.config_manager:
            QMessageBox.information(self, "提示", "当前上下文无法编辑常用字体")
            return
        favs = self.config_manager.get('favorite_fonts', [])
        current_text = '\n'.join(favs)
        text, ok = QInputDialog.getMultiLineText(
            self, "编辑常用字体",
            "每行一个字体名称（可从左侧列表复制字体名）:",
            current_text
        )
        if ok:
            new_favs = [f.strip() for f in text.splitlines() if f.strip()]
            self.config_manager.set('favorite_fonts', new_favs)
            self._refresh_fav_fonts()

    # ── 编辑常用字号 ──────────────────────────────────────────────────────────
    def _edit_fav_sizes(self):
        if not self.config_manager:
            QMessageBox.information(self, "提示", "当前上下文无法编辑常用字号")
            return
        sizes = self.config_manager.get('favorite_sizes', [])
        current_text = ' '.join(str(s) for s in sizes)
        text, ok = QInputDialog.getText(
            self, "编辑常用字号",
            "用空格或逗号分隔字号（如: 10 12 14 18 24 36 72）:",
            text=current_text
        )
        if ok:
            import re
            parts = re.split(r'[\s,]+', text.strip())
            new_sizes = []
            for p in parts:
                try:
                    v = int(p)
                    if 4 <= v <= 400:
                        new_sizes.append(v)
                except ValueError:
                    pass
            new_sizes = sorted(set(new_sizes))
            if new_sizes:
                self.config_manager.set('favorite_sizes', new_sizes)
                self._refresh_fav_sizes()

    # ── 结果获取 ──────────────────────────────────────────────────────────────
    def selected_font(self) -> QFont:
        f = QFont(self._selected_family)
        f.setPointSize(self._selected_size)
        return f

    @staticmethod
    def get_font(current_font: QFont, config_manager, parent=None, title="选择字体"):
        """替代 QFontDialog.getFont 的静态方法，返回 (font, ok)"""
        dlg = FontPickerDialog(current_font, config_manager, parent)
        dlg.setWindowTitle(title)
        ok = dlg.exec() == QDialog.DialogCode.Accepted
        return dlg.selected_font(), ok


class ToastNotification(QWidget):
    """右下角弹出通知，自动消失"""
    def __init__(self, parent=None):
        super().__init__(parent, Qt.WindowType.ToolTip | Qt.WindowType.FramelessWindowHint |
                         Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(8)

        self._icon_label = QLabel("⚠")
        self._icon_label.setStyleSheet("font-size:18px;")
        layout.addWidget(self._icon_label)

        self._msg_label = QLabel()
        self._msg_label.setWordWrap(True)
        self._msg_label.setMaximumWidth(280)
        self._msg_label.setStyleSheet("font-size:13px; color:#323130;")
        layout.addWidget(self._msg_label)

        self.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
                border: 1px solid rgba(0,0,0,0.15);
                border-radius: 8px;
            }
            QLabel { border: none; background: transparent; }
        """)

        self._timer = QTimer(self)
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self._fade_out)

        self._anim = QPropertyAnimation(self, b"windowOpacity")
        self._anim.setDuration(400)
        self._anim.finished.connect(self._on_anim_finished)

    def show_message(self, message, icon="⚠", duration=4000):
        self._icon_label.setText(icon)
        self._msg_label.setText(message)
        self.adjustSize()
        self._position_bottom_right()
        self.setWindowOpacity(0.0)
        self.show()
        self._anim.stop()
        self._anim.setStartValue(0.0)
        self._anim.setEndValue(1.0)
        self._anim.start()
        self._timer.start(duration)

    def _position_bottom_right(self):
        screen = QApplication.primaryScreen().availableGeometry()
        margin = 20
        x = screen.right() - self.width() - margin
        y = screen.bottom() - self.height() - margin
        self.move(x, y)

    def _fade_out(self):
        self._anim.stop()
        self._anim.setStartValue(self.windowOpacity())
        self._anim.setEndValue(0.0)
        self._anim.start()

    def _on_anim_finished(self):
        if self.windowOpacity() == 0.0:
            self.hide()

    def mousePressEvent(self, event):
        """点击立即关闭"""
        self._timer.stop()
        self._fade_out()


class GroupAssetPreviewPopup(QWidget):
    """组合素材悬停预览浮窗"""
    def __init__(self, parent=None):
        super().__init__(parent, Qt.WindowType.ToolTip | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setWindowFlags(Qt.WindowType.ToolTip | Qt.WindowType.FramelessWindowHint |
                            Qt.WindowType.WindowStaysOnTopHint)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        self._container = QWidget(self)
        self._container.setStyleSheet("""
            QWidget {
                background: rgba(255,255,255,0.97);
                border: 1px solid rgba(0,0,0,0.15);
                border-radius: 8px;
            }
        """)
        inner = QVBoxLayout(self._container)
        inner.setContentsMargins(10, 10, 10, 10)
        inner.setSpacing(6)

        self._title = QLabel()
        self._title.setStyleSheet("font-weight:bold; font-size:13px; color:#323130;")
        inner.addWidget(self._title)

        self._img_row = QHBoxLayout()
        self._img_row.setSpacing(4)
        inner.addLayout(self._img_row)

        self._text_label = QLabel()
        self._text_label.setStyleSheet("font-size:12px; color:#605e5c;")
        self._text_label.setWordWrap(True)
        inner.addWidget(self._text_label)

        layout.addWidget(self._container)
        self.setFixedWidth(260)

    def show_asset(self, asset, global_pos):
        # 清空旧图片
        while self._img_row.count():
            w = self._img_row.takeAt(0).widget()
            if w:
                w.deleteLater()

        name = asset.get('name', '')
        items = asset.get('items', [])
        text_count = sum(1 for d in items if d['type'] == 'VTextItem')
        image_count = sum(1 for d in items if d['type'] == 'VImageItem')
        self._title.setText(f"{name}  ({image_count}图 {text_count}文)")

        # 优先显示缩略图
        thumb_path = asset.get('thumb_path', '')
        if thumb_path and os.path.exists(thumb_path):
            pix = QPixmap(thumb_path)
            if not pix.isNull():
                lbl = QLabel()
                scaled = pix.scaled(220, 300,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation)
                lbl.setPixmap(scaled)
                lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                lbl.setStyleSheet("border:1px solid #eee; border-radius:4px; background:#fafaf5;")
                self._img_row.addWidget(lbl)
        else:
            # 没有缩略图时显示各图片小图
            img_items = [d for d in items if d['type'] == 'VImageItem']
            for d in img_items[:4]:
                path = d.get('path', '')
                lbl = QLabel()
                lbl.setFixedSize(52, 52)
                lbl.setStyleSheet("border:1px solid #ddd; border-radius:4px; background:#f5f5f5;")
                if os.path.exists(path):
                    pix = QPixmap(path).scaled(50, 50,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation)
                    lbl.setPixmap(pix)
                    lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                else:
                    lbl.setText("?")
                    lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self._img_row.addWidget(lbl)
            if image_count > 4:
                more = QLabel(f"+{image_count-4}")
                more.setAlignment(Qt.AlignmentFlag.AlignCenter)
                more.setStyleSheet("color:#888; font-size:11px;")
                self._img_row.addWidget(more)
        self._img_row.addStretch()

        # 显示文字内容（最多3条，每条最多20字）
        text_items = [d for d in items if d['type'] == 'VTextItem']
        lines = []
        for d in text_items[:3]:
            t = d.get('text', '').replace('\n', ' ')
            lines.append(t[:20] + ('…' if len(t) > 20 else ''))
        if text_count > 3:
            lines.append(f"…共{text_count}条文字")
        self._text_label.setText('\n'.join(lines) if lines else '')

        self.adjustSize()
        # 定位：在鼠标右侧，避免超出屏幕
        screen = QApplication.primaryScreen().availableGeometry()
        x = global_pos.x() + 16
        y = global_pos.y()
        if x + self.width() > screen.right():
            x = global_pos.x() - self.width() - 8
        if y + self.height() > screen.bottom():
            y = screen.bottom() - self.height()
        self.move(x, y)
        self.show()


class AssetLibraryDockWidget(QDockWidget):
    """素材库停靠面板"""
    def __init__(self, asset_manager, main_window):
        super().__init__("素材库", main_window)
        self.asset_manager = asset_manager
        self.main_window = main_window
        self.setAllowedAreas(Qt.DockWidgetArea.LeftDockWidgetArea | Qt.DockWidgetArea.RightDockWidgetArea)
        self.setFeatures(QDockWidget.DockWidgetFeature.DockWidgetMovable |
                        QDockWidget.DockWidgetFeature.DockWidgetClosable |
                        QDockWidget.DockWidgetFeature.DockWidgetFloatable)
        self.content_widget = QWidget()
        self.setWidget(self.content_widget)
        self.setup_ui()
        self.refresh_assets()

    def setup_ui(self):
        layout = QVBoxLayout(self.content_widget)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # 标签页1：组合素材
        group_tab = QWidget()
        gl = QVBoxLayout(group_tab)
        gl.setContentsMargins(4, 4, 4, 4)
        self.group_list = QListWidget()
        self.group_list.setDragEnabled(True)
        self.group_list.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        self.group_list.setDefaultDropAction(Qt.DropAction.CopyAction)
        self.group_list.itemDoubleClicked.connect(self.rename_group_asset)
        self.group_list.startDrag = self._start_group_drag
        self.group_list.setMouseTracking(True)
        self.group_list.mouseMoveEvent = self._on_group_list_mouse_move
        self.group_list.leaveEvent = self._on_group_list_leave
        self.group_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.group_list.customContextMenuRequested.connect(self._on_group_list_context_menu)
        self._preview_popup = GroupAssetPreviewPopup()
        self._preview_timer = QTimer()
        self._preview_timer.setSingleShot(True)
        self._preview_timer.timeout.connect(self._show_preview)
        self._preview_pos = None
        self._preview_asset = None
        gl.addWidget(self.group_list)
        gb = QHBoxLayout()
        btn_use = QPushButton("使用"); btn_use.setMaximumHeight(30); btn_use.setProperty("class", "primary")
        btn_use.clicked.connect(self.use_group_asset_selected); gb.addWidget(btn_use)
        btn_edit = QPushButton("编辑"); btn_edit.setMaximumHeight(30)
        btn_edit.clicked.connect(self.edit_group_asset_selected); gb.addWidget(btn_edit)
        btn_ren = QPushButton("重命名"); btn_ren.setMaximumHeight(30)
        btn_ren.clicked.connect(self.rename_group_asset_selected); gb.addWidget(btn_ren)
        btn_del = QPushButton("删除"); btn_del.setMaximumHeight(30)
        btn_del.setProperty("class", "danger")
        btn_del.clicked.connect(self.delete_group_asset); gb.addWidget(btn_del)
        gl.addLayout(gb)
        self.tabs.addTab(group_tab, "组合素材")

        # 标签页2：常用文字
        text_tab = QWidget()
        tl = QVBoxLayout(text_tab)
        tl.setContentsMargins(4, 4, 4, 4)
        tl.addWidget(QLabel("单击插入到光标处:"))
        self.snippet_list = QListWidget()
        # 设置不抢焦点，避免编辑器失焦
        self.snippet_list.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.snippet_list.itemClicked.connect(self._insert_snippet)
        tl.addWidget(self.snippet_list)
        tb = QHBoxLayout()
        btn_add = QPushButton("添加"); btn_add.setMaximumHeight(30); btn_add.setProperty("class", "primary")
        btn_add.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_add.clicked.connect(self._add_snippet); tb.addWidget(btn_add)
        btn_edit = QPushButton("编辑"); btn_edit.setMaximumHeight(30)
        btn_edit.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_edit.clicked.connect(self._edit_snippet); tb.addWidget(btn_edit)
        btn_sdel = QPushButton("删除"); btn_sdel.setMaximumHeight(30)
        btn_sdel.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_sdel.setProperty("class", "danger")
        btn_sdel.clicked.connect(self._delete_snippet); tb.addWidget(btn_sdel)
        tl.addLayout(tb)
        self.tabs.addTab(text_tab, "常用文字")
        self._refresh_snippets()

    def _get_snippets(self):
        return self.main_window.scene.config_manager.get(
            'text_snippets', ['（一）', '（二）', '第一章', '注：'])

    def _save_snippets(self, snippets):
        self.main_window.scene.config_manager.set('text_snippets', snippets)

    def _refresh_snippets(self):
        self.snippet_list.clear()
        for s in self._get_snippets():
            self.snippet_list.addItem(QListWidgetItem(s))

    def _insert_snippet(self, item):
        text = item.text()
        mw = self.main_window
        # 找正在编辑的 VTextItem
        for scene_item in mw.scene.items():
            if isinstance(scene_item, VTextItem) and scene_item.is_editing:
                editor = scene_item.inline_editor
                if editor and editor.isVisible():
                    editor.insertPlainText(text)
                    return
        mw.status_bar.showMessage("请先双击文字元素进入编辑状态，再单击片段插入", 3000)

    def _add_snippet(self):
        text, ok = QInputDialog.getMultiLineText(self, "添加常用文字", "输入文字片段:")
        if ok and text.strip():
            snippets = self._get_snippets()
            snippets.append(text.strip())
            self._save_snippets(snippets)
            self._refresh_snippets()

    def _edit_snippet(self):
        item = self.snippet_list.currentItem()
        if not item:
            return
        idx = self.snippet_list.row(item)
        text, ok = QInputDialog.getMultiLineText(self, "编辑常用文字", "修改文字片段:", item.text())
        if ok and text.strip():
            snippets = self._get_snippets()
            snippets[idx] = text.strip()
            self._save_snippets(snippets)
            self._refresh_snippets()

    def _delete_snippet(self):
        item = self.snippet_list.currentItem()
        if not item:
            return
        snippets = self._get_snippets()
        snippets.pop(self.snippet_list.row(item))
        self._save_snippets(snippets)
        self._refresh_snippets()

    def _start_group_drag(self, supported_actions):
        item = self.group_list.currentItem()
        if not item:
            return
        asset = item.data(Qt.ItemDataRole.UserRole)
        if not asset:
            return
        mime = QMimeData()
        mime.setData('application/x-group-asset-id', str(asset['id']).encode())
        drag = QDrag(self.group_list)
        drag.setMimeData(mime)
        pixmap = QPixmap(120, 30)
        pixmap.fill(QColor(0, 120, 215, 180))
        painter = QPainter(pixmap)
        painter.setPen(Qt.GlobalColor.white)
        painter.drawText(pixmap.rect(), Qt.AlignmentFlag.AlignCenter, "放置组合素材")
        painter.end()
        drag.setPixmap(pixmap)
        drag.setHotSpot(QPoint(60, 15))
        drag.exec(Qt.DropAction.CopyAction)

    def _on_group_list_mouse_move(self, event):
        """鼠标在组合列表上移动时，延迟显示预览"""
        item = self.group_list.itemAt(event.pos())
        if item:
            asset = item.data(Qt.ItemDataRole.UserRole)
            if asset != self._preview_asset:
                self._preview_asset = asset
                self._preview_pos = self.group_list.mapToGlobal(event.pos())
                self._preview_timer.start(500)  # 悬停 500ms 后显示
        else:
            self._hide_preview()
        QListWidget.mouseMoveEvent(self.group_list, event)

    def _on_group_list_leave(self, event):
        self._hide_preview()
        QListWidget.leaveEvent(self.group_list, event)

    def _show_preview(self):
        if self._preview_asset and self._preview_pos:
            self._preview_popup.show_asset(self._preview_asset, self._preview_pos)

    def _hide_preview(self):
        self._preview_timer.stop()
        self._preview_asset = None
        self._preview_popup.hide()

    def refresh_assets(self):
        self.asset_manager.load_assets()
        self.group_list.clear()
        for asset in self.asset_manager.get_group_assets():
            item = QListWidgetItem(asset['name'])
            item.setData(Qt.ItemDataRole.UserRole, asset)
            text_count = sum(1 for d in asset['items'] if d['type'] == 'VTextItem')
            image_count = sum(1 for d in asset['items'] if d['type'] == 'VImageItem')
            item.setToolTip(f"文字: {text_count}  图片: {image_count}  连接: {len(asset['image_text_connections'])} 条")
            self.group_list.addItem(item)

    def use_group_asset(self, item):
        asset = item.data(Qt.ItemDataRole.UserRole)
        if asset:
            center = self.main_window.view.mapToScene(self.main_window.view.viewport().rect().center())
            self._place_group_at(asset, center)

    def _place_group_at(self, asset, base_pos):
        if not asset or not asset['items']:
            return
        min_x = min(d['scene_pos'][0] for d in asset['items'])
        min_y = min(d['scene_pos'][1] for d in asset['items'])
        new_items = []
        new_items_by_asset_index = [None] * len(asset['items'])
        scene = self.main_window.scene
        sub_commands = []
        for idx, item_data in enumerate(asset['items']):
            new_item = None
            if item_data['type'] == 'VTextItem':
                new_item = VTextItem(item_data['text'], item_data['font_size'], item_data['box_height'])
                new_item.font_family = item_data['font_family']
                new_item.text_color = QColor(item_data['text_color'])
                for k in ('chars_per_column', 'column_spacing', 'character_spacing', 'auto_height', 'manual_line_break', 'layer_eye_color'):
                    if k in item_data:
                        setattr(new_item, k, item_data[k])
                new_item.rebuild()
            elif item_data['type'] == 'VImageItem':
                image_path = item_data.get('path', '')
                if image_path and not os.path.isabs(image_path):
                    image_path = os.path.join(APP_DIR, image_path)
                if image_path and os.path.exists(image_path):
                    new_item = VImageItem(image_path, item_data['width'])
            if new_item:
                off_x = item_data['scene_pos'][0] - min_x
                off_y = item_data['scene_pos'][1] - min_y
                new_item.setPos(base_pos.x() + off_x, base_pos.y() + off_y)
                new_item.setZValue(10)
                cmd = AddItemCommand(scene, new_item)
                cmd.execute()
                sub_commands.append(cmd)
                if item_data.get('connection_point_deleted', False):
                    new_item.delete_connection_point()
                elif 'connection_point_visible' in item_data and new_item.connection_point:
                    new_item.connection_point.setVisible(item_data['connection_point_visible'])
                new_items.append(new_item)
                new_items_by_asset_index[idx] = new_item
        for idx, item_data in enumerate(asset['items']):
            parent_index = item_data.get('parent_index', -1)
            if parent_index != -1 and parent_index < len(new_items_by_asset_index):
                child = new_items_by_asset_index[idx]
                parent = new_items_by_asset_index[parent_index]
                if not child or not parent:
                    continue
                sp = child.scenePos()
                child.setParentItem(parent)
                child.setPos(parent.mapFromScene(sp))
                scene.add_connector(parent, child)
        for img_idx, text_idx in asset['image_text_connections']:
            if img_idx < len(new_items_by_asset_index) and text_idx < len(new_items_by_asset_index):
                item1 = new_items_by_asset_index[img_idx]
                item2 = new_items_by_asset_index[text_idx]
                if not item1 or not item2:
                    continue
                conn_cmd = scene._make_connector_command(item1, item2)
                if conn_cmd:
                    conn_cmd.execute()
                    sub_commands.append(conn_cmd)
        # 最终统一应用连接点可见性/删除状态，防止父子关系建立/连线创建的回调将其覆盖
        for idx, item_data in enumerate(asset['items']):
            cloned = new_items_by_asset_index[idx]
            if cloned:
                if item_data.get('connection_point_deleted', False):
                    cloned.delete_connection_point()
                elif 'connection_point_visible' in item_data and cloned.connection_point:
                    cloned.connection_point.setVisible(item_data['connection_point_visible'])
        if sub_commands:
            scene.undo_stack.push(MacroCommand(scene, sub_commands))
        scene.clearSelection()
        for i in new_items:
            i.setSelected(True)
        scene.update()
        return new_items

    def delete_group_asset(self):
        current_item = self.group_list.currentItem()
        if current_item:
            asset = current_item.data(Qt.ItemDataRole.UserRole)
            reply = QMessageBox.question(self, "确认删除", f"确定要删除组合素材'{asset['name']}' 吗？")
            if reply == QMessageBox.StandardButton.Yes:
                self.asset_manager.remove_group_asset(asset['id'])
                self.refresh_assets()

    def use_group_asset_selected(self):
        current_item = self.group_list.currentItem()
        if current_item:
            self.use_group_asset(current_item)

    def rename_group_asset(self, item):
        asset = item.data(Qt.ItemDataRole.UserRole)
        if not asset:
            return
        new_name, ok = QInputDialog.getText(self, "重命名", "请输入新名称:", text=asset['name'])
        if ok and new_name.strip():
            asset['name'] = new_name.strip()
            for a in self.asset_manager.assets['groups']:
                if a['id'] == asset['id']:
                    a['name'] = new_name.strip()
                    break
            self.asset_manager.save_assets()
            self.refresh_assets()

    def rename_group_asset_selected(self):
        current_item = self.group_list.currentItem()
        if current_item:
            self.rename_group_asset(current_item)

    def _on_group_list_context_menu(self, pos):
        item = self.group_list.itemAt(pos)
        if not item:
            return
        asset = item.data(Qt.ItemDataRole.UserRole)
        menu = QMenu(self)
        menu.addAction("使用").triggered.connect(lambda: self.use_group_asset(item))
        menu.addAction("编辑组合").triggered.connect(lambda: self._edit_group_asset(asset))
        menu.addAction("重命名").triggered.connect(lambda: self.rename_group_asset(item))
        menu.addSeparator()
        menu.addAction("删除").triggered.connect(self.delete_group_asset)
        menu.exec(self.group_list.mapToGlobal(pos))

    def edit_group_asset_selected(self):
        current_item = self.group_list.currentItem()
        if current_item:
            asset = current_item.data(Qt.ItemDataRole.UserRole)
            self._edit_group_asset(asset)

    def _edit_group_asset(self, asset):
        """把组合放到画布上进入编辑模式"""
        if not asset:
            return
        scene = self.main_window.scene
        view = self.main_window.view
        # 放置到视图中心
        center = view.mapToScene(view.viewport().rect().center())
        # 计算偏移使组合居中
        if asset['items']:
            min_x = min(d['scene_pos'][0] for d in asset['items'])
            min_y = min(d['scene_pos'][1] for d in asset['items'])
            max_x = max(d['scene_pos'][0] for d in asset['items'])
            max_y = max(d['scene_pos'][1] for d in asset['items'])
            grp_w = max_x - min_x
            grp_h = max_y - min_y
            base_pos = QPointF(center.x() - grp_w / 2, center.y() - grp_h / 2)
        else:
            base_pos = center

        placed = self._place_group_at(asset, base_pos)

        scene._editing_group_asset_id = asset['id']
        scene._editing_group_items = [i for i in (placed or []) if isinstance(i, (VImageItem, VTextItem))]

        self.main_window.status_bar.showMessage(
            f"正在编辑组合「{asset['name']}」— 编辑完成后右键画布选择「更新到素材库」", 0
        )

class AssetLibraryWidget(QWidget):
    """素材库窗口"""
    def __init__(self, asset_manager, main_window):
        super().__init__()
        self.asset_manager = asset_manager
        self.main_window = main_window
        self.setup_ui()
        self.refresh_assets()
    
    def setup_ui(self):
        """设置界面"""
        self.setWindowTitle("素材库")
        self.setGeometry(100, 100, 400, 600)
        
        layout = QVBoxLayout()
        
        # 标签页
        self.tab_widget = QTabWidget()
        
        # 文字素材标签页
        self.text_tab = QWidget()
        text_layout = QVBoxLayout()
        
        # 文字素材列表
        self.text_list = QListWidget()
        self.text_list.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        self.text_list.itemDoubleClicked.connect(self.use_text_asset)
        text_layout.addWidget(QLabel("文字素材:"))
        text_layout.addWidget(self.text_list)
        
        # 文字操作按钮
        text_btn_layout = QHBoxLayout()
        btn_delete_text = QPushButton("删除选中")
        btn_delete_text.clicked.connect(self.delete_text_asset)
        text_btn_layout.addWidget(btn_delete_text)
        text_layout.addLayout(text_btn_layout)
        
        self.text_tab.setLayout(text_layout)
        self.tab_widget.addTab(self.text_tab, "文字")
        
        # 图片素材标签页
        self.image_tab = QWidget()
        image_layout = QVBoxLayout()
        
        # 图片素材列表
        self.image_list = QListWidget()
        self.image_list.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        self.image_list.setViewMode(QListView.ViewMode.IconMode)
        self.image_list.setIconSize(QSize(80, 80))
        self.image_list.itemDoubleClicked.connect(self.use_image_asset)
        image_layout.addWidget(QLabel("图片素材:"))
        image_layout.addWidget(self.image_list)
        
        # 图片操作按钮
        image_btn_layout = QHBoxLayout()
        btn_delete_image = QPushButton("删除选中")
        btn_delete_image.clicked.connect(self.delete_image_asset)
        image_btn_layout.addWidget(btn_delete_image)
        image_layout.addLayout(image_btn_layout)
        
        self.image_tab.setLayout(image_layout)
        self.tab_widget.addTab(self.image_tab, "图片")
        
        # 组合素材标签页
        self.group_tab = QWidget()
        group_layout = QVBoxLayout()
        
        # 组合素材列表
        self.group_list = QListWidget()
        self.group_list.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        self.group_list.itemDoubleClicked.connect(self.use_group_asset)
        group_layout.addWidget(QLabel("组合素材:"))
        group_layout.addWidget(self.group_list)
        
        # 组合操作按钮
        group_btn_layout = QHBoxLayout()
        btn_delete_group = QPushButton("删除选中")
        btn_delete_group.clicked.connect(self.delete_group_asset)
        group_btn_layout.addWidget(btn_delete_group)
        group_layout.addLayout(group_btn_layout)
        
        self.group_tab.setLayout(group_layout)
        self.tab_widget.addTab(self.group_tab, "组合")
        
        layout.addWidget(self.tab_widget)
        
        # 底部按钮
        btn_layout = QHBoxLayout()
        btn_refresh = QPushButton("刷新")
        btn_refresh.clicked.connect(self.refresh_assets)
        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.close)
        btn_layout.addWidget(btn_refresh)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def refresh_assets(self):
        """刷新素材列表"""
        print("开始刷新素材库显示")
        
        # 重新加载素材数据
        self.asset_manager.load_assets()
        
        # 刷新文字素材
        self.text_list.clear()
        text_assets = self.asset_manager.get_text_assets()
        print(f"加载文字素材: {len(text_assets)} 条")
        for asset in text_assets:
            item = QListWidgetItem(asset['name'])
            item.setData(Qt.ItemDataRole.UserRole, asset)
            item.setToolTip(f"文字: {asset['text'][:50]}...\n字体: {asset['font_family']}\n大小: {asset['font_size']}")
            self.text_list.addItem(item)
        
        # 刷新图片素材
        self.image_list.clear()
        image_assets = self.asset_manager.get_image_assets()
        print(f"加载图片素材: {len(image_assets)} 条")
        for asset in image_assets:
            item = QListWidgetItem(asset['name'])
            item.setData(Qt.ItemDataRole.UserRole, asset)
            
            # 设置缩略图
            if os.path.exists(asset['path']):
                pixmap = QPixmap(asset['path'])
                if not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(80, 80, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                    item.setIcon(QIcon(scaled_pixmap))
            
            item.setToolTip(f"图片: {asset['name']}\n尺寸: {asset['width']}px")
            self.image_list.addItem(item)
        
        # 刷新组合素材
        self.group_list.clear()
        group_assets = self.asset_manager.get_group_assets()
        print(f"加载组合素材: {len(group_assets)} 条")
        for asset in group_assets:
            item = QListWidgetItem(asset['name'])
            item.setData(Qt.ItemDataRole.UserRole, asset)
            
            # 创建详细信息
            details = f"包含 {asset['item_count']} 个元素\n"
            text_count = sum(1 for item_data in asset['items'] if item_data['type'] == 'VTextItem')
            image_count = sum(1 for item_data in asset['items'] if item_data['type'] == 'VImageItem')
            details += f"文字: {text_count} 条 图片: {image_count} 个\n"
            details += f"连接: {len(asset['image_text_connections'])} 条"
            
            item.setToolTip(details)
            self.group_list.addItem(item)
        
        print("素材库刷新完成")
    
    def use_text_asset(self, item):
        """使用文字素材"""
        asset = item.data(Qt.ItemDataRole.UserRole)
        if asset:
            # 创建文字项目
            text_item = VTextItem(
                asset['text'],
                asset['font_size'],
                asset['box_height']
            )
            text_item.font_family = asset['font_family']
            text_item.text_color = QColor(asset['text_color'])
            for key in ('chars_per_column', 'column_spacing', 'character_spacing', 'auto_height', 'manual_line_break'):
                if key in asset:
                    setattr(text_item, key, asset[key])
            text_item.rebuild()
            
            # 添加到画布中央
            center = self.main_window.view.mapToScene(self.main_window.view.viewport().rect().center())
            text_item.setPos(center)
            self.main_window.scene.add_item_with_undo(text_item)
            print(f"已添加文字素材 {asset['name']}")
    
    def use_image_asset(self, item):
        """使用图片素材"""
        asset = item.data(Qt.ItemDataRole.UserRole)
        if asset and os.path.exists(asset['path']):
            # 创建图片项目
            image_item = VImageItem(asset['path'], asset['width'])
            
            # 添加到画布中央
            center = self.main_window.view.mapToScene(self.main_window.view.viewport().rect().center())
            image_item.setPos(center)
            self.main_window.scene.add_item_with_undo(image_item)
            print(f"已添加图片素材 {asset['name']}")
    
    def use_group_asset(self, item):
        """使用组合素材"""
        asset = item.data(Qt.ItemDataRole.UserRole)
        if asset:
            # 获取粘贴位置
            center = self.main_window.view.mapToScene(self.main_window.view.viewport().rect().center())
            
            # 计算所有项目的边界框，用于确定粘贴位置
            if asset['items']:
                min_x = min(item_data['scene_pos'][0] for item_data in asset['items'])
                min_y = min(item_data['scene_pos'][1] for item_data in asset['items'])
            else:
                min_x = min_y = 0
            
            base_x, base_y = center.x(), center.y()
            new_items = []
            
            # 第一步：创建所有项目
            for idx, item_data in enumerate(asset['items']):
                new_item = None
                
                if item_data['type'] == 'VTextItem':
                    new_item = VTextItem(
                        item_data['text'],
                        item_data['font_size'],
                        item_data['box_height']
                    )
                    new_item.font_family = item_data['font_family']
                    new_item.text_color = QColor(item_data['text_color'])
                    
                    # 恢复其他属性
                    if 'chars_per_column' in item_data:
                        new_item.chars_per_column = item_data['chars_per_column']
                    if 'column_spacing' in item_data:
                        new_item.column_spacing = item_data['column_spacing']
                    if 'character_spacing' in item_data:
                        new_item.character_spacing = item_data['character_spacing']
                    if 'auto_height' in item_data:
                        new_item.auto_height = item_data['auto_height']
                    if 'manual_line_break' in item_data:
                        new_item.manual_line_break = item_data['manual_line_break']
                    if 'layer_eye_color' in item_data:
                        new_item.layer_eye_color = item_data['layer_eye_color']
                    
                    new_item.rebuild()
                        
                elif item_data['type'] == 'VImageItem':
                    if os.path.exists(item_data['path']):
                        new_item = VImageItem(item_data['path'], item_data['width'])
                
                if new_item:
                    # 计算相对于原始组合的偏移量，然后应用到新的基准位置
                    offset_x = item_data['scene_pos'][0] - min_x
                    offset_y = item_data['scene_pos'][1] - min_y
                    new_item.setPos(base_x + offset_x, base_y + offset_y)
                    
                    # 使用撤销系统添加元素
                    command = AddItemCommand(self.main_window.scene, new_item)
                    self.main_window.scene.undo_stack.push(command)
                    
                    # 在AddItemCommand执行后，重新设置连接点可见性
                    # 因为AddItemCommand.execute()会使用场景的全局设置覆盖个别设置
                    if item_data.get('connection_point_deleted', False):
                        new_item.delete_connection_point()
                    elif 'connection_point_visible' in item_data and new_item.connection_point:
                        new_item.connection_point.setVisible(item_data['connection_point_visible'])
                    
                    new_items.append(new_item)
            
            # 第二步：恢复父子关系
            for idx, item_data in enumerate(asset['items']):
                if item_data['parent_index'] != -1 and item_data['parent_index'] < len(new_items):
                    child_item = new_items[idx]
                    parent_item = new_items[item_data['parent_index']]
                    
                    # 保存当前场景坐标
                    current_scene_pos = child_item.scenePos()
                    # 设置父子关系
                    child_item.setParentItem(parent_item)
                    # 将场景坐标转换为父级的本地坐标
                    local_pos = parent_item.mapFromScene(current_scene_pos)
                    child_item.setPos(local_pos)
                    
                    # 创建父子连接线
                    self.main_window.scene.add_connector(parent_item, child_item)
            
            # 第三步：恢复图文连接
            for img_idx, text_idx in asset['image_text_connections']:
                if img_idx < len(new_items) and text_idx < len(new_items):
                    img_item = new_items[img_idx]
                    text_item = new_items[text_idx]
                    # 确保是正确的类型
                    if isinstance(img_item, VImageItem) and isinstance(text_item, VTextItem):
                        self.main_window.scene.add_image_text_connector(img_item, text_item)
                    elif isinstance(img_item, VTextItem) and isinstance(text_item, VImageItem):
                        self.main_window.scene.add_image_text_connector(text_item, img_item)
            
            print(f"已添加组合素材 {asset['name']} ({len(new_items)} 个元素)")
    
    def delete_text_asset(self):
        """删除选中的文字素材"""
        current_item = self.text_list.currentItem()
        if current_item:
            asset = current_item.data(Qt.ItemDataRole.UserRole)
            reply = QMessageBox.question(self, "确认删除", f"确定要删除文字素材'{asset['name']}' 吗？")
            if reply == QMessageBox.StandardButton.Yes:
                self.asset_manager.remove_text_asset(asset['id'])
                self.refresh_assets()
    
    def delete_image_asset(self):
        """删除选中的图片素材"""
        current_item = self.image_list.currentItem()
        if current_item:
            asset = current_item.data(Qt.ItemDataRole.UserRole)
            reply = QMessageBox.question(self, "确认删除", f"确定要删除图片素材'{asset['name']}' 吗？")
            if reply == QMessageBox.StandardButton.Yes:
                self.asset_manager.remove_image_asset(asset['id'])
                self.refresh_assets()
    
    def delete_group_asset(self):
        """删除选中的组合素材"""
        current_item = self.group_list.currentItem()
        if current_item:
            asset = current_item.data(Qt.ItemDataRole.UserRole)
            reply = QMessageBox.question(self, "确认删除", f"确定要删除组合素材'{asset['name']}' 吗？")
            if reply == QMessageBox.StandardButton.Yes:
                self.asset_manager.remove_group_asset(asset['id'])
                self.refresh_assets()

class ProjectData:
    """Helper to serialize/deserialize project"""
    @staticmethod
    def scene_to_dict(scene):
        project_data = {
            'version': '2.0',
            'items': [],
            'connectors': [],
            'image_text_connectors': [],
            'free_connection_points': [],
            'scene_rect': [
                scene.sceneRect().x(),
                scene.sceneRect().y(),
                scene.sceneRect().width(),
                scene.sceneRect().height()
            ]
        }
        
        # Store items with IDs to reconstruct hierarchy
        item_map = {}  # item -> id
        
        # Assign IDs
        for idx, item in enumerate(scene.items()):
            if isinstance(item, (VTextItem, VImageItem)):
                item_map[item] = idx
        
        # Save items
        for item, item_id in item_map.items():
            data = {
                'id': item_id,
                'type': item.__class__.__name__,
                'x': item.x(),
                'y': item.y(),
                'scene_x': item.scenePos().x(),
                'scene_y': item.scenePos().y(),
                'parent_id': item_map.get(item.parentItem(), -1) if isinstance(item.parentItem(), (VTextItem, VImageItem)) else -1,
                'z': item.zValue()
            }
            
            if isinstance(item, VTextItem):
                data['text'] = item.full_text
                data['font_size'] = item.font_size
                data['box_height'] = item.box_height
                data['font_family'] = item.font_family
                data['text_color'] = item.text_color.name()
                data['chars_per_column'] = item.chars_per_column
                data['column_spacing'] = item.column_spacing
                data['character_spacing'] = item.character_spacing
                data['auto_height'] = item.auto_height
                data['manual_line_break'] = item.manual_line_break
                data['layer_eye_color'] = getattr(item, 'layer_eye_color', None)
                if item.connection_point:
                    data['connection_point_visible'] = item.connection_point.isVisible()
                else:
                    data['connection_point_deleted'] = True
            elif isinstance(item, VImageItem):
                data['path'] = item.file_path
                data['width'] = item.target_width
                data['opacity'] = item.image_opacity
                data['locked'] = item.locked
                data['visible'] = item.isVisible()
                if item.connection_point:
                    data['connection_point_visible'] = item.connection_point.isVisible()
                else:
                    data['connection_point_deleted'] = True
            data['custom_connection_points'] = [
                [point.pos().x(), point.pos().y()]
                for point in getattr(item, 'custom_connection_points', [])
                if point.scene() is scene
            ]
            
            project_data['items'].append(data)

        project_data['free_connection_points'] = [
            [point.pos().x(), point.pos().y()]
            for point in getattr(scene, 'free_connection_points', [])
            if point.scene() is scene
        ]
        
        # Save parent-child connectors
        for conn in scene.connectors:
            if hasattr(conn, 'parent_element') and hasattr(conn, 'child_element'):
                parent_id = item_map.get(conn.parent_element, -1)
                child_id = item_map.get(conn.child_element, -1)
                if parent_id != -1 and child_id != -1:
                    project_data['connectors'].append({
                        'parent_id': parent_id,
                        'child_id': child_id
                    })
        
        # Save image-text connectors
        for conn in scene.image_text_connectors:
            conn_data = {}
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                # VImageTextConnector
                img_id = item_map.get(conn.image_item, -1)
                text_id = item_map.get(conn.text_item, -1)
                if img_id != -1 and text_id != -1:
                    conn_data = {
                        'type': 'VImageTextConnector',
                        'image_id': img_id,
                        'text_id': text_id,
                        'line_width': conn.line_width if hasattr(conn, 'line_width') else 3
                    }
                    if getattr(conn, 'point1', None) is not None and getattr(conn, 'point1', None).point_type == 'custom':
                        conn_data['point1_custom_index'] = getattr(conn.item1, 'custom_connection_points', []).index(conn.point1)
                    if getattr(conn, 'point2', None) is not None and getattr(conn, 'point2', None).point_type == 'custom':
                        conn_data['point2_custom_index'] = getattr(conn.item2, 'custom_connection_points', []).index(conn.point2)
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                # VGenericConnector
                item1_id = item_map.get(conn.item1, -1)
                item2_id = item_map.get(conn.item2, -1)
                has_point_refs = getattr(conn, 'point1', None) is not None and getattr(conn, 'point2', None) is not None
                if (item1_id != -1 and item2_id != -1) or has_point_refs:
                    conn_data = {
                        'type': 'VGenericConnector',
                        'item1_id': item1_id,
                        'item2_id': item2_id,
                        'connection_type': conn.connection_type if hasattr(conn, 'connection_type') else 'generic',
                        'line_width': conn.line_width if hasattr(conn, 'line_width') else 3
                    }
                    if getattr(conn, 'point1', None) is not None and conn.point1.point_type == 'custom':
                        if conn.point1.parent_element is None:
                            conn_data['free_point1_index'] = scene.free_connection_points.index(conn.point1)
                        else:
                            conn_data['point1_custom_index'] = conn.item1.custom_connection_points.index(conn.point1)
                    elif getattr(conn, 'point1', None) is not None:
                        conn_data['point1_default'] = True
                    if getattr(conn, 'point2', None) is not None and conn.point2.point_type == 'custom':
                        if conn.point2.parent_element is None:
                            conn_data['free_point2_index'] = scene.free_connection_points.index(conn.point2)
                        else:
                            conn_data['point2_custom_index'] = conn.item2.custom_connection_points.index(conn.point2)
                    elif getattr(conn, 'point2', None) is not None:
                        conn_data['point2_default'] = True
            
            if conn_data:
                project_data['image_text_connectors'].append(conn_data)
        
        print(f"工程已保存: {len(project_data['items'])} 个元素, {len(project_data['connectors'])} 个父子连接, {len(project_data['image_text_connectors'])} 个图文连接")
        return project_data

    @staticmethod
    def save(scene, filepath):
        """保存旧版单场景格式，供外部调用兼容。"""
        project_data = ProjectData.scene_to_dict(scene)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(project_data, f, indent=2, ensure_ascii=False)

    @staticmethod
    def save_documents(documents, filepath):
        """保存一个工程文件中的多个文档。"""
        project_data = {
            'version': '3.0',
            'documents': [
                {
                    'name': document.name,
                    'data': ProjectData.scene_to_dict(document.scene)
                }
                for document in documents
            ]
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(project_data, f, indent=2, ensure_ascii=False)
        print(f"多文档工程已保存: {len(project_data['documents'])} 个文档")

    @staticmethod
    def read_documents(filepath):
        """读取多文档工程；旧单文档格式会自动包装成一个文档。"""
        with open(filepath, 'r', encoding='utf-8') as f:
            project_data = json.load(f)

        if isinstance(project_data, dict) and isinstance(project_data.get('documents'), list):
            documents = []
            for index, document in enumerate(project_data['documents']):
                if not isinstance(document, dict):
                    continue
                data = document.get('data', {})
                if not isinstance(data, dict):
                    continue
                documents.append({
                    'name': document.get('name') or f'文档 {index + 1}',
                    'data': data
                })
            return documents or [{'name': '文档 1', 'data': {}}]

        name = os.path.splitext(os.path.basename(filepath))[0] or '文档 1'
        return [{'name': name, 'data': project_data}]

    @staticmethod
    def load(scene, filepath):
        """读取旧版单场景格式，或读取文件中的第一个文档。"""
        documents = ProjectData.read_documents(filepath)
        ProjectData.load_scene(scene, documents[0].get('data', {}))

    @staticmethod
    def load_scene(scene, project_data):
        scene.clear()
        scene.connectors = []
        scene.image_text_connectors = []
        scene.free_connection_points = []
        scene.selection_order = []
        scene.undo_stack.clear()

        scene_rect = project_data.get('scene_rect') if isinstance(project_data, dict) else None
        if isinstance(scene_rect, (list, tuple)) and len(scene_rect) == 4:
            scene.setSceneRect(
                float(scene_rect[0]),
                float(scene_rect[1]),
                float(scene_rect[2]),
                float(scene_rect[3])
            )
        
        # 兼容旧版本格式
        if isinstance(project_data, list):
            data = project_data
            connectors_data = []
            image_text_connectors_data = []
        else:
            data = project_data.get('items', [])
            connectors_data = project_data.get('connectors', [])
            image_text_connectors_data = project_data.get('image_text_connectors', [])
        
        # First pass: Create items
        id_map = {}  # id -> item
        delayed_parents = []
        
        for d in data:
            item = None
            if d['type'] == 'VTextItem':
                item = VTextItem(d['text'], d['font_size'], d['box_height'])
                if 'font_family' in d:
                    item.font_family = d['font_family']
                if 'text_color' in d:
                    item.text_color = QColor(d['text_color'])
                if 'chars_per_column' in d:
                    item.chars_per_column = d['chars_per_column']
                if 'column_spacing' in d:
                    item.column_spacing = d['column_spacing']
                if 'character_spacing' in d:
                    item.character_spacing = d['character_spacing']
                if 'auto_height' in d:
                    item.auto_height = d['auto_height']
                if 'manual_line_break' in d:
                    item.manual_line_break = d['manual_line_break']
                item.layer_eye_color = d.get('layer_eye_color')
                item.rebuild()
            elif d['type'] == 'VImageItem':
                item = VImageItem(d['path'], d['width'])
                if 'opacity' in d:
                    item.set_opacity(d['opacity'])
                if d.get('locked', False):
                    item.set_locked(True)
                if 'visible' in d:
                    item.set_image_visible(d['visible'])
            
            if item:
                # 使用场景坐标（如果有的话）
                if 'scene_x' in d and 'scene_y' in d:
                    item.setPos(d['scene_x'], d['scene_y'])
                else:
                    item.setPos(d['x'], d['y'])
                
                item.setZValue(d.get('z', 0))
                scene.addItem(item)
                
                # 恢复连接点可见性 / 删除状态
                if d.get('connection_point_deleted', False):
                    item.delete_connection_point()
                elif 'connection_point_visible' in d and item.connection_point:
                    item.connection_point.setVisible(d['connection_point_visible'])
                for point_pos in d.get('custom_connection_points', []):
                    if isinstance(point_pos, (list, tuple)) and len(point_pos) == 2:
                        point = item.create_custom_connection_point(item.mapToScene(
                            QPointF(float(point_pos[0]), float(point_pos[1]))
                        ))
                        point.setVisible(scene.show_connection_points)
                
                id_map[d['id']] = item
                if d['parent_id'] != -1:
                    delayed_parents.append((item, d['parent_id']))
        
        # Second pass: Restore hierarchy
        for item, pid in delayed_parents:
            if pid in id_map:
                parent = id_map[pid]
                # 保存当前场景坐标
                curr_scene_pos = item.scenePos()
                # 设置父级
                item.setParentItem(parent)
                # 将场景坐标转换为父级的本地坐标并设置
                local_pos = parent.mapFromScene(curr_scene_pos)
                item.setPos(local_pos)

        for point_pos in project_data.get('free_connection_points', []):
            if isinstance(point_pos, (list, tuple)) and len(point_pos) == 2:
                scene.add_free_connection_point(QPointF(float(point_pos[0]), float(point_pos[1])))
        
        # Third pass: Restore parent-child connectors
        for conn_data in connectors_data:
            parent_id = conn_data.get('parent_id', -1)
            child_id = conn_data.get('child_id', -1)
            if parent_id in id_map and child_id in id_map:
                scene.add_connector(id_map[parent_id], id_map[child_id])
        
        # Fourth pass: Restore image-text connectors
        for conn_data in image_text_connectors_data:
            conn_type = conn_data.get('type', 'VImageTextConnector')
            line_width = conn_data.get('line_width', 3)
            
            if conn_type == 'VImageTextConnector':
                img_id = conn_data.get('image_id', -1)
                text_id = conn_data.get('text_id', -1)
                if img_id in id_map and text_id in id_map:
                    conn = VImageTextConnector(id_map[img_id], id_map[text_id], line_width)
                    scene.addItem(conn)
                    scene.image_text_connectors.append(conn)
                    conn.update_path()
                    scene.sync_connectors_visibility(id_map[img_id])
            elif conn_type == 'VGenericConnector':
                item1_id = conn_data.get('item1_id', -1)
                item2_id = conn_data.get('item2_id', -1)
                connection_type = conn_data.get('connection_type', 'generic')
                item1 = id_map.get(item1_id)
                item2 = id_map.get(item2_id)
                free_points = getattr(scene, 'free_connection_points', [])
                point1 = free_points[conn_data['free_point1_index']] if (
                    'free_point1_index' in conn_data
                    and 0 <= conn_data['free_point1_index'] < len(free_points)
                ) else None
                point2 = free_points[conn_data['free_point2_index']] if (
                    'free_point2_index' in conn_data
                    and 0 <= conn_data['free_point2_index'] < len(free_points)
                ) else None
                if point1 is None and item1 is not None:
                    points1 = getattr(item1, 'custom_connection_points', [])
                    point1 = points1[conn_data['point1_custom_index']] if (
                        'point1_custom_index' in conn_data
                        and 0 <= conn_data['point1_custom_index'] < len(points1)
                    ) else None
                if point2 is None and item2 is not None:
                    points2 = getattr(item2, 'custom_connection_points', [])
                    point2 = points2[conn_data['point2_custom_index']] if (
                        'point2_custom_index' in conn_data
                        and 0 <= conn_data['point2_custom_index'] < len(points2)
                    ) else None
                if point1 is None and conn_data.get('point1_default') and item1 is not None:
                    point1 = getattr(item1, 'connection_point', None)
                if point2 is None and conn_data.get('point2_default') and item2 is not None:
                    point2 = getattr(item2, 'connection_point', None)
                if (item1 is not None and item2 is not None) or (point1 is not None and point2 is not None):
                    conn = VGenericConnector(item1, item2, connection_type, line_width, point1, point2)
                    scene.addItem(conn)
                    scene.image_text_connectors.append(conn)
                    for point in (point1, point2):
                        if point is not None and conn not in point.connected_lines:
                            point.connected_lines.append(conn)
                    conn.update_path()
                    scene.sync_connectors_visibility()
        
        print(f"工程已加载: {len(id_map)} 个元素, {len(connectors_data)} 个父子连接, {len(image_text_connectors_data)} 个图文连接")

# --- Undo/Redo System ---

class UndoCommand:
    """撤销命令基类"""
    def __init__(self, scene):
        self.scene = scene
    
    def execute(self):
        """执行命令"""
        pass
    
    def undo(self):
        """撤销命令"""
        pass

class AddItemCommand(UndoCommand):
    """添加元素命令 - 增加了父子项安全检查"""
    def __init__(self, scene, item):
        super().__init__(scene)
        self.item = item

    def execute(self):
        """执行添加操作：只有当元素既没有父级也不在场景中时才添加"""
        # 如果元素已经有父级了，它会随父级自动进入场景，无需手动 addItem
        if not self.item.scene() and not self.item.parentItem():
            self.scene.addItem(self.item)
            
        if isinstance(self.item, (VTextItem, VImageItem)):
            self.item.set_connection_points_visible(self.scene.show_connection_points)
        if isinstance(self.item, VTextItem):
            self.item.update_hover_tooltip()

    def undo(self):
        """撤销添加操作：增加场景归属判断，防止重复删除报错"""
        # 检查元素是否还在当前场景中（可能已随父级被移除）
        if self.item.scene() == self.scene:
            self.scene.remove_all_connectors_for_item(self.item)
            self.scene.remove_image_text_connectors(self.item)
            self.scene.removeItem(self.item)

class DeleteItemCommand(UndoCommand):
    """删除元素命令"""
    def __init__(self, scene, item):
        super().__init__(scene)
        self.item = item
        self.item_data = None
        self.parent_item = None
        self.child_items = []  # 保存子元素
        self.child_connectors = []  # 保存作为父级的连接器
        self.parent_connector = None  # 保存作为子级的连接器
        self.image_text_connectors = []  # 保存图文连接器
        self.save_item_state()
    
    def execute(self):
        # 保存父子关系和连接器信息
        self.save_relationships()
        
        # 删除与此元素相关的所有连接器
        self.scene.remove_all_connectors_for_item(self.item)
        # 删除与此元素相关的图文连接器
        self.scene.remove_image_text_connectors(self.item)
        # 删除元素本身
        self.scene.removeItem(self.item)
    
    def undo(self):
        # 重新创建元素
        if self.item_data['type'] == 'VTextItem':
            new_item = VTextItem(
                self.item_data['text'],
                self.item_data['font_size'],
                self.item_data['box_height']
            )
            new_item.font_family = self.item_data['font_family']
            new_item.text_color = QColor(self.item_data['text_color'])
            new_item.chars_per_column = self.item_data.get('chars_per_column', new_item.chars_per_column)
            new_item.column_spacing = self.item_data.get('column_spacing', new_item.column_spacing)
            new_item.character_spacing = self.item_data.get('character_spacing', 0)
            new_item.auto_height = self.item_data.get('auto_height', new_item.auto_height)
            new_item.manual_line_break = self.item_data.get('manual_line_break', new_item.manual_line_break)
            new_item.rebuild()
        elif self.item_data['type'] == 'VImageItem':
            new_item = VImageItem(
                self.item_data['path'],
                self.item_data['width']
            )
        
        # 设置位置
        if self.parent_item and self.parent_item.scene():
            # 如果有父级，恢复父子关系
            new_item.setParentItem(self.parent_item)
            new_item.setPos(self.item_data['pos'][0], self.item_data['pos'][1])
            # 重新创建父子连接器
            self.scene.add_connector(self.parent_item, new_item)
        else:
            # 没有父级，直接添加到场景
            new_item.setPos(self.item_data['scene_pos'][0], self.item_data['scene_pos'][1])
            self.scene.addItem(new_item)
        
        # 恢复子元素的父子关系
        for child_item in self.child_items:
            if child_item.scene():
                # 保存子元素当前场景位置
                child_scene_pos = child_item.scenePos()
                # 设置父子关系
                child_item.setParentItem(new_item)
                # 转换为本地坐标
                child_local_pos = new_item.mapFromScene(child_scene_pos)
                child_item.setPos(child_local_pos)
                # 重新创建连接器
                self.scene.add_connector(new_item, child_item)
        
        # 恢复图文连接器（这里需要更复杂的逻辑，暂时跳过）
        
        self.item = new_item
    
    def save_relationships(self):
        """保存父子关系和连接器"""
        # 保存父级关系
        if self.item.parentItem() and isinstance(self.item.parentItem(), BaseElement):
            self.parent_item = self.item.parentItem()
        
        # 保存子元素
        for child in self.item.childItems():
            if isinstance(child, BaseElement):
                self.child_items.append(child)
        
        # 保存相关的连接器
        self.child_connectors = [c for c in self.scene.connectors if c.parent_element == self.item]
        self.parent_connector = next((c for c in self.scene.connectors if c.child_element == self.item), None)
        
        # 保存图文连接器
        self.image_text_connectors = [c for c in self.scene.image_text_connectors 
                                    if (hasattr(c, 'image_item') and (c.image_item == self.item or c.text_item == self.item)) or
                                       (hasattr(c, 'item1') and (c.item1 == self.item or c.item2 == self.item))]
    
    def save_item_state(self):
        """保存元素状态"""
        if isinstance(self.item, VTextItem):
            self.item_data = {
                'type': 'VTextItem',
                'text': self.item.full_text,
                'font_size': self.item.font_size,
                'box_height': self.item.box_height,
                'font_family': self.item.font_family,
                'text_color': self.item.text_color.name(),
                'chars_per_column': self.item.chars_per_column,
                'column_spacing': self.item.column_spacing,
                'character_spacing': self.item.character_spacing,
                'auto_height': self.item.auto_height,
                'manual_line_break': self.item.manual_line_break,
                'pos': (self.item.x(), self.item.y()),
                'scene_pos': (self.item.scenePos().x(), self.item.scenePos().y())
            }
        elif isinstance(self.item, VImageItem):
            self.item_data = {
                'type': 'VImageItem',
                'path': self.item.file_path,
                'width': self.item.target_width,
                'pos': (self.item.x(), self.item.y()),
                'scene_pos': (self.item.scenePos().x(), self.item.scenePos().y())
            }

class SetParentCommand(UndoCommand):
    """设置父子关系命令"""
    def __init__(self, scene, child_item, new_parent, old_parent=None):
        super().__init__(scene)
        self.child_item = child_item
        self.new_parent = new_parent
        self.old_parent = old_parent
        self.old_pos = child_item.pos()
        self.old_scene_pos = child_item.scenePos()
    
    def execute(self):
        # 移除旧的连接器
        if self.old_parent:
            self.scene.remove_child_connectors(self.child_item)
        
        # 保存当前场景坐标
        current_scene_pos = self.child_item.scenePos()
        
        # 设置新的父子关系
        if self.new_parent:
            self.child_item.setParentItem(self.new_parent)
            # 转换为父级的本地坐标
            local_pos = self.new_parent.mapFromScene(current_scene_pos)
            self.child_item.setPos(local_pos)
            # 创建新的连接器
            self.scene.add_connector(self.new_parent, self.child_item)
        else:
            # 移除父子关系
            self.child_item.setParentItem(None)
            self.child_item.setPos(current_scene_pos)
    
    def undo(self):
        # 移除当前连接器
        self.scene.remove_child_connectors(self.child_item)
        
        # 恢复旧的父子关系
        if self.old_parent and self.old_parent.scene():
            self.child_item.setParentItem(self.old_parent)
            self.child_item.setPos(self.old_pos)
            # 重新创建旧的连接器
            self.scene.add_connector(self.old_parent, self.child_item)
        else:
            # 恢复为无父级状态
            self.child_item.setParentItem(None)
            self.child_item.setPos(self.old_scene_pos)

class MoveItemCommand(UndoCommand):
    """移动元素命令"""
    def __init__(self, scene, item, old_pos, new_pos):
        super().__init__(scene)
        self.item = item
        self.old_scene_pos = old_pos
        self.new_scene_pos = new_pos

    def execute(self):
        """redo：移到新位置"""
        if self.item.parentItem():
            self.item.setPos(self.item.parentItem().mapFromScene(self.new_scene_pos))
        else:
            self.item.setPos(self.new_scene_pos)
        self._update_connectors()

    def undo(self):
        # 只需移动父级，子元素作为子项会自动跟随
        if self.item.parentItem():
            local_pos = self.item.parentItem().mapFromScene(self.old_scene_pos)
            self.item.setPos(local_pos)
        else:
            self.item.setPos(self.old_scene_pos)

        # 更新所有相关连线
        if self.item.scene():
            self.item.scene().update_connectors(self.item)
            self.item.scene().update_image_text_connectors(self.item)
            for child in self.item.childItems():
                if isinstance(child, BaseElement):
                    self.item.scene().update_connectors(child)
                    self.item.scene().update_image_text_connectors(child)

    def _update_connectors(self):
        if self.item.scene():
            self.item.scene().update_connectors(self.item)
            self.item.scene().update_image_text_connectors(self.item)
            for child in self.item.childItems():
                if isinstance(child, BaseElement):
                    self.item.scene().update_connectors(child)
                    self.item.scene().update_image_text_connectors(child)

class EditTextCommand(UndoCommand):
    """文字编辑命令"""
    def __init__(self, scene, item, old_text, new_text):
        super().__init__(scene)
        self.item = item
        self.old_text = old_text
        self.new_text = new_text

    def _restore_scene_pos(self, scene_pos):
        if self.item.parentItem():
            self.item.setPos(self.item.parentItem().mapFromScene(scene_pos))
        else:
            self.item.setPos(scene_pos)

    def execute(self):
        scene_pos = self.item.scenePos()
        self.item.full_text = self.new_text
        self.item.rebuild(preserve_position=True)
        self._restore_scene_pos(scene_pos)
        if self.item.scene():
            self.item.scene().update_connectors(self.item)
            self.item.scene().update_image_text_connectors(self.item)

    def undo(self):
        scene_pos = self.item.scenePos()
        self.item.full_text = self.old_text
        self.item.rebuild(preserve_position=True)
        self._restore_scene_pos(scene_pos)
        if self.item.scene():
            self.item.scene().update_connectors(self.item)
            self.item.scene().update_image_text_connectors(self.item)


class AddConnectorCommand(UndoCommand):
    """添加连接线命令（支持父子连接线和图文连接线）"""
    def __init__(self, scene, connector):
        super().__init__(scene)
        self.connector = connector

    def execute(self):
        # 处理父子红线连接器
        if isinstance(self.connector, VConnector):
            if self.connector not in self.scene.connectors:
                self.scene.addItem(self.connector)
                self.scene.connectors.append(self.connector)
            self.connector.update_path()
            self.scene.sync_connectors_visibility()
        # 处理图文/通用连接器
        else:
            if self.connector not in self.scene.image_text_connectors:
                self.scene.addItem(self.connector)
                self.scene.image_text_connectors.append(self.connector)
            self.connector.update_path()
            self.scene.sync_connectors_visibility()

    def undo(self):
        if isinstance(self.connector, VConnector):
            if self.connector in self.scene.connectors:
                self.scene.removeItem(self.connector)
                self.scene.connectors.remove(self.connector)
        else:
            if self.connector in self.scene.image_text_connectors:
                self.scene.removeItem(self.connector)
                self.scene.image_text_connectors.remove(self.connector)


class MacroCommand(UndoCommand):
    """批量命令：将多个命令打包为一次撤销/重做"""
    def __init__(self, scene, commands):
        super().__init__(scene)
        self.commands = commands  # 已经执行过的命令列表

    def execute(self):
        for cmd in self.commands:
            cmd.execute()

    def undo(self):
        for cmd in reversed(self.commands):
            cmd.undo()


class UndoStack:
    """撤销栈管理器"""
    def __init__(self, max_size=50):
        self.commands = []
        self.current_index = -1
        self.max_size = max_size

    def push(self, command):
        """记录已执行的命令（不再调用 execute）"""
        self.commands = self.commands[:self.current_index + 1]
        self.commands.append(command)
        self.current_index += 1
        if len(self.commands) > self.max_size:
            self.commands.pop(0)
            self.current_index -= 1

    def push_and_execute(self, command):
        """执行命令并记录（用于需要立即执行的场景）"""
        command.execute()
        self.push(command)

    def undo(self):
        if self.can_undo():
            self.commands[self.current_index].undo()
            self.current_index -= 1
            return True
        return False

    def redo(self):
        if self.can_redo():
            self.current_index += 1
            self.commands[self.current_index].execute()
            return True
        return False

    def can_undo(self):
        return self.current_index >= 0

    def can_redo(self):
        return self.current_index < len(self.commands) - 1

    def clear(self):
        self.commands.clear()
        self.current_index = -1

# --- Graphics Items ---

class AnchorHandle(QGraphicsRectItem):
    """An anchor point for connectors"""
    def __init__(self, parent, role="bottom"):
        super().__init__(-4, -4, 8, 8, parent)
        self.setBrush(QBrush(QColor("red")))
        self.setPen(Qt.PenStyle.NoPen)
        self.role = role # top, bottom, left, right
        self.setVisible(False) # Show only when needed or strictly mainly logic
        
    def get_scene_pos(self):
        return self.mapToScene(0, 0)

class ConnectionPoint(QGraphicsEllipseItem):
    """可视化连接点：增强版（解决层级遮挡和多线连接问题）"""
    HIT_RADIUS = 2
    BOUNDS_RADIUS = 7
    BASE_RADIUS = 2
    HOVER_RADIUS = 5

    def __init__(self, parent_item, point_type="image_top"):
        super().__init__()
        self.parent_element = parent_item
        self.point_type = point_type
        self._custom_position = point_type == "custom"
        self._press_scene_pos = None
        self.connected_lines = []
        self._hovered = False
        self.base_brush = QBrush(QColor(255, 100, 100, 200))
        self.base_pen = QPen(QColor(200, 50, 50), 3)
        self.hover_brush = QBrush(QColor(0, 255, 120, 240)) # 悬停时的绿色
        self.hover_pen = QPen(QColor(0, 200, 80), 3)

        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, self._custom_position)
        # 画布独立点需要能够被点击选中、拖动和通过右键删除；元素上的固定连接点仍不可选。
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, self._custom_position)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges, True)

        # --- 核心修改：将 ZValue 设为极高，确保点永远在连线之上 ---
        self.setZValue(2500) 
        if parent_item is not None:
            self.setParentItem(parent_item)
        self.update_position()
        self.setAcceptHoverEvents(True)

    def boundingRect(self):
        r = self.BOUNDS_RADIUS
        return QRectF(-r, -r, r * 2, r * 2)

    def shape(self):
        path = QPainterPath()
        r = self.HIT_RADIUS
        path.addEllipse(QRectF(-r, -r, r * 2, r * 2))
        return path

    def paint(self, painter, option, widget):
        lod = option.levelOfDetailFromTransform(painter.worldTransform())
        lod = max(lod, 0.001)

        # 判断当前点是否是选中的“连线起点”
        is_active_source = (self.scene() and getattr(self.scene(), 'connection_source_point', None) == self)

        radius = self.HOVER_RADIUS if (self._hovered or is_active_source) else self.BASE_RADIUS
        min_screen_radius = self.HOVER_RADIUS if (self._hovered or is_active_source) else self.BASE_RADIUS
        if lod < 1.0:
            radius = max(radius, min_screen_radius / lod)

        rect = QRectF(-radius, -radius, radius * 2, radius * 2)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        
        if is_active_source:
            painter.setBrush(QBrush(Qt.GlobalColor.yellow)) # 选中起点变黄
            painter.setPen(QPen(QColor(200, 200, 0), 3))
        else:
            # 正常悬停变绿，平时变红
            painter.setBrush(self.hover_brush if self._hovered else self.base_brush)
            painter.setPen(self.hover_pen if self._hovered else self.base_pen)
            
        painter.drawEllipse(rect)

    def update_position(self):
        if not self.parent_element: return
        rect = self.parent_element.boundingRect()
        if self.point_type == "image_top":
            self.setPos(rect.center().x(), rect.top())
        elif self.point_type == "text_bottom":
            self.setPos(rect.center().x(), rect.bottom())

    def boundingRect(self):
        """扩大重绘范围，避免缩小时视觉被裁切"""
        r = max(self.BOUNDS_RADIUS, 10 if self._custom_position else self.BOUNDS_RADIUS)
        return QRectF(-r, -r, r * 2, r * 2)

    def shape(self):
        """扩大点击/悬停感应区到约 40x40 像素"""
        path = QPainterPath()
        r = 8 if self._custom_position else self.HIT_RADIUS
        path.addEllipse(QRectF(-r, -r, r * 2, r * 2))
        return path

    def paint(self, painter, option, widget):
        """圆心始终保持在(0,0)，缩小时仍保持足够可见"""
        lod = option.levelOfDetailFromTransform(painter.worldTransform())
        lod = max(lod, 0.001)

        radius = self.HOVER_RADIUS if self._hovered else self.BASE_RADIUS
        min_screen_radius = self.HOVER_RADIUS if self._hovered else self.BASE_RADIUS
        if lod < 1.0:
            radius = max(radius, min_screen_radius / lod)

        rect = QRectF(-radius, -radius, radius * 2, radius * 2)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setBrush(self.hover_brush if self._hovered else self.base_brush)
        painter.setPen(self.hover_pen if self._hovered else self.base_pen)
        painter.drawEllipse(rect)

    def update_position(self):
        """更新连接点位置（仅更新自身坐标，连线由父元素的 itemChange 统一更新）"""
        if not self.parent_element:
            return
        if self._custom_position:
            return
        rect = self.parent_element.boundingRect()
        if self.point_type == "image_top":
            self.setPos(rect.center().x(), rect.top())
        elif self.point_type == "text_bottom":
            self.setPos(rect.center().x(), rect.bottom())

    def hoverEnterEvent(self, event):
        """鼠标悬停进入"""
        self._hovered = True
        self.update()
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        """鼠标悬停离开"""
        self._hovered = False
        self.update()
        super().hoverLeaveEvent(event)

    def mousePressEvent(self, event):
        """鼠标按下开始连接"""
        if event.button() == Qt.MouseButton.LeftButton:
            if self._custom_position:
                if self.scene() and getattr(self.scene(), 'free_connection_point_mode', 'move') == 'connect':
                    self.scene().start_connection_from_point(self)
                    event.accept()
                    return
                self._press_scene_pos = event.scenePos()
                self.setSelected(True)
                if self.scene():
                    self.scene().hide_temp_alignment_guide()
                super().mousePressEvent(event)
                return
            self.setSelected(False)
            if self.scene():
                self.scene().start_connection_from_point(self)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self._custom_position and event.button() == Qt.MouseButton.LeftButton:
            super().mouseReleaseEvent(event)
            self._press_scene_pos = None
            if self.scene():
                self.scene().hide_temp_alignment_guide()
            return
        super().mouseReleaseEvent(event)

    def mouseDoubleClickEvent(self, event):
        """独立连接点单击用于选中/拖动，双击才开始或完成连线。"""
        if (self._custom_position and event.button() == Qt.MouseButton.LeftButton
                and self.scene()
                and getattr(self.scene(), 'free_connection_point_mode', 'move') == 'connect'):
            self.setSelected(True)
            self.scene().start_connection_from_point(self)
            event.accept()
            return
        super().mouseDoubleClickEvent(event)

    def contextMenuEvent(self, event):
        if self._custom_position:
            self._show_context_menu(event.screenPos())
            event.accept()
            return
        super().contextMenuEvent(event)

    def _show_context_menu(self, global_pos):
        """显示独立连接点菜单；视图拦截右键时也通过此入口调用。"""
        menu = QMenu()
        delete_action = menu.addAction("删除此连接点")
        if menu.exec(global_pos) != delete_action or self.scene() is None:
            return
        scene = self.scene()
        # 先移除依赖此点的连线，避免删除后留下无法更新的端点引用。
        for connector in list(self.connected_lines):
            scene.remove_connector_item(connector)
        self.connected_lines.clear()
        if self.parent_element is not None:
            points = getattr(self.parent_element, 'custom_connection_points', [])
            if self in points:
                points.remove(self)
        free_points = getattr(scene, 'free_connection_points', [])
        if self in free_points:
            free_points.remove(self)
        scene.removeItem(self)

    def itemChange(self, change, value):
        if (self._custom_position
                and change == QGraphicsItem.GraphicsItemChange.ItemPositionChange
                and self.scene()):
            parent = self.parentItem()
            scene_pos = parent.mapToScene(value) if parent else value
            threshold = self.scene().snap_threshold
            snap_x, snap_y = scene_pos.x(), scene_pos.y()
            best_x = best_y = threshold + 1
            horizontal_point_alignment_y = None
            for guide in self.scene().guides:
                if not guide.isVisible():
                    continue
                if guide.orientation == Qt.Orientation.Vertical:
                    delta = guide.pos_value - snap_x
                    if abs(delta) < abs(best_x):
                        best_x = delta
                else:
                    delta = guide.pos_value - snap_y
                    if abs(delta) < abs(best_y):
                        best_y = delta
            for item in self.scene().items():
                if not isinstance(item, ConnectionPoint) or item is self or not item.isVisible():
                    continue
                target = item.get_scene_center()
                dx, dy = target.x() - snap_x, target.y() - snap_y
                if abs(dx) < abs(best_x):
                    best_x = dx
                if abs(dy) < abs(best_y):
                    best_y = dy
                    horizontal_point_alignment_y = target.y()
            if abs(best_x) <= threshold:
                snap_x += best_x
            if abs(best_y) <= threshold:
                snap_y += best_y

            # 仅在独立连接点与另一个连接点水平吸附时显示临时蓝线。
            # 对齐到永久辅助线时，该辅助线本身已提供可见反馈，无需重复显示。
            if horizontal_point_alignment_y is not None and abs(best_y) <= threshold:
                self.scene().show_temp_alignment_guide(
                    Qt.Orientation.Horizontal, horizontal_point_alignment_y
                )
            else:
                self.scene().hide_temp_alignment_guide()
            scene_pos = QPointF(snap_x, snap_y)
            value = parent.mapFromScene(scene_pos) if parent else scene_pos
        elif (self._custom_position
              and change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged):
            for connector in self.connected_lines:
                if connector.scene():
                    connector.update_path()
        return super().itemChange(change, value)

    def get_scene_center(self):
        """获取连接点在场景中的中心位置"""
        return self.mapToScene(0, 0)

class VGenericConnector(QGraphicsPathItem):
    """通用连接线 - 支持任意两个元素之间的连接"""
    def __init__(self, item1, item2, connection_type="generic", line_width=None,
                 point1=None, point2=None):
        super().__init__()
        self.item1 = item1
        self.item2 = item2
        self.point1 = point1
        self.point2 = point2
        self.connection_type = connection_type  # "image-image", "text-text", "generic"
        self.line_width = line_width if line_width is not None else DEFAULT_LINE_WIDTH  # 线条粗细
        self.base_color = QColor(255, 0, 0, 200)  # 基础颜色
        self.setZValue(50)  # 显示在图片前面
        
        # 统一使用红色连接线
        pen = QPen(self.base_color)
        pen.setWidth(self.line_width)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, True)  # 可选中
        self.setAcceptHoverEvents(True)  # 接受悬停事件

    # --- 新增：重写 shape 方法以扩大点击区域 ---
    def shape(self):
        """
        核心修改：返回一个加宽后的路径用于碰撞检测。
        视觉上保持 line_width 宽度，但点击判定区域扩大到 20 像素。
        """
        path = self.path()
        if path.isEmpty():
            return super().shape()
            
        stroker = QPainterPathStroker()
        stroker.setWidth(20)  # 点击感应宽度设置为 20 像素
        stroker.setCapStyle(Qt.PenCapStyle.RoundCap)
        stroker.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
        return stroker.createStroke(path)
    # ----------------------------------------
    
    def set_line_width(self, width):
        """设置线条粗细"""
        self.line_width = width
        pen = self.pen()
        pen.setWidth(width)
        self.setPen(pen)
        self.update()
    
    def hoverEnterEvent(self, event):
        """鼠标悬停时高亮"""
        pen = QPen(QColor(255, 150, 0, 255))  # 橙色高亮
        pen.setWidth(self.line_width + 1)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        super().hoverEnterEvent(event)
    
    def hoverLeaveEvent(self, event):
        """鼠标离开时恢复"""
        pen = QPen(self.base_color)
        pen.setWidth(self.line_width)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        super().hoverLeaveEvent(event)
    
    def paint(self, painter, option, widget):
        """绘制连接线，选中时显示高亮"""
        # 如果设置了背景图片在连线之上，则在 drawBackground 中手动绘制，此处跳过
        if self.scene() and self.scene().config_manager.get('bg_above_connectors', False):
            return
            
        if self.isSelected():
            # 选中时使用橙色粗线
            pen = QPen(QColor(255, 140, 0), self.line_width + 2, Qt.PenStyle.SolidLine)
            self.setPen(pen)
        else:
            # 连线被取消选中时，强制恢复到基础画笔颜色
            pen = QPen(self.base_color)
            pen.setWidth(self.line_width)
            pen.setStyle(Qt.PenStyle.SolidLine)
            self.setPen(pen)
        super().paint(painter, option, widget)
    
    def contextMenuEvent(self, event):
        """右键菜单"""
        menu = QMenu()
        
        # 线条粗细子菜单
        width_menu = menu.addMenu("线条粗细")
        width_1 = width_menu.addAction("细 (1px)")
        width_2 = width_menu.addAction("较细 (2px)")
        width_3 = width_menu.addAction("正常 (3px)")
        width_4 = width_menu.addAction("较粗 (4px)")
        width_5 = width_menu.addAction("粗 (5px)")
        width_8 = width_menu.addAction("很粗 (8px)")
        width_custom = width_menu.addAction("自定义...")
        
        # 当前粗细标记
        width_actions = {1: width_1, 2: width_2, 3: width_3, 4: width_4, 5: width_5, 8: width_8}
        if self.line_width in width_actions:
            width_actions[self.line_width].setCheckable(True)
            width_actions[self.line_width].setChecked(True)
        
        menu.addSeparator()
        delete_action = menu.addAction("删除连接线")
        
        action = menu.exec(event.screenPos())
        
        if action == width_1:
            self.set_line_width(1)
        elif action == width_2:
            self.set_line_width(2)
        elif action == width_3:
            self.set_line_width(3)
        elif action == width_4:
            self.set_line_width(4)
        elif action == width_5:
            self.set_line_width(5)
        elif action == width_8:
            self.set_line_width(8)
        elif action == width_custom:
            width, ok = QInputDialog.getInt(None, "自定义线条粗细", "线条粗细 (像素):", self.line_width, 1, 20)
            if ok:
                self.set_line_width(width)
        elif action == delete_action:
            if self.scene():
                self.scene().remove_connector_item(self)
        
    def update_path(self):
        if self.point1 is not None or self.point2 is not None:
            if (self.point1 is None or self.point2 is None
                    or self.point1.scene() is None or self.point2.scene() is None):
                return
            anchor1 = self.point1.get_scene_center()
            anchor2 = self.point2.get_scene_center()
        else:
            if not self.item1 or not self.item2 or not self.item1.scene() or not self.item2.scene():
                return

            # 获取两个元素的连接点
            point1 = self.get_connection_point(self.item1)
            point2 = self.get_connection_point(self.item2)

            if not point1 or not point2:
                rect1 = self.item1.boundingRect()
                pos1 = self.item1.scenePos()
                anchor1 = pos1 + QPointF(rect1.width()/2, rect1.height()/2)
                rect2 = self.item2.boundingRect()
                pos2 = self.item2.scenePos()
                anchor2 = pos2 + QPointF(rect2.width()/2, rect2.height()/2)
            else:
                anchor1 = point1.get_scene_center()
                anchor2 = point2.get_scene_center()
        
        path = QPainterPath()
        path.moveTo(anchor1)
        
        # 计算控制点，创建优美的曲线
        distance = (anchor2 - anchor1).manhattanLength()
        curve_offset = min(distance * 0.3, 80)
        
        # 根据相对位置调整控制点
        dx = anchor2.x() - anchor1.x()
        dy = anchor2.y() - anchor1.y()
        
        if abs(dx) > abs(dy):  # 水平方向为主
            ctrl1 = anchor1 + QPointF(curve_offset if dx > 0 else -curve_offset, 0)
            ctrl2 = anchor2 - QPointF(curve_offset if dx > 0 else -curve_offset, 0)
        else:  # 垂直方向为主
            ctrl1 = anchor1 + QPointF(0, curve_offset if dy > 0 else -curve_offset)
            ctrl2 = anchor2 - QPointF(0, curve_offset if dy > 0 else -curve_offset)
        
        path.cubicTo(ctrl1, ctrl2, anchor2)
        self.setPath(path)
    
    def get_connection_point(self, item):
        """获取元素的连接点"""
        # 优先使用属性
        if hasattr(item, 'connection_point') and item.connection_point:
            return item.connection_point
            
        # 备选：从子项目中查找
        for child in item.childItems():
            if isinstance(child, ConnectionPoint):
                return child
        return None

class VImageTextConnector(QGraphicsPathItem):
    """图文连接线- 连接图片顶部中点和文字底部中点"""
    def __init__(self, image_item, text_item, line_width=None):
        super().__init__()
        self.image_item = image_item
        self.text_item = text_item
        self.line_width = line_width if line_width is not None else DEFAULT_LINE_WIDTH  # 线条粗细
        self.base_color = QColor(255, 100, 100, 200)  # 基础颜色
        self.setZValue(55)  # 显示在图片前面
        
        pen = QPen(self.base_color)
        pen.setWidth(self.line_width)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, True)  # 可选中
        self.setAcceptHoverEvents(True)  # 接受悬停事件

    # --- 新增：重写 shape 方法以扩大点击区域 ---
    def shape(self):
        """
        核心修改：利用 stroker 将‘线’扩展为‘面’进行拾取判定。
        即使在画布缩小时，也能轻松选中连线。
        """
        path = self.path()
        if path.isEmpty():
            return super().shape()
            
        stroker = QPainterPathStroker()
        stroker.setWidth(20)  # 点击感应宽度设置为 20 像素
        return stroker.createStroke(path)
    # ----------------------------------------
    
    def set_line_width(self, width):
        """设置线条粗细"""
        self.line_width = width
        pen = self.pen()
        pen.setWidth(width)
        self.setPen(pen)
        self.update()
    
    def hoverEnterEvent(self, event):
        """鼠标悬停时高亮"""
        pen = QPen(QColor(255, 150, 0, 255))  # 橙色高亮
        pen.setWidth(self.line_width + 1)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        super().hoverEnterEvent(event)
    
    def hoverLeaveEvent(self, event):
        """鼠标离开时恢复"""
        pen = QPen(self.base_color)
        pen.setWidth(self.line_width)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        super().hoverLeaveEvent(event)
    
    def paint(self, painter, option, widget):
        """绘制连接线，选中时显示高亮"""
        # 如果设置了背景图片在连线之上，则在 drawBackground 中手动绘制，此处跳过
        if self.scene() and self.scene().config_manager.get('bg_above_connectors', False):
            return
            
        if self.isSelected():
            # 选中时使用橙色粗线
            pen = QPen(QColor(255, 140, 0), self.line_width + 2, Qt.PenStyle.SolidLine)
            self.setPen(pen)
        else:
            # 连线被取消选中时，强制恢复到基础画笔颜色
            pen = QPen(self.base_color)
            pen.setWidth(self.line_width)
            pen.setStyle(Qt.PenStyle.SolidLine)
            self.setPen(pen)
        super().paint(painter, option, widget)
        
    def contextMenuEvent(self, event):
        """右键菜单"""
        menu = QMenu()
        
        # 线条粗细子菜单
        width_menu = menu.addMenu("线条粗细")
        width_1 = width_menu.addAction("细 (1px)")
        width_2 = width_menu.addAction("较细 (2px)")
        width_3 = width_menu.addAction("正常 (3px)")
        width_4 = width_menu.addAction("较粗 (4px)")
        width_5 = width_menu.addAction("粗 (5px)")
        width_8 = width_menu.addAction("很粗 (8px)")
        width_custom = width_menu.addAction("自定义...")
        
        # 当前粗细标记
        width_actions = {1: width_1, 2: width_2, 3: width_3, 4: width_4, 5: width_5, 8: width_8}
        if self.line_width in width_actions:
            width_actions[self.line_width].setCheckable(True)
            width_actions[self.line_width].setChecked(True)
        
        menu.addSeparator()
        delete_action = menu.addAction("删除连接线")
        
        action = menu.exec(event.screenPos())
        
        if action == width_1:
            self.set_line_width(1)
        elif action == width_2:
            self.set_line_width(2)
        elif action == width_3:
            self.set_line_width(3)
        elif action == width_4:
            self.set_line_width(4)
        elif action == width_5:
            self.set_line_width(5)
        elif action == width_8:
            self.set_line_width(8)
        elif action == width_custom:
            width, ok = QInputDialog.getInt(None, "自定义线条粗细", "线条粗细 (像素):", self.line_width, 1, 20)
            if ok:
                self.set_line_width(width)
        elif action == delete_action:
            if self.scene():
                self.scene().remove_connector_item(self)
        
    def update_path(self):
        if not self.image_item.scene() or not self.text_item.scene():
            return

        # 从连接点获取位置
        img_point = getattr(self.image_item, 'connection_point', None)
        text_point = getattr(self.text_item, 'connection_point', None)
        
        if not img_point or not text_point:
            # 如果没有找到属性，尝试从子项目中查找
            if not img_point:
                for child in self.image_item.childItems():
                    if isinstance(child, ConnectionPoint) and child.point_type == "image_top":
                        img_point = child
                        break
            
            if not text_point:
                for child in self.text_item.childItems():
                    if isinstance(child, ConnectionPoint) and child.point_type == "text_bottom":
                        text_point = child
                        break
        
        if not img_point or not text_point:
            # 如果没有连接点，使用原来的计算方条
            img_rect = self.image_item.boundingRect()
            img_pos = self.image_item.scenePos()
            img_anchor = img_pos + QPointF(img_rect.width()/2, 0)
            
            text_rect = self.text_item.boundingRect()
            text_pos = self.text_item.scenePos()
            text_anchor = text_pos + QPointF(text_rect.width()/2, text_rect.height())
        else:
            # 使用连接点的位置
            img_anchor = img_point.get_scene_center()
            text_anchor = text_point.get_scene_center()
        
        path = QPainterPath()
        path.moveTo(img_anchor)
        
        # 计算控制点，创建优美的曲条
        distance = (text_anchor - img_anchor).manhattanLength()
        curve_offset = min(distance * 0.3, 100)  # 曲线弯曲程度
        
        # 根据相对位置调整控制条
        if text_anchor.y() > img_anchor.y():  # 文字在图片下条
            ctrl1 = img_anchor + QPointF(0, curve_offset)
            ctrl2 = text_anchor - QPointF(0, curve_offset)
        else:  # 文字在图片上条
            ctrl1 = img_anchor - QPointF(0, curve_offset)
            ctrl2 = text_anchor + QPointF(0, curve_offset)
        
        path.cubicTo(ctrl1, ctrl2, text_anchor)
        self.setPath(path)

class VConnector(QGraphicsPathItem):
    """Dynamic Red Line Connector"""
    def __init__(self, parent_item, child_item):
        super().__init__()
        self.parent_element = parent_item
        self.child_element = child_item
        self.setZValue(45)  # 显示在图片前面
        
        pen = QPen(QColor(255, 0, 0, 150))
        pen.setWidth(3)  # 更粗的线条
        pen.setStyle(Qt.PenStyle.DashLine)
        self.setPen(pen)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
        
    def update_path(self):
        if not self.parent_element.scene() or not self.child_element.scene():
            return

        # Simple logic: closest points roughly
        # Or specifically: Parent Bottom -> Child Top
        p_anchor = self.parent_element.scenePos() + QPointF(self.parent_element.boundingRect().width()/2, self.parent_element.boundingRect().height())
        c_anchor = self.child_element.scenePos() + QPointF(self.child_element.boundingRect().width()/2, 0)
        
        if isinstance(self.parent_element, VTextItem):
            # Text grows down/left. Anchor at bottom of last col? Or center?
            # Let's use Bottom-Center of bounding rect for now
            pass
            
        path = QPainterPath()
        path.moveTo(p_anchor)
        # Bezier curve for smooth look
        ctrl1 = p_anchor + QPointF(0, 50)
        ctrl2 = c_anchor - QPointF(0, 50)
        path.cubicTo(ctrl1, ctrl2, c_anchor)
        self.setPath(path)

    def paint(self, painter, option, widget):
        # 如果设置了背景图片在连线之上，则在 drawBackground 中手动绘制，此处跳过
        if self.scene() and self.scene().config_manager.get('bg_above_connectors', False):
            return
        super().paint(painter, option, widget)

class BaseElement(QGraphicsItem):
    """Common base for Text and Image elements"""
    def __init__(self):
        super().__init__()
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges)
        self.connectors = []
        self.custom_connection_points = []
        self._drag_start_pos_scene = QPointF() # 记录拖动开始时的场景位置 

    def itemChange(self, change, value):
        if change == QGraphicsItem.GraphicsItemChange.ItemPositionChange:
            # 吸附到辅助线
            if self.scene() and self.scene().guides:
                value = self._snap_to_guides(value)
            
            # 水平移动模式：锁定Y + 右边缘X吸附
            horizontal_lock = getattr(self.scene(), '_horizontal_move_lock_y', {}) if self.scene() else {}
            if self in horizontal_lock:
                parent = self.parentItem()
                new_scene_pos = parent.mapToScene(value) if parent else value
                snap_dx = self._horizontal_image_snap_offset(new_scene_pos, horizontal_lock)
                locked_scene_pos = QPointF(
                    new_scene_pos.x() + snap_dx,
                    horizontal_lock[self]
                )
                value = parent.mapFromScene(locked_scene_pos) if parent else locked_scene_pos
            else:
                # 非水平移动模式：顶部Y吸附（如果开启）
                if isinstance(self, VImageItem):
                    parent = self.parentItem()
                    new_scene_pos = parent.mapToScene(value) if parent else value
                    snap_dy = self._calc_top_edge_snap(new_scene_pos)
                    if abs(snap_dy) > 0.01:
                        snapped_scene_pos = QPointF(new_scene_pos.x(), new_scene_pos.y() + snap_dy)
                        value = parent.mapFromScene(snapped_scene_pos) if parent else snapped_scene_pos

        if change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged:
            # 先更新连接点位置，再更新连线，确保连线使用最新的连接点坐标
            if hasattr(self, 'connection_point') and self.connection_point:
                self.connection_point.update_position()
            if self.scene():
                self.scene().update_connectors(self)
                self.scene().update_image_text_connectors(self)
                self._update_children_connectors()

        return super().itemChange(change, value)

    def _horizontal_image_snap_offset(self, new_scene_pos, horizontal_lock):
        """Align the moving image's right edge X with the nearest image."""
        scene = self.scene()
        def set_indicator(indicator):
            if scene and getattr(scene, '_image_right_edge_snap_indicator', None) != indicator:
                scene._image_right_edge_snap_indicator = indicator
                scene.update()

        owner = getattr(scene, '_horizontal_move_lock_owner', None) if scene else None
        lock_x = getattr(scene, '_horizontal_move_lock_x', {}) if scene else {}
        owner_right_x = getattr(scene, '_horizontal_move_owner_right_x', None) if scene else None
        if (not scene
                or not scene.config_manager.get('image_right_edge_snap_enabled', False)
                or not isinstance(owner, VImageItem)
                or owner not in horizontal_lock
                or self not in lock_x
                or owner_right_x is None):
            set_indicator(None)
            return 0.0

        moving_items = {
            item for item in scene.selectedItems()
            if isinstance(item, BaseElement) and item in horizontal_lock
        }
        moving_items.add(owner)

        def belongs_to_moving_group(candidate):
            for moving_item in moving_items:
                current = candidate
                while current is not None:
                    if current is moving_item:
                        return True
                    current = current.parentItem()

                current = moving_item.parentItem()
                while current is not None:
                    if current is candidate:
                        return True
                    current = current.parentItem()
            return False

        drag_dx = new_scene_pos.x() - lock_x[self]
        moving_right_x = owner_right_x + drag_dx
        threshold = scene.snap_threshold
        best_offset = None
        best_candidate = None
        for candidate in scene.items():
            if (not isinstance(candidate, VImageItem)
                    or not candidate.isVisible()
                    or getattr(candidate, '_was_hidden', False)
                    or belongs_to_moving_group(candidate)):
                continue
            candidate_right_x = candidate.sceneBoundingRect().right()
            offset = candidate_right_x - moving_right_x
            if abs(offset) <= threshold and (
                    best_offset is None or abs(offset) < abs(best_offset)):
                best_offset = offset
                best_candidate = candidate

        if best_candidate is None:
            set_indicator(None)
            return 0.0

        owner_rect = owner.sceneBoundingRect()
        candidate_rect = best_candidate.sceneBoundingRect()
        aligned_x = candidate_rect.right()
        set_indicator((
            aligned_x,
            min(owner_rect.top(), candidate_rect.top()),
            max(owner_rect.bottom(), candidate_rect.bottom()),
            owner_rect.top(),
            candidate_rect.top(),
        ))

        return best_offset

    def _calc_top_edge_snap(self, new_scene_pos):
        """计算顶部Y吸附偏移量，行为与右边缘X吸附保持一致。"""
        scene = self.scene()
        def set_indicator(indicator):
            if scene and getattr(scene, '_image_top_edge_snap_indicator', None) != indicator:
                scene._image_top_edge_snap_indicator = indicator
                scene.update()

        if (not scene
                or getattr(scene, '_batch_importing', False)
                or not scene.config_manager.get('image_top_edge_snap_enabled', False)
                or not isinstance(self, VImageItem)):
            set_indicator(None)
            return 0.0

        # 获取所有正在移动的元素
        moving_items = {
            item for item in scene.selectedItems()
            if isinstance(item, BaseElement)
        }
        if self not in moving_items:
            moving_items.add(self)

        def belongs_to_moving_group(candidate):
            for moving_item in moving_items:
                current = candidate
                while current is not None:
                    if current is moving_item:
                        return True
                    current = current.parentItem()
                current = moving_item.parentItem()
                while current is not None:
                    if current is candidate:
                        return True
                    current = current.parentItem()
            return False

        # 与右边缘X吸附一样，按场景包围盒边缘计算，兼容父子元素和非零局部边界。
        current_scene_pos = self.scenePos()
        moving_top_y = (
            self.sceneBoundingRect().top()
            + new_scene_pos.y() - current_scene_pos.y()
        )
        threshold = scene.snap_threshold
        best_offset = None
        best_candidate = None

        for candidate in scene.items():
            if (not isinstance(candidate, VImageItem)
                    or not candidate.isVisible()
                    or getattr(candidate, '_was_hidden', False)
                    or belongs_to_moving_group(candidate)):
                continue
            candidate_top_y = candidate.sceneBoundingRect().top()
            offset = candidate_top_y - moving_top_y
            if abs(offset) <= threshold and (
                    best_offset is None or abs(offset) < abs(best_offset)):
                best_offset = offset
                best_candidate = candidate

        if best_candidate is None:
            set_indicator(None)
            return 0.0

        owner_rect = self.sceneBoundingRect()
        candidate_rect = best_candidate.sceneBoundingRect()
        aligned_y = candidate_rect.top()
        set_indicator((
            aligned_y,
            min(owner_rect.left(), candidate_rect.left()),
            max(owner_rect.right(), candidate_rect.right()),
            owner_rect.left(),
            candidate_rect.left(),
        ))

        return best_offset


    def _snap_to_guides(self, new_pos, threshold=None):
        """将位置吸附到最近的永久辅助线。"""
        scene = self.scene()
        if threshold is None:
            threshold = scene.snap_threshold
        rect = self.boundingRect()

        # 如果是子级元素，new_pos 是父级局部坐标，需转为场景坐标
        parent = self.parentItem()
        if parent:
            scene_pos = parent.mapToScene(new_pos)
        else:
            scene_pos = new_pos

        x, y = scene_pos.x(), scene_pos.y()

        x_edges = [x, x + rect.width(), x + rect.width() / 2]
        y_edges = [y, y + rect.height(), y + rect.height() / 2]

        # 加入连接点的场景坐标作为额外吸附候选点
        if hasattr(self, 'connection_point') and self.connection_point and self.connection_point.isVisible():
            current_scene_pos = self.scenePos()
            cp_scene = scene_pos + (self.connection_point.get_scene_center() - current_scene_pos)
            x_edges.append(cp_scene.x())
            y_edges.append(cp_scene.y())

        best_dx, best_dy = threshold + 1, threshold + 1
        best_guide_x, best_guide_y = None, None

        # 检查永久辅助线
        for guide in scene.guides:
            if not guide.isVisible():
                continue
            if guide.orientation == Qt.Orientation.Vertical:
                gx = guide.pos_value
                for ex in x_edges:
                    d = abs(ex - gx)
                    if d < abs(best_dx):
                        best_dx = gx - ex
                        best_guide_x = gx
            else:
                gy = guide.pos_value
                for ey in y_edges:
                    d = abs(ey - gy)
                    if d < abs(best_dy):
                        best_dy = gy - ey
                        best_guide_y = gy

        # 画布存在永久辅助线时，只使用永久辅助线吸附，不显示元素间临时水平线。
        scene.hide_temp_alignment_guide()

        if abs(best_dx) <= threshold:
            x += best_dx
        if abs(best_dy) <= threshold:
            y += best_dy

        snapped_scene = QPointF(x, y)
        # 转回父级局部坐标
        if parent:
            return parent.mapFromScene(snapped_scene)
        return snapped_scene

    def _update_children_connectors(self):
        """递归更新所有子元素的连线"""
        for child in self.childItems():
            if isinstance(child, BaseElement):
                if hasattr(child, 'connection_point') and child.connection_point:
                    child.connection_point.update_position()
                if self.scene():
                    self.scene().update_connectors(child)
                    self.scene().update_image_text_connectors(child)
                child._update_children_connectors()

    def mouseReleaseEvent(self, event):
        """记录移动命令到撤销栈"""
        if event.button() == Qt.MouseButton.LeftButton:
            scene = self.scene()
            if scene is None:
                super().mouseReleaseEvent(event)
                return
            
            # 隐藏临时对齐辅助线
            scene.hide_temp_alignment_guide()
            scene._image_right_edge_snap_indicator = None
            scene._image_top_edge_snap_indicator = None
            scene.update()
            
            current_pos_scene = self.scenePos()
            if (current_pos_scene - self._drag_start_pos_scene).manhattanLength() > 2.0:
                command = MoveItemCommand(scene, self, self._drag_start_pos_scene, current_pos_scene)
                scene.undo_stack.push(command)
            if getattr(scene, '_horizontal_move_lock_owner', None) is self:
                scene._horizontal_move_lock_y = {}
                scene._horizontal_move_lock_x = {}
                scene._horizontal_move_lock_owner = None
                scene._horizontal_move_owner_right_x = None

        super().mouseReleaseEvent(event)

    def mousePressEvent(self, event):
        """记录拖动起始位置"""
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_start_pos_scene = self.scenePos()
            scene = self.scene()
            if scene:
                scene._image_right_edge_snap_indicator = None
                scene._image_top_edge_snap_indicator = None
                scene.update()
            if scene and scene.config_manager.get('horizontal_move_only', False):
                selected_items = [
                    item for item in scene.selectedItems()
                    if isinstance(item, BaseElement)
                ]
                if self not in selected_items:
                    selected_items.append(self)
                scene._horizontal_move_lock_y = {
                    item: item.scenePos().y()
                    for item in selected_items
                }
                scene._horizontal_move_lock_x = {
                    item: item.scenePos().x()
                    for item in selected_items
                }
                scene._horizontal_move_lock_owner = self
                scene._horizontal_move_owner_right_x = self.sceneBoundingRect().right()
                scene._image_right_edge_snap_indicator = None
        super().mousePressEvent(event)

    
    def _show_context_menu(self, global_pos):
        """供外部直接调用的右键菜单入口"""
        self._build_base_context_menu(global_pos)

    def create_custom_connection_point(self, scene_pos=None):
        """在元素上创建一个可独立拖动的连接点。"""
        if scene_pos is None:
            scene_pos = self.mapToScene(self.boundingRect().center())
        local_pos = self.mapFromScene(scene_pos)
        point = ConnectionPoint(self, "custom")
        point.setPos(local_pos)
        point.setVisible(self.scene().show_connection_points if self.scene() else True)
        self.custom_connection_points.append(point)
        if self.scene():
            self.scene().update()
        return point

    def _build_base_context_menu(self, global_pos):
        menu = QMenu()

        # 添加隐藏/显示/删除连接点选项
        if hasattr(self, 'connection_point') and self.connection_point:
            if self.connection_point.isVisible():
                toggle_connection_point_action = menu.addAction("隐藏连接点")
            else:
                toggle_connection_point_action = menu.addAction("显示连接点")
            delete_connection_point_action = menu.addAction("删除连接点（永久）")
            menu.addSeparator()
        else:
            toggle_connection_point_action = None
            delete_connection_point_action = None
        
        # 复制和删除
        copy_action = menu.addAction("复制 (Copy)")
        delete_action = menu.addAction("删除 (Delete)")
        save_as_asset_action = menu.addAction("保存组合")
        menu.addSeparator()
        
        # 对齐功能
        selected_items = [item for item in self.scene().selectedItems() if isinstance(item, BaseElement)]
        if len(selected_items) >= 2:
            align_menu = menu.addMenu("对齐 (Align)")
            align_top_action = align_menu.addAction("顶部对齐")
            align_right_action = align_menu.addAction("右对齐")
            align_center_h_action = align_menu.addAction("水平居中对齐")
            align_center_v_action = align_menu.addAction("垂直居中对齐")
            menu.addSeparator()
        
        unbind_action = menu.addAction("解除父级绑定 (Unbind)")
        set_parent_action = menu.addAction("设置父级 (Set Parent)")
        
        # 图文连接选项
        menu.addSeparator()
        connect_image_text_action = menu.addAction("图文连接 (Connect to Image/Text)")
        disconnect_image_text_action = menu.addAction("断开图文连接 (Disconnect Image/Text)")
        
        # 批量图文连接选项
        if len(selected_items) >= 2:
            batch_menu = menu.addMenu("批量连接 (Batch Connect)")
            chain_connect_action = batch_menu.addAction("批量连线")
            auto_connect_action = batch_menu.addAction("智能连接")
            position_connect_action = batch_menu.addAction("位置连接")
            connect_to_text_action = batch_menu.addAction("连到文字")
            connect_to_image_action = batch_menu.addAction("连到图片")
            batch_menu.addSeparator()
            clear_connections_action = batch_menu.addAction("清除所有连接")
        
        action = menu.exec(global_pos)

        # 处理隐藏/显示/删除连接点
        if toggle_connection_point_action and action == toggle_connection_point_action:
            self.toggle_connection_point()
        elif delete_connection_point_action and action == delete_connection_point_action:
            self.delete_connection_point()
        elif action == copy_action:
            if self.scene():
                self.scene().copy_item(self)
        elif action == delete_action:
            if self.scene():
                self.scene().delete_item(self)
        elif action == save_as_asset_action:
            if self.scene():
                self.scene().save_group_as_asset()
        elif len(selected_items) >= 2:
            if action == align_top_action:
                self.scene().align_top(selected_items)
            elif action == align_right_action:
                self.scene().align_right(selected_items)
            elif action == align_center_h_action:
                self.scene().align_center_horizontal(selected_items)
            elif action == align_center_v_action:
                self.scene().align_center_vertical(selected_items)
            elif action == chain_connect_action:
                self.scene().batch_chain_connect_selected_items()
            elif action == auto_connect_action:
                self.scene().auto_connect_selected_items()
            elif action == position_connect_action:
                self.scene().connect_by_position()
            elif action == connect_to_text_action:
                self.scene().connect_all_images_to_text()
            elif action == connect_to_image_action:
                self.scene().connect_all_texts_to_image()
            elif action == clear_connections_action:
                self.scene().remove_all_image_text_connections()
        elif action == connect_image_text_action:
            if self.scene():
                self.scene().start_image_text_binding(self)
        elif action == disconnect_image_text_action:
            if self.scene():
                self.scene().remove_image_text_connectors(self)
        
        if action == unbind_action:
            # 使用撤销命令解除父级绑定
            if self.scene():
                old_parent = self.parentItem() if isinstance(self.parentItem(), BaseElement) else None
                if old_parent:
                    command = SetParentCommand(self.scene(), self, None, old_parent)
                    command.execute()
                    self.scene().undo_stack.push(command)
        elif action == set_parent_action:
             if self.scene(): self.scene().start_binding_mode(self)

    def contextMenuEvent(self, event):
        self._build_base_context_menu(event.screenPos())

    def paint(self, painter, option, widget):
        if self.isSelected():
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, False)
            rect = self.boundingRect()

            # ── CDR 风格：无填充背景，1px 蓝色虚线外框 ──────────────────
            dash_pen = QPen(QColor(0, 120, 215, 220), 1, Qt.PenStyle.DashLine)
            dash_pen.setCosmetic(True)
            dash_pen.setDashPattern([4, 3])
            painter.setPen(dash_pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(rect)

            # ── 四角白色锚点节点（4×4px，Cosmetic） ─────────────────────
            node_pen = QPen(QColor(0, 100, 200), 1)
            node_pen.setCosmetic(True)
            node_brush = QBrush(QColor(255, 255, 255))
            painter.setPen(node_pen)
            painter.setBrush(node_brush)
            # 节点尺寸在场景坐标中对应屏幕 4px
            if self.scene() and self.scene().views():
                scale = self.scene().views()[0].transform().m11()
                ns = 4.0 / scale if scale > 0 else 4.0
            else:
                ns = 4.0
            hs = ns / 2
            for cx, cy in [
                (rect.left(),  rect.top()),
                (rect.right(), rect.top()),
                (rect.left(),  rect.bottom()),
                (rect.right(), rect.bottom()),
            ]:
                painter.drawRect(QRectF(cx - hs, cy - hs, ns, ns))
    
    def toggle_connection_point(self):
        """切换连接点的可见性"""
        if hasattr(self, 'connection_point') and self.connection_point:
            current_visible = self.connection_point.isVisible()
            self.connection_point.setVisible(not current_visible)
            element_type = "图片" if isinstance(self, VImageItem) else "文字"
            if current_visible:
                print(f"{element_type}连接点已隐藏")
            else:
                print(f"{element_type}连接点已显示")

    def delete_connection_point(self):
        """永久删除连接点"""
        if hasattr(self, 'connection_point') and self.connection_point:
            cp = self.connection_point
            # 先断开该连接点关联的所有连线
            if self.scene():
                self.scene().remove_image_text_connectors(self)
            # 从场景移除并销毁
            if cp.scene():
                cp.scene().removeItem(cp)
            else:
                cp.setParentItem(None)
            self.connection_point = None
            element_type = "图片" if isinstance(self, VImageItem) else "文字"
            print(f"{element_type}连接点已永久删除")

class InlineTextEditor(QTextEdit):
    """原位内联编辑器：透明覆盖层，只负责接收键盘输入，视觉由 VTextItem.paint 负责。
    
    重构亮点：
    1. inputMethodQuery 重写，强制 IME 候选框对齐到当前竖排光标的物理位置
    2. cursorPositionChanged 实时通知 VTextItem 重绘，消除光标滞后感
    3. 完全透明，零视觉干扰，所有绘制权交给 VTextItem.paint
    """
    editingFinished = pyqtSignal(str)
    editingCancelled = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        # 完全透明，无边框——视觉效果由 VTextItem 自己画
        self.setStyleSheet("""
            QTextEdit {
                background: transparent;
                border: none;
                color: rgba(0,0,0,0);
                selection-background-color: transparent;
                selection-color: transparent;
            }
        """)
        # 隐藏 QTextEdit 自带竖向光标，竖排文字光标由 VTextItem.paint 自绘
        self.setCursorWidth(0)
        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0, 0))
        palette.setColor(QPalette.ColorRole.Base, QColor(0, 0, 0, 0))
        palette.setColor(QPalette.ColorRole.Window, QColor(0, 0, 0, 0))
        self.setPalette(palette)
        self.original_text = ""
        self.text_item = None
        # 每次文字变化时通知 VTextItem 实时预览
        self.textChanged.connect(self._on_text_changed)
        # 光标移动时也要立即重绘（任务一：消除选区/光标滞后）
        self.cursorPositionChanged.connect(self._on_cursor_moved)

    def _on_text_changed(self):
        if self.text_item and self.text_item.is_editing:
            self.text_item._editing_text = self.toPlainText()
            self.text_item._update_editing_rect()
            self.text_item.update()

    def _on_cursor_moved(self):
        """光标移动/选区变化时立即触发 VTextItem 重绘，消除卡顿感"""
        if self.text_item and self.text_item.is_editing:
            self.text_item.update()

    # ── 任务二：IME 候选框物理对齐 ─────────────────────────────────────────
    def inputMethodQuery(self, query):
        """重写输入法查询，将光标物理矩形返回给输入法，使候选框贴准当前字符位置。"""
        if query == Qt.InputMethodQuery.ImCursorRectangle and self.text_item and self.text_item.is_editing:
            try:
                rect = self._get_cursor_rect_in_viewport()
                if rect is not None:
                    return rect
            except Exception:
                pass
        return super().inputMethodQuery(query)

    def _get_cursor_rect_in_viewport(self):
        """计算当前光标在视图坐标系中的物理矩形，供 IME 精确定位候选框。"""
        text_item = self.text_item
        if not text_item or not text_item.scene():
            return None
        views = text_item.scene().views()
        if not views:
            return None
        view = views[0]

        text = self.toPlainText()
        cursor_pos = self.textCursor().position()
        positions, _, main_fm, char_h, col_step = text_item._editing_layout(text)

        # 找到当前光标对应的位置记录
        cur_entry = None
        for entry in positions:
            if entry['index'] == cursor_pos:
                cur_entry = entry
                break
        if cur_entry is None and positions:
            cur_entry = positions[-1]
        if cur_entry is None:
            return None

        # 光标位于该字符的顶端横线处（竖排）
        item_x = cur_entry['rect'].left()
        item_y = cur_entry['cursor_y']
        item_w = col_step
        item_h = char_h

        # 将 item 本地坐标 → 场景坐标 → 视图坐标
        scene_pt_tl = text_item.mapToScene(QPointF(item_x, item_y))
        scene_pt_br = text_item.mapToScene(QPointF(item_x + item_w, item_y + item_h))
        view_tl = view.mapFromScene(scene_pt_tl)
        view_br = view.mapFromScene(scene_pt_br)
        return QRect(view_tl, view_br)

    def insert_text_at_cursor(self, text):
        cursor = self.textCursor()
        cursor.insertText(text)
        self.setFocus()

    def _set_vertical_cursor_from_viewport_pos(self, pos, keep_anchor=False):
        if not self.text_item or not self.text_item.scene():
            return False
        views = self.text_item.scene().views()
        if not views:
            return False
        view = views[0]
        scene_pos = view.mapToScene(self.mapToParent(pos))
        item_pos = self.text_item.mapFromScene(scene_pos)
        cursor_pos = self.text_item._cursor_position_from_point(item_pos)
        cursor = self.textCursor()
        if keep_anchor:
            cursor.setPosition(cursor_pos, QTextCursor.MoveMode.KeepAnchor)
        else:
            cursor.setPosition(cursor_pos)
        self.setTextCursor(cursor)
        self.text_item.update()
        return True

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton and self._set_vertical_cursor_from_viewport_pos(event.pos()):
            self.setFocus()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if event.buttons() & Qt.MouseButton.LeftButton:
            if self._set_vertical_cursor_from_viewport_pos(event.pos(), keep_anchor=True):
                event.accept()
                return
        super().mouseMoveEvent(event)

    def start_editing(self, text_item, text, cursor_position=None, select_all=False):
        self.text_item = text_item
        self.original_text = text

        # 设置字体（与竖排渲染一致，影响光标高度）
        font = QFont(text_item.font_family)
        font.setPointSize(text_item.font_size)
        self.setFont(font)

        self.blockSignals(True)
        self.setPlainText(text)
        self.blockSignals(False)

        # QTextEdit 只负责接收输入，视觉光标由 VTextItem.paint 自绘
        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0, 0))
        palette.setColor(QPalette.ColorRole.Base, QColor(0, 0, 0, 0))
        palette.setColor(QPalette.ColorRole.Window, QColor(0, 0, 0, 0))
        self.setPalette(palette)
        self.setCursorWidth(0)

        scene = text_item.scene()
        if scene and scene.views():
            view = scene.views()[0]
            # 精确对齐到文字包围盒
            item_rect = text_item.boundingRect()
            scene_rect = text_item.mapRectToScene(item_rect)
            view_rect = view.mapFromScene(scene_rect).boundingRect()

            self.setParent(view.viewport())
            self.setGeometry(view_rect)
            self.show()
            self.setFocus()
            cursor = self.textCursor()
            if select_all:
                self.selectAll()
            else:
                pos = len(text) if cursor_position is None else max(0, min(cursor_position, len(text)))
                cursor.setPosition(pos)
                self.setTextCursor(cursor)

    def insertFromMimeData(self, source):
        super().insertFromMimeData(source)
        if self.text_item and self.text_item.scene():
            config_manager = getattr(self.text_item.scene(), 'config_manager', None)
            if config_manager and config_manager.get('auto_exit_after_paste', False):
                QTimer.singleShot(0, self.finish_editing)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self.editingCancelled.emit()
            self.hide()
            self.text_item = None
            self.original_text = ""
        elif event.key() == Qt.Key.Key_Return and event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            self.finish_editing()
        elif event.key() in (Qt.Key.Key_Up, Qt.Key.Key_Down, Qt.Key.Key_Left, Qt.Key.Key_Right):
            # 竖排方向键逻辑：上下=同列移动，左右=跨列移动
            self._handle_vertical_arrow(event)
        else:
            super().keyPressEvent(event)

    def _handle_vertical_arrow(self, event):
        """竖排方向键：上下=同列上下移动，左=移到左边列（col_idx更大），右=移到右边列（col_idx更小）"""
        if not self.text_item:
            return
        text = self.toPlainText()
        cursor = self.textCursor()
        pos = cursor.position()
        positions, _, _, char_h, col_step = self.text_item._editing_layout(text)

        # 找到当前光标对应的位置信息（优先精确匹配 index，退而求其次用末尾）
        cur_item = next((p for p in positions if p['index'] == pos), None)
        if cur_item is None and positions:
            cur_item = positions[-1]

        key = event.key()
        new_pos = pos

        if key == Qt.Key.Key_Up:
            same_col = [p for p in positions if p['col_idx'] == cur_item['col_idx'] and p['index'] < pos]
            if same_col:
                new_pos = same_col[-1]['index']

        elif key == Qt.Key.Key_Down:
            same_col = [p for p in positions if p['col_idx'] == cur_item['col_idx'] and p['index'] > pos]
            if same_col:
                new_pos = same_col[0]['index']

        elif key == Qt.Key.Key_Left:
            # 竖排从右到左：Left = 移到左边列（col_idx 更大的列）
            next_col = cur_item['col_idx'] + 1
            next_col_items = [p for p in positions if p['col_idx'] == next_col]
            if next_col_items:
                target_y = cur_item['cursor_y']
                closest = min(next_col_items, key=lambda p: abs(p['cursor_y'] - target_y))
                new_pos = closest['index']

        elif key == Qt.Key.Key_Right:
            # 竖排从右到左：Right = 移到右边列（col_idx 更小的列）
            prev_col = cur_item['col_idx'] - 1
            if prev_col >= 0:
                prev_col_items = [p for p in positions if p['col_idx'] == prev_col]
                if prev_col_items:
                    target_y = cur_item['cursor_y']
                    closest = min(prev_col_items, key=lambda p: abs(p['cursor_y'] - target_y))
                    new_pos = closest['index']

        if new_pos != pos:
            cursor.setPosition(new_pos)
            self.setTextCursor(cursor)
            if self.text_item:
                self.text_item.update()

    def focusOutEvent(self, event):
        if getattr(self, '_inserting_snippet', False):
            super().focusOutEvent(event)
            return
        self.finish_editing()
        super().focusOutEvent(event)

    def finish_editing(self):
        new_text = self.toPlainText()
        self.editingFinished.emit(new_text)
        self.hide()
        self.text_item = None
        self.original_text = ""

class VTextItem(BaseElement):
    """Vertical Text Engine (Right-to-Left columns)"""
    def __init__(self, text="请输入文本", font_size=DEFAULT_FONT_SIZE, box_height=400):
        super().__init__()
        self.full_text = text
        self.font_size = font_size
        self.font_family = DEFAULT_FONT
        self.text_color = QColor(Qt.GlobalColor.black)
        self.box_height = box_height  # 保留作为最大高度限制
        self.chars_per_column = 15  # 每列字符数，可以调整
        self.auto_height = True  # 是否自动调整高度
        self.manual_line_break = True  # 是否启用手动换行（响应\n字符）
        self.layer_eye_color = None  # 用户指定的层级眼睛颜色：yellow/red/green/None
        
        # 列间距属性
        self.column_spacing = COLUMN_SPACING  # 列间距（所有列间距相同）
        self.character_spacing = 0  # 同一列中字符之间的额外间距
        
        self._rect = QRectF(0, 0, 100, 100)  # 初始值，会在rebuild中重新计算
        self.connection_point = None  # 连接线
        
        # 内联编辑器
        self.inline_editor = None
        self.is_editing = False
        self._editing_text = ""  # 编辑中的实时文字（用于黑底白字预览）
        
        self.rebuild()
        self.create_connection_point()

    def rebuild(self, preserve_position=False):
        # 记录旧宽度用于位置补偿
        old_width = self._rect.width()
        old_right = self._rect.right()
        
        # 1. 清理旧的子项
        scene = self.scene()
        for child in self.childItems():
            if not isinstance(child, (BaseElement, ConnectionPoint)):
                if scene: scene.removeItem(child)
                else: child.setParentItem(None)
            
        # 2. 准备基础字体
        main_font = QFont(self.font_family)
        main_font.setPointSize(self.font_size)
        small_font = QFont(self.font_family)
        small_font.setPointSize(max(1, int(self.font_size * 0.5)))
        
        main_fm = QFontMetrics(main_font)
        small_fm = QFontMetrics(small_font)
        
        char_h_main = main_fm.height()
        char_h_small = small_fm.height()
        
        # 列宽依然以主字号为准
        col_step = self.font_size + self.column_spacing
        
        # 换行高度限制
        if self.auto_height:
            char_step_main = max(1.0, char_h_main * LINE_HEIGHT_RATIO + self.character_spacing)
            effective_height = char_h_main + (self.chars_per_column - 1) * char_step_main
        else:
            effective_height = self.box_height
        
        cursor_y = 0
        col_idx = 0
        generated_items = []
        
        # --- 核心逻辑：精准匹配“十一”和“十二” ---
        text_content = self.full_text
        i = 0
        SPECIAL_WORDS = ["十一", "十二"]

        while i < len(text_content):
            char = text_content[i]
            
            # 处理换行
            if char == '\n':
                if self.manual_line_break:
                    cursor_y = 0
                    col_idx += 1
                i += 1
                continue

            # 检查当前位置是否是“十一”或“十二”
            is_special = False
            if i + 1 < len(text_content) and text_content[i:i+2] in SPECIAL_WORDS:
                is_special = True
            elif i > 0 and text_content[i-1:i+1] in SPECIAL_WORDS:
                is_special = True

            # 根据是否是特殊字符选择字体和高度
            if is_special:
                current_font = small_font
                current_h = char_h_small
            else:
                current_font = main_font
                current_h = char_h_main

            # 换列判断
            if cursor_y + current_h > effective_height:
                cursor_y = 0
                col_idx += 1
            
            # 创建字符项
            t = QGraphicsSimpleTextItem(char)
            t.setFont(current_font)
            t.setBrush(QBrush(self.text_color))
            
            # 计算 X 偏移（相对于右侧第一列）
            x_local = -(col_idx * col_step)
            
            # 旋转特殊符号
            if char in ROTATE_CHARS:
                t.setTransformOriginPoint(t.boundingRect().center())
                t.setRotation(90)
            
            # 基础坐标
            final_x = x_local
            final_y = cursor_y
            
            # 标点符号偏移
            if char in OFFSET_CHARS:
                final_x += self.font_size * 0.4
                final_y -= self.font_size * 0.4
                
            # --- 关键：缩小字号后的居中修正 ---
            if is_special:
                # 使缩小的字在当前列的横向中间
                final_x += (self.font_size - t.boundingRect().width()) / 2

            t.setParentItem(self)
            t.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
            t.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
            t.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
            t.setPos(final_x, final_y)
            generated_items.append(t)
            
            # 累加 Y 坐标
            cursor_y += max(1.0, current_h * LINE_HEIGHT_RATIO + self.character_spacing)
            i += 1

        # 3. 重新对齐并计算方框 (Bounding Rect)
        if generated_items:
            combined_rect = QRectF()
            for item in generated_items:
                combined_rect = combined_rect.united(item.mapRectToParent(item.boundingRect()))

            if preserve_position:
                # 对话框编辑：对象 scenePos 不变，右侧文字锚点不动；
                # 新增列向左扩展，删除列从左侧收回。
                target_rect = combined_rect.adjusted(-2, -2, 2, 2)
                offset_x = old_right - target_rect.right()
                for item in generated_items:
                    item.setX(item.x() + offset_x)
                combined_rect.translate(offset_x, 0)
            else:
                # 普通重建：保持旧行为，将本地左边界归零，再用 moveBy 固定右边界。
                min_x = min(item.pos().x() for item in generated_items)
                for item in generated_items:
                    item.setX(item.x() - min_x)
                combined_rect = QRectF()
                for item in generated_items:
                    combined_rect = combined_rect.united(item.mapRectToParent(item.boundingRect()))

            self._rect = combined_rect.adjusted(-2, -2, 2, 2)
        else:
            if preserve_position:
                self._rect = QRectF(old_right - col_step, 0, col_step, char_h_main)
            else:
                self._rect = QRectF(0, 0, col_step, char_h_main)

        # 4. 锚点补偿：如果文字变宽/变窄，自动调整位置保持右边界不动
        if not preserve_position:
            new_right = self._rect.right()   # 本地坐标系下的右边界
            old_right_local = old_right       # 上次的右边界（本地坐标）
            dx = old_right_local - new_right
            if abs(dx) > 0.1:
                self.moveBy(dx, 0)
        
        # 5. 通知场景并更新连接点
        self.prepareGeometryChange()
        if self.connection_point:
            self.connection_point.update_position()
            if self.scene():
                self.scene().update_image_text_connectors(self)
        self.update_hover_tooltip()

    def update_hover_tooltip(self):
        scene = self.scene()
        enabled = False
        if scene and hasattr(scene, 'config_manager'):
            enabled = scene.config_manager.get('show_text_hover_tooltip', False)
        if enabled and self.full_text:
            lines = [xml_escape(line) for line in self.full_text.splitlines()]
            text = '<br>'.join(lines)
            self.setToolTip(
                '<html><body>'
                '<div style="width:520px; font-size:14px; line-height:1.45; '
                'white-space:pre-wrap;">'
                f'{text}'
                '</div></body></html>'
            )
        else:
            self.setToolTip('')
    
    def create_connection_point(self):
        """创建文字的连接点(底部中点)"""
        if not self.connection_point:
            self.connection_point = ConnectionPoint(self, "text_bottom")
            if self.scene():
                visible = self.scene().show_connection_points
                self.connection_point.setVisible(visible)
            else:
                self.connection_point.setVisible(True)
    
    def set_connection_points_visible(self, visible):
        """设置连接点可见性"""
        if self.connection_point:
            self.connection_point.setVisible(visible)
        for point in getattr(self, 'custom_connection_points', []):
            point.setVisible(visible)
        
    def _show_context_menu(self, global_pos, item_pos=None):
        self._build_text_context_menu(global_pos, item_pos)

    def contextMenuEvent(self, event):
        self._build_text_context_menu(event.screenPos(), event.pos())

    def _build_text_context_menu(self, global_pos, item_pos=None):
        menu = QMenu()
        clicked_cursor_position = self._cursor_position_from_point(item_pos) if item_pos is not None else None
        
        # 编辑选项
        action_inline_edit = menu.addAction("直接编辑 (Inline Edit)")
        action_dialog_edit = menu.addAction("对话框编辑 (Dialog Edit)")
        action_reset_editing = menu.addAction("重置编辑状态 (Reset Editing)")
        menu.addSeparator()
        
        action_font = menu.addAction("设置字体 (Font)")
        action_color = menu.addAction("设置颜色 (Color)")
        action_chars_per_col = menu.addAction("设置每列字数 (Chars per Column)")
        action_column_spacing = menu.addAction("设置列间距 (Column Spacing)")
        action_character_spacing = menu.addAction("设置字间距 (Character Spacing)")
        menu.addSeparator()
        
        # 添加隐藏/显示/删除连接点选项
        if self.connection_point and self.connection_point.isVisible():
            toggle_connection_point_action = menu.addAction("隐藏连接点")
        else:
            toggle_connection_point_action = menu.addAction("显示连接点")
        if self.connection_point:
            delete_cp_action = menu.addAction("删除连接点（永久）")
        else:
            delete_cp_action = None
        menu.addSeparator()
        
        copy_action = menu.addAction("复制 (Copy)")
        delete_action = menu.addAction("删除 (Delete)")
        save_as_asset_action = menu.addAction("保存组合")
        menu.addSeparator()

        main_window = None
        if self.scene() and self.scene().views():
            view = self.scene().views()[0]
            main_window = getattr(view, '_main_window', None)
        update_group_asset_action = None
        cancel_group_edit_action = None
        if main_window and getattr(self.scene(), '_editing_group_asset_id', None) is not None:
            update_group_asset_action = menu.addAction("✅ 更新到素材库（完成编辑）")
            cancel_group_edit_action = menu.addAction("❌ 取消编辑组合")
            menu.addSeparator()
        
        selected_items = [item for item in self.scene().selectedItems() if isinstance(item, BaseElement)]
        if len(selected_items) >= 2:
            align_menu = menu.addMenu("对齐 (Align)")
            align_top_action = align_menu.addAction("顶部对齐")
            align_right_action = align_menu.addAction("右对齐")
            align_center_h_action = align_menu.addAction("水平居中对齐")
            align_center_v_action = align_menu.addAction("垂直居中对齐")
            menu.addSeparator()
        
        unbind_action = menu.addAction("解除父级绑定 (Unbind)")
        set_parent_action = menu.addAction("设置父级 (Set Parent)")
        menu.addSeparator()
        connect_image_text_action = menu.addAction("图文连接 (Connect to Image/Text)")
        disconnect_image_text_action = menu.addAction("断开图文连接 (Disconnect Image/Text)")
        
        if len(selected_items) >= 2:
            batch_menu = menu.addMenu("批量连接 (Batch Connect)")
            auto_connect_action = batch_menu.addAction("智能连接")
            position_connect_action = batch_menu.addAction("位置连接")
            connect_to_text_action = batch_menu.addAction("连到文字")
            connect_to_image_action = batch_menu.addAction("连到图片")
            batch_menu.addSeparator()
            clear_connections_action = batch_menu.addAction("清除所有连接")
        
        action = menu.exec(global_pos)
        
        if update_group_asset_action and action == update_group_asset_action:
            main_window.finish_edit_group_asset()
            return
        if cancel_group_edit_action and action == cancel_group_edit_action:
            main_window.cancel_edit_group_asset()
            return

        if action == action_inline_edit:
            self.start_inline_editing(cursor_position=clicked_cursor_position, select_all=False)
        elif action == action_dialog_edit:
            self.start_dialog_editing()
        elif action == action_reset_editing:
            self.reset_editing_state()
        elif action == action_font:
            self.change_font_settings()
        elif action == action_color:
            self.change_color_settings()
        elif action == action_chars_per_col:
            self.change_chars_per_column_settings()
        elif action == action_column_spacing:
            self.change_column_spacing_settings()
        elif action == action_character_spacing:
            self.change_character_spacing_settings()
        elif action == toggle_connection_point_action:
            self.toggle_connection_point()
        elif delete_cp_action and action == delete_cp_action:
            self.delete_connection_point()
        elif action == copy_action:
            if self.scene():
                self.scene().copy_item(self)
        elif action == delete_action:
            if self.scene():
                self.scene().delete_item(self)
        elif action == save_as_asset_action:
            if self.scene():
                self.scene().save_group_as_asset()
        elif len(selected_items) >= 2:
            if action == align_top_action:
                self.scene().align_top(selected_items)
            elif action == align_right_action:
                self.scene().align_right(selected_items)
            elif action == align_center_h_action:
                self.scene().align_center_horizontal(selected_items)
            elif action == align_center_v_action:
                self.scene().align_center_vertical(selected_items)
            elif action == auto_connect_action:
                self.scene().auto_connect_selected_items()
            elif action == position_connect_action:
                self.scene().connect_by_position()
            elif action == connect_to_text_action:
                self.scene().connect_all_images_to_text()
            elif action == connect_to_image_action:
                self.scene().connect_all_texts_to_image()
            elif action == clear_connections_action:
                self.scene().remove_all_image_text_connections()
        elif action == connect_image_text_action:
            if self.scene():
                self.scene().start_image_text_binding(self)
        elif action == disconnect_image_text_action:
            if self.scene():
                self.scene().remove_image_text_connectors(self)
        
        if action == unbind_action:
            if self.scene():
                old_parent = self.parentItem() if isinstance(self.parentItem(), BaseElement) else None
                if old_parent:
                    command = SetParentCommand(self.scene(), self, None, old_parent)
                    command.execute()
                    self.scene().undo_stack.push(command)
        elif action == set_parent_action:
            if self.scene(): self.scene().start_binding_mode(self)

    def change_font_settings(self):
        current_font = QFont(self.font_family)
        current_font.setPointSize(self.font_size)
        cfg = self.scene().config_manager if self.scene() else None
        font, ok = FontPickerDialog.get_font(current_font, cfg, None, "选择字体")
        if ok:
            self.font_family = font.family()
            self.font_size = font.pointSize()
            self.rebuild()
            if self.scene(): 
                self.scene().update_connectors(self)
                self.scene().update_image_text_connectors(self)

    def change_color_settings(self):
        color = QColorDialog.getColor(self.text_color, None, "选择颜色")
        if color.isValid():
            self.text_color = color
            self.rebuild()
    
    def change_chars_per_column_settings(self):
        """设置每列字符数"""
        chars_count, ok = QInputDialog.getInt(None, "设置每列字符数", "每列字符数", self.chars_per_column, 5, 50)
        if ok:
            self.chars_per_column = chars_count
            self.rebuild()
            if self.scene(): 
                self.scene().update_connectors(self)
                self.scene().update_image_text_connectors(self)
    
    def change_column_spacing_settings(self):
        """设置列间距"""
        spacing, ok = QInputDialog.getInt(None, "设置列间距", "列间距 (像素):", self.column_spacing, 0, 200)
        if ok:
            self.column_spacing = spacing
            self.rebuild()
            if self.scene():
                self.scene().update_connectors(self)
                self.scene().update_image_text_connectors(self)
            print(f"列间距已设置为: {self.column_spacing}px")

    def change_character_spacing_settings(self):
        """设置同一竖列中相邻字符之间的额外间距。"""
        spacing, ok = QInputDialog.getInt(
            None, "设置字间距", "字间距 (像素):", self.character_spacing, -200, 200
        )
        if ok:
            self.character_spacing = spacing
            self.rebuild()
            if self.scene():
                self.scene().update_connectors(self)
                self.scene().update_image_text_connectors(self)
    
    def toggle_connection_point(self):
        """切换连接点的可见性"""
        if self.connection_point:
            current_visible = self.connection_point.isVisible()
            self.connection_point.setVisible(not current_visible)
            if current_visible:
                print("文字连接点已隐藏")
            else:
                print("文字连接点已显示")

    def mouseDoubleClickEvent(self, event):
        """双击像 Photoshop 文字层一样原位编辑，并尽量把光标放到点击处"""
        print(f"双击文字，当前编辑状态: {self.is_editing}")
        cursor_position = self._cursor_position_from_point(event.pos())
        select_all = bool(event.modifiers() & Qt.KeyboardModifier.ControlModifier)
        if self.is_editing and self.inline_editor:
            if select_all:
                self.inline_editor.selectAll()
            else:
                cursor = self.inline_editor.textCursor()
                cursor.setPosition(cursor_position)
                self.inline_editor.setTextCursor(cursor)
            self.inline_editor.setFocus()
        else:
            self.start_inline_editing(cursor_position=cursor_position, select_all=select_all)
        event.accept()

    def mousePressEvent(self, event):
        """编辑状态下点击定位光标，非编辑状态走父类拖拽逻辑"""
        if self.is_editing and event.button() == Qt.MouseButton.LeftButton:
            pos = self._cursor_position_from_point(event.pos())
            if self.inline_editor:
                cursor = self.inline_editor.textCursor()
                cursor.setPosition(pos)
                self.inline_editor.setTextCursor(cursor)
                self.inline_editor.setFocus()
            self._drag_select_start = pos
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        """编辑状态下拖拽选中文字（任务一：实时同步选区，零卡顿）"""
        if self.is_editing and event.buttons() & Qt.MouseButton.LeftButton:
            pos = self._cursor_position_from_point(event.pos())
            if self.inline_editor and self._drag_select_start is not None:
                cursor = self.inline_editor.textCursor()
                # anchor 固定在拖拽起点，position 跟随鼠标实时更新
                cursor.setPosition(self._drag_select_start)
                cursor.setPosition(pos, QTextCursor.MoveMode.KeepAnchor)
                self.inline_editor.setTextCursor(cursor)
                # 立即重绘，选区蓝块跟手
                self.update()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self.is_editing:
            self._drag_select_start = None
            event.accept()
            return
        super().mouseReleaseEvent(event)
    
    def start_inline_editing(self, cursor_position=None, select_all=False):
        """开始内联编辑：原位显示文字层、直接输入，体验接近 Photoshop"""
        if self.is_editing:
            return

        self.is_editing = True
        self._editing_text = self.full_text  # 初始化为当前文字
        # 记录初始右边界（场景坐标），用 mapRectToScene 确保精确
        scene_rect = self.mapRectToScene(self._rect)
        self._editing_right = scene_rect.right()
        self._editing_scene_y = scene_rect.top()
        self._editing_local_right = self._rect.right()
        self._editing_origin_scene_pos = self.scenePos()  # 记录进入编辑时的原始场景位置
        self._cursor_visible = True
        self._cursor_opacity = 1.0   # 任务三：用于平滑淡入淡出的透明度值
        self._drag_select_start = None  # 拖拽选择起始位置

        # 任务三：与系统光标同步的 500ms 闪烁定时器
        if not hasattr(self, '_cursor_timer'):
            self._cursor_timer = QTimer()
            self._cursor_timer.timeout.connect(self._blink_cursor)
        # 从系统设置读取闪烁间隔（fallback 500ms）
        blink_rate = QApplication.cursorFlashTime() // 2
        if blink_rate <= 0:
            blink_rate = 500
        self._cursor_timer.start(blink_rate)

        # 隐藏子字符，编辑期间由 paint 直接绘制预览
        for child in self.childItems():
            if isinstance(child, QGraphicsSimpleTextItem):
                child.setVisible(False)

        # 创建内联编辑器
        if not self.inline_editor:
            self.inline_editor = InlineTextEditor()
            self.inline_editor.editingFinished.connect(self.finish_inline_editing)
            self.inline_editor.editingCancelled.connect(self.cancel_inline_editing)

        # 触发文字层编辑态重绘
        self.update()

        # 启动透明覆盖编辑器
        self.inline_editor.start_editing(self, self.full_text, cursor_position, select_all)

    def insert_text(self, text):
        """如果正在编辑，则在编辑器光标处插入文本"""
        if self.is_editing and self.inline_editor:
            self.inline_editor.insert_text_at_cursor(text)
            return True
        return False
    
    def finish_inline_editing(self, new_text):
        """完成内联编辑：恢复正常显示并保存，保持对象当前位置不变"""
        if hasattr(self, '_cursor_timer'):
            self._cursor_timer.stop()
        # 先保存原始位置的副本
        origin_scene_pos = getattr(self, '_editing_origin_scene_pos', None) or self.scenePos()
        
        self.is_editing = False
        self._editing_text = ""
        self._editing_right = None
        self._editing_scene_y = None
        self._editing_local_right = None
        self._editing_origin_scene_pos = None
        
        # 恢复子字符可见性和颜色
        for child in self.childItems():
            if isinstance(child, QGraphicsSimpleTextItem):
                child.setVisible(True)
                child.setBrush(QBrush(self.text_color))
        self.update()

        if new_text != self.full_text and self.scene():
            old_text = self.full_text
            cmd = EditTextCommand(self.scene(), self, old_text, new_text)
            cmd.execute()
            self.scene().undo_stack.push(cmd)
        else:
            self.rebuild(preserve_position=True)
        
        # 使用之前保存的位置副本
        if self.parentItem():
            self.setPos(self.parentItem().mapFromScene(origin_scene_pos))
        else:
            self.setPos(origin_scene_pos)

        if self.scene():
            self.scene().update_connectors(self)
            self.scene().update_image_text_connectors(self)

    def cancel_inline_editing(self):
        """取消内联编辑：恢复原始文字和颜色，保持对象当前位置不变"""
        if hasattr(self, '_cursor_timer'):
            self._cursor_timer.stop()
        # 先保存原始位置的副本
        origin_scene_pos = getattr(self, '_editing_origin_scene_pos', None) or self.scenePos()
        
        self.is_editing = False
        self._editing_text = ""
        self._editing_right = None
        self._editing_scene_y = None
        self._editing_local_right = None
        self._editing_origin_scene_pos = None
        
        self.rebuild(preserve_position=True)
        
        # 使用之前保存的位置副本
        if self.parentItem():
            self.setPos(self.parentItem().mapFromScene(origin_scene_pos))
        else:
            self.setPos(origin_scene_pos)
            
        for child in self.childItems():
            if isinstance(child, QGraphicsSimpleTextItem):
                child.setVisible(True)
                child.setBrush(QBrush(self.text_color))
        self.update()
        if self.inline_editor:
            self.inline_editor.hide()

    def reset_editing_state(self):
        """强制重置编辑状态"""
        if hasattr(self, '_cursor_timer'):
            self._cursor_timer.stop()
        self.is_editing = False
        self._editing_text = ""
        self._editing_right = None
        self._editing_scene_y = None
        for child in self.childItems():
            if isinstance(child, QGraphicsSimpleTextItem):
                child.setVisible(True)
                child.setBrush(QBrush(self.text_color))
        self.update()
        if self.inline_editor:
            self.inline_editor.hide()
            self.inline_editor.text_item = None
            self.inline_editor.original_text = ""

    def start_dialog_editing(self):
        """开始对话框编辑"""
        text, ok = QInputDialog.getMultiLineText(None, "编辑文本", "请输入排版内容", self.full_text)
        if ok and text != self.full_text and self.scene():
            old_text = self.full_text
            cmd = EditTextCommand(self.scene(), self, old_text, text)
            cmd.execute()
            self.scene().undo_stack.push(cmd)
    
    def keyPressEvent(self, event):
        """处理键盘事件"""
        if event.key() == Qt.Key.Key_F2 or event.key() == Qt.Key.Key_Return:
            # F2 或 Enter 键开始内联编辑
            self.start_inline_editing()
            event.accept()
        else:
            super().keyPressEvent(event)

    def _blink_cursor(self):
        """光标闪烁（任务三：平滑切换，与系统节奏同步）"""
        self._cursor_visible = not getattr(self, '_cursor_visible', True)
        self.update()

    def _editing_layout(self, text):
        """返回竖排编辑预览的字符位置，供绘制光标和双击定位复用"""
        main_font = QFont(self.font_family)
        main_font.setPointSize(self.font_size)
        main_fm = QFontMetrics(main_font)
        char_h = main_fm.height()
        col_step = self.font_size + self.column_spacing
        char_step = max(1.0, char_h * LINE_HEIGHT_RATIO + self.character_spacing)
        effective_height = (
            char_h + (self.chars_per_column - 1) * char_step
            if self.auto_height else self.box_height
        )
        # 去掉 rebuild 里 adjusted(-2,-2,2,2) 带来的 2px 右边偏移，保持与非编辑状态对齐
        right_edge = self._rect.right() - 2
        cursor_y = 0
        col_idx = 0
        positions = []

        for index, ch in enumerate(text):
            if ch == '\n':
                col_x_left = right_edge - (col_idx + 1) * col_step
                positions.append({
                    'index': index,
                    'char': ch,
                    'rect': QRectF(col_x_left, cursor_y, col_step, char_h),
                    'cursor_y': cursor_y,
                    'col_idx': col_idx,
                    'x': col_x_left,
                })
                if self.manual_line_break:
                    cursor_y = 0
                    col_idx += 1
                continue

            if cursor_y + char_h > effective_height:
                cursor_y = 0
                col_idx += 1

            col_x_left = right_edge - (col_idx + 1) * col_step
            x = col_x_left + (col_step - self.font_size) / 2
            positions.append({
                'index': index,
                'char': ch,
                'rect': QRectF(col_x_left, cursor_y, col_step, char_h),
                'cursor_y': cursor_y,
                'col_idx': col_idx,
                'x': x,
            })
            cursor_y += char_step

        col_x_left = right_edge - (col_idx + 1) * col_step
        positions.append({
            'index': len(text),
            'char': '',
            'rect': QRectF(col_x_left, cursor_y, col_step, char_h),
            'cursor_y': cursor_y,
            'col_idx': col_idx,
            'x': col_x_left,
        })
        return positions, main_font, main_fm, char_h, col_step

    def _cursor_position_from_point(self, point):
        """把点击位置精确换算成字符索引（任务一重构）。
        
        竖排逻辑：
        - 遍历每个字符的精确物理矩形（QRectF）
        - 落在矩形内：上半 → 置于字符前，下半 → 置于字符后
        - 落在矩形外：按列+Y方向的综合距离找最近字符
        """
        text = self._editing_text if self.is_editing else self.full_text
        positions, _, _, char_h, _ = self._editing_layout(text)
        if not positions:
            return 0

        px, py = point.x(), point.y()

        # ── 第一优先级：点落在某字符矩形内部，精确判断上下半区 ──────────
        for entry in positions:
            rect = entry['rect']
            # 稍微扩展判定区域，避免列间隙漏判
            hit_rect = rect.adjusted(-1, 0, 1, 0)
            if hit_rect.contains(QPointF(px, py)):
                mid_y = rect.top() + rect.height() * 0.5
                if py <= mid_y:
                    # 点在字符上半部 → 光标置于字符之前
                    return entry['index']
                else:
                    # 点在字符下半部 → 光标置于字符之后
                    idx = entry['index']
                    # 末尾哨兵（index == len(text)）不能 +1
                    return idx + 1 if idx < len(text) else idx

        # ── 第二优先级：落在字符矩形外，找综合距离最近的字符 ─────────────
        # 竖排优先按列（X轴）判断所属列，再在列内按Y查找最近行
        best_index = len(text)
        best_dist = float('inf')

        for entry in positions:
            rect = entry['rect']
            # 计算点到矩形的切比雪夫距离（列宽方向权重更大，优先归属到正确列）
            cx = max(rect.left(), min(px, rect.right()))
            cy = max(rect.top(), min(py, rect.bottom()))
            dx = abs(px - cx) * 1.8   # X 方向加权，优先列归属
            dy = abs(py - cy)
            dist = dx + dy
            if dist < best_dist:
                best_dist = dist
                mid_y = rect.top() + rect.height() * 0.5
                idx = entry['index']
                if py > mid_y and idx < len(text):
                    best_index = idx + 1
                else:
                    best_index = idx

        return max(0, min(best_index, len(text)))

    def _update_editing_rect(self):
        """编辑状态下根据当前文字实时计算包围盒，位置固定，向左扩展/收回"""
        text = self._editing_text
        positions, _, _, char_h, col_step = self._editing_layout(text)
        max_col = max((item['col_idx'] for item in positions), default=0)
        max_y = max((item['cursor_y'] + char_h for item in positions), default=char_h)
        new_w = (max_col + 1) * col_step
        new_h = max_y

        local_right = getattr(self, '_editing_local_right', self._rect.right())

        self.prepareGeometryChange()
        self._rect = QRectF(local_right - new_w, 0, new_w, new_h)

        # 同步更新透明编辑器的位置和大小
        if self.inline_editor and self.inline_editor.isVisible() and self.scene() and self.scene().views():
            view = self.scene().views()[0]
            scene_rect = self.mapRectToScene(self._rect)
            view_rect = view.mapFromScene(scene_rect).boundingRect()
            self.inline_editor.setGeometry(view_rect)

    def paint(self, painter, option, widget):
        """编辑状态下绘制原位文字层预览：真·所见即所得（透明背景 + 原色文字 + 半透明选区）"""
        if self.is_editing:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            preview_text = self._editing_text if self._editing_text != "" or self.full_text == "" else self.full_text
            positions, main_font, main_fm, char_h, col_step = self._editing_layout(preview_text)

            # ── 背景完全透明，不绘制任何底色 ─────────────────────────────
            right_edge = self._rect.right() - 2

            # ── 获取光标和选区信息 ─────────────────────────────────────────
            sel_start, sel_end = -1, -1
            cursor_char_pos = -1
            if self.inline_editor and self.inline_editor.isVisible():
                tc = self.inline_editor.textCursor()
                cursor_char_pos = tc.position()
                if tc.hasSelection():
                    sel_start = tc.selectionStart()
                    sel_end = tc.selectionEnd()

            # ── 第一层：半透明蓝色选区（绘制在文字下方） ─────────────────
            if sel_start >= 0 and sel_end > sel_start:
                col_sel = {}
                for entry in positions:
                    if entry['char'] == '\n' or entry['index'] >= len(preview_text):
                        continue
                    if sel_start <= entry['index'] < sel_end:
                        cidx = entry['col_idx']
                        top = entry['cursor_y']
                        bot = entry['cursor_y'] + char_h
                        if cidx not in col_sel:
                            col_sel[cidx] = (top, bot)
                        else:
                            col_sel[cidx] = (min(col_sel[cidx][0], top),
                                             max(col_sel[cidx][1], bot))
                for cidx, (top, bot) in col_sel.items():
                    col_x = right_edge - (cidx + 1) * col_step
                    # alpha=70：半透明，文字颜色在选区上仍清晰可读
                    painter.fillRect(QRectF(col_x, top, col_step, bot - top),
                                     QColor(0, 120, 215, 70))

            # ── 第二层：用原始文字颜色绘制字符 ───────────────────────────
            painter.setFont(main_font)
            painter.setPen(self.text_color)
            for entry in positions:
                if entry['index'] >= len(preview_text) or entry['char'] == '\n':
                    continue
                painter.drawText(
                    QPointF(entry['x'],
                            entry['cursor_y'] + char_h - main_fm.descent()),
                    entry['char']
                )

            # ── 第三层：I-Beam 光标，颜色跟随 self.text_color ─────────────
            if cursor_char_pos >= 0 and self.inline_editor:
                show_cursor = getattr(self, '_cursor_visible', True)
                if show_cursor and cursor_char_pos < len(positions):
                    entry = positions[cursor_char_pos]
                    cy   = entry['cursor_y']
                    cidx = entry['col_idx']
                    col_x_left  = right_edge - (cidx + 1) * col_step
                    col_x_right = col_x_left + col_step
                    lx0 = col_x_left  + 2
                    lx1 = col_x_right - 2

                    beam_color = QColor(self.text_color)   # 与文字颜色一致
                    pen_main = QPen(beam_color, 2)
                    pen_main.setCosmetic(True)
                    painter.setPen(pen_main)
                    painter.drawLine(QPointF(lx0, cy), QPointF(lx1, cy))

                    serif_h = 4
                    pen_serif = QPen(beam_color, 1.5)
                    pen_serif.setCosmetic(True)
                    painter.setPen(pen_serif)
                    painter.drawLine(QPointF(lx0, cy - serif_h * 0.5),
                                     QPointF(lx0, cy + serif_h * 0.5))
                    painter.drawLine(QPointF(lx1, cy - serif_h * 0.5),
                                     QPointF(lx1, cy + serif_h * 0.5))
        else:
            for child in self.childItems():
                if isinstance(child, QGraphicsSimpleTextItem):
                    child.setBrush(QBrush(self.text_color))
            super().paint(painter, option, widget)

    def boundingRect(self):
        return self._rect

class ResizeHandle(QGraphicsItem):
    """图片缩放控制点：角点=方形(等比)，边中点=圆形(单向)"""
    SIZE = 10

    TOOLTIPS = {
        'tl': '等比缩放', 'tr': '等比缩放',
        'bl': '等比缩放', 'br': '等比缩放',
        'tc': '上下拉伸', 'bc': '上下拉伸',
        'ml': '左右拉伸', 'mr': '左右拉伸',
    }

    def __init__(self, parent_image, role):
        super().__init__(parent_image)
        self.parent_image = parent_image
        self.role = role
        self.is_corner = role in ('tl', 'tr', 'bl', 'br')
        self._hovered = False
        self.setZValue(200)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
        self.setAcceptHoverEvents(True)
        self.setToolTip(self.TOOLTIPS.get(role, ''))
        self._dragging = False
        self._orig_rect = QRectF()
        self._update_cursor()

    def boundingRect(self):
        s = self._screen_size()
        return QRectF(-s/2, -s/2, s, s)

    def _screen_size(self):
        """计算在当前视图缩放下，屏幕上固定10px对应的场景尺寸"""
        if self.scene() and self.scene().views():
            scale = self.scene().views()[0].transform().m11()
            if scale > 0:
                return self.SIZE / scale
        return self.SIZE

    def paint(self, painter, option, widget):
        s = self._screen_size()
        fill = QColor(0, 120, 215) if self._hovered else QColor(255, 255, 255)
        border = QColor(0, 90, 180)
        pen = QPen(border, 1.5 / (self.scene().views()[0].transform().m11() if self.scene() and self.scene().views() else 1))
        painter.setPen(pen)
        painter.setBrush(QBrush(fill))
        if self.is_corner:
            painter.drawRect(QRectF(-s/2, -s/2, s, s))
        else:
            painter.drawEllipse(QRectF(-s/2, -s/2, s, s))

    def _update_cursor(self):
        cursors = {
            'tl': Qt.CursorShape.SizeFDiagCursor, 'br': Qt.CursorShape.SizeFDiagCursor,
            'tr': Qt.CursorShape.SizeBDiagCursor, 'bl': Qt.CursorShape.SizeBDiagCursor,
            'tc': Qt.CursorShape.SizeVerCursor,   'bc': Qt.CursorShape.SizeVerCursor,
            'ml': Qt.CursorShape.SizeHorCursor,   'mr': Qt.CursorShape.SizeHorCursor,
        }
        self.setCursor(cursors.get(self.role, Qt.CursorShape.SizeAllCursor))

    def hoverEnterEvent(self, event):
        self._hovered = True
        self.update()
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        self._hovered = False
        self.update()
        super().hoverLeaveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = True
            self._orig_rect = QRectF(self.parent_image._rect)
            self._orig_pos = self.parent_image.scenePos()
            self._orig_pos_local = self.parent_image.pos()
            # 记录按下时在图片本地坐标系的位置
            self._drag_start_local = self.parent_image.mapFromScene(event.scenePos())
            event.accept()

    def mouseMoveEvent(self, event):
        if not self._dragging:
            return
        img = self.parent_image
        # 用原始场景坐标系计算 delta，避免 img 位置变化导致坐标系偏移
        scene_start = self._drag_start_local  # 这是原始本地坐标
        # 将当前鼠标场景坐标转到原始位置的本地坐标系
        # 用 orig_pos 重建变换，避免 img 已经移动的影响
        cur_scene = event.scenePos()
        # 直接用场景坐标差值（不依赖 img 当前位置）
        orig_scene = self._orig_pos + scene_start  # 近似：orig_scene_pos + local_offset
        dx = cur_scene.x() - orig_scene.x()
        dy = cur_scene.y() - orig_scene.y()

        r = QRectF(self._orig_rect)

        r = QRectF(self._orig_rect)

        # 根据角色调整矩形
        if 'r' in self.role:
            r.setRight(r.right() + dx)
        if 'l' in self.role:
            r.setLeft(r.left() + dx)
        if 'b' in self.role:
            r.setBottom(r.bottom() + dy)
        if 't' in self.role:
            r.setTop(r.top() + dy)

        # 角点：等比缩放
        if self.role in ('tl', 'tr', 'bl', 'br'):
            new_w = max(20, r.width())
            ratio = img._orig_ratio
            new_h = new_w * ratio
            if self.role == 'tl':
                r = QRectF(r.right() - new_w, r.bottom() - new_h, new_w, new_h)
            elif self.role == 'tr':
                r = QRectF(r.left(), r.bottom() - new_h, new_w, new_h)
            elif self.role == 'bl':
                r = QRectF(r.right() - new_w, r.top(), new_w, new_h)
            elif self.role == 'br':
                r = QRectF(r.left(), r.top(), new_w, new_h)
        else:
            r.setWidth(max(20, r.width()))
            r.setHeight(max(20, r.height()))

        img._apply_resize(r, self._orig_pos)
        event.accept()

    def mouseReleaseEvent(self, event):
        self._dragging = False
        event.accept()


class VImageItem(BaseElement):
    """Image Item that fits into columns"""
    def __init__(self, path, target_width=DEFAULT_FONT_SIZE, target_height=0):
        super().__init__()
        self.file_path = path
        self.target_width = target_width
        self.connection_point = None
        self._handles = []
        self._orig_ratio = 1.0
        self.image_opacity = 1.0
        self.locked = False

        pix = QPixmap(path)
        if not pix.isNull():
            self._orig_ratio = pix.height() / pix.width()
            if target_height and target_height > 0:
                # 强制指定宽高
                target_h = target_height
                self.p_item = QGraphicsPixmapItem(pix.scaled(
                    int(target_width), int(target_h),
                    Qt.AspectRatioMode.IgnoreAspectRatio,
                    Qt.TransformationMode.SmoothTransformation))
            else:
                target_h = target_width * self._orig_ratio
                self.p_item = QGraphicsPixmapItem(pix.scaled(
                    int(target_width), int(target_h),
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation))
            self.p_item.setParentItem(self)
            self._rect = QRectF(0, 0, target_width, target_h)
        else:
            self._rect = QRectF(0, 0, target_width, target_width)

        self.create_connection_point()
        self._create_handles()
        self._update_handles()
        self._show_handles(False)

    def set_opacity(self, opacity):
        """设置图片透明度 0.0-1.0"""
        self.image_opacity = max(0.0, min(1.0, opacity))
        if hasattr(self, 'p_item'):
            self.p_item.setOpacity(self.image_opacity)
        self.update()

    def set_locked(self, locked):
        """锁定/解锁图片"""
        self.locked = locked
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, not locked)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, not locked)
        self._show_handles(False)
        # 锁定时显示半透明遮罩提示
        if hasattr(self, 'p_item'):
            self.p_item.setOpacity(self.image_opacity * (0.6 if locked else 1.0))
        self.update()

    def set_image_visible(self, visible):
        """设置图片显示/隐藏，同步相关连线可见性"""
        self.setVisible(visible)
        if not visible:
            self.setSelected(False)
            self._show_handles(False)
        if not self.scene():
            return
        sc = self.scene()
        # 同步父子连线
        for conn in sc.connectors:
            if conn.parent_element == self or conn.child_element == self:
                both_visible = conn.parent_element.isVisible() and conn.child_element.isVisible()
                conn.setVisible(sc.show_connectors and both_visible)
        # 同步图文连线
        for conn in sc.image_text_connectors:
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                i1, i2 = conn.image_item, conn.text_item
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                i1, i2 = conn.item1, conn.item2
            else:
                continue
            if i1 == self or i2 == self:
                both_visible = i1.isVisible() and i2.isVisible()
                conn.setVisible(sc.show_image_text_connectors and both_visible)
        sc.update_connectors(self)
        sc.update_image_text_connectors(self)

    def toggle_image_visible(self):
        """切换图片显示/隐藏"""
        self.set_image_visible(not self.isVisible())

    def _create_handles(self):
        for role in ('tl', 'tc', 'tr', 'ml', 'mr', 'bl', 'bc', 'br'):
            h = ResizeHandle(self, role)
            self._handles.append(h)

    def _update_handles(self):
        r = self._rect
        cx, cy = r.center().x(), r.center().y()
        positions = {
            'tl': (r.left(),  r.top()),
            'tc': (cx,        r.top()),
            'tr': (r.right(), r.top()),
            'ml': (r.left(),  cy),
            'mr': (r.right(), cy),
            'bl': (r.left(),  r.bottom()),
            'bc': (cx,        r.bottom()),
            'br': (r.right(), r.bottom()),
        }
        for h in self._handles:
            x, y = positions[h.role]
            h.setPos(x, y)

    def _show_handles(self, visible):
        for h in self._handles:
            h.setVisible(visible)

    def _apply_resize(self, new_rect, orig_scene_pos):
        """应用新尺寸，重新缩放图片"""
        w = max(20, new_rect.width())
        h = max(20, new_rect.height())

        pix = QPixmap(self.file_path)
        if not pix.isNull():
            scaled = pix.scaled(int(w), int(h),
                                Qt.AspectRatioMode.IgnoreAspectRatio,
                                Qt.TransformationMode.SmoothTransformation)
            self.p_item.setPixmap(scaled)

        # 位置补偿：拖左/上边时，元素需要移动以保持右/下边不动
        # new_rect 的 topLeft 是相对于原始 (0,0) 的偏移
        offset_local = new_rect.topLeft()
        if abs(offset_local.x()) > 0.1 or abs(offset_local.y()) > 0.1:
            # 将本地偏移转换为场景偏移
            if self.parentItem():
                offset_scene = self.parentItem().mapToScene(self.pos() + offset_local) - self.parentItem().mapToScene(self.pos())
                new_scene_pos = orig_scene_pos + offset_scene
                self.setPos(self.parentItem().mapFromScene(new_scene_pos))
            else:
                self.setPos(orig_scene_pos + offset_local)

        self.prepareGeometryChange()
        self._rect = QRectF(0, 0, w, h)
        self.target_width = w
        self._update_handles()

        if self.connection_point:
            self.connection_point.update_position()
        if self.scene():
            self.scene().update_image_text_connectors(self)

    def itemChange(self, change, value):
        if change == QGraphicsItem.GraphicsItemChange.ItemSelectedHasChanged:
            # 只有在调整大小模式下才随选中状态显示控制点
            if self.scene() and getattr(self.scene(), 'resize_mode', False):
                self._show_handles(bool(value))
            else:
                self._show_handles(False)
        return super().itemChange(change, value)
    
    def create_connection_point(self):
        """创建图片的连接点(顶部中点)"""
        if not self.connection_point:
            self.connection_point = ConnectionPoint(self, "image_top")
            if self.scene():
                visible = self.scene().show_connection_points
                self.connection_point.setVisible(visible)
            else:
                self.connection_point.setVisible(True)
    
    def set_connection_points_visible(self, visible):
        """设置连接点可见性"""
        if self.connection_point:
            self.connection_point.setVisible(visible)
        for point in getattr(self, 'custom_connection_points', []):
            point.setVisible(visible)
    
    def toggle_connection_point(self):
        """切换连接点的可见性"""
        if self.connection_point:
            current_visible = self.connection_point.isVisible()
            self.connection_point.setVisible(not current_visible)
            if current_visible:
                print("图片连接点已隐藏")
            else:
                print("图片连接点已显示")

    def _get_child_texts(self):
        """获取直接子文字元素列表"""
        return [c for c in self.childItems() if isinstance(c, VTextItem)]

    def _build_image_context_menu(self, global_pos):
        menu = QMenu()
        child_texts = self._get_child_texts()

        # 子文字编辑区
        if child_texts:
            text_menu = menu.addMenu(f"编辑子文字 ({len(child_texts)} 个)")
            for idx, txt in enumerate(child_texts):
                label = txt.full_text[:12] + ('…' if len(txt.full_text) > 12 else '')
                t_menu = text_menu.addMenu(f"[{idx+1}] {label}")
                t_menu.addAction("内联编辑").triggered.connect(lambda _, t=txt: t.start_inline_editing())
                t_menu.addAction("对话框编辑").triggered.connect(lambda _, t=txt: t.start_dialog_editing())
                t_menu.addSeparator()
                t_menu.addAction("设置字体").triggered.connect(lambda _, t=txt: t.change_font_settings())
                t_menu.addAction("设置颜色").triggered.connect(lambda _, t=txt: t.change_color_settings())
                t_menu.addAction("每列字数").triggered.connect(lambda _, t=txt: t.change_chars_per_column_settings())
                t_menu.addAction("列间距").triggered.connect(lambda _, t=txt: t.change_column_spacing_settings())
                t_menu.addAction("字间距").triggered.connect(lambda _, t=txt: t.change_character_spacing_settings())
                t_menu.addSeparator()
                t_menu.addAction("解除父级绑定").triggered.connect(
                    lambda _, t=txt: (
                        lambda cmd: (cmd.execute(), self.scene().undo_stack.push(cmd))
                    )(SetParentCommand(self.scene(), t, None, self))
                    if self.scene() else None
                )
            menu.addSeparator()

        # 锁定/解锁
        lock_lbl = "🔓 解锁图片" if self.locked else "🔒 锁定图片"
        menu.addAction(lock_lbl).triggered.connect(lambda: self.set_locked(not self.locked))
        visible_lbl = "隐藏图片" if self.isVisible() else "显示图片"
        menu.addAction(visible_lbl).triggered.connect(self.toggle_image_visible)
        menu.addSeparator()

        # 图片自身操作
        if self.connection_point:
            lbl = "隐藏连接点" if self.connection_point.isVisible() else "显示连接点"
            menu.addAction(lbl).triggered.connect(self.toggle_connection_point)
            menu.addAction("删除连接点（永久）").triggered.connect(self.delete_connection_point)
            menu.addSeparator()

        # 图片透明度
        menu.addAction(f"设置透明度 (当前 {int(self.image_opacity*100)}%)").triggered.connect(self._change_opacity)
        menu.addSeparator()

        menu.addAction("复制").triggered.connect(lambda: self.scene().copy_item(self) if self.scene() else None)
        menu.addAction("删除").triggered.connect(lambda: self.scene().delete_item(self) if self.scene() else None)
        menu.addAction("保存组合").triggered.connect(lambda: self.scene().save_group_as_asset() if self.scene() else None)
        menu.addSeparator()

        main_window = None
        if self.scene() and self.scene().views():
            view = self.scene().views()[0]
            main_window = getattr(view, '_main_window', None)
        if main_window and getattr(self.scene(), '_editing_group_asset_id', None) is not None:
            menu.addAction("✅ 更新到素材库（完成编辑）").triggered.connect(main_window.finish_edit_group_asset)
            menu.addAction("❌ 取消编辑组合").triggered.connect(main_window.cancel_edit_group_asset)
            menu.addSeparator()

        selected_items = [i for i in self.scene().selectedItems() if isinstance(i, BaseElement)] if self.scene() else []
        if len(selected_items) >= 2:
            align_menu = menu.addMenu("对齐")
            align_menu.addAction("顶部对齐").triggered.connect(lambda: self.scene().align_top(selected_items))
            align_menu.addAction("右对齐").triggered.connect(lambda: self.scene().align_right(selected_items))
            align_menu.addAction("水平居中").triggered.connect(lambda: self.scene().align_center_horizontal(selected_items))
            align_menu.addAction("垂直居中").triggered.connect(lambda: self.scene().align_center_vertical(selected_items))
            batch_menu = menu.addMenu("批量连接")
            batch_menu.addAction("智能连接").triggered.connect(lambda: self.scene().auto_connect_selected_items())
            batch_menu.addAction("清除所有连接").triggered.connect(lambda: self.scene().remove_all_image_text_connections())
            menu.addSeparator()

        menu.addAction("解除父级绑定").triggered.connect(
            lambda: (
                lambda cmd: (cmd.execute(), self.scene().undo_stack.push(cmd))
            )(SetParentCommand(self.scene(), self, None, self.parentItem()))
            if self.scene() and isinstance(self.parentItem(), BaseElement) else None
        )
        menu.addAction("设置父级").triggered.connect(lambda: self.scene().start_binding_mode(self) if self.scene() else None)
        menu.addSeparator()
        menu.addAction("图文连接").triggered.connect(lambda: self.scene().start_image_text_binding(self) if self.scene() else None)
        menu.addAction("断开图文连接").triggered.connect(lambda: self.scene().remove_image_text_connectors(self) if self.scene() else None)

        menu.exec(global_pos)

    def _change_opacity(self):
        """弹出对话框设置透明度"""
        val, ok = QInputDialog.getInt(
            None, "设置透明度", "透明度 (0-100)%:",
            int(self.image_opacity * 100), 0, 100
        )
        if ok:
            self.set_opacity(val / 100.0)

    def _show_context_menu(self, global_pos):
        self._build_image_context_menu(global_pos)

    def contextMenuEvent(self, event):
        if self.locked:
            event.ignore()
            return
        self._build_image_context_menu(event.screenPos())

    def paint(self, painter, option, widget):
        """图片管理模式下，隐藏的图片显示为半透明虚框"""
        scene = self.scene()
        if scene and getattr(scene, 'image_manage_mode', False) and not self.isVisible():
            # 强制绘制半透明虚框（即使 isVisible() 为 False，paint 不会被调用）
            # 这里通过 drawForeground 或直接在 scene 里绘制，见下方 drawForeground
            pass
        super().paint(painter, option, widget)

    def boundingRect(self):
        return self._rect


class CorelSvgExporter:
    """Write editable SVG from scene objects instead of rendering a flat snapshot."""

    @staticmethod
    def export(scene, filepath):
        content_rect = CorelSvgExporter._content_rect(scene)
        if content_rect.isEmpty():
            raise ValueError("No visible content to export.")

        # Keep the SVG page aligned with the editor canvas.  Cropping the page
        # to content_rect makes every object appear shifted toward the SVG
        # origin, so the exported layout no longer matches the canvas.
        scene_rect = scene.sceneRect()
        if scene_rect.isEmpty():
            scene_rect = content_rect
        view = QRectF(scene_rect)
        view.moveTo(0, 0)
        offset = scene_rect.topLeft()
        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            (
                '<svg xmlns="http://www.w3.org/2000/svg" '
                'xmlns:xlink="http://www.w3.org/1999/xlink" version="1.1" '
                f'width="{CorelSvgExporter._num(CorelSvgExporter._px_to_mm(view.width()))}mm" '
                f'height="{CorelSvgExporter._num(CorelSvgExporter._px_to_mm(view.height()))}mm" '
                f'viewBox="0 0 {CorelSvgExporter._num(view.width())} {CorelSvgExporter._num(view.height())}">'
            ),
            '<rect x="0" y="0" width="100%" height="100%" fill="#ffffff"/>',
        ]

        items = [i for i in scene.items() if CorelSvgExporter._is_export_item(i)]
        items.sort(key=lambda i: i.zValue())
        exported_connectors = set()
        exported_groups = set()
        group_number = 0

        export_dir = os.path.splitext(filepath)[0] + "_assets"

        for item in items:
            group_root = CorelSvgExporter._group_root(item)
            if group_root is not None:
                if group_root in exported_groups:
                    continue
                grouped_items = [
                    candidate for candidate in items
                    if CorelSvgExporter._group_root(candidate) is group_root
                ]
                if len(grouped_items) > 1:
                    group_number += 1
                    exported_groups.add(group_root)
                    lines.append(f'<g id="layout-group-{group_number}">')
                    for grouped_item in grouped_items:
                        lines.extend(
                            CorelSvgExporter._item_to_svg(
                                grouped_item, offset, export_dir, filepath, exported_connectors
                            )
                        )
                    lines.append('</g>')
                    continue

            lines.extend(
                CorelSvgExporter._item_to_svg(
                    item, offset, export_dir, filepath, exported_connectors
                )
            )

        lines.append('</svg>')

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

    @staticmethod
    def _is_export_item(item):
        if isinstance(item, (VTextItem, VImageItem)):
            return item.isVisible()
        if isinstance(item, (VImageTextConnector, VGenericConnector)):
            return item.isVisible() and not item.path().isEmpty()
        return False

    @staticmethod
    def _group_root(item):
        """返回绑定组合的根元素；连线不参与 SVG 组合。"""
        if not isinstance(item, BaseElement):
            return None
        root = item
        while isinstance(root.parentItem(), BaseElement):
            root = root.parentItem()
        return root

    @staticmethod
    def _item_to_svg(item, offset, export_dir, filepath, exported_connectors):
        """将单个场景元素转换为 SVG，并复用同一套绝对场景坐标。"""
        if isinstance(item, VImageItem):
            return CorelSvgExporter._image_to_svg(item, offset, export_dir, filepath)
        if isinstance(item, VTextItem):
            return CorelSvgExporter._text_to_svg(item, offset)
        if isinstance(item, (VImageTextConnector, VGenericConnector)):
            if id(item) in exported_connectors:
                return []
            path_line = CorelSvgExporter._connector_to_svg(item, offset)
            exported_connectors.add(id(item))
            return [path_line] if path_line else []
        return []

    @staticmethod
    def _content_rect(scene):
        rect = QRectF()
        for item in scene.items():
            if not CorelSvgExporter._is_export_item(item):
                continue
            item_rect = CorelSvgExporter._scene_rect(item)
            rect = item_rect if rect.isNull() else rect.united(item_rect)
        return rect

    @staticmethod
    def _scene_rect(item):
        """Return bounds of the pixels that will actually be written to SVG."""
        if isinstance(item, VTextItem):
            rect = QRectF()
            for child in item.childItems():
                if not isinstance(child, QGraphicsSimpleTextItem) or not child.isVisible():
                    continue
                child_rect = child.sceneTransform().mapRect(child.boundingRect())
                rect = child_rect if rect.isNull() else rect.united(child_rect)
            return rect
        return item.sceneTransform().mapRect(item.boundingRect())

    @staticmethod
    def _image_to_svg(item, offset, export_dir, svg_path):
        path = item.file_path
        if not os.path.isabs(path):
            path = os.path.abspath(path)

        href = path.replace('\\', '/')
        try:
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)
            pix = QPixmap(path)
            if not pix.isNull():
                safe_name = f"image_{abs(hash((path, item.scenePos().x(), item.scenePos().y())))}.png"
                out_path = os.path.join(export_dir, safe_name)
                pix.save(out_path, "PNG")
                href = os.path.relpath(out_path, os.path.dirname(svg_path)).replace('\\', '/')
        except Exception:
            href = path.replace('\\', '/')

        # 关键：rect 是 item 局部坐标的 (0,0,w,h)，SVG <image> 在没有 x/y 时
        # 默认以 (0,0) 为锚点绘制，再延伸到 width/height。
        # sceneTransform 的 dx/dy 表示"局部 (0,0) 映射到的场景坐标"，
        # SVG 的 matrix(a,b,c,d,e,f) 中 (e,f) 也是相对锚点的最终平移，
        # 二者语义一致，所以直接把 sceneTransform 的 dx/dy 扣掉偏移即可。
        # offset 这里传的是画布 scene_rect 的左上角，
        # 让场景坐标直接对齐到 SVG viewBox。
        # 不要再叠加 x/y，否则旋转/缩放会以 (0,0) 为锚点，导致图片跑到错位置。
        rect = item.boundingRect()
        transform = item.sceneTransform()
        m11 = transform.m11()
        m12 = transform.m12()
        m21 = transform.m21()
        m22 = transform.m22()
        dx = transform.dx() - offset.x()
        dy = transform.dy() - offset.y()
        matrix = ' '.join(
            CorelSvgExporter._num(value)
            for value in (m11, m12, m21, m22, dx, dy)
        )
        opacity = getattr(item, 'image_opacity', 1.0)
        attrs = [
            f'width="{CorelSvgExporter._num(rect.width())}"',
            f'height="{CorelSvgExporter._num(rect.height())}"',
            f'transform="matrix({matrix})"',
            'preserveAspectRatio="none"',
            f'href="{CorelSvgExporter._attr(href)}"',
            f'xlink:href="{CorelSvgExporter._attr(href)}"',
        ]
        if opacity < 1.0:
            attrs.append(f'opacity="{CorelSvgExporter._num(opacity)}"')
        return [f'<image {" ".join(attrs)}/>']

    @staticmethod
    def _text_to_svg(item, offset):
        lines = []
        children = [
            child for child in item.childItems()
            if isinstance(child, QGraphicsSimpleTextItem) and child.isVisible()
        ]
        children.sort(key=lambda c: (c.pos().x(), c.pos().y()))

        lines.append('<g class="editable-text">')
        for child in children:
            text = child.text()
            if not text:
                continue
            glyph_path = QPainterPath()
            # QGraphicsSimpleTextItem positions its text box from y=0, but
            # QPainterPath.addText() treats the supplied y as the baseline.
            # Use the same ascent that Qt uses for the item so exported glyphs
            # occupy the same scene coordinates as they do on the canvas.
            baseline_y = QFontMetrics(child.font()).ascent()
            glyph_path.addText(QPointF(0, baseline_y), child.font(), text)
            # 先将字形转换到场景坐标，再在场景坐标中减去 SVG 视图偏移。
            # 不能把 offset 直接加入 sceneTransform，否则旋转字符会把
            # 偏移量一起旋转，造成文字与图片的相对位置发生变化。
            glyph_path = child.sceneTransform().map(glyph_path)
            glyph_path.translate(-offset.x(), -offset.y())
            d = CorelSvgExporter._path_data(glyph_path, QPointF(0, 0))
            if not d:
                continue
            fill = child.brush().color().name()
            lines.append(
                f'<path d="{CorelSvgExporter._attr(d)}" '
                f'fill="{CorelSvgExporter._attr(fill)}" stroke="none"/>'
            )
        lines.append('</g>')
        return lines

    @staticmethod
    def _connector_to_svg(item, offset):
        # item.path() 返回 item 局部坐标，必须先 mapToScene
        # 再扣 offset，否则连线端点与它连接的图片/文字完全对不上。
        path = item.mapToScene(item.path())
        d = CorelSvgExporter._path_data(path, offset)
        if not d:
            return ''
        pen = item.pen()
        color = pen.color()
        opacity = color.alphaF()
        attrs = [
            f'd="{CorelSvgExporter._attr(d)}"',
            'fill="none"',
            f'stroke="{CorelSvgExporter._attr(color.name())}"',
            f'stroke-width="{CorelSvgExporter._num(pen.widthF() or pen.width() or 1)}"',
            'stroke-linecap="round"',
            'stroke-linejoin="round"',
        ]
        if opacity < 1.0:
            attrs.append(f'stroke-opacity="{CorelSvgExporter._num(opacity)}"')
        if pen.style() == Qt.PenStyle.DashLine:
            attrs.append('stroke-dasharray="8 5"')
        return f'<path {" ".join(attrs)}/>'

    @staticmethod
    def _path_data(path, offset):
        data = []
        i = 0
        while i < path.elementCount():
            e = path.elementAt(i)
            x = e.x - offset.x()
            y = e.y - offset.y()
            if e.type == QPainterPath.ElementType.MoveToElement:
                data.append(f'M {CorelSvgExporter._num(x)} {CorelSvgExporter._num(y)}')
            elif e.type == QPainterPath.ElementType.LineToElement:
                data.append(f'L {CorelSvgExporter._num(x)} {CorelSvgExporter._num(y)}')
            elif e.type == QPainterPath.ElementType.CurveToElement and i + 2 < path.elementCount():
                c1 = e
                c2 = path.elementAt(i + 1)
                end = path.elementAt(i + 2)
                data.append(
                    'C '
                    f'{CorelSvgExporter._num(c1.x - offset.x())} {CorelSvgExporter._num(c1.y - offset.y())} '
                    f'{CorelSvgExporter._num(c2.x - offset.x())} {CorelSvgExporter._num(c2.y - offset.y())} '
                    f'{CorelSvgExporter._num(end.x - offset.x())} {CorelSvgExporter._num(end.y - offset.y())}'
                )
                i += 2
            i += 1
        return ' '.join(data)

    @staticmethod
    def _num(value):
        value = float(value)
        if abs(value) < 0.0001:
            value = 0.0
        return f'{value:.3f}'.rstrip('0').rstrip('.')

    @staticmethod
    def _px_to_mm(value):
        return float(value) * 25.4 / CORELDRAW_EXPORT_DPI

    @staticmethod
    def _attr(value):
        return xml_escape(str(value), {'"': '&quot;'})

    @staticmethod
    def _text(value):
        return xml_escape(str(value))

# --- Canvas & Scene ---

class GuideItem(QGraphicsItem):
    """辅助线（水平或垂直），可拖动，双击删除"""
    RULER_SIZE = 20  # 标尺宽度（像素，视图坐标）

    def __init__(self, orientation, pos_value, scene_rect):
        super().__init__()
        # orientation: Qt.Orientation.Horizontal | Vertical
        self.orientation = orientation
        self.pos_value = pos_value      # 场景坐标中的位置值
        self.scene_rect = scene_rect    # 用于计算绘制范围
        self._dragging = False

        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIgnoresTransformations, False)
        self.setAcceptHoverEvents(True)
        self.setZValue(1000)  # 始终在最上层

        self._hovered = False
        self._selected_for_edit = False
        self._update_pos()

    def _update_pos(self):
        r = self.scene_rect
        if self.orientation == Qt.Orientation.Horizontal:
            self.setPos(r.left(), self.pos_value)
        else:
            self.setPos(self.pos_value, r.top())

    def boundingRect(self):
        r = self.scene_rect
        if self.orientation == Qt.Orientation.Horizontal:
            return QRectF(0, -4, r.width(), 8)
        else:
            return QRectF(-4, 0, 8, r.height())

    def paint(self, painter, option, widget):
        if getattr(self, "_selected_for_edit", False):
            color = QColor(255, 165, 0, 255)
        else:
            color = QColor(0, 210, 255, 255) if self._hovered else QColor(0, 180, 255, 220)
        pen = QPen(color, 2, Qt.PenStyle.SolidLine)
        painter.setPen(pen)
        r = self.scene_rect
        if self.orientation == Qt.Orientation.Horizontal:
            painter.drawLine(QPointF(0, 0), QPointF(r.width(), 0))
        else:
            painter.drawLine(QPointF(0, 0), QPointF(0, r.height()))

    def hoverEnterEvent(self, event):
        self._hovered = True
        if self.orientation == Qt.Orientation.Horizontal:
            self.setCursor(Qt.CursorShape.SizeVerCursor)
        else:
            self.setCursor(Qt.CursorShape.SizeHorCursor)
        self.update()
        super().hoverEnterEvent(event)

    def hoverLeaveEvent(self, event):
        self._hovered = False
        self.update()
        super().hoverLeaveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            if self.scene():
                self.scene().select_guide(self)
            self._dragging = True
            self._drag_start = event.scenePos()
            self._start_value = self.pos_value
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._dragging:
            delta = event.scenePos() - self._drag_start
            if self.orientation == Qt.Orientation.Horizontal:
                self.pos_value = self._start_value + delta.y()
            else:
                self.pos_value = self._start_value + delta.x()
            self._update_pos()
            self.update()
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._dragging = False
        event.accept()

    def mouseDoubleClickEvent(self, event):
        """双击删除辅助线"""
        if self.scene():
            self.scene().remove_guide(self)
        event.accept()

    def contextMenuEvent(self, event):
        if self.scene():
            self.scene().select_guide(self)
        menu = QMenu()
        menu.addAction("删除此辅助线").triggered.connect(
            lambda: self.scene().remove_guide(self) if self.scene() else None
        )
        menu.exec(event.screenPos())


class NavigatorWidget(QWidget):
    """导航器：显示画布全景缩略图，支持点击跳转和拖动视口框"""

    NAV_W = 220
    NAV_H = 160

    def __init__(self, view, scene, parent=None):
        super().__init__(parent)
        self.view = view
        self.scene = scene
        self.setFixedSize(self.NAV_W, self.NAV_H)
        self.setWindowTitle("导航器")

        self._thumb = QPixmap()
        self._scale = 1.0
        self._scene_rect = QRectF()
        self._dragging = False
        self._drag_offset = QPointF()
        self._dirty = True

        self._refresh_timer = QTimer(self)
        self._refresh_timer.setInterval(400)
        self._refresh_timer.timeout.connect(self._do_refresh)
        self._refresh_timer.start()

        self.view.transformChanged.connect(self.update)
        self.view.horizontalScrollBar().valueChanged.connect(self.update)
        self.view.verticalScrollBar().valueChanged.connect(self.update)
        self.scene.changed.connect(self._mark_dirty)

        self.setMouseTracking(True)
        self.setCursor(Qt.CursorShape.CrossCursor)

    def set_document(self, view, scene):
        """切换导航器所跟踪的文档。"""
        if self.view is view and self.scene is scene:
            return
        try:
            self.view.transformChanged.disconnect(self.update)
            self.view.horizontalScrollBar().valueChanged.disconnect(self.update)
            self.view.verticalScrollBar().valueChanged.disconnect(self.update)
            self.scene.changed.disconnect(self._mark_dirty)
        except (RuntimeError, TypeError):
            pass
        self.view = view
        self.scene = scene
        self._thumb = QPixmap()
        self._scene_rect = QRectF()
        self._dirty = True
        self.view.transformChanged.connect(self.update)
        self.view.horizontalScrollBar().valueChanged.connect(self.update)
        self.view.verticalScrollBar().valueChanged.connect(self.update)
        self.scene.changed.connect(self._mark_dirty)
        self.update()

    def _mark_dirty(self):
        self._dirty = True

    def _do_refresh(self):
        if not self._dirty:
            return
        self._dirty = False
        self._render_thumb()
        self.update()

    def _render_thumb(self):
        sr = self.scene.sceneRect()
        if sr.isEmpty():
            return
        self._scene_rect = sr
        sx = self.NAV_W / sr.width()
        sy = self.NAV_H / sr.height()
        self._scale = min(sx, sy)
        tw = max(1, int(sr.width()  * self._scale))
        th = max(1, int(sr.height() * self._scale))
        self._thumb = QPixmap(tw, th)
        self._thumb.fill(Qt.GlobalColor.white)
        painter = QPainter(self._thumb)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        # 用标志位通知 drawBackground 跳过网格/辅助线，避免修改场景状态
        self.scene._rendering_thumb = True
        self.scene.render(painter, QRectF(0, 0, tw, th), sr)
        self.scene._rendering_thumb = False
        painter.end()

    def _thumb_offset(self):
        return QPoint((self.NAV_W - self._thumb.width())  // 2,
                      (self.NAV_H - self._thumb.height()) // 2)

    def _viewport_rect_in_nav(self, ox, oy):
        if self._scale <= 0 or self._scene_rect.isEmpty():
            return QRectF()
        vp = self.view.viewport().rect()
        tl = self.view.mapToScene(vp.topLeft())
        br = self.view.mapToScene(vp.bottomRight())
        sr = self._scene_rect
        x = ox + (tl.x() - sr.left()) * self._scale
        y = oy + (tl.y() - sr.top())  * self._scale
        w = (br.x() - tl.x()) * self._scale
        h = (br.y() - tl.y()) * self._scale
        return QRectF(x, y, w, h).normalized()

    def _nav_to_scene(self, nav_pos):
        if self._scale <= 0 or self._scene_rect.isEmpty():
            return QPointF()
        off = self._thumb_offset()
        sx = self._scene_rect.left() + (nav_pos.x() - off.x()) / self._scale
        sy = self._scene_rect.top()  + (nav_pos.y() - off.y()) / self._scale
        return QPointF(sx, sy)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.fillRect(self.rect(), QColor(40, 40, 44))
        if self._thumb.isNull():
            self._render_thumb()
        if not self._thumb.isNull():
            off = self._thumb_offset()
            painter.drawPixmap(off.x(), off.y(), self._thumb)
            vp_rect = self._viewport_rect_in_nav(off.x(), off.y())
            if vp_rect.isValid():
                painter.setPen(QPen(QColor(0, 180, 255), 2))
                painter.setBrush(QBrush(QColor(0, 180, 255, 40)))
                painter.drawRect(vp_rect)
        painter.setPen(QPen(QColor(80, 80, 85), 1))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRect(self.rect().adjusted(0, 0, -1, -1))

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            off = self._thumb_offset()
            vp_rect = self._viewport_rect_in_nav(off.x(), off.y())
            if vp_rect.contains(QPointF(event.pos())):
                self._dragging = True
                self._drag_offset = QPointF(event.pos()) - vp_rect.center()
            else:
                self.view.centerOn(self._nav_to_scene(event.pos()))
            self.update()

    def mouseMoveEvent(self, event):
        if self._dragging:
            target = QPointF(event.pos()) - self._drag_offset
            self.view.centerOn(self._nav_to_scene(target))
            self.update()

    def mouseReleaseEvent(self, event):
        self._dragging = False

    def force_refresh(self):
        self._dirty = True
        self._do_refresh()


class LayoutScene(QGraphicsScene):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setBackgroundBrush(QBrush(QColor(45, 45, 48))) 
        self.grid_pen = QPen(QColor(220, 220, 220, 100))
        self.binding_source = None
        self.connectors = []
        self.image_text_connectors = []  
        self.show_grid = True  
        self.show_connectors = True  
        self.show_image_text_connectors = True  
        self.undo_stack = UndoStack()  
        self._horizontal_move_lock_y = {}
        self._horizontal_move_lock_x = {}
        self._horizontal_move_lock_owner = None
        self._horizontal_move_owner_right_x = None
        self._image_right_edge_snap_indicator = None
        self._image_top_edge_snap_indicator = None  # 顶部吸附指示器
        self.clipboard_items = []  
        self.clipboard_image_text_connections = []  
        self.show_connection_points = True  
        self.connection_mode = False  
        self.connection_source_point = None  
        self.asset_manager = AssetManager()
        self.config_manager = ConfigManager()  # 配置管理器
        self.image_text_binding_mode = False  
        self.image_text_source = None
        self.selection_order = []  # 记录选中顺序
        self.last_selection_by_marquee = False  # 最后一次选中是否为框选
        self.background_pixmap = None  # 背景图片缓存
        self.guides = []          # 辅助线列表
        self._selected_guide = None
        self.show_guides = True   # 辅助线显示开关
        self.snap_threshold = self.config_manager.get('snap_threshold', 20)  # 辅助线吸附距离（场景像素）
        self._temp_alignment_guide = None  # 临时对齐辅助线
        self.free_connection_points = []  # 画布上的独立连接点
        self.free_connection_point_mode = 'connect'  # 独立连接点模式：move / connect
        self._batch_importing = False  # 批量导入标志，禁止显示临时对齐线
        self.resize_mode = False  # 图片调整大小模式
        self.stamping_session = None  # 盖章式批量复制会话
        self.image_manage_mode = False  # 图片管理模式（Alt+, 切换）
        self.align_reference_mode = None  # 对齐基准点选模式
        self.align_reference_candidates = []
        self._pending_selection_click_item = None
        self._editing_group_asset_id = None  # 正在编辑的组合素材ID
        self._editing_group_items = []       # 正在编辑的组合元素列表
        
        # 连接选择改变信号
        self.selectionChanged.connect(self.on_selection_changed_track)
        
        # 加载背景图片
        self.load_background_image()

         # --- 新增：初始化辅助线管理器 ---
        #self.guides = guides_manager.GuidesManager(self)

    def load_background_image(self):
        """加载背景图片"""
        bg_path = self.config_manager.get('default_background_image', '')
        if bg_path and os.path.exists(bg_path):
            self.background_pixmap = QPixmap(bg_path)
            if self.background_pixmap.isNull():
                print(f"无法加载背景图片: {bg_path}")
                self.background_pixmap = None
            else:
                print(f"背景图片已加载: {bg_path}")
        else:
            self.background_pixmap = None
    
    def set_background_image(self, image_path):
        """设置背景图片"""
        if image_path and os.path.exists(image_path):
            self.config_manager.set('default_background_image', image_path)
            self.load_background_image()
            self.update()
            return True
        else:
            # 清除背景图片
            self.config_manager.set('default_background_image', '')
            self.background_pixmap = None
            self.update()
            return True

    # --- 辅助线管理 ---
    def add_guide(self, orientation, pos_value):
        """添加一条辅助线"""
        guide = GuideItem(orientation, pos_value, self.sceneRect())
        self.addItem(guide)
        self.guides.append(guide)
        guide.setVisible(self.show_guides)
        return guide

    def add_free_connection_point(self, scene_pos):
        """在画布上创建一个不依附任何元素的连接点。"""
        point = ConnectionPoint(None, "custom")
        self.addItem(point)
        point.setPos(scene_pos)
        point.setVisible(self.show_connection_points)
        self.free_connection_points.append(point)
        return point

    def add_free_connection_point_at_nearest_guide_intersection(self, scene_pos):
        """在右键位置最近的可见横、竖辅助线交叉点创建独立连接点。"""
        vertical_guides = [
            guide for guide in self.guides
            if guide.isVisible() and guide.orientation == Qt.Orientation.Vertical
        ]
        horizontal_guides = [
            guide for guide in self.guides
            if guide.isVisible() and guide.orientation == Qt.Orientation.Horizontal
        ]
        if not vertical_guides or not horizontal_guides:
            return None

        nearest_vertical = min(
            vertical_guides,
            key=lambda guide: abs(guide.pos_value - scene_pos.x())
        )
        nearest_horizontal = min(
            horizontal_guides,
            key=lambda guide: abs(guide.pos_value - scene_pos.y())
        )
        return self.add_free_connection_point(
            QPointF(nearest_vertical.pos_value, nearest_horizontal.pos_value)
        )

    def set_free_connection_point_mode(self, mode, cancel_pending_connection=True):
        """切换独立连接点的移动/连接模式。"""
        if mode not in ('move', 'connect'):
            return
        if mode == 'move' and cancel_pending_connection and self.connection_mode:
            self.cancel_connection_mode()
        self.free_connection_point_mode = mode
        label = '移动模式' if mode == 'move' else '连接模式'
        self._show_status_message(f'独立连接点：{label}', 2500)

    def add_startup_horizontal_guides(self):
        """按配置添加启动横向辅助线，列表中的每项为场景Y坐标。"""
        if self.guides:
            return
        rect = self.sceneRect()
        if rect.isEmpty():
            return
        positions = self.config_manager.get('startup_horizontal_guides', [])
        for value in positions:
            try:
                y = float(value)
            except (TypeError, ValueError):
                continue
            if rect.top() <= y <= rect.bottom():
                self.add_guide(Qt.Orientation.Horizontal, y)

    def select_guide(self, guide):
        """高亮当前操作的辅助线，并显示删除提示。"""
        if guide not in self.guides:
            return
        if self._selected_guide is not guide:
            if self._selected_guide:
                self._selected_guide._selected_for_edit = False
                self._selected_guide.update()
            self._selected_guide = guide
            guide._selected_for_edit = True
            guide.update()

        main_window = self.parent()
        if main_window and hasattr(main_window, "status_bar"):
            main_window.status_bar.showMessage("已选中辅助线，双击删除这条辅助线")

    def clear_guide_selection(self):
        """取消辅助线高亮和删除提示。"""
        if not self._selected_guide:
            return
        self._selected_guide._selected_for_edit = False
        self._selected_guide.update()
        self._selected_guide = None
        main_window = self.parent()
        if main_window and hasattr(main_window, "status_bar"):
            main_window.status_bar.clearMessage()

    def remove_guide(self, guide):
        """删除指定辅助线"""
        if guide in self.guides:
            if guide is self._selected_guide:
                self.clear_guide_selection()
            self.removeItem(guide)
            self.guides.remove(guide)

    def clear_guides(self):
        """清除所有辅助线"""
        self.clear_guide_selection()
        for g in self.guides[:]:
            self.removeItem(g)
        self.guides.clear()

    def set_guides_visible(self, visible):
        """显示/隐藏所有辅助线"""
        self.show_guides = visible
        for g in self.guides:
            g.setVisible(visible)

    def show_temp_alignment_guide(self, orientation, pos_value):
        """显示临时对齐辅助线（拖动时使用）"""
        # 如果已存在临时辅助线，更新位置
        if self._temp_alignment_guide is not None:
            self._temp_alignment_guide.pos_value = pos_value
            self._temp_alignment_guide.orientation = orientation
            self._temp_alignment_guide.update()
        else:
            # 创建新的临时辅助线
            self._temp_alignment_guide = GuideItem(orientation, pos_value, self.sceneRect())
            # 设置临时辅助线样式（可以设置不同颜色或虚线样式）
            self._temp_alignment_guide._is_temp = True
            self.addItem(self._temp_alignment_guide)
            self._temp_alignment_guide.setVisible(True)
            self._temp_alignment_guide.setZValue(100)  # 确保在最上层显示

    def hide_temp_alignment_guide(self):
        """隐藏临时对齐辅助线"""
        if self._temp_alignment_guide is not None:
            self.removeItem(self._temp_alignment_guide)
            self._temp_alignment_guide = None

    def set_resize_mode(self, enabled):
        """切换图片调整大小模式"""
        self.resize_mode = enabled
        if enabled:
            # 进入调整模式时记录所有选中图片的原始状态
            self._resize_snapshot = {}
            for item in self.items():
                if isinstance(item, VImageItem) and item.isSelected():
                    self._resize_snapshot[item] = {
                        'rect': QRectF(item._rect),
                        'pos': item.scenePos(),
                        'width': item.target_width,
                    }
                    item._show_handles(True)
        else:
            # 退出时清除记录
            self._resize_snapshot = {}
            for item in self.items():
                if isinstance(item, VImageItem):
                    item._show_handles(False)

    def confirm_resize(self):
        """确认调整：清除记录，退出模式"""
        self._resize_snapshot = {}
        self.set_resize_mode(False)
        if self.views():
            mw = self.views()[0].window()
            if hasattr(mw, 'btn_resize'):
                mw.btn_resize.setChecked(False)
            if hasattr(mw, 'status_bar'):
                mw.status_bar.showMessage("调整已确认", 3000)

    def cancel_resize(self):
        """取消调整：还原所有图片到进入调整模式前的状态"""
        snapshot = getattr(self, '_resize_snapshot', {})
        for item, state in snapshot.items():
            if item.scene() == self:
                # 还原尺寸
                item._apply_resize(state['rect'], state['pos'])
                # 还原位置
                if item.parentItem():
                    item.setPos(item.parentItem().mapFromScene(state['pos']))
                else:
                    item.setPos(state['pos'])
        self._resize_snapshot = {}
        self.set_resize_mode(False)
        if self.views():
            mw = self.views()[0].window()
            if hasattr(mw, 'btn_resize'):
                mw.btn_resize.setChecked(False)
            if hasattr(mw, 'status_bar'):
                mw.status_bar.showMessage("调整已取消，已还原原始大小和位置", 3000)

    def toggle_image_manage_mode(self):
        """切换图片管理模式（Alt+,）：隐藏的图片显示为半透明虚框，点击切换显示/隐藏"""
        self.image_manage_mode = not self.image_manage_mode

        for item in self.items():
            if isinstance(item, VImageItem):
                if self.image_manage_mode:
                    if not item.isVisible():
                        item._was_hidden = True
                        # 临时显示，设为半透明，禁止选中和移动避免误操作
                        item.setVisible(True)
                        if hasattr(item, 'p_item') and item.p_item:
                            item.p_item.setOpacity(0.01)  # 几乎透明，让遮罩层负责视觉
                        item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
                        item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
                    else:
                        item._was_hidden = False
                        item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
                        item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
                else:
                    # 退出管理模式：恢复原始状态
                    was_hidden = getattr(item, '_was_hidden', False)
                    item.setVisible(not was_hidden)
                    if hasattr(item, 'p_item') and item.p_item:
                        item.p_item.setOpacity(item.image_opacity)
                    item._was_hidden = False
                    # 恢复交互标志
                    item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, not item.locked)
                    item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, not item.locked)

        self.update()
        if self.views():
            mw = self.views()[0].window()
            if self.image_manage_mode:
                if hasattr(mw, 'status_bar'):
                    mw.status_bar.showMessage('图片管理模式：点击图片切换显示/隐藏，再按 Alt+, 退出', 0)
                self.views()[0].setCursor(Qt.CursorShape.PointingHandCursor)
            else:
                if hasattr(mw, 'status_bar'):
                    mw.status_bar.showMessage('已退出图片管理模式', 3000)
                self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)

    def drawForeground(self, painter, rect):
        """图片管理模式下，给原本隐藏的图片画虚线边框和标注"""
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        if self.image_manage_mode:
            for item in self.items():
                if isinstance(item, VImageItem) and getattr(item, '_was_hidden', False):
                    scene_rect = item.mapRectToScene(item.boundingRect())
                    # 半透明蓝色遮罩
                    painter.setOpacity(0.25)
                    painter.fillRect(scene_rect, QColor(0, 120, 215))
                    painter.setOpacity(1.0)
                    # 虚线边框
                    pen = QPen(QColor(0, 120, 215), 2, Qt.PenStyle.DashLine)
                    pen.setCosmetic(True)
                    painter.setPen(pen)
                    painter.setBrush(Qt.BrushStyle.NoBrush)
                    painter.drawRect(scene_rect)
                    # 文件名提示
                    painter.setPen(QColor(255, 255, 255))
                    font = QFont('Arial', 9)
                    font.setBold(True)
                    painter.setFont(font)
                    name = os.path.basename(item.file_path)
                    painter.drawText(scene_rect, Qt.AlignmentFlag.AlignCenter, f'点击显示\n{name}')

        indicator = self._image_right_edge_snap_indicator
        if indicator and not getattr(self, '_rendering_pdf', False):
            x, top, bottom, moving_top, target_top = indicator
            scale = abs(painter.worldTransform().m11()) or 1.0
            padding = 8.0 / scale
            radius = 5.0 / scale
            color = QColor(255, 45, 100)
            pen = QPen(color, 3, Qt.PenStyle.SolidLine)
            pen.setCosmetic(True)
            painter.setOpacity(1.0)
            painter.setPen(pen)
            painter.setBrush(QBrush(color))
            painter.drawLine(QPointF(x, top - padding), QPointF(x, bottom + padding))
            painter.drawEllipse(QPointF(x, moving_top), radius, radius)
            painter.drawEllipse(QPointF(x, target_top), radius, radius)

        # 绘制顶部Y吸附指示线
        top_indicator = self._image_top_edge_snap_indicator
        if top_indicator and not getattr(self, '_rendering_pdf', False):
            y, left, right, moving_left, target_left = top_indicator
            scale = abs(painter.worldTransform().m11()) or 1.0
            padding = 8.0 / scale
            radius = 5.0 / scale
            color = QColor(50, 200, 255)  # 蓝色，与右边缘的红色区分
            pen = QPen(color, 3, Qt.PenStyle.SolidLine)
            pen.setCosmetic(True)
            painter.setOpacity(1.0)
            painter.setPen(pen)
            painter.setBrush(QBrush(color))
            painter.drawLine(QPointF(left - padding, y), QPointF(right + padding, y))
            painter.drawEllipse(QPointF(moving_left, y), radius, radius)
            painter.drawEllipse(QPointF(target_left, y), radius, radius)

    def drawBackground(self, painter, rect):
        # PDF导出时：直接填白色，跳过灰色外框和网格
        if getattr(self, '_rendering_pdf', False):
            painter.fillRect(rect, QColor(255, 255, 255))
            # 仍然绘制背景图片
            if self.background_pixmap and not self.background_pixmap.isNull():
                opacity = self.config_manager.get('background_opacity', 1.0)
                scale_mode = self.config_manager.get('background_scale_mode', 'fit')
                canvas_rect = self.sceneRect()
                old_opacity = painter.opacity()
                painter.setOpacity(opacity)
                if scale_mode == 'fit':
                    scaled = self.background_pixmap.scaled(
                        canvas_rect.size().toSize(),
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation)
                    x = canvas_rect.x() + (canvas_rect.width() - scaled.width()) / 2
                    y = canvas_rect.y() + (canvas_rect.height() - scaled.height()) / 2
                    painter.drawPixmap(int(x), int(y), scaled)
                elif scale_mode == 'fill':
                    scaled = self.background_pixmap.scaled(
                        canvas_rect.size().toSize(),
                        Qt.AspectRatioMode.KeepAspectRatioByExpanding,
                        Qt.TransformationMode.SmoothTransformation)
                    x = canvas_rect.x() + (canvas_rect.width() - scaled.width()) / 2
                    y = canvas_rect.y() + (canvas_rect.height() - scaled.height()) / 2
                    painter.drawPixmap(int(x), int(y), scaled)
                elif scale_mode == 'stretch':
                    scaled = self.background_pixmap.scaled(
                        canvas_rect.size().toSize(),
                        Qt.AspectRatioMode.IgnoreAspectRatio,
                        Qt.TransformationMode.SmoothTransformation)
                    painter.drawPixmap(canvas_rect.toRect(), scaled)
                elif scale_mode == 'tile':
                    painter.drawTiledPixmap(canvas_rect.toRect(), self.background_pixmap)
                painter.setOpacity(old_opacity)
            return

        # 绘制外部背景
        painter.fillRect(rect, QColor(60, 60, 60))
        
        canvas_rect = self.sceneRect()
        
        # PDF导出时跳过灰色编辑器背景、阴影、画布底色（由PDF导出代码自己填白色）
        is_pdf_export = getattr(self, '_rendering_pdf', False)
        if not is_pdf_export:
            # 绘制阴影
            shadow_rect = canvas_rect.translated(5, 5)
            painter.fillRect(shadow_rect, QColor(30, 30, 30, 150))
            # 绘制画布背景色
            painter.fillRect(canvas_rect, QColor(250, 250, 245))
        
        # 缩略图渲染时跳过网格和辅助线
        is_thumb = getattr(self, '_rendering_thumb', False)

        # 绘制网格（在背景图片之前）
        if self.show_grid and not is_thumb:
            painter.setPen(self.grid_pen)
            c_left = int(canvas_rect.left())
            c_right = int(canvas_rect.right())
            c_top = int(canvas_rect.top())
            c_bottom = int(canvas_rect.bottom())
            step = 50
            for x in range(c_left, c_right + 1, step):
                painter.drawLine(x, c_top, x, c_bottom)
            for y in range(c_top, c_bottom + 1, step):
                painter.drawLine(c_left, y, c_right, y)
        
        # --- 新增：手动绘制连线层（如果开启了背景图片覆盖连线功能） ---
        if self.config_manager.get('bg_above_connectors', False):
            # 绘制父子连线
            for conn in self.connectors:
                if conn.isVisible():
                    painter.save()
                    painter.setPen(conn.pen())
                    painter.drawPath(conn.path())
                    painter.restore()
            
            # 绘制图文连线
            for conn in self.image_text_connectors:
                if conn.isVisible():
                    painter.save()
                    # VGenericConnector 可能需要根据选中状态更新 pen
                    # 这里简化处理，直接绘制
                    painter.setPen(conn.pen())
                    painter.drawPath(conn.path())
                    painter.restore()

        # 绘制背景图片（在网格之上，元素之下）
        if self.background_pixmap and not self.background_pixmap.isNull():
            opacity = self.config_manager.get('background_opacity', 0.3)
            scale_mode = self.config_manager.get('background_scale_mode', 'fit')
            
            # 保存当前透明度
            old_opacity = painter.opacity()
            painter.setOpacity(opacity)
            
            if scale_mode == 'fit':
                # 适应画布，保持宽高比
                scaled_pixmap = self.background_pixmap.scaled(
                    canvas_rect.size().toSize(),
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation
                )
                # 居中绘制
                x = canvas_rect.x() + (canvas_rect.width() - scaled_pixmap.width()) / 2
                y = canvas_rect.y() + (canvas_rect.height() - scaled_pixmap.height()) / 2
                painter.drawPixmap(int(x), int(y), scaled_pixmap)
            
            elif scale_mode == 'fill':
                # 填充画布，保持宽高比，可能裁剪
                scaled_pixmap = self.background_pixmap.scaled(
                    canvas_rect.size().toSize(),
                    Qt.AspectRatioMode.KeepAspectRatioByExpanding,
                    Qt.TransformationMode.SmoothTransformation
                )
                x = canvas_rect.x() + (canvas_rect.width() - scaled_pixmap.width()) / 2
                y = canvas_rect.y() + (canvas_rect.height() - scaled_pixmap.height()) / 2
                painter.drawPixmap(int(x), int(y), scaled_pixmap)
            
            elif scale_mode == 'stretch':
                # 拉伸填充，不保持宽高比
                scaled_pixmap = self.background_pixmap.scaled(
                    canvas_rect.size().toSize(),
                    Qt.AspectRatioMode.IgnoreAspectRatio,
                    Qt.TransformationMode.SmoothTransformation
                )
                painter.drawPixmap(canvas_rect.toRect(), scaled_pixmap)
            
            elif scale_mode == 'tile':
                # 平铺
                painter.drawTiledPixmap(canvas_rect.toRect(), self.background_pixmap)
            
            # 恢复透明度
            painter.setOpacity(old_opacity)
        
        # 绘制画布边框（最后绘制）
        painter.setPen(QPen(QColor(180, 180, 180), 1))
        painter.drawRect(canvas_rect)
    
    def on_selection_changed_track(self):
        """追踪选中顺序：支持点击、框选、Ctrl+A 等所有选中方式"""
        current_selected = set(
            item for item in self.selectedItems()
            if isinstance(item, (VImageItem, VTextItem))
        )
        previous_set = set(self.selection_order)

        # 移除已取消选中的元素
        self.selection_order = [item for item in self.selection_order if item in current_selected]

        newly_selected = current_selected - previous_set
        if newly_selected:
            clicked_item = self._pending_selection_click_item
            # 单击/Ctrl加选：标记为非框选
            if clicked_item in newly_selected and len(newly_selected) == 1:
                self.last_selection_by_marquee = False
                self.selection_order.append(clicked_item)
            else:
                # 框选/Ctrl+A 等批量选中：由 LayoutView 的实时扫描顺序决定，这里只补充未被扫到的元素
                if clicked_item in newly_selected:
                    sorted_new = [clicked_item] + [i for i in newly_selected if i != clicked_item]
                else:
                    sorted_new = list(newly_selected)
                self.selection_order.extend(sorted_new)
            
    def start_binding_mode(self, item):
        self.binding_source = item
        views = self.views()
        if views:
            views[0].setCursor(Qt.CursorShape.CrossCursor)
        print("Select parent for binding...")

    def contextMenuEvent(self, event):
        """右键菜单：盖章或对齐基准选择过程中屏蔽菜单"""
        if self.stamping_session or self.align_reference_mode:
            event.accept()
            return

        item = self.itemAt(event.scenePos(), QTransform())
        
        # 如果右键点击的是普通物体（文字/图片），调用 BaseElement 的右键菜单
        if item and isinstance(item, BaseElement):
            super().contextMenuEvent(event)

    def mousePressEvent(self, event):
        """鼠标按下事件：处理元素点击、父子绑定、对齐基准选择和盖章会话"""
        # 图片管理模式：点击图片切换显示/隐藏
        if self.image_manage_mode and event.button() == Qt.MouseButton.LeftButton:
            click_pos = event.scenePos()
            # 手动遍历所有图片，检查点击位置是否在包围盒内
            clicked_img = None
            for item in self.items():
                if isinstance(item, VImageItem):
                    scene_rect = item.mapRectToScene(item.boundingRect())
                    if scene_rect.contains(click_pos):
                        clicked_img = item
                        break
            if clicked_img:
                if getattr(clicked_img, '_was_hidden', False):
                    # 原来是隐藏的，现在显示它
                    clicked_img._was_hidden = False
                    if hasattr(clicked_img, 'p_item') and clicked_img.p_item:
                        clicked_img.p_item.setOpacity(clicked_img.image_opacity)
                else:
                    # 原来是显示的，现在隐藏它
                    clicked_img._was_hidden = True
                    if hasattr(clicked_img, 'p_item') and clicked_img.p_item:
                        clicked_img.p_item.setOpacity(0.01)
                self.update()
                event.accept()
                return

        if event.button() == Qt.MouseButton.LeftButton:
            raw = self.itemAt(event.scenePos(), QTransform())
            clicked_item = raw
            while clicked_item is not None and not isinstance(clicked_item, BaseElement):
                clicked_item = clicked_item.parentItem()
            self._pending_selection_click_item = clicked_item

        # 盖章模式拦截：拖拽中按右键触发盖章
        if event.button() == Qt.MouseButton.RightButton and self.stamping_session:
            self.stamp_current_selection()
            event.accept()
            return

        # 对齐基准选择模式：点选当前已选对象中的一个作为基准
        if self.align_reference_mode and event.button() == Qt.MouseButton.LeftButton:
            raw = self.itemAt(event.scenePos(), QTransform())
            target_item = raw
            while target_item is not None and not isinstance(target_item, BaseElement):
                target_item = target_item.parentItem()

            if target_item and target_item in self.align_reference_candidates:
                self._execute_alignment_with_reference(self.align_reference_mode, target_item, self.align_reference_candidates)
                self._cancel_align_reference_mode()
            else:
                print("请点击当前已选对象中的一个作为对齐基准，按 ESC 可取消")
            event.accept()
            return

        if self.binding_source:
            # itemAt 可能返回字符子项/连接点等，需向上找到真正的 BaseElement
            raw = self.itemAt(event.scenePos(), QTransform())
            target_item = raw
            while target_item is not None and not isinstance(target_item, BaseElement):
                target_item = target_item.parentItem()

            if target_item and isinstance(target_item, BaseElement) and target_item != self.binding_source:
                command = SetParentCommand(self, self.binding_source, target_item)
                command.execute()
                self.undo_stack.push(command)
                print(f"已设置父级: {type(target_item).__name__}")
            else:
                print("已取消设置父级")

            self.binding_source = None
            if self.views():
                self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)
            return

        super().mousePressEvent(event)

        # 启动盖章会话追踪：如果按下左键后有选中元素，则开启会话
        if event.button() == Qt.MouseButton.LeftButton and self.selectedItems():
            # --- 新增：递归获取所有子级元素的逻辑 ---
            all_items_to_process = []
            def collect_all_children(item):
                if isinstance(item, BaseElement):
                    if item not in all_items_to_process:
                        all_items_to_process.append(item)
                # 遍历当前项的所有子项目
                for child in item.childItems():
                    collect_all_children(child)

            # 对每一个选中的物体执行递归搜寻
            for selected_item in self.selectedItems():
                collect_all_children(selected_item)
            # ----------------------------------------

            self.stamping_session = {
                'stamps': [],
                'initial_items': all_items_to_process,
                'moved': False  # 是否真正拖动过
            }

    def mouseReleaseEvent(self, event):
        """鼠标松开：结束盖章会话"""
        if self.stamping_session:
            self.finish_stamping_session()
        super().mouseReleaseEvent(event)

    def add_connector(self, parent, child):
        self.remove_child_connectors(child)
        conn = VConnector(parent, child)
        self.addItem(conn)
        self.connectors.append(conn)
        conn.update_path()
        self.sync_connectors_visibility(child)

    def remove_child_connectors(self, child):
        to_rem = [c for c in self.connectors if c.child_element == child]
        for c in to_rem:
            self.removeItem(c)
            self.connectors.remove(c)
    
    def remove_all_connectors_for_item(self, item):
        """移除与指定元素相关的所有连接器（父子关系连接器）"""
        to_rem = [c for c in self.connectors if c.parent_element == item or c.child_element == item]
        for c in to_rem:
            self.removeItem(c)
            self.connectors.remove(c)

    def stamp_current_selection(self):
        """修复版盖章函数 - 每次复制单独推入撤销栈"""
        if not self.stamping_session:
            return
            
        items_to_clone = self.stamping_session['initial_items']
        if not items_to_clone:
            return
            
        # 1. 接收三个返回值
        all_clones, internal_conns, roots_only = self.clone_items(items_to_clone)
        
        batch_cmds = []
        
        # 2. 【核心修复】只把顶层父级加入场景
        for root_item in roots_only:
            root_item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
            root_item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
            
            # 使用命令添加 root_item，子级会自动跟进去，不会再报 scene 不同步的错
            cmd = AddItemCommand(self, root_item)
            cmd.execute()
            batch_cmds.append(cmd)
            
        # 3. 添加连线
        for conn in internal_conns:
            cmd = AddConnectorCommand(self, conn)
            cmd.execute()
            batch_cmds.append(cmd)
            
        # 4. 【核心修复】此时所有元素都在场景里了，绝对坐标已生成，更新连线路径
        # 这样连线就不会从 (0,0) 射出来了
        for conn in internal_conns:
            conn.update_path()
            
        # 恢复所有副本的交互属性
        for cmd in batch_cmds:
            if isinstance(cmd, AddItemCommand):
                cmd.item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, True)
                cmd.item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, True)
        
        # 5. 【修改点】每次复制都单独推入撤销栈，而不是等到会话结束
        if batch_cmds:
            # 如果这是第一次复制，创建一个新的宏命令
            if len(self.stamping_session['stamps']) == 0:
                macro = MacroCommand(self, batch_cmds)
                self.undo_stack.push(macro)
            else:
                # 对于后续的复制，创建一个新的宏命令单独推入撤销栈
                # 这样每次复制都可以单独撤销
                macro = MacroCommand(self, batch_cmds)
                self.undo_stack.push(macro)
            
            self.stamping_session['stamps'].append(batch_cmds)
            print(f"已复制 {len(roots_only)} 个元素，可按 Ctrl+Z 单独撤销此次复制")

    def finish_stamping_session(self):
        """正常结束盖章会话 - 现在每次复制都已单独推入撤销栈"""
        if not self.stamping_session:
            return
            
        print(f"盖章会话结束，共生成 {len(self.stamping_session['stamps'])} 串副本")
        print("提示：每次复制都已单独记录到撤销栈，可按 Ctrl+Z 逐一撤销")
            
        self.stamping_session = None

    def abort_stamping_session(self):
        """放弃盖章会话，撤销本次拖拽过程中生成的所有副本"""
        if not self.stamping_session:
            return
            
        print("放弃盖章，正在移除副本...")
        for batch in reversed(self.stamping_session['stamps']):
            for cmd in reversed(batch):
                cmd.undo()
                
        self.stamping_session = None

    def clone_items(self, items):
        """克隆一组元素及其内部连接关系，解决子级跳位并支持安全添加"""
        if not items:
            return [], [], [] # 注意：现在返回三个值
            
        item_map = {item: idx for idx, item in enumerate(items)}
        clones_list = []  # 用于存储所有生成的副本对象
        id_to_clone = {}  # 原件索引 -> 副本对象的映射
        
        # --- 第一步：仅创建副本对象（不设置坐标，不设置父子关系） ---
        for item in items:
            clone = None
            if isinstance(item, VTextItem):
                clone = VTextItem(item.full_text, item.font_size, item.box_height)
                clone.font_family = item.font_family
                clone.text_color = QColor(item.text_color)
                clone.chars_per_column = item.chars_per_column
                clone.column_spacing = item.column_spacing
                clone.character_spacing = item.character_spacing
                clone.auto_height = item.auto_height
                clone.manual_line_break = item.manual_line_break
                clone.layer_eye_color = getattr(item, 'layer_eye_color', None)
                clone.rebuild()
                # 同步连接点可见性
                if item.connection_point and clone.connection_point:
                    clone.connection_point.setVisible(item.connection_point.isVisible())
            elif isinstance(item, VImageItem):
                clone = VImageItem(item.file_path, item.target_width)
                clone.set_opacity(item.image_opacity)
                if item.locked:
                    clone.set_locked(True)
                # 同步连接点可见性
                if item.connection_point and clone.connection_point:
                    clone.connection_point.setVisible(item.connection_point.isVisible())
            
            if clone:
                clone.setZValue(item.zValue())
                clones_list.append(clone)
                id_to_clone[item_map[item]] = clone
                
        # --- 第二步：建立副本之间的父子层级关系 ---
        new_internal_conns = []
        for item in items:
            parent = item.parentItem()
            # 只有当原件的父级也在本次选中的克隆列表里，才建立副本间的父子关系
            if parent in item_map:
                c_item = id_to_clone[item_map[item]]
                c_parent = id_to_clone[item_map[parent]]
                
                c_item.prepareGeometryChange() # 刷新图形状态
                c_item.setParentItem(c_parent) # 核心：建立层级
                
                # 为这对副本创建一条对应的红虚线连接器
                new_conn = VConnector(c_parent, c_item)
                new_internal_conns.append(new_conn)

        # --- 第三步：设置坐标并筛选出顶层父级 ---
        roots_only = []
        for item in items:
            c_item = id_to_clone.get(item_map[item])
            if not c_item:
                continue
                
            if c_item.parentItem():
                # 如果这个副本是有父级的，直接复制原件相对于父级的【本地偏移】
                # 这样父级移动到哪，它就跟到哪，不会乱跳
                c_item.setPos(item.pos())
            else:
                # 如果它是独立元素或最顶层父级，复制它在画布上的【绝对位置】
                c_item.setPos(item.scenePos())
                roots_only.append(c_item) # 记录下来：它是这组克隆体的“根”
        
        # 返回：所有副本（用于属性操作）、内部连线（用于后续更新路径）、根父级列表（用于 addItem）
        return clones_list, new_internal_conns, roots_only

    def update_connectors(self, item_moved):
        for c in self.connectors:
            if c.parent_element == item_moved or c.child_element == item_moved:
                c.update_path()

    def update_all_connectors(self):
        for c in self.connectors:
            c.update_path()
        self.sync_connectors_visibility()

    def _connector_items_visible(self, *items):
        return all(item is not None and item.isVisible() for item in items)

    def _image_text_connector_items(self, conn):
        if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
            return conn.image_item, conn.text_item
        if hasattr(conn, 'item1') and hasattr(conn, 'item2'):
            return conn.item1, conn.item2
        return None, None

    def _image_text_connector_visible(self, conn):
        """独立连接点连线没有两个元素端点，按连线显示总开关处理。"""
        item1, item2 = self._image_text_connector_items(conn)
        if item1 is None or item2 is None:
            return self.show_image_text_connectors
        return self.show_image_text_connectors and self._connector_items_visible(item1, item2)

    def sync_connectors_visibility(self, changed_item=None):
        for conn in self.connectors:
            if changed_item is not None and conn.parent_element != changed_item and conn.child_element != changed_item:
                continue
            conn.setVisible(
                self.show_connectors and
                self._connector_items_visible(conn.parent_element, conn.child_element)
            )

        for conn in self.image_text_connectors:
            item1, item2 = self._image_text_connector_items(conn)
            if (changed_item is not None
                    and item1 is not None and item2 is not None
                    and item1 != changed_item and item2 != changed_item):
                continue
            conn.setVisible(self._image_text_connector_visible(conn))
    
    def set_connectors_visible(self, visible):
        """控制所有连接器的可见性"""
        self.show_connectors = visible
        for c in self.connectors:
            c.setVisible(visible and self._connector_items_visible(c.parent_element, c.child_element))
    
    def set_image_text_connectors_visible(self, visible):
        """控制图文连接器的可见性"""
        self.show_image_text_connectors = visible
        for c in self.image_text_connectors:
            c.setVisible(self._image_text_connector_visible(c))
    
    def set_connection_points_visible(self, visible):
        """控制所有连接点的可见性"""
        self.show_connection_points = visible
        for item in self.items():
            if isinstance(item, (VTextItem, VImageItem)):
                item.set_connection_points_visible(visible)
        for point in getattr(self, 'free_connection_points', []):
            if point.scene() is self:
                point.setVisible(visible)

    def _show_status_message(self, message, timeout=0):
        main_window = self.parent()
        if main_window and hasattr(main_window, 'status_bar'):
            main_window.status_bar.showMessage(message, timeout)
        else:
            print(message)
    
    def start_connection_from_point(self, point):
        """从连接点开始连接"""
        if self.connection_mode and self.connection_source_point:
            source_point = self.connection_source_point
            if self.complete_connection(self.connection_source_point, point):
                self.connection_mode = False
                self.connection_source_point = None
                if self.views():
                    self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)
                source_point.update()
                point.update()
                self._show_status_message("连线已创建。点击任意连接点开始下一条连线。", 4000)
        else:
            self.connection_mode = True
            self.connection_source_point = point
            if self.views():
                self.views()[0].setCursor(Qt.CursorShape.CrossCursor)
            point.update()
            self._show_status_message("连线模式：请选择第二个连接点，按 Esc 取消。")
            print("连接模式：点击另一个连接点完成连接，或按ESC键取消")
    
    def cancel_connection_mode(self):
        """取消连接模式"""
        if self.connection_mode:
            self.connection_mode = False
            self.connection_source_point = None
            if self.views():
                self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)
            self._show_status_message("已取消连线模式。", 3000)
            print("已取消连接模式")
    
    def complete_connection(self, source_point, target_point):
        """完成两个连接点之间的连接"""
        if source_point == target_point:
            print("不能连接到自身")
            return False
        
        source_item = source_point.parent_element
        target_item = target_point.parent_element
        if (getattr(source_point, 'point_type', '') == 'custom'
                or getattr(target_point, 'point_type', '') == 'custom') and (
                    source_item is None or target_item is None):
            if source_point.scene() is None or target_point.scene() is None:
                return False
            conn = VGenericConnector(
                source_item, target_item, 'generic',
                self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH),
                source_point, target_point
            )
            self.addItem(conn)
            self.image_text_connectors.append(conn)
            source_point.connected_lines.append(conn)
            target_point.connected_lines.append(conn)
            conn.update_path()
            self.sync_connectors_visibility()
            return True
        if (isinstance(source_item, BaseElement)
                and isinstance(target_item, BaseElement)
                and (getattr(source_point, 'point_type', '') == 'custom'
                     or getattr(target_point, 'point_type', '') == 'custom')):
            connection_type = (
                "image-image" if isinstance(source_item, VImageItem) and isinstance(target_item, VImageItem)
                else "text-text" if isinstance(source_item, VTextItem) and isinstance(target_item, VTextItem)
                else "generic"
            )
            conn = VGenericConnector(
                source_item, target_item, connection_type,
                self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH),
                source_point, target_point
            )
            self.addItem(conn)
            self.image_text_connectors.append(conn)
            source_point.connected_lines.append(conn)
            target_point.connected_lines.append(conn)
            conn.update_path()
            self.sync_connectors_visibility()
            return True

        if isinstance(source_item, VImageItem) and isinstance(target_item, VTextItem):
            self.add_image_text_connector(source_item, target_item)
        elif isinstance(source_item, VTextItem) and isinstance(target_item, VImageItem):
            self.add_image_text_connector(target_item, source_item)
        elif isinstance(source_item, VImageItem) and isinstance(target_item, VImageItem):
            # 图片和图片之间的连接
            self.add_image_image_connector(source_item, target_item)
        elif isinstance(source_item, VTextItem) and isinstance(target_item, VTextItem):
            # 文字和文字之间的连接
            self.add_text_text_connector(source_item, target_item)
        else:
            print("连接类型不支持")
            return False
        return True
    
    def toggle_connection_points(self):
        """切换连接点显示状态"""
        self.show_connection_points = not self.show_connection_points
        self.set_connection_points_visible(self.show_connection_points)
        print(f"连接点显示 {'开启' if self.show_connection_points else '关闭'}")
    
    def save_item_as_asset(self, item):
        """保存元素为素材"""
        if isinstance(item, VTextItem):
            asset = self.asset_manager.add_text_asset(item)
            if asset:
                print(f"文字素材已保存 {asset['name']}")
                self.refresh_asset_library()
        elif isinstance(item, VImageItem):
            asset = self.asset_manager.add_image_asset(item)
            if asset:
                print(f"图片素材已保存 {asset['name']}")
                self.refresh_asset_library()
            else:
                print("保存图片素材失败")
    
    def save_group_as_asset(self, items=None):
        """保存组合为素材"""
        if items is None:
            items = [item for item in self.selectedItems() if isinstance(item, BaseElement)]
        
        if len(items) < 2:
            print("请选择至少两个元素来保存组合素材")
            return
        
        asset = self.asset_manager.add_group_asset(items, self)
        if asset:
            print(f"组合素材已保存 {asset['name']} (包含 {len(items)} 个元素)")
            self.refresh_asset_library()
        else:
            print("保存组合素材失败")
    
    def refresh_asset_library(self):
        """刷新素材库窗口"""
        if self.views():
            view = self.views()[0]
            widget = view
            while widget:
                # 刷新新的停靠面板版本
                if hasattr(widget, 'asset_library_dock'):
                    widget.asset_library_dock.refresh_assets()
                    break
                # 兼容旧版本窗口版本
                elif hasattr(widget, 'asset_library'):
                    if widget.asset_library and widget.asset_library.isVisible():
                        widget.asset_library.refresh_assets()
                    break
                widget = widget.parent()
    
    def start_image_text_binding(self, item):
        """开始图文连接模式"""
        if isinstance(item, (VImageItem, VTextItem)):
            self.image_text_binding_mode = True
            self.image_text_source = item
            views = self.views()
            if views:
                views[0].setCursor(Qt.CursorShape.CrossCursor)
            print(f"图文连接模式：请选择要连接的{'文字' if isinstance(item, VImageItem) else '图片'}")
    
    def add_image_text_connector(self, image_item, text_item):
        """添加图文连接线"""
        for conn in self.image_text_connectors:
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                if ((conn.image_item == image_item and conn.text_item == text_item) or
                        (conn.image_item == text_item and conn.text_item == image_item)):
                    print("这两个元素已经连接")
                    return
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                if ((conn.item1 == image_item and conn.item2 == text_item) or
                        (conn.item1 == text_item and conn.item2 == image_item)):
                    print("这两个元素已经连接")
                    return

        conn = VImageTextConnector(image_item, text_item, self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH))
        self.undo_stack.push_and_execute(AddConnectorCommand(self, conn))
        print("图文连接已创建")
    
    def add_image_image_connector(self, image1, image2):
        """添加图片-图片连接线"""
        for conn in self.image_text_connectors:
            if hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                if ((conn.item1 == image1 and conn.item2 == image2) or
                        (conn.item1 == image2 and conn.item2 == image1)):
                    print("这两个图片已经连接")
                    return
            elif hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                if ((conn.image_item == image1 and conn.text_item == image2) or
                        (conn.image_item == image2 and conn.text_item == image1)):
                    print("这两个图片已经连接")
                    return

        conn = VGenericConnector(image1, image2, "image-image", self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH))
        self.undo_stack.push_and_execute(AddConnectorCommand(self, conn))
        print("图片-图片连接已创建")

    def add_text_text_connector(self, text1, text2):
        """添加文字-文字连接线"""
        for conn in self.image_text_connectors:
            if hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                if ((conn.item1 == text1 and conn.item2 == text2) or
                        (conn.item1 == text2 and conn.item2 == text1)):
                    print("这两个文字已经连接")
                    return
            elif hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                if ((conn.image_item == text1 and conn.text_item == text2) or
                        (conn.image_item == text2 and conn.text_item == text1)):
                    print("这两个文字已经连接")
                    return

        conn = VGenericConnector(text1, text2, "text-text", self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH))
        self.undo_stack.push_and_execute(AddConnectorCommand(self, conn))
        print("文字-文字连接已创建")
    
    def remove_image_text_connectors(self, item):
        """移除与指定元素相关的所有连接线"""
        to_remove = []
        for conn in self.image_text_connectors:
            # 检查图文连接器
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                if conn.image_item == item or conn.text_item == item:
                    to_remove.append(conn)
            # 检查通用连接器
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                if conn.item1 == item or conn.item2 == item:
                    to_remove.append(conn)
        
        for conn in to_remove:
            self.removeItem(conn)
            self.image_text_connectors.remove(conn)
    
    def update_image_text_connectors(self, item):
        """更新与指定元素相关的所有连接线"""
        for conn in self.image_text_connectors:
            # 检查图文连接器
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                if conn.image_item == item or conn.text_item == item:
                    conn.update_path()
            # 检查通用连接器
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                if conn.item1 == item or conn.item2 == item:
                    conn.update_path()
    
    def update_all_image_text_connectors(self):
        """更新所有图文连接器"""
        for conn in self.image_text_connectors:
            conn.update_path()
        self.sync_connectors_visibility()
    
    def auto_connect_selected_items(self):
        """智能连接：
        - 框选后调用：按空间位置（从右到左、从上到下）排序，跳过父子关系
        - Ctrl加选后调用：按选中顺序连线，跳过父子关系
        """
        selected = [item for item in self.selectedItems()
                    if isinstance(item, (VImageItem, VTextItem)) and item.scene() == self]

        if len(selected) < 2:
            self._show_status_message("请至少选中两个元素进行智能连接", 3000)
            return

        def has_ancestor_relation(a, b):
            p = a.parentItem()
            while p:
                if p == b:
                    return True
                p = p.parentItem()
            p = b.parentItem()
            while p:
                if p == a:
                    return True
                p = p.parentItem()
            return False

        if self.last_selection_by_marquee:
            # 框选：按连接点位置从右到左、从上到下排序
            def sort_key(it):
                cp = getattr(it, 'connection_point', None)
                if cp and cp.isVisible():
                    p = cp.get_scene_center()
                else:
                    p = it.scenePos()
                return (p.y(), -p.x())
            items = sorted(selected, key=sort_key)
        else:
            # Ctrl加选：按选中顺序
            ordered = list(self.selection_order)
            for item in selected:
                if item not in ordered:
                    ordered.append(item)
            items = [item for item in ordered if item in selected]

        commands = []
        skipped_parent_child = 0
        skipped_existing = 0

        for item1, item2 in zip(items, items[1:]):
            if item1 == item2:
                continue
            if has_ancestor_relation(item1, item2):
                skipped_parent_child += 1
                continue

            cmd = self._make_connector_command(item1, item2)
            if cmd is None:
                skipped_existing += 1
                continue

            cmd.execute()
            commands.append(cmd)

        if commands:
            self.undo_stack.push(MacroCommand(self, commands))

        msg = f"智能连接完成：新增 {len(commands)} 条连线"
        details = []
        if skipped_parent_child:
            details.append(f"跳过父子关系 {skipped_parent_child} 对")
        if skipped_existing:
            details.append(f"跳过已存在 {skipped_existing} 对")
        if details:
            msg += "，" + "，".join(details)
        print(msg)
        self._show_status_message(msg, 4000)
    
    def group_chain_connect(self):
        """组合连接：将选中的图片+子文字视为组合，按位置顺序依次连接 b点→a点。
        
        每个组合：
          a点 = 图片连接点（顶部）
          b点 = 子文字连接点（底部），若无子文字则用图片连接点
        连接规则：第N组b点 → 第N+1组a点
        排序：按图片连接点位置，从上到下、同行从右到左
        """
        selected = [item for item in self.selectedItems()
                    if isinstance(item, VImageItem) and item.scene() == self]

        if len(selected) < 2:
            self._show_status_message("请至少选中两张图片进行组合连接", 3000)
            return

        # 识别每个组合的 a、b 点
        groups = []
        for img in selected:
            # a 点：图片连接点（需可见）
            point_a = img

            # b 点：找连接点可见的子文字，没有则用图片本身
            child_texts = [c for c in img.childItems() if isinstance(c, VTextItem)]
            point_b = img  # 默认用图片
            for ct in child_texts:
                cp = getattr(ct, 'connection_point', None)
                if cp and cp.isVisible():
                    point_b = ct
                    break

            groups.append({
                'image': img,
                'point_a': point_a,
                'point_b': point_b,
            })

        # 按图片连接点位置排序：从上到下（Y升序），同行从右到左（X降序）
        def sort_key(g):
            cp = getattr(g['image'], 'connection_point', None)
            if cp and cp.isVisible():
                p = cp.get_scene_center()
            else:
                p = g['image'].scenePos()
            return (p.y(), -p.x())

        groups.sort(key=sort_key)

        # 依次连接：第N组b点 → 第N+1组a点
        commands = []
        skipped = 0

        for i in range(len(groups) - 1):
            src = groups[i]['point_b']    # 当前组 b 点
            dst = groups[i + 1]['point_a']  # 下一组 a 点

            if src == dst:
                skipped += 1
                continue

            cmd = self._make_connector_command(src, dst)
            if cmd is None:
                skipped += 1
                continue

            cmd.execute()
            commands.append(cmd)

        if commands:
            self.undo_stack.push(MacroCommand(self, commands))

        msg = f"组合连接完成：新增 {len(commands)} 条连线"
        if skipped:
            msg += f"，跳过 {skipped} 对"
        print(msg)
        self._show_status_message(msg, 4000)

    def connect_image_points_right_to_left(self):
        """连接图片a点：按从右到左顺序，依次连接选中图片的连接点"""
        images = [item for item in self.selectedItems()
                  if isinstance(item, VImageItem) and item.scene() == self]

        if len(images) < 2:
            self._show_status_message("请至少选中两张图片", 3000)
            return

        # 按连接点X降序（从右到左），同X按Y升序（从上到下）
        def sort_key(img):
            cp = getattr(img, 'connection_point', None)
            if cp and cp.isVisible():
                p = cp.get_scene_center()
            else:
                p = img.scenePos()
            return (-p.x(), p.y())

        images.sort(key=sort_key)

        commands = []
        skipped = 0
        for img1, img2 in zip(images, images[1:]):
            cmd = self._make_connector_command(img1, img2)
            if cmd is None:
                skipped += 1
                continue
            cmd.execute()
            commands.append(cmd)

        if commands:
            self.undo_stack.push(MacroCommand(self, commands))

        msg = f"图片连接点连线完成：新增 {len(commands)} 条"
        if skipped:
            msg += f"，跳过 {skipped} 对"
        print(msg)
        self._show_status_message(msg, 4000)

    def connect_by_position(self):
        """按位置连接：上下相邻的图片和文字自动连接"""
        selected = self.selectedItems()
        items = [item for item in selected if isinstance(item, (VImageItem, VTextItem))]
        if len(items) < 2: return
        items.sort(key=lambda item: item.scenePos().y())
        connections_made = 0
        for i in range(len(items) - 1):
            current = items[i]
            next_item = items[i + 1]
            if ((isinstance(current, VImageItem) and isinstance(next_item, VTextItem)) or
                (isinstance(current, VTextItem) and isinstance(next_item, VImageItem))):
                distance = abs(next_item.scenePos().y() - current.scenePos().y())
                if distance < 200:
                    if isinstance(current, VImageItem):
                        self.add_image_text_connector(current, next_item)
                    else:
                        self.add_image_text_connector(next_item, current)
                    connections_made += 1
        print(f"按位置创建了 {connections_made} 个图文连接")
    
    def connect_all_images_to_text(self):
        """将所有选中的图片连接到一个文字"""
        selected = self.selectedItems()
        images = [item for item in selected if isinstance(item, VImageItem)]
        texts = [item for item in selected if isinstance(item, VTextItem)]
        if len(texts) != 1 or not images: return
        target_text = texts[0]
        for img in images:
            self.add_image_text_connector(img, target_text)
        print(f"已将选中图片连接到文字")
    
    def connect_all_texts_to_image(self):
        """将所有选中的文字连接到一个图片"""
        selected = self.selectedItems()
        images = [item for item in selected if isinstance(item, VImageItem)]
        texts = [item for item in selected if isinstance(item, VTextItem)]
        if len(images) != 1 or not texts: return
        target_image = images[0]
        for text in texts:
            self.add_image_text_connector(target_image, text)
        print(f"已将选中文字连接到图片")
    
    def remove_all_image_text_connections(self):
        """移除所有图文连接"""
        count = len(self.image_text_connectors)
        for conn in self.image_text_connectors[:]:
            self.removeItem(conn)
        self.image_text_connectors.clear()
        print(f"已移除 {count} 个图文连接")
    
    def remove_connector_item(self, connector):
        """删除单个连接线"""
        # 从连接点的反向引用中同步移除，防止点被移动或删除时访问已失效的连线。
        for point in (getattr(connector, 'point1', None), getattr(connector, 'point2', None)):
            if point is not None and connector in getattr(point, 'connected_lines', []):
                point.connected_lines.remove(connector)
        if connector in self.image_text_connectors:
            self.removeItem(connector)
            self.image_text_connectors.remove(connector)
            print("已删除连接线")
        elif connector in self.connectors:
            self.removeItem(connector)
            self.connectors.remove(connector)
            print("已删除父子连接线")
    
    def copy_items(self, items):
        """复制多个元素到剪贴板"""
        if not items: return
        self.clipboard_items = []
        self.clipboard_image_text_connections = []
        item_to_index = {item: idx for idx, item in enumerate(items)}
        
        for idx, item in enumerate(items):
            if isinstance(item, VTextItem):
                # 保存连接点可见性状态
                connection_point_visible = item.connection_point.isVisible() if item.connection_point else True
                
                item_data = {
                    'type': 'VTextItem',
                    'text': item.full_text,
                    'font_size': item.font_size,
                    'box_height': item.box_height,
                    'font_family': item.font_family,
                    'text_color': item.text_color.name(),
                    'chars_per_column': item.chars_per_column,
                    'column_spacing': item.column_spacing,
                    'character_spacing': item.character_spacing,
                    'auto_height': item.auto_height,
                    'manual_line_break': item.manual_line_break,
                    'layer_eye_color': getattr(item, 'layer_eye_color', None),
                    'connection_point_visible': connection_point_visible,
                    'scene_pos': (item.scenePos().x(), item.scenePos().y()),
                    'local_pos': (item.x(), item.y()),
                    'parent_index': item_to_index.get(item.parentItem(), -1) if isinstance(item.parentItem(), BaseElement) else -1
                }
                self.clipboard_items.append(item_data)
            elif isinstance(item, VImageItem):
                # 保存连接点可见性状态
                connection_point_visible = item.connection_point.isVisible() if item.connection_point else True
                
                item_data = {
                    'type': 'VImageItem',
                    'path': item.file_path,
                    'width': item.target_width,
                    'connection_point_visible': connection_point_visible,
                    'scene_pos': (item.scenePos().x(), item.scenePos().y()),
                    'local_pos': (item.x(), item.y()),
                    'parent_index': item_to_index.get(item.parentItem(), -1) if isinstance(item.parentItem(), BaseElement) else -1
                }
                self.clipboard_items.append(item_data)
        
        for conn in self.image_text_connectors:
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                img_idx = item_to_index.get(conn.image_item, -1)
                text_idx = item_to_index.get(conn.text_item, -1)
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                img_idx = item_to_index.get(conn.item1, -1)
                text_idx = item_to_index.get(conn.item2, -1)
            else:
                continue
            if img_idx != -1 and text_idx != -1:
                self.clipboard_image_text_connections.append((img_idx, text_idx))
        print(f"已复制 {len(self.clipboard_items)} 个元素到剪贴板")
    
    def copy_item(self, item):
        self.copy_items([item])
    
    def paste_items(self, pos=None):
        """粘贴剪贴板中的所有元素（整体作为一次撤销）"""
        if not self.clipboard_items:
            return []

        min_x = min(item['scene_pos'][0] for item in self.clipboard_items)
        min_y = min(item['scene_pos'][1] for item in self.clipboard_items)

        if pos is None:
            if self.views():
                view = self.views()[0]
                center = view.mapToScene(view.viewport().rect().center())
                base_x, base_y = center.x(), center.y()
            else:
                base_x, base_y = 100, 100
        else:
            base_x, base_y = pos.x(), pos.y()

        new_items = []
        sub_commands = []  # 收集所有子命令

        for idx, item_data in enumerate(self.clipboard_items):
            new_item = None
            if item_data['type'] == 'VTextItem':
                new_item = VTextItem(item_data['text'], item_data['font_size'], item_data['box_height'])
                new_item.font_family = item_data['font_family']
                new_item.text_color = QColor(item_data['text_color'])
                for k in ('chars_per_column', 'column_spacing', 'character_spacing', 'auto_height', 'manual_line_break', 'layer_eye_color'):
                    if k in item_data:
                        setattr(new_item, k, item_data[k])
                new_item.rebuild()
            elif item_data['type'] == 'VImageItem':
                new_item = VImageItem(item_data['path'], item_data['width'])

            if new_item:
                offset_x = item_data['scene_pos'][0] - min_x
                offset_y = item_data['scene_pos'][1] - min_y
                new_item.setPos(base_x + offset_x, base_y + offset_y)

                # 直接执行，不 push 到栈
                cmd = AddItemCommand(self, new_item)
                cmd.execute()
                sub_commands.append(cmd)

                if item_data.get('connection_point_deleted', False):
                    new_item.delete_connection_point()
                elif 'connection_point_visible' in item_data and new_item.connection_point:
                    new_item.connection_point.setVisible(item_data['connection_point_visible'])

                new_items.append(new_item)

        # 恢复父子关系
        for idx, item_data in enumerate(self.clipboard_items):
            if item_data['parent_index'] != -1 and item_data['parent_index'] < len(new_items):
                child_item = new_items[idx]
                parent_item = new_items[item_data['parent_index']]
                current_scene_pos = child_item.scenePos()
                child_item.setParentItem(parent_item)
                child_item.setPos(parent_item.mapFromScene(current_scene_pos))
                self.add_connector(parent_item, child_item)

        # 恢复图文连接（直接执行，收集命令）
        for img_idx, text_idx in self.clipboard_image_text_connections:
            if img_idx < len(new_items) and text_idx < len(new_items):
                conn_cmd = self._make_connector_command(new_items[img_idx], new_items[text_idx])
                if conn_cmd:
                    conn_cmd.execute()
                    sub_commands.append(conn_cmd)

        # 最终统一应用连接点可见性/删除状态，防止父子关系建立/连线创建的回调将其覆盖
        for idx, item_data in enumerate(self.clipboard_items):
            if idx < len(new_items):
                item = new_items[idx]
                if item:
                    if item_data.get('connection_point_deleted', False):
                        item.delete_connection_point()
                    elif 'connection_point_visible' in item_data and item.connection_point:
                        item.connection_point.setVisible(item_data['connection_point_visible'])

        # 整体打包为一次撤销
        if sub_commands:
            self.undo_stack.push(MacroCommand(self, sub_commands))

        return new_items

    def _make_connector_command(self, item1, item2):
        """创建连接线命令但不 push 到栈"""
        # 检查是否已连接
        for conn in self.image_text_connectors:
            if hasattr(conn, 'image_item'):
                if (conn.image_item == item1 and conn.text_item == item2) or \
                   (conn.image_item == item2 and conn.text_item == item1):
                    return None
            elif hasattr(conn, 'item1'):
                if (conn.item1 == item1 and conn.item2 == item2) or \
                   (conn.item1 == item2 and conn.item2 == item1):
                    return None
        if isinstance(item1, VImageItem) and isinstance(item2, VTextItem):
            return AddConnectorCommand(self, VImageTextConnector(item1, item2,
                self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH)))
        elif isinstance(item1, VTextItem) and isinstance(item2, VImageItem):
            return AddConnectorCommand(self, VImageTextConnector(item2, item1,
                self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH)))
        elif isinstance(item1, VImageItem) and isinstance(item2, VImageItem):
            return AddConnectorCommand(self, VGenericConnector(item1, item2, "image-image",
                self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH)))
        elif isinstance(item1, VTextItem) and isinstance(item2, VTextItem):
            return AddConnectorCommand(self, VGenericConnector(item1, item2, "text-text",
                self.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH)))
        return None

    def _has_parent_child_relation(self, item1, item2):
        return item1.parentItem() == item2 or item2.parentItem() == item1

    def batch_chain_connect_selected_items(self):
        """Connect selected image/text items pair-by-pair in click selection order."""
        items = [
            item for item in self.selection_order
            if isinstance(item, (VImageItem, VTextItem)) and item.isSelected() and item.scene() == self
        ]

        if len(items) < 2:
            print("请按顺序选中至少两个图片或文字元素")
            return

        commands = []
        skipped_parent_child = 0
        skipped_existing = 0
        skipped_unsupported = 0

        for item1, item2 in zip(items, items[1:]):
            if item1 == item2:
                skipped_existing += 1
                continue
            if self._has_parent_child_relation(item1, item2):
                skipped_parent_child += 1
                continue

            command = self._make_connector_command(item1, item2)
            if command is None:
                skipped_existing += 1
                continue

            command.execute()
            commands.append(command)

        if commands:
            self.undo_stack.push(MacroCommand(self, commands))

        message = f"批量连线完成：新增 {len(commands)} 条"
        details = []
        if skipped_parent_child:
            details.append(f"跳过父子关系 {skipped_parent_child} 对")
        if skipped_existing:
            details.append(f"跳过已存在/重复 {skipped_existing} 对")
        if skipped_unsupported:
            details.append(f"跳过不支持 {skipped_unsupported} 对")
        if details:
            message += "，" + "，".join(details)
        print(message)
        self._show_status_message(message, 4000)
    
    def paste_item(self, pos=None):
        items = self.paste_items(pos)
        return items[0] if items else None
    
    def delete_item(self, item):
        self.undo_stack.push_and_execute(DeleteItemCommand(self, item))

    def add_item_with_undo(self, item):
        """添加元素到场景并记录撤销"""
        self.addItem(item)
        if isinstance(item, (VTextItem, VImageItem)):
            item.set_connection_points_visible(self.show_connection_points)
        self.undo_stack.push(AddItemCommand(self, item))
    
    def undo(self):
        self.undo_stack.undo()
    
    def redo(self):
        self.undo_stack.redo()
    
    def keyPressEvent(self, event):
        # 对齐基准点选模式：ESC 取消
        if self.align_reference_mode and event.key() == Qt.Key.Key_Escape:
            self._cancel_align_reference_mode()
            event.accept()
            return

        # 盖章会话按键拦截：空格盖章（需要鼠标正在按下拖动中），ESC放弃
        if self.stamping_session:
            if event.key() == Qt.Key.Key_Space:
                # 只有鼠标正在按下时才允许空格盖章，避免普通拖动时误触
                if QApplication.mouseButtons() & Qt.MouseButton.LeftButton:
                    self.stamp_current_selection()
                event.accept()
                return
            elif event.key() == Qt.Key.Key_Escape:
                self.abort_stamping_session()
                event.accept()
                return

        # ESC键取消连接模式
        if event.key() == Qt.Key.Key_Escape:
            if self.resize_mode:
                self.cancel_resize()
                event.accept()
                return
            if self.connection_mode:
                self.cancel_connection_mode()
                event.accept()
                return
            elif self.image_text_binding_mode:
                self.image_text_binding_mode = False
                self.image_text_source = None
                if self.views():
                    self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)
                print("已取消图文连接模式")
                event.accept()
                return
            elif self.binding_source:
                self.binding_source = None
                if self.views():
                    self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)
                print("已取消父子绑定模式")
                event.accept()
                return
            
            # 清除所有选中状态，包括连接线
            if self.selectedItems():
                self.clearSelection()
                print("已清除所有选中状态")
                event.accept()
                return
        
        # 调整图片模式：回车确认
        if self.resize_mode and event.key() == Qt.Key.Key_Return:
            self.confirm_resize()
            event.accept()
            return

        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.key() == Qt.Key.Key_Z:
                self.undo()
            elif event.key() == Qt.Key.Key_Y:
                self.redo()
            elif event.key() == Qt.Key.Key_C:
                selected = self.selectedItems()
                base_elements = [item for item in selected if isinstance(item, BaseElement)]
                if base_elements: self.copy_items(base_elements)
            elif event.key() == Qt.Key.Key_V:
                self.paste_items()
        elif event.key() == Qt.Key.Key_Delete:
            selected = self.selectedItems()
            for item in selected:
                if isinstance(item, BaseElement):
                    self.delete_item(item)
                elif isinstance(item, (VImageTextConnector, VGenericConnector)):
                    # 删除连接线
                    self.remove_connector_item(item)
        elif event.key() in (Qt.Key.Key_Left, Qt.Key.Key_Right, Qt.Key.Key_Up, Qt.Key.Key_Down):
            items = [i for i in self.selectedItems() if isinstance(i, BaseElement)]
            if items:
                # Shift键使用大步长，否则1px
                if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                    step = self.config_manager.get('nudge_large_step', 10)
                else:
                    step = 1
                dx = dy = 0
                if event.key() == Qt.Key.Key_Left:  dx = -step
                elif event.key() == Qt.Key.Key_Right: dx = step
                elif event.key() == Qt.Key.Key_Up:   dy = -step
                elif event.key() == Qt.Key.Key_Down:  dy = step
                if self.config_manager.get('horizontal_move_only', False):
                    dy = 0

                move_commands = []
                for item in items:
                    old_pos = item.scenePos()
                    new_pos = old_pos + QPointF(dx, dy)
                    if item.parentItem():
                        item.setPos(item.parentItem().mapFromScene(new_pos))
                    else:
                        item.setPos(new_pos)
                    move_commands.append(MoveItemCommand(self, item, old_pos, new_pos))

                if move_commands:
                    self.undo_stack.push(MacroCommand(self, move_commands))
                event.accept()
                return
        else:
            super().keyPressEvent(event)
    
    def _start_align_reference_mode(self, align_mode, items, tip_text):
        self.align_reference_mode = align_mode
        self.align_reference_candidates = list(items)
        if self.views():
            self.views()[0].setCursor(Qt.CursorShape.CrossCursor)
        print(tip_text)

    def _cancel_align_reference_mode(self):
        self.align_reference_mode = None
        self.align_reference_candidates = []
        if self.views():
            self.views()[0].setCursor(Qt.CursorShape.ArrowCursor)
        print("已取消对齐基准选择")

    def _execute_alignment_with_reference(self, align_mode, ref, items):
        # 收集所有移动命令
        move_commands = []
        
        if align_mode == 'top':
            ref_value = ref.scenePos().y()
            for item in items:
                if item == ref:  # 跳过基准对象本身
                    continue
                current_pos = item.scenePos()
                new_scene_pos = QPointF(current_pos.x(), ref_value)
                # 创建移动命令
                cmd = MoveItemCommand(self, item, current_pos, new_scene_pos)
                cmd.execute()  # 执行移动
                move_commands.append(cmd)
        elif align_mode == 'right':
            ref_value = ref.scenePos().x() + ref.boundingRect().width()
            for item in items:
                if item == ref:  # 跳过基准对象本身
                    continue
                current_pos = item.scenePos()
                new_x = ref_value - item.boundingRect().width()
                new_scene_pos = QPointF(new_x, current_pos.y())
                # 创建移动命令
                cmd = MoveItemCommand(self, item, current_pos, new_scene_pos)
                cmd.execute()  # 执行移动
                move_commands.append(cmd)
        elif align_mode == 'center_h':
            ref_value = ref.scenePos().x() + ref.boundingRect().width() / 2
            for item in items:
                if item == ref:  # 跳过基准对象本身
                    continue
                current_pos = item.scenePos()
                new_x = ref_value - item.boundingRect().width() / 2
                new_scene_pos = QPointF(new_x, current_pos.y())
                # 创建移动命令
                cmd = MoveItemCommand(self, item, current_pos, new_scene_pos)
                cmd.execute()  # 执行移动
                move_commands.append(cmd)
        elif align_mode == 'center_v':
            ref_value = ref.scenePos().y() + ref.boundingRect().height() / 2
            for item in items:
                if item == ref:  # 跳过基准对象本身
                    continue
                current_pos = item.scenePos()
                new_y = ref_value - item.boundingRect().height() / 2
                new_scene_pos = QPointF(current_pos.x(), new_y)
                # 创建移动命令
                cmd = MoveItemCommand(self, item, current_pos, new_scene_pos)
                cmd.execute()  # 执行移动
                move_commands.append(cmd)

        # 将所有移动命令打包为一个宏命令，以便一次撤销
        if move_commands:
            macro_cmd = MacroCommand(self, move_commands)
            self.undo_stack.push(macro_cmd)

        print(f"已完成对齐，基准对象: {type(ref).__name__}")

    def _top_level_selected(self):
        """只取顶层选中元素，子级跟着父级移动不需要单独对齐"""
        all_selected = [item for item in self.selectedItems() if isinstance(item, BaseElement)]
        selected_set = set(all_selected)
        return [item for item in all_selected
                if not isinstance(item.parentItem(), BaseElement) or
                item.parentItem() not in selected_set]

    def align_top(self, items=None):
        if items is None:
            items = self._top_level_selected()
        if len(items) < 2:
            return
        self._start_align_reference_mode('top', items, '请在画布中点选一个已选对象作为"顶部对齐"基准，按 ESC 可取消')

    def align_right(self, items=None):
        if items is None:
            items = self._top_level_selected()
        if len(items) < 2:
            return
        self._start_align_reference_mode('right', items, '请在画布中点选一个已选对象作为"右对齐"基准，按 ESC 可取消')

    def align_center_horizontal(self, items=None):
        if items is None:
            items = self._top_level_selected()
        if len(items) < 2:
            return
        self._start_align_reference_mode('center_h', items, '请在画布中点选一个已选对象作为"水平居中"基准，按 ESC 可取消')

    def align_center_vertical(self, items=None):
        if items is None:
            items = self._top_level_selected()
        if len(items) < 2:
            return
        self._start_align_reference_mode('center_v', items, '请在画布中点选一个已选对象作为“垂直居中”基准，按 ESC 可取消')

    def batch_copy(self, params):
        """步长和重复复制"""
        items = [i for i in self.selectedItems() if isinstance(i, BaseElement)]
        if not items:
            return

        if params['step_x'] == 0 and params['step_y'] == 0:
            print("警告：X和Y偏移均为0，副本将叠在原位")

        # 先把当前选中存入剪贴板
        self.copy_items(items)

        # 计算选中组合左上角作为基准
        combined = QRectF()
        for item in items:
            combined = combined.united(QRectF(item.scenePos(), item.boundingRect().size()))

        step_x = params['step_x']
        step_y = params['step_y']
        count  = params['count']

        new_items = []
        for n in range(1, count + 1):
            paste_pos = QPointF(combined.left() + step_x * n,
                                combined.top()  + step_y * n)
            pasted = self.paste_items(paste_pos)
            new_items.extend(pasted)

        self.clearSelection()
        for i in new_items:
            i.setSelected(True)
        print(f"步长复制完成：{len(new_items)} 个新元素，step=({step_x}, {step_y})")


class LayoutView(QGraphicsView):
    transformChanged = pyqtSignal()  # 变换改变信号
    RULER_SIZE = 20  # 标尺厚度（像素）

    def __init__(self, scene):
        super().__init__(scene)
        # --- 在这里插入/确保这两行存在 ---
        self.setViewportUpdateMode(QGraphicsView.ViewportUpdateMode.FullViewportUpdate)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setDragMode(QGraphicsView.DragMode.NoDrag)
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.setAcceptDrops(True)
        self._is_panning = False
        self._marquee_origin = QPoint()
        self._marquee_active = False
        self._marquee_mode = Qt.ItemSelectionMode.ContainsItemShape
        self._marquee_band = QRubberBand(QRubberBand.Shape.Rectangle, self.viewport())
        self._marquee_sweep_order = []  # 框选过程中鼠标扫过的元素顺序
        self._free_point_f5_move_active = False
        self._pan_start = QPoint()
        self._main_window = None
        self._smart_brush_enabled = False
        self._smart_brush_active = False
        self._smart_brush_radius = 18  # 会在 set_main_window 后从配置更新
        self._smart_brush_pos = QPoint()
        self._smart_brush_seen = set()
        # 辅助线拖拽状态
        self._guide_dragging = False       # 正在从标尺拖出辅助线
        self._guide_orientation = None     # 拖出方向
        self._guide_preview = None         # 预览辅助线对象
        # 为标尺留出边距
        self._hovered_connection_point = None
        self.setViewportMargins(self.RULER_SIZE, self.RULER_SIZE, 0, 0)

    def set_main_window(self, mw):
        self._main_window = mw
        # 从配置读取笔刷默认大小
        if hasattr(mw, 'scene') and mw.scene:
            r = mw.scene.config_manager.get('smart_brush_radius', 18)
            self._smart_brush_radius = max(4, min(120, int(r)))

    def set_smart_brush_enabled(self, enabled):
        self._smart_brush_enabled = enabled
        self._smart_brush_active = False
        self._smart_brush_seen.clear()
        if enabled:
            self.setCursor(Qt.CursorShape.BlankCursor)  # 隐藏系统光标，用圆形笔刷圈代替
            msg = "智能笔刷：涂抹对象加入选区；按住 Alt 涂抹从选区中减去"
        else:
            self.setCursor(Qt.CursorShape.ArrowCursor)
            msg = "已退出智能笔刷"
        if self._main_window and hasattr(self._main_window, 'status_bar'):
            self._main_window.status_bar.showMessage(msg, 0 if enabled else 3000)
        self.viewport().update()

    def is_smart_brush_enabled(self):
        return self._smart_brush_enabled

    def set_smart_brush_radius(self, radius):
        self._smart_brush_radius = max(4, min(120, int(radius)))
        if self._main_window and hasattr(self._main_window, 'smart_brush_size_spin'):
            spin = self._main_window.smart_brush_size_spin
            if spin.value() != self._smart_brush_radius:
                spin.blockSignals(True)
                spin.setValue(self._smart_brush_radius)
                spin.blockSignals(False)
        self.viewport().update()
        if self._smart_brush_enabled and self._main_window and hasattr(self._main_window, 'status_bar'):
            self._main_window.status_bar.showMessage(f"智能笔刷大小：{self._smart_brush_radius}px", 1200)

    def adjust_smart_brush_radius(self, delta):
        self.set_smart_brush_radius(self._smart_brush_radius + delta)

    def _update_smart_brush_overlay(self, old_pos=None):
        """只刷新智能笔刷光标覆盖的区域，避免每次移动重绘整个视口。"""
        radius = self._smart_brush_radius + 4
        rect = QRectF(self._smart_brush_pos.x() - radius,
                      self._smart_brush_pos.y() - radius,
                      radius * 2, radius * 2)
        if old_pos is not None:
            old_rect = QRectF(old_pos.x() - radius, old_pos.y() - radius,
                              radius * 2, radius * 2)
            rect = rect.united(old_rect)
        self.viewport().update(rect.toAlignedRect())

    def _smart_brush_scene_path(self, pos):
        center = self.mapToScene(pos)
        scale = max(abs(self.transform().m11()), 0.001)
        radius = self._smart_brush_radius / scale
        path = QPainterPath()
        path.addEllipse(center, radius, radius)
        return path

    def _smart_brush_allows_item(self, item):
        if not isinstance(item, (VImageItem, VTextItem)):
            return False
        if getattr(item, 'locked', False):
            return False
        if not (item.flags() & QGraphicsItem.GraphicsItemFlag.ItemIsSelectable):
            return False
        mode = self.scene().config_manager.get('marquee_mode', 'all') if self.scene() else 'all'
        if mode == 'images':
            return isinstance(item, VImageItem)
        if mode == 'connected':
            cp = getattr(item, 'connection_point', None)
            return bool(cp and cp.isVisible())
        return True

    def _items_under_smart_brush(self, pos):
        if not self.scene():
            return []
        path = self._smart_brush_scene_path(pos)
        raw_items = self.scene().items(path, Qt.ItemSelectionMode.IntersectsItemShape,
                                       Qt.SortOrder.DescendingOrder, self.transform())
        found = []
        seen = set()
        for item in raw_items:
            element = item
            while element and not isinstance(element, (VImageItem, VTextItem)):
                element = element.parentItem()
            if element and element not in seen and self._smart_brush_allows_item(element):
                found.append(element)
                seen.add(element)
        return found

    def _apply_smart_brush_at(self, pos, subtract=False):
        scene = self.scene()
        if not scene:
            return
        changed = False
        for item in self._items_under_smart_brush(pos):
            key = (item, subtract)
            if key in self._smart_brush_seen:
                continue
            self._smart_brush_seen.add(key)
            if subtract:
                if item.isSelected():
                    item.setSelected(False)
                    if item in scene.selection_order:
                        scene.selection_order = [i for i in scene.selection_order if i != item]
                    changed = True
            else:
                if not item.isSelected():
                    item.setSelected(True)
                    if item in scene.selection_order:
                        scene.selection_order = [i for i in scene.selection_order if i != item]
                    scene.selection_order.append(item)
                    changed = True
        if changed:
            scene.last_selection_by_marquee = False
            if self._main_window and hasattr(self._main_window, 'status_bar'):
                action = "减去" if subtract else "加入"
                self._main_window.status_bar.showMessage(
                    f"智能笔刷：已{action}对象，当前选中 {len(scene.selectedItems())} 个", 1500
                )

    def _connection_point_at(self, pos, tolerance=7, exclude_point=None):
        if not self.scene():
            return None

        hit_offset_y = -6
        nearest = None
        nearest_dist = tolerance
        for item in self.scene().items():
            if not isinstance(item, ConnectionPoint) or not item.isVisible() or item == exclude_point:
                continue
            view_pos = self.mapFromScene(item.get_scene_center())
            dist = math.hypot(view_pos.x() - pos.x(), view_pos.y() + hit_offset_y - pos.y())
            if dist <= nearest_dist:
                nearest = item
                nearest_dist = dist
        return nearest

    def _sync_connection_point_hover(self, pos):
        hovered = self._connection_point_at(pos)

        if hovered == self._hovered_connection_point:
            return

        if self._hovered_connection_point:
            self._hovered_connection_point._hovered = False
            self._hovered_connection_point.update()

        self._hovered_connection_point = hovered

        if self._hovered_connection_point:
            self._hovered_connection_point._hovered = True
            self._hovered_connection_point.update()

    def _in_h_ruler(self, pos):
        """鼠标是否在水平标尺区域（顶部）"""
        return pos.y() < self.RULER_SIZE and pos.x() >= self.RULER_SIZE

    def _in_v_ruler(self, pos):
        """鼠标是否在垂直标尺区域（左侧）"""
        return pos.x() < self.RULER_SIZE and pos.y() >= self.RULER_SIZE

    def _update_marquee_style(self, current_pos):
        left_to_right = current_pos.x() >= self._marquee_origin.x()
        if left_to_right:
            self._marquee_mode = Qt.ItemSelectionMode.ContainsItemShape
            self._marquee_band.setStyleSheet("background: rgba(0, 120, 215, 0.18); border: 1px solid rgba(0, 120, 215, 0.95);")
        else:
            self._marquee_mode = Qt.ItemSelectionMode.IntersectsItemShape
            self._marquee_band.setStyleSheet("background: rgba(0, 200, 120, 0.18); border: 1px solid rgba(0, 200, 120, 0.95);")

    def _apply_marquee_selection(self):
        rect = self._marquee_band.geometry()
        if rect.width() < 3 or rect.height() < 3:
            return

        scene_polygon = self.mapToScene(rect)
        path = QPainterPath()
        path.addPolygon(scene_polygon)

        scene = self.scene()
        mode = scene.config_manager.get('marquee_mode', 'all')
        scene.clearSelection()
        items = scene.items(path, self._marquee_mode, Qt.SortOrder.DescendingOrder, self.transform())

        has_connector_in_rect = False
        for item in items:
            if (isinstance(item, ConnectionPoint)
                    and getattr(item, 'point_type', '') == 'custom'
                    and getattr(scene, 'free_connection_point_mode', 'move') == 'move'
                    and item.isVisible()
                    and (item.flags() & QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)):
                item.setSelected(True)
            elif mode == 'images':
                if isinstance(item, VImageItem) and not getattr(item, 'locked', False):
                    item.setSelected(True)
                elif isinstance(item, (VImageTextConnector, VGenericConnector)):
                    has_connector_in_rect = True
            elif mode == 'connected':
                if isinstance(item, (VImageItem, VTextItem)) and not getattr(item, 'locked', False):
                    cp = getattr(item, 'connection_point', None)
                    if cp and cp.isVisible():
                        item.setSelected(True)
            else:
                if item.flags() & QGraphicsItem.GraphicsItemFlag.ItemIsSelectable:
                    item.setSelected(True)

        # 独立连接点按圆心是否落入选框判定。这样左到右的“完全包含”框选
        # 不会因为点的可点击范围较大而漏选。
        if getattr(scene, 'free_connection_point_mode', 'move') == 'move':
            for point in getattr(scene, 'free_connection_points', []):
                if point.scene() is scene and point.isVisible() and path.contains(point.get_scene_center()):
                    point.setSelected(True)

        # 仅图片模式下框选到连线时弹出提示
        if mode == 'images' and has_connector_in_rect:
            mw = self._main_window
            if mw and hasattr(mw, '_toast'):
                mw._toast.show_message(
                    "框选范围内有连线未被选中\n当前模式：仅图片\nAlt+M 切换为「全选」模式", "⚠", 4000)

    def paintEvent(self, event):
        super().paintEvent(event)
        # 标尺画在 viewport 上
        painter = QPainter(self.viewport())
        self._draw_rulers(painter)
        if self._smart_brush_enabled:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            subtract = QApplication.keyboardModifiers() & Qt.KeyboardModifier.AltModifier
            color = QColor(230, 70, 70, 210) if subtract else QColor(0, 150, 255, 210)
            fill = QColor(color)
            fill.setAlpha(35)
            painter.setPen(QPen(color, 2))
            painter.setBrush(QBrush(fill))
            painter.drawEllipse(self._smart_brush_pos, self._smart_brush_radius, self._smart_brush_radius)
        painter.end()

    def scrollContentsBy(self, dx, dy):
        """平移时强制重绘整个 viewport 以清除标尺残影"""
        super().scrollContentsBy(dx, dy)
        self.viewport().update()

    def _draw_rulers(self, painter):
        """在 viewport 上绘制标尺"""
        R = self.RULER_SIZE
        vp = self.viewport().rect()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, False)

        # 背景
        ruler_color = QColor(50, 50, 55)
        tick_color = QColor(180, 180, 180)
        text_color = QColor(200, 200, 200)
        corner_color = QColor(40, 40, 45)

        # 水平标尺（顶部）
        painter.fillRect(R, 0, vp.width() - R, R, ruler_color)
        # 垂直标尺（左侧）
        painter.fillRect(0, R, R, vp.height() - R, ruler_color)
        # 左上角
        painter.fillRect(0, 0, R, R, corner_color)

        painter.setPen(QPen(tick_color, 1))
        font = QFont("Arial", 7)
        painter.setFont(font)
        painter.setPen(text_color)
        display_unit = self.scene().config_manager.get('display_unit', 'mm')
        painter.drawText(3, R - 5, "mm" if display_unit == 'mm' else "px")

        # 水平刻度
        scene_left = self.mapToScene(QPoint(R, 0)).x()
        scene_right = self.mapToScene(QPoint(vp.width(), 0)).x()
        step = self._ruler_step()
        x = math.floor(scene_left / step) * step
        while x <= scene_right:
            vx = self.mapFromScene(QPointF(x, 0)).x()
            painter.setPen(QPen(tick_color, 1))
            painter.drawLine(vx, R - 6, vx, R)
            if vx > R + 2:
                painter.setPen(text_color)
                painter.drawText(vx + 2, R - 2, self._format_ruler_label(x))
            x += step

        # 垂直刻度
        scene_top = self.mapToScene(QPoint(0, R)).y()
        scene_bottom = self.mapToScene(QPoint(0, vp.height())).y()
        y = math.floor(scene_top / step) * step
        while y <= scene_bottom:
            vy = self.mapFromScene(QPointF(0, y)).y()
            painter.setPen(QPen(tick_color, 1))
            painter.drawLine(R - 6, vy, R, vy)
            if vy > R + 2:
                painter.setPen(text_color)
                painter.drawText(2, vy - 1, self._format_ruler_label(y))
            y += step

        # 鼠标位置十字线（在标尺上）
        cursor_vp = self.viewport().mapFromGlobal(self.cursor().pos())
        if vp.contains(cursor_vp):
            painter.setPen(QPen(QColor(0, 180, 255, 200), 1))
            painter.drawLine(cursor_vp.x(), 0, cursor_vp.x(), R)
            painter.drawLine(0, cursor_vp.y(), R, cursor_vp.y())

    def _ruler_step(self):
        """根据当前缩放计算合适的标尺刻度间距（内部返回场景像素）。"""
        scale = self.transform().m11()
        if self.scene().config_manager.get('display_unit', 'mm') == 'px':
            for step_px in [10, 20, 50, 100, 200, 500, 1000, 2000]:
                if step_px * scale >= 40:
                    return step_px
            return 2000
        px_per_mm = CORELDRAW_EXPORT_DPI / 25.4
        for step_mm in [1, 2, 5, 10, 20, 50, 100, 200, 500]:
            step_px = step_mm * px_per_mm
            if step_px * scale >= 40:
                return step_px
        return 500 * px_per_mm

    def _format_ruler_label(self, scene_value):
        if self.scene().config_manager.get('display_unit', 'mm') == 'px':
            return str(int(round(scene_value)))
        value_mm = scene_value * 25.4 / CORELDRAW_EXPORT_DPI
        if abs(value_mm) >= 100 or abs(value_mm - round(value_mm)) < 0.01:
            return str(int(round(value_mm)))
        return f"{value_mm:.1f}".rstrip('0').rstrip('.')

    def contextMenuEvent(self, event):
        """右键：点到元素交给元素处理，空白处弹画布菜单"""
        scene_pos = self.mapToScene(event.pos())

        # 遍历该点所有 items（包括锁定的不可选元素）
        items_at = self.scene().items(scene_pos, Qt.ItemSelectionMode.IntersectsItemShape,
                                      Qt.SortOrder.DescendingOrder, self.transform())
        target = None
        # 独立连接点位于画布顶层，优先交给连接点自身处理右键菜单。
        for it in items_at:
            if isinstance(it, ConnectionPoint) and getattr(it, 'point_type', '') == 'custom':
                it._show_context_menu(event.globalPos())
                event.accept()
                return
        for it in items_at:
            if isinstance(it, VTextItem):
                target = it
                break
        if target is None:
            for it in items_at:
                if isinstance(it, VImageItem):
                    target = it
                    break
        if target is None:
            for it in items_at:
                if isinstance(it, BaseElement):
                    target = it
                    break

        if target is not None:
            if isinstance(target, VTextItem):
                target._show_context_menu(event.globalPos(), target.mapFromScene(scene_pos))
            else:
                target._show_context_menu(event.globalPos())
            event.accept()
            return

        if self._main_window:
            self._main_window.show_canvas_context_menu(event.pos())
        event.accept()

    def mousePressEvent(self, event):
        pos = event.pos()
        if self._smart_brush_enabled and event.button() == Qt.MouseButton.LeftButton:
            self._smart_brush_active = True
            self._smart_brush_seen.clear()
            self._smart_brush_pos = pos
            self._apply_smart_brush_at(pos, bool(event.modifiers() & Qt.KeyboardModifier.AltModifier))
            self.viewport().update()
            event.accept()
            return
        # 处理中键平移
        if event.button() == Qt.MouseButton.MiddleButton:
            self._is_panning = True
            self._pan_start = event.pos()
            self.setCursor(Qt.CursorShape.ClosedHandCursor)
            event.accept()
            return

        if event.button() == Qt.MouseButton.LeftButton and self.scene():
            exclude_point = self.scene().connection_source_point if self.scene().connection_mode else None
            point = self._connection_point_at(pos, exclude_point=exclude_point)
            if point:
                is_custom = getattr(point, 'point_type', '') == 'custom'
                point_mode = getattr(self.scene(), 'free_connection_point_mode', 'move')
                if is_custom and point_mode == 'move':
                    # 移动模式交给 QGraphicsItem 处理，支持单击、拖动和框选。
                    super().mousePressEvent(event)
                else:
                    # 图片/文字连接点保持原有单击连线；独立点仅在连接模式下连线。
                    self.scene().start_connection_from_point(point)
                    event.accept()
                event.accept()
                return

        if event.button() in (Qt.MouseButton.LeftButton, Qt.MouseButton.RightButton):
            # 1. 优先处理标尺点击逻辑
            if self._in_h_ruler(pos) or self._in_v_ruler(pos):
                if self._in_h_ruler(pos):
                    self._guide_dragging = True
                    self._guide_orientation = Qt.Orientation.Horizontal
                    scene_y = self.mapToScene(pos).y()
                    self._guide_preview = self.scene().add_guide(Qt.Orientation.Horizontal, scene_y)
                elif self._in_v_ruler(pos):
                    self._guide_dragging = True
                    self._guide_orientation = Qt.Orientation.Vertical
                    scene_x = self.mapToScene(pos).x()
                    self._guide_preview = self.scene().add_guide(Qt.Orientation.Vertical, scene_x)
                event.accept()
                return

            # 2. 检查点击位置的物体状态（实现“穿透”锁定物体的关键）
            raw_hit = self.itemAt(pos)

            # 如果点击的是辅助线，直接交给场景处理，不触发框选
            if isinstance(raw_hit, GuideItem):
                super(LayoutView, self).mousePressEvent(event)
                return
            self.scene().clear_guide_selection()

            hit_element = raw_hit
            # 向上追溯，直到找到 BaseElement（即我们的 VImageItem 或 VTextItem）
            while hit_element and not isinstance(hit_element, BaseElement):
                hit_element = hit_element.parentItem()
            
            # 判断是否为锁定状态
            is_locked = False
            if hit_element:
                if isinstance(hit_element, VImageItem) and hit_element.locked:
                    is_locked = True
                # 如果未来有锁定的文字，也可以在这里增加判断

            # 3. 开启框选逻辑的条件：点在空白处 OR 点在锁定的物体上
            if hit_element is None or is_locked:
                self._marquee_active = True
                self._marquee_origin = pos
                self._marquee_sweep_order = []  # 开始新的框选，清空扫过记录
                self._update_marquee_style(pos)
                self._marquee_band.setGeometry(QRect(self._marquee_origin, QSize()))
                self._marquee_band.show()
                # 记录是哪个键触发的框选，以便在 Release 事件中对应
                self._marquee_trigger_button = event.button()
                event.accept()
                return
            
            # 4. 仅图片模式下，点击连线时阻止选中并弹提示
            if hit_element is None and raw_hit is not None:
                # raw_hit 可能是连线
                if isinstance(raw_hit, (VImageTextConnector, VGenericConnector)):
                    mode = self.scene().config_manager.get('marquee_mode', 'all')
                    if mode == 'images':
                        if self._main_window and hasattr(self._main_window, '_toast'):
                            self._main_window._toast.show_message(
                                "当前模式：仅图片\n连线无法选中\nAlt+M 切换为「全选」模式", "⚠", 4000)
                        event.accept()
                        return

            # 5. 如果点击的是普通非锁定物体，执行系统默认的选择/拖拽逻辑
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        # 智能笔刷模式下不需要连接点悬停扫描；该扫描会遍历整个场景。
        if not self._smart_brush_enabled:
            self._sync_connection_point_hover(event.pos())
        if self._smart_brush_enabled:
            old_pos = self._smart_brush_pos
            self._smart_brush_pos = event.pos()
            if self._smart_brush_active and event.buttons() & Qt.MouseButton.LeftButton:
                self._apply_smart_brush_at(event.pos(), bool(event.modifiers() & Qt.KeyboardModifier.AltModifier))
                self._update_smart_brush_overlay(old_pos)
                event.accept()
                return
            self._update_smart_brush_overlay(old_pos)

        # 标尺辅助线拖拽预览
        if self._guide_dragging and self._guide_preview:
            scene_pos = self.mapToScene(event.pos())
            if self._guide_orientation == Qt.Orientation.Horizontal:
                self._guide_preview.pos_value = scene_pos.y()
            else:
                self._guide_preview.pos_value = scene_pos.x()
            self._guide_preview._update_pos()
            self._guide_preview.update()
            self.update() 
            event.accept()
            return

        # CAD 风格选框更新
        if self._marquee_active:
            self._update_marquee_style(event.pos())
            current_rect = QRect(self._marquee_origin, event.pos()).normalized()
            self._marquee_band.setGeometry(current_rect)

            # 实时追踪鼠标扫过的元素顺序
            scene = self.scene()
            if scene:
                scene_polygon = self.mapToScene(current_rect)
                path = QPainterPath()
                path.addPolygon(scene_polygon)
                mode = scene.config_manager.get('marquee_mode', 'all')
                items_in_rect = scene.items(path, self._marquee_mode,
                                            Qt.SortOrder.DescendingOrder, self.transform())
                newly_entered = []
                for item in items_in_rect:
                    if item in self._marquee_sweep_order:
                        continue
                    if (isinstance(item, ConnectionPoint)
                            and getattr(item, 'point_type', '') == 'custom'
                            and getattr(scene, 'free_connection_point_mode', 'move') == 'move'
                            and item.isVisible()
                            and (item.flags() & QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)):
                        newly_entered.append(item)
                    elif mode == 'images':
                        if isinstance(item, VImageItem) and not getattr(item, 'locked', False):
                            newly_entered.append(item)
                    elif mode == 'connected':
                        if isinstance(item, (VImageItem, VTextItem)) and not getattr(item, 'locked', False):
                            cp = getattr(item, 'connection_point', None)
                            if cp and cp.isVisible():
                                newly_entered.append(item)
                    else:
                        if isinstance(item, (VImageItem, VTextItem)) and \
                                (item.flags() & QGraphicsItem.GraphicsItemFlag.ItemIsSelectable):
                            newly_entered.append(item)

                # 新进入的元素按连接点到鼠标距离排序（最近的排前面，模拟扫过顺序）
                if newly_entered:
                    mouse_scene = self.mapToScene(event.pos())

                    def dist_to_mouse(it):
                        cp = getattr(it, 'connection_point', None)
                        if cp and cp.isVisible():
                            p = cp.get_scene_center()
                        else:
                            p = it.scenePos()
                        return (p.x() - mouse_scene.x()) ** 2 + (p.y() - mouse_scene.y()) ** 2

                    newly_entered.sort(key=dist_to_mouse)
                    self._marquee_sweep_order.extend(newly_entered)

            self.viewport().update()
            event.accept()
            return

        # 视图平移
        if self._is_panning:
            delta = event.pos() - self._pan_start
            self._pan_start = event.pos()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value() - delta.x())
            self.verticalScrollBar().setValue(self.verticalScrollBar().value() - delta.y())
            event.accept()
        else:
            # 更新标尺的光标跟随线
            self.update()
            super().mouseMoveEvent(event)
            self._sync_connection_point_hover(event.pos())

    def leaveEvent(self, event):
        if self._hovered_connection_point:
            self._hovered_connection_point._hovered = False
            self._hovered_connection_point.update()
            self._hovered_connection_point = None
        self._smart_brush_active = False
        super().leaveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._smart_brush_enabled and event.button() == Qt.MouseButton.LeftButton:
            self._smart_brush_active = False
            self._smart_brush_seen.clear()
            self.viewport().update()
            # 完成选择后自动退出笔刷模式
            if self._main_window:
                self._main_window.toggle_smart_brush(False)
            event.accept()
            return

        # 结束辅助线拖拽
        if self._guide_dragging:
            self._guide_dragging = False
            self._guide_preview = None
            self._guide_orientation = None
            event.accept()
            return

        # 结束框选并应用选择逻辑
        if self._marquee_active and event.button() == self._marquee_trigger_button:
            self._marquee_active = False
            self._apply_marquee_selection()
            # 用框选过程中实时追踪的扫过顺序覆盖 selection_order
            scene = self.scene()
            if scene and self._marquee_sweep_order:
                # 只保留最终确实被选中的元素，顺序不变
                selected_set = set(scene.selectedItems())
                scene.selection_order = [item for item in self._marquee_sweep_order
                                         if item in selected_set]
            scene.last_selection_by_marquee = True  # 标记为框选
            self._marquee_sweep_order = []
            self._marquee_band.hide()
            event.accept()
            return

        # 结束平移
        if event.button() == Qt.MouseButton.MiddleButton:
            self._is_panning = False
            self.setCursor(Qt.CursorShape.ArrowCursor)
            event.accept()
        else:
            super().mouseReleaseEvent(event)

    def keyPressEvent(self, event):
        """按住 F5 时，独立连接点临时进入移动模式。"""
        if event.key() == Qt.Key.Key_F5 and self.scene():
            scene = self.scene()
            if (getattr(scene, 'free_connection_point_mode', 'connect') == 'connect'
                    and not getattr(self, '_free_point_f5_move_active', False)):
                self._free_point_f5_move_active = True
                scene.set_free_connection_point_mode('move', cancel_pending_connection=False)
            event.accept()
            return
        super().keyPressEvent(event)

    def keyReleaseEvent(self, event):
        if event.key() == Qt.Key.Key_F5 and self.scene():
            if getattr(self, '_free_point_f5_move_active', False):
                self._free_point_f5_move_active = False
                self.scene().set_free_connection_point_mode('connect', cancel_pending_connection=False)
            event.accept()
            return
        super().keyReleaseEvent(event)
    
    def mouseDoubleClickEvent(self, event):
        """双击空白处：调整图片模式下确认调整"""
        if event.button() == Qt.MouseButton.LeftButton and self.scene():
            if self.scene().resize_mode:
                # 检查是否双击在空白处（不是图片上）
                item = self.itemAt(event.pos())
                element = item
                while element and not isinstance(element, VImageItem):
                    element = element.parentItem() if element else None
                if not element:
                    self.scene().confirm_resize()
                    event.accept()
                    return
        super().mouseDoubleClickEvent(event)

    def wheelEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ShiftModifier:
            # Shift+滚轮 保留原始滚动行为
            super().wheelEvent(event)
        else:
            # 滚轮直接缩放（以鼠标位置为中心）
            old_pos = self.mapToScene(event.position().toPoint())
            scale = 1.1 if event.angleDelta().y() > 0 else 0.9
            self.scale(scale, scale)
            new_pos = self.mapToScene(event.position().toPoint())
            delta = new_pos - old_pos
            self.translate(delta.x(), delta.y())
            self.transformChanged.emit()
    
    def fit_in_view(self):
        """合适屏幕 - 让整个画布区域适合视图"""
        scene = self.scene()
        if scene:
            # 始终适配整个场景矩形（画布区域）
            self.fitInView(scene.sceneRect(), Qt.AspectRatioMode.KeepAspectRatio)
            self.transformChanged.emit()
    
    def fill_view(self):
        """填充屏幕 - 让画布内容填满整个视图"""
        scene = self.scene()
        if scene:
            # 获取所有可见元素的边界
            items_rect = scene.itemsBoundingRect()
            if not items_rect.isEmpty():
                self.fitInView(items_rect, Qt.AspectRatioMode.KeepAspectRatioByExpanding)
            else:
                # 如果没有元素，填充整个场景
                self.fitInView(scene.sceneRect(), Qt.AspectRatioMode.KeepAspectRatioByExpanding)
            self.transformChanged.emit()
    
    def actual_size(self):
        """实际大小 - 100%缩放"""
        self.resetTransform()
        self.transformChanged.emit()
    
    def zoom_in(self):
        """放大"""
        self.scale(1.25, 1.25)
        self.transformChanged.emit()
    
    def zoom_out(self):
        """缩小"""
        self.scale(0.8, 0.8)
        self.transformChanged.emit()
    
    def zoom_to_selection(self):
        """缩放到选中内容"""
        scene = self.scene()
        if scene:
            selected_items = scene.selectedItems()
            if selected_items:
                # 计算选中项目的边界
                selection_rect = QRectF()
                for item in selected_items:
                    if hasattr(item, 'boundingRect'):
                        item_rect = item.boundingRect()
                        scene_rect = QRectF(item.scenePos(), item_rect.size())
                        if selection_rect.isEmpty():
                            selection_rect = scene_rect
                        else:
                            selection_rect = selection_rect.united(scene_rect)
                
                if not selection_rect.isEmpty():
                    # 添加边距
                    margin = 30
                    selection_rect.adjust(-margin, -margin, margin, margin)
                    self.fitInView(selection_rect, Qt.AspectRatioMode.KeepAspectRatio)
                    self.transformChanged.emit()
    def dragEnterEvent(self, event):
        mime = event.mimeData()
        if (mime.hasFormat('application/x-group-asset-id') or
                mime.hasUrls() or
                mime.hasImage() or
                mime.hasFormat('image/png') or
                mime.hasFormat('image/jpeg')):
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        mime = event.mimeData()
        if (mime.hasFormat('application/x-group-asset-id') or
                mime.hasUrls() or
                mime.hasImage() or
                mime.hasFormat('image/png') or
                mime.hasFormat('image/jpeg')):
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)

    def dropEvent(self, event):
        # 处理素材库组合
        if event.mimeData().hasFormat('application/x-group-asset-id'):
            data = event.mimeData().data('application/x-group-asset-id')
            asset_id = int(data.data().decode())
            
            # 转换为画布精确坐标
            drop_pos = self.mapToScene(event.position().toPoint())
            
            scene = self.scene()
            asset = next((a for a in scene.asset_manager.get_group_assets() if a['id'] == asset_id), None)
            
            if asset:
                main_win = self.window()
                while main_win and not hasattr(main_win, 'asset_library_dock'):
                    main_win = main_win.parent()
                
                if main_win:
                    # 放置素材
                    main_win.asset_library_dock._place_group_at(asset, drop_pos)
                    # 自动让画面跳到放置点
                    self.centerOn(drop_pos)
                    scene.update()
            
            event.acceptProposedAction()
            return

        # 处理文件拖入
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                path = url.toLocalFile()
                
                # 处理 Excel 文件
                if path.lower().endswith(('.xlsx', '.xls')):
                    mw = self.window()
                    while mw and not hasattr(mw, 'open_family_tree_import_dialog'):
                        mw = mw.parent()
                    if mw:
                        # 使用延迟调用，避免拖放事件处理中的模态对话框问题
                        QTimer.singleShot(100, lambda: mw._open_excel_import_with_path(path))
                    event.acceptProposedAction()
                    return
                
                # 处理图片文件
                elif path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
                    pix = QPixmap(path)
                    mw = self.window()
                    w, h = mw._calc_insert_size(pix, self.scene()) if not pix.isNull() else (DEFAULT_FONT_SIZE * 4, 0)
                    img = VImageItem(path, target_width=w, target_height=h)
                    if self.scene().config_manager.get('insert_image_fit_canvas', False):
                        img.setPos(self.scene().sceneRect().topLeft())
                    else:
                        img.setPos(self.mapToScene(event.position().toPoint()))
                    if self.scene().config_manager.get('insert_image_to_bottom', False):
                        img.setZValue(-1)
                    self.scene().add_item_with_undo(img)
            event.acceptProposedAction()
            return

        # 处理从其他软件拖入的图片数据（image/png、image/jpeg 等）
        mime = event.mimeData()
        pix = None
        if mime.hasImage():
            img_data = mime.imageData()
            if img_data:
                pix = QPixmap.fromImage(img_data)
        elif mime.hasFormat('image/png') or mime.hasFormat('image/jpeg'):
            fmt = 'image/png' if mime.hasFormat('image/png') else 'image/jpeg'
            ba = mime.data(fmt)
            pix = QPixmap()
            pix.loadFromData(ba)

        if pix and not pix.isNull():
            # 保存为临时文件
            import tempfile
            suffix = '.png'
            tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False,
                                              dir=ASSETS_DIR, prefix='dropped_')
            tmp_path = tmp.name
            tmp.close()
            pix.save(tmp_path, 'PNG')

            pos = self.mapToScene(event.position().toPoint())
            mw = self.window()
            target_w = mw._calc_insert_width(pix, self.scene()) if not pix.isNull() else DEFAULT_FONT_SIZE * 4
            img = VImageItem(tmp_path, target_width=target_w)
            if self.scene().config_manager.get('insert_image_fit_canvas', False):
                img.setPos(self.scene().sceneRect().topLeft())
            else:
                img.setPos(pos)
            if self.scene().config_manager.get('insert_image_to_bottom', False):
                img.setZValue(-1)
            self.scene().add_item_with_undo(img)
            event.acceptProposedAction()
    

   

        

# --- Main Window ---

class FamilyTreeImportDialog(QDialog):
    """从 Excel 批量导入族谱成员对话框"""

    def __init__(self, scene, parent=None):
        super().__init__(parent)
        self.scene = scene
        self.setWindowTitle("批量导入族谱成员")
        self.setMinimumWidth(520)
        self._build_ui()
        self._load_settings()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(10)

        # Excel 文件路径
        file_group = QGroupBox("Excel 文件")
        fl = QHBoxLayout(file_group)
        self.txt_excel = QLineEdit()
        self.txt_excel.setPlaceholderText("请选择 Excel 文件...")
        fl.addWidget(self.txt_excel)
        btn_browse = QPushButton("浏览...")
        btn_browse.setFixedWidth(70)
        btn_browse.clicked.connect(self._browse_excel)
        fl.addWidget(btn_browse)
        root.addWidget(file_group)

        # Excel 列映射说明
        hint = QLabel(
            "Excel 默认格式：A列=辈分  B列=主文字  C列=副文字1  D列=副文字2  E列=图片路径\n"
            "使用组合模板时，模板中文字内容需与下方占位符一致"
        )
        hint.setStyleSheet("color: #666; font-size: 11px;")
        hint.setWordWrap(True)
        root.addWidget(hint)

        # 模板素材选择
        tmpl_group = QGroupBox("组合素材模板（可选）")
        tl = QHBoxLayout(tmpl_group)
        tl.addWidget(QLabel("输入/选择模板名称："))
        self.combo_template = QComboBox()
        self.combo_template.setEditable(True)
        self.combo_template.addItem("— 不使用模板，纯文字 —", None)
        for asset in self.scene.asset_manager.get_group_assets():
            self.combo_template.addItem(asset['name'], asset)
        tl.addWidget(self.combo_template, 1)
        root.addWidget(tmpl_group)

        template_map_group = QGroupBox("模板文字对象指定")
        tm = QFormLayout(template_map_group)
        self.combo_template_t1 = QComboBox()
        self.combo_template_t2 = QComboBox()
        self.combo_template_t3 = QComboBox()
        tm.addRow("主文字填入:", self.combo_template_t1)
        tm.addRow("副文字1填入:", self.combo_template_t2)
        tm.addRow("副文字2填入:", self.combo_template_t3)
        self.btn_save_template_map = QPushButton("保存此模板指定")
        self.btn_save_template_map.clicked.connect(self._save_template_map_clicked)
        tm.addRow("", self.btn_save_template_map)
        root.addWidget(template_map_group)
        self._template_map_loading = False
        self._current_template_key = None
        self.combo_template.currentIndexChanged.connect(self._on_template_changed)
        if self.combo_template.lineEdit():
            self.combo_template.lineEdit().editingFinished.connect(self._on_template_name_edited)
        for combo in (self.combo_template_t1, self.combo_template_t2, self.combo_template_t3):
            combo.currentIndexChanged.connect(self._save_current_template_text_map)

        # 布局参数
        layout_group = QGroupBox("布局参数")
        lg = QFormLayout(layout_group)
        lg.setHorizontalSpacing(12)

        self.combo_unit = QComboBox()
        self.combo_unit.addItem("像素 px", "px")
        self.combo_unit.addItem("毫米 mm", "mm")
        self._current_unit = "px"
        self._loading_unit = False
        self.combo_unit.currentIndexChanged.connect(self._on_unit_changed)
        lg.addRow("尺寸单位:", self.combo_unit)

        self.spin_start_x = QDoubleSpinBox()
        self.spin_start_x.setRange(-99999, 99999)
        self.spin_start_x.setValue(500)
        self.spin_start_x.setSuffix(" px")
        lg.addRow("起始 X（第1人右侧）:", self.spin_start_x)

        self.spin_spacing = QDoubleSpinBox()
        self.spin_spacing.setRange(1, 9999)
        self.spin_spacing.setValue(200)
        self.spin_spacing.setSuffix(" px")
        lg.addRow("节点间距（X方向）:", self.spin_spacing)

        root.addWidget(layout_group)

        # 各辈分 Y 坐标
        y_group = QGroupBox("各辈分 Y 坐标")
        yg = QFormLayout(y_group)
        yg.setHorizontalSpacing(12)
        self.y_spins = []
        defaults = [500, 1000, 1500, 2000, 2500]
        for i in range(5):
            sp = QDoubleSpinBox()
            sp.setRange(-99999, 99999)
            sp.setValue(defaults[i])
            sp.setSuffix(" px")
            yg.addRow(f"第 {i+1} 代 / {'一二三四五'[i]}代:", sp)
            self.y_spins.append(sp)
        root.addWidget(y_group)

        # 按钮
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("导入")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)
        self._refresh_template_text_map()
        self._update_unit_suffix()

    def _template_asset_key(self, asset):
        if not asset:
            return ''
        return str(asset.get('name', '')).strip()

    def _find_template_asset_by_name(self, name):
        name = str(name).strip()
        if not name:
            return None
        for asset in self.scene.asset_manager.get_group_assets():
            if str(asset.get('name', '')).strip() == name:
                return asset
        return None

    def _current_template_asset(self):
        text_asset = self._find_template_asset_by_name(self.combo_template.currentText())
        if text_asset:
            return text_asset
        return self.combo_template.currentData()

    def _on_template_name_edited(self):
        asset = self._find_template_asset_by_name(self.combo_template.currentText())
        if not asset:
            self._refresh_template_text_map()
            return
        idx = self.combo_template.findText(asset['name'])
        if idx >= 0:
            self.combo_template.setCurrentIndex(idx)
        else:
            self._refresh_template_text_map()

    def _on_template_changed(self):
        self._refresh_template_text_map()

    def _saved_template_maps(self):
        maps = self.scene.config_manager.get('ftimport_template_text_maps', {})
        return maps if isinstance(maps, dict) else {}

    def _save_current_template_text_map(self):
        if getattr(self, '_template_map_loading', False):
            return
        key = self._current_template_key
        if not key:
            return
        maps = self._saved_template_maps()
        maps[key] = {
            't1': self.combo_template_t1.currentData(),
            't2': self.combo_template_t2.currentData(),
            't3': self.combo_template_t3.currentData(),
        }
        self.scene.config_manager.set('ftimport_template_text_maps', maps)

    def _save_template_map_clicked(self):
        self._on_template_name_edited()
        self._save_current_template_text_map()

    def _restore_combo_data(self, combo, value):
        for idx in range(combo.count()):
            if combo.itemData(idx) == value:
                combo.setCurrentIndex(idx)
                return

    def _update_unit_suffix(self):
        unit = self.combo_unit.currentData() or "px"
        suffix = f" {unit}"
        for spin in (self.spin_start_x, self.spin_spacing, *self.y_spins):
            spin.setSuffix(suffix)

    def _on_unit_changed(self):
        new_unit = self.combo_unit.currentData() or "px"
        old_unit = getattr(self, '_current_unit', 'px')
        if getattr(self, '_loading_unit', False) or new_unit == old_unit:
            self._current_unit = new_unit
            self._update_unit_suffix()
            return

        px_per_mm = CORELDRAW_EXPORT_DPI / 25.4
        if old_unit == 'px' and new_unit == 'mm':
            factor = 1 / px_per_mm
        elif old_unit == 'mm' and new_unit == 'px':
            factor = px_per_mm
        else:
            factor = 1

        for spin in (self.spin_start_x, self.spin_spacing, *self.y_spins):
            spin.setValue(spin.value() * factor)

        self._current_unit = new_unit
        self._update_unit_suffix()

    def _refresh_template_text_map(self):
        self._save_current_template_text_map()
        asset = self._current_template_asset()
        self._current_template_key = self._template_asset_key(asset)
        combos = (self.combo_template_t1, self.combo_template_t2, self.combo_template_t3)
        self._template_map_loading = True
        for combo in combos:
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("按占位符自动识别", None)

        if asset:
            text_count = 0
            for idx, item_data in enumerate(asset.get('items', [])):
                if item_data.get('type') != 'VTextItem':
                    continue
                text_count += 1
                text = item_data.get('text', '').replace('\n', '↵')
                if len(text) > 24:
                    text = text[:24] + '...'
                label = f"[{idx}] {text or '空文字'}"
                for combo in combos:
                    combo.addItem(label, idx)
            if text_count == 0:
                for combo in combos:
                    combo.setItemText(0, "当前模板没有文字对象")

        for combo in combos:
            combo.blockSignals(False)
        saved_map = self._saved_template_maps().get(self._current_template_key, {})
        if isinstance(saved_map, dict):
            self._restore_combo_data(self.combo_template_t1, saved_map.get('t1'))
            self._restore_combo_data(self.combo_template_t2, saved_map.get('t2'))
            self._restore_combo_data(self.combo_template_t3, saved_map.get('t3'))
        self._template_map_loading = False

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)"
        )
        if path:
            self.txt_excel.setText(path)
            self._save_settings()

    def _load_settings(self):
        cfg = self.scene.config_manager
        self.txt_excel.setText(cfg.get('ftimport_excel_path', ''))
        unit = cfg.get('ftimport_unit', 'px')
        idx = self.combo_unit.findData(unit)
        self._loading_unit = True
        if idx >= 0:
            self.combo_unit.setCurrentIndex(idx)
        self._current_unit = unit
        self._loading_unit = False
        self.spin_start_x.setValue(cfg.get('ftimport_start_x', 500))
        self.spin_spacing.setValue(cfg.get('ftimport_spacing', 200))
        defaults = [500, 1000, 1500, 2000, 2500]
        for i, sp in enumerate(self.y_spins):
            sp.setValue(cfg.get(f'ftimport_y{i+1}', defaults[i]))
        template_name = cfg.get('ftimport_template_name', '')
        if template_name:
            idx = self.combo_template.findText(template_name)
            if idx >= 0:
                self.combo_template.setCurrentIndex(idx)
            else:
                self.combo_template.setCurrentText(template_name)
        self._refresh_template_text_map()
        self._update_unit_suffix()

    def _save_settings(self):
        cfg = self.scene.config_manager
        cfg.set('ftimport_excel_path', self.txt_excel.text())
        cfg.set('ftimport_template_name', self.combo_template.currentText().strip())
        self._save_current_template_text_map()
        cfg.set('ftimport_unit', self.combo_unit.currentData() or 'px')
        cfg.set('ftimport_start_x', self.spin_start_x.value())
        cfg.set('ftimport_spacing', self.spin_spacing.value())
        for i, sp in enumerate(self.y_spins):
            cfg.set(f'ftimport_y{i+1}', sp.value())

    def _get_t1_font(self):
        """从模板的B组（t1）文字对象取字号和字体，找不到时用系统默认"""
        cfg = self.scene.config_manager
        default_size = cfg.get('default_font_size', DEFAULT_FONT_SIZE)
        default_family = cfg.get('default_font_family', DEFAULT_FONT)
        asset = self._current_template_asset()
        if not asset:
            return default_size, default_family
        t1_index = self.combo_template_t1.currentData()
        if t1_index is None:
            # 没有手动指定时，取第一个文字对象
            for item_data in asset.get('items', []):
                if item_data.get('type') == 'VTextItem':
                    return item_data.get('font_size', default_size), item_data.get('font_family', default_family)
            return default_size, default_family
        items = asset.get('items', [])
        if t1_index < len(items) and items[t1_index].get('type') == 'VTextItem':
            d = items[t1_index]
            return d.get('font_size', default_size), d.get('font_family', default_family)
        return default_size, default_family

    def get_params(self):
        self._save_settings()
        unit = self.combo_unit.currentData() or 'px'
        scale = CORELDRAW_EXPORT_DPI / 25.4 if unit == 'mm' else 1.0
        def to_px(value):
            return value * scale
        font_size, font_family = self._get_t1_font()
        return {
            'excel_path': self.txt_excel.text().strip(),
            'unit': unit,
            'start_x': to_px(self.spin_start_x.value()),
            'spacing': to_px(self.spin_spacing.value()),
            'y_coords': [to_px(sp.value()) for sp in self.y_spins],
            'font_size': font_size,
            'font_family': font_family,
            'img_width': 150,
            'auto_connect': True,
            'template_name': self.combo_template.currentText().strip(),
            'template_asset': self._current_template_asset(),
            'placeholders': {
                't1': 't1',
                't2': 't2',
                't3': 't3',
            },
            'columns': {
                'generation': 'A',
                't1': 'B',
                't2': 'C',
                't3': 'D',
                'image': 'E',
                'name': '内容',
                'group': '组',
            },
            'group_match': {
                'enabled': True,
                't1': 'a',
                't2': 'c',
                't3': 'b',
                'extra_text': 'd',
            },
            'template_text_map': {
                't1': self.combo_template_t1.currentData(),
                't2': self.combo_template_t2.currentData(),
                't3': self.combo_template_t3.currentData(),
            },
        }


class DocumentState:
    """窗口内一个独立文档的运行时状态。"""
    def __init__(self, name, scene, view):
        self.name = name
        self.scene = scene
        self.view = view
        self.dirty = False


class PlusTabBar(QTabBar):
    """标签栏：最后一个标签后面紧跟一个 + 按钮"""
    addTabRequested = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._plus_btn = None
        # Keep tabs at their content width so the trailing + button remains
        # inside the tab bar instead of being pushed outside its clipped area.
        self.setExpanding(False)

    def _ensure_plus_btn(self):
        """延迟创建 + 按钮，确保父窗口已设置"""
        # Before QTabWidget.setTabBar() the tab bar has no container parent;
        # defer creation until it can be placed outside the clipped tab bar.
        button_parent = self.parentWidget()
        if button_parent is None:
            return
        if self._plus_btn is None:
            self._plus_btn = QToolButton(button_parent)
            self._plus_btn.setText("+")
            self._plus_btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
            self._plus_btn.setFixedSize(28, 28)
            self._plus_btn.setToolTip("新建文档 (Ctrl+N)")
            self._plus_btn.setStyleSheet(
                "QToolButton { "
                "   border: 1px solid #ccc; "
                "   border-radius: 4px; "
                "   font-size: 16px; "
                "   font-weight: bold; "
                "   background: #f0f0f0; "
                "   color: #333; "
                "   padding: 0px; "
                "   margin: 0px; "
                "   min-width: 0px; "
                "   min-height: 0px; "
                "}"
                "QToolButton:hover { "
                "   background: #e0e0e0; "
                "   border-color: #999; "
                "}"
                "QToolButton:pressed { "
                "   background: #d0d0d0; "
                "}"
            )
            self._plus_btn.clicked.connect(self.addTabRequested)
            self._plus_btn.setVisible(True)
            self._plus_btn.raise_()
            print("加号按钮已创建")

    def _move_plus_btn(self):
        """将 + 按钮放在最后一个标签的右侧"""
        self._ensure_plus_btn()
        if self._plus_btn is None:
            return
        count = self.count()
        if count > 0:
            last_rect = self.tabRect(count - 1)
            x = last_rect.right() + 2
            y = last_rect.top() + (last_rect.height() - self._plus_btn.height()) // 2
            print(f"移动加号按钮到: x={x}, y={y}, 标签数={count}, 标签宽度={last_rect.width()}")
        else:
            x, y = 4, 4
            print(f"无标签，加号按钮位置: x={x}, y={y}")
        button_parent = self._plus_btn.parentWidget()
        if button_parent is self:
            self._plus_btn.move(x, y)
        else:
            self._plus_btn.move(self.mapTo(button_parent, QPoint(x, y)))
        self._plus_btn.setVisible(True)
        self._plus_btn.raise_()

    def tabLayoutChange(self):
        super().tabLayoutChange()
        # QTabBar can invoke this virtual method during its base
        # constructor, before our instance attributes have been initialized.
        if hasattr(self, "_plus_btn"):
            self._move_plus_btn()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._move_plus_btn()

    def showEvent(self, event):
        super().showEvent(event)
        self._move_plus_btn()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._startup_fit_done = False
        self._startup_fit_last_viewport_size = None
        self._startup_fit_attempts = 0
        self.setWindowTitle("VertiLayout Pro - 竖排排版引擎")
        self.setGeometry(100, 100, 1400, 900)
        self.setMinimumSize(800, 600)  # 设置最小窗口尺寸，允许自由调整
        # 应用Fluent Design样式
        self.apply_fluent_design_style()
        
        self._documents = []
        self._asset_manager = None
        self._loading_documents = False
        self._current_project_path = None  # 当前工程文件路径
        self.scene = LayoutScene(self)
        self.scene.setSceneRect(0, 0, 7054, 5021)
        if self.scene.config_manager.get('startup_horizontal_guides_enabled', False):
            self.scene.add_startup_horizontal_guides()
        self.scene.selectionChanged.connect(self.on_selection_changed)
        self._asset_manager = self.scene.asset_manager
        self._toast = ToastNotification()
        self._last_selected_images = []  # 缓存最后一次选中的图片列表
        self.view = LayoutView(self.scene)
        self.view.set_main_window(self)
        self.view.transformChanged.connect(self.update_zoom_display)
        self.scene.changed.connect(lambda *_args, scene=self.scene: self._mark_scene_dirty(scene))
        self._documents.append(DocumentState("文档 1", self.scene, self.view))
        
        # 创建停靠面板
        # 右侧：层级 & 属性面板
        sidebar = QDockWidget("层级 & 属性", self)
        sidebar.setObjectName("dock_hierarchy")
        sidebar.setAllowedAreas(Qt.DockWidgetArea.LeftDockWidgetArea | Qt.DockWidgetArea.RightDockWidgetArea)
        self._eye_visible_icon = self._make_eye_icon(True)
        self._eye_hidden_icon = self._make_eye_icon(False)
        self.tree_widget = QTreeWidget()
        self.tree_widget.setHeaderLabels(["", "排版元素", "位置"])
        self.tree_widget.setColumnWidth(0, 50)
        self.tree_widget.setColumnWidth(1, 180)
        self.tree_widget.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tree_widget.setAllColumnsShowFocus(True)
        self.tree_widget.itemClicked.connect(self._on_tree_item_clicked)
        self.sidebar_tabs = QTabWidget()
        self.sidebar_tabs.setObjectName("sidebar_tabs")
        self.sidebar_tabs.addTab(self.tree_widget, "层级")
        self.property_panel = self.create_property_panel()
        self.sidebar_tabs.addTab(self.property_panel, "属性")
        sidebar.setWidget(self.sidebar_tabs)
        self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, sidebar)
        self._tree_item_counter = 0  # 全局编号计数器
        self._tree_updating = False   # 防止 itemChanged 递归触发
        self._tree_nodes_by_item = {}
        
        # 左侧：素材库面板
        self.asset_library_dock = AssetLibraryDockWidget(self.scene.asset_manager, self)
        self.asset_library_dock.setObjectName("dock_assets")
        self.asset_library_dock.setMinimumWidth(250)  # 设置最小宽度
        self.asset_library_dock.setMaximumWidth(400)  # 设置最大宽度
        self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, self.asset_library_dock)
        
        # 保持对旧版本素材库的引用（用于兼容性）
        self.asset_library = None
        
        self.document_tabs = QTabWidget()
        self.document_tabs.setMovable(False)
        self.document_tabs.setDocumentMode(False)
        # 使用自定义标签栏，+ 按钮紧跟在标签后面
        plus_tab_bar = PlusTabBar()
        plus_tab_bar.addTabRequested.connect(self.new_document)
        plus_tab_bar.setTabsClosable(True)
        self.document_tabs.setTabBar(plus_tab_bar)
        self.document_tabs.setTabsClosable(True)
        self.document_tabs.tabCloseRequested.connect(self.close_document)
        self.document_tabs.tabBarDoubleClicked.connect(self.rename_document)
        self.document_tabs.addTab(self.view, "文档 1")
        self._set_document_tab_close_text(0)
        self.document_tabs.setCurrentIndex(0)
        self.document_tabs.currentChanged.connect(self._on_document_changed)

        self.setCentralWidget(self.document_tabs)
        self.create_menu_bar()
        self.create_toolbars()
        # 同步框选模式 UI 初始状态
        self._sync_marquee_mode_ui()
        
        # 创建状态栏
        self.status_bar = self.statusBar()
        self.zoom_label = QLabel("缩放: 100%")
        self.status_bar.addPermanentWidget(self.zoom_label)
        
        # 快捷键提示
        hint = QLabel("  放大: Ctrl+=   缩小: Ctrl+-   合适屏幕: Ctrl+0   撤销: Ctrl+Z   保存: Ctrl+S   导出: Ctrl+E   素材库: F9")
        hint.setStyleSheet("color: gray; font-size: 11px;")
        hint.setText("  当前模式：编辑   选中对象：0")
        self.status_bar.addWidget(hint)
        
        # 导航器停靠面板
        nav_dock = QDockWidget("导航器", self)
        nav_dock.setObjectName("dock_navigator")
        nav_dock.setAllowedAreas(Qt.DockWidgetArea.LeftDockWidgetArea |
                                  Qt.DockWidgetArea.RightDockWidgetArea |
                                  Qt.DockWidgetArea.BottomDockWidgetArea)
        nav_dock.setFeatures(QDockWidget.DockWidgetFeature.DockWidgetMovable |
                              QDockWidget.DockWidgetFeature.DockWidgetClosable |
                              QDockWidget.DockWidgetFeature.DockWidgetFloatable)
        self.navigator = NavigatorWidget(self.view, self.scene)
        nav_dock.setWidget(self.navigator)
        nav_dock.setMaximumHeight(self.navigator.NAV_H + 30)  # 改为最大高度，不固定
        self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, nav_dock)
        
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.refresh_ui)
        self.timer.start(2000)
        
        print("Vertical Layout Engine Started...")
        # 恢复上次窗口大小和停靠面板布局
        self._restore_window_state()
        self._update_document_tab_title(self._documents[0])
        
        # 延迟触发加号按钮显示
        QTimer.singleShot(100, lambda: self.document_tabs.tabBar()._move_plus_btn())

    def _create_document(self, name):
        """创建一个独立的场景和视图。"""
        scene = LayoutScene(self)
        scene.asset_manager = self._asset_manager
        scene.setSceneRect(0, 0, 7054, 5021)
        if scene.config_manager.get('startup_horizontal_guides_enabled', False):
            scene.add_startup_horizontal_guides()
        scene.selectionChanged.connect(self.on_selection_changed)
        scene.changed.connect(lambda *_args, scene=scene: self._mark_scene_dirty(scene))
        view = LayoutView(scene)
        view.set_main_window(self)
        view.transformChanged.connect(self.update_zoom_display)
        return DocumentState(name, scene, view)

    def _document_has_content(self, document):
        return any(
            isinstance(item, (VImageItem, VTextItem))
            for item in document.scene.items()
        )

    def _update_document_tab_title(self, document):
        if not hasattr(self, 'document_tabs'):
            return
        try:
            index = self._documents.index(document)
        except ValueError:
            return
        title = document.name + (" *" if document.dirty else "")
        self.document_tabs.setTabText(index, title)
        if index == self.document_tabs.currentIndex():
            suffix = f" - {document.name}" if document.name else ""
            if self._current_project_path:
                suffix = f" - {os.path.basename(self._current_project_path)}{suffix}"
            self.setWindowTitle(f"VertiLayout Pro{suffix}")

    def _set_document_tab_close_text(self, index):
        """将文档标签关闭按钮的提示改为中文。"""
        if not hasattr(self, 'document_tabs'):
            return
        button = self.document_tabs.tabBar().tabButton(
            index,
            QTabBar.ButtonPosition.RightSide
        )
        if button:
            button.setToolTip("关闭文档")
            button.setAccessibleName("关闭文档")

    def _mark_scene_dirty(self, scene):
        if self._loading_documents:
            return
        for document in self._documents:
            if document.scene is scene:
                if not document.dirty:
                    document.dirty = True
                    self._update_document_tab_title(document)
                return

    def _on_document_changed(self, index):
        if index < 0 or index >= len(self._documents):
            return
        document = self._documents[index]
        self.scene = document.scene
        self.view = document.view
        if hasattr(self, 'navigator'):
            self.navigator.set_document(self.view, self.scene)
        self._last_selected_images = []
        self._update_document_tab_title(document)
        if hasattr(self, 'bg_above_connectors_action'):
            self.bg_above_connectors_action.blockSignals(True)
            self.bg_above_connectors_action.setChecked(
                self.scene.config_manager.get('bg_above_connectors', False)
            )
            self.bg_above_connectors_action.blockSignals(False)
        if hasattr(self, 'horizontal_move_only_action'):
            self.horizontal_move_only_action.blockSignals(True)
            self.horizontal_move_only_action.setChecked(
                self.scene.config_manager.get('horizontal_move_only', False)
            )
            self.horizontal_move_only_action.blockSignals(False)
        if hasattr(self, 'image_right_edge_snap_action'):
            self.image_right_edge_snap_action.blockSignals(True)
            self.image_right_edge_snap_action.setChecked(
                self.scene.config_manager.get('image_right_edge_snap_enabled', False)
            )
            self.image_right_edge_snap_action.blockSignals(False)
        if hasattr(self, 'text_preset_combo'):
            self._refresh_text_preset_combo()
        if hasattr(self, 'marquee_mode_combo'):
            self._sync_marquee_mode_ui()
        if hasattr(self, 'refresh_ui'):
            self.refresh_ui()
        if hasattr(self, 'fit_in_view'):
            self.view.transformChanged.emit()

    def new_document(self):
        """新建一个文档标签页。"""
        index = len(self._documents) + 1
        document = self._create_document(f"文档 {index}")
        self._documents.append(document)
        self.document_tabs.addTab(document.view, document.name)
        self._set_document_tab_close_text(self.document_tabs.indexOf(document.view))
        self.document_tabs.setCurrentWidget(document.view)
        self.view.fit_in_view()
        self.status_bar.showMessage(f"已新建 {document.name}", 2000)

    def rename_document(self, index=None):
        """重命名指定文档，默认重命名当前文档。"""
        if index is None:
            index = self.document_tabs.currentIndex()
        if index < 0 or index >= len(self._documents):
            return

        document = self._documents[index]
        new_name, ok = QInputDialog.getText(
            self,
            "重命名文档",
            "请输入文档名称：",
            text=document.name
        )
        new_name = new_name.strip()
        if ok and new_name and new_name != document.name:
            document.name = new_name
            document.dirty = True
            self._update_document_tab_title(document)
            self.status_bar.showMessage(f"文档已重命名为：{new_name}", 3000)

    def _ask_save_before_action(self, title):
        has_unsaved_content = any(
            document.dirty and self._document_has_content(document)
            for document in self._documents
        )
        if not has_unsaved_content:
            return True
        reply = QMessageBox.question(
            self,
            title,
            "当前工程有未保存的修改，是否先保存？",
            QMessageBox.StandardButton.Save |
            QMessageBox.StandardButton.Discard |
            QMessageBox.StandardButton.Cancel
        )
        if reply == QMessageBox.StandardButton.Cancel:
            return False
        if reply == QMessageBox.StandardButton.Save:
            return self.quick_save_proj()
        return True

    def close_document(self, index=None):
        """关闭一个文档标签页，至少保留一个空白文档。"""
        if index is None:
            index = self.document_tabs.currentIndex()
        if index < 0 or index >= len(self._documents):
            return
        document = self._documents[index]
        if document.dirty:
            reply = QMessageBox.question(
                self,
                "关闭文档",
                f"{document.name} 有未保存的修改，是否先保存整个工程？",
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel
            )
            if reply == QMessageBox.StandardButton.Cancel:
                return
            if reply == QMessageBox.StandardButton.Save and not self.quick_save_proj():
                return

        # Keep the document model in sync before removeTab() emits
        # currentChanged; otherwise the navigator can rebind to this view
        # immediately before it is deleted.
        self._documents.pop(index)
        self.document_tabs.removeTab(index)
        document.view.deleteLater()
        if not self._documents:
            self.new_document()
        else:
            self.document_tabs.setCurrentIndex(min(index, len(self._documents) - 1))
            self._on_document_changed(self.document_tabs.currentIndex())

    def _replace_documents(self, document_data):
        """用加载结果替换当前窗口内的全部文档。"""
        old_documents = self._documents[:]
        self.document_tabs.blockSignals(True)
        while self.document_tabs.count():
            self.document_tabs.removeTab(0)
        self.document_tabs.blockSignals(False)
        self._documents = []
        for document in old_documents:
            document.view.deleteLater()

        self._loading_documents = True
        try:
            for index, data in enumerate(document_data):
                document = self._create_document(data.get('name') or f"文档 {index + 1}")
                ProjectData.load_scene(document.scene, data.get('data', {}))
                document.dirty = False
                self._documents.append(document)
                self.document_tabs.addTab(document.view, document.name)
                self._set_document_tab_close_text(self.document_tabs.indexOf(document.view))
        finally:
            self._loading_documents = False

        if not self._documents:
            self.new_document()
            return
        self.document_tabs.setCurrentIndex(0)
        self._on_document_changed(0)

    def showEvent(self, event):
        super().showEvent(event)
        if not self._startup_fit_done:
            self._startup_fit_done = True
            self._fit_window_to_available_screen()
            self._startup_fit_last_viewport_size = None
            self._startup_fit_attempts = 0
            QTimer.singleShot(0, self._fit_startup_view)

    def _fit_startup_view(self):
        """Fit the current document once the visible viewport has a size."""
        if not self.isVisible() or not hasattr(self, 'view'):
            return
        viewport = self.view.viewport()
        if viewport.width() <= 0 or viewport.height() <= 0:
            QTimer.singleShot(100, self._fit_startup_view)
            return
        viewport_size = viewport.size()
        size_is_stable = viewport_size == self._startup_fit_last_viewport_size
        self._startup_fit_last_viewport_size = QSize(viewport_size)
        self._startup_fit_attempts += 1
        self.view.fit_in_view()
        if not size_is_stable and self._startup_fit_attempts < 10:
            QTimer.singleShot(100, self._fit_startup_view)

    def changeEvent(self, event):
        super().changeEvent(event)
        if event.type() == QEvent.Type.WindowStateChange:
            # 窗口状态改变后延迟执行，确保布局已完成
            if self.isMaximized() and self._startup_fit_done:
                QTimer.singleShot(100, self.fit_in_view)

    def _fit_window_to_available_screen(self):
        """启动时确保窗口落在当前屏幕可用区域内。"""
        if self.isMaximized() or self.isFullScreen():
            return
        screen = self.screen() or QApplication.primaryScreen()
        if not screen:
            return
        available = screen.availableGeometry()
        max_w = max(800, int(available.width() * 0.94))
        max_h = max(600, int(available.height() * 0.94))
        new_w = min(self.width(), max_w)
        new_h = min(self.height(), max_h)
        if new_w != self.width() or new_h != self.height():
            self.resize(new_w, new_h)
        frame = self.frameGeometry()
        if not available.contains(frame):
            frame.moveCenter(available.center())
            if frame.left() < available.left():
                frame.moveLeft(available.left())
            if frame.top() < available.top():
                frame.moveTop(available.top())
            if frame.right() > available.right():
                frame.moveRight(available.right())
            if frame.bottom() > available.bottom():
                frame.moveBottom(available.bottom())
            self.move(frame.topLeft())

    def _save_window_state(self):
        s = QSettings("VertiLayout", "VertiLayoutPro")
        s.setValue("geometry", self.saveGeometry())
        s.setValue("windowState", self.saveState())
        # 保存层级面板列宽
        s.setValue("tree_col0_width", self.tree_widget.columnWidth(0))
        s.setValue("tree_col1_width", self.tree_widget.columnWidth(1))
        s.setValue("tree_col2_width", self.tree_widget.columnWidth(2))

    def _restore_window_state(self):
        s = QSettings("VertiLayout", "VertiLayoutPro")
        geom = s.value("geometry")
        state = s.value("windowState")
        if geom:
            self.restoreGeometry(geom)
        if state:
            self.restoreState(state)
        # 恢复层级面板列宽
        if s.contains("tree_col0_width"):
            self.tree_widget.setColumnWidth(0, int(s.value("tree_col0_width")))
        if s.contains("tree_col1_width"):
            self.tree_widget.setColumnWidth(1, int(s.value("tree_col1_width")))
        if s.contains("tree_col2_width"):
            self.tree_widget.setColumnWidth(2, int(s.value("tree_col2_width")))

    def closeEvent(self, event):
        """窗口关闭时提示保存，停止定时器，保存窗口状态"""
        if any(document.dirty for document in self._documents):
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("保存工程")
            msg_box.setText("当前工程有未保存的修改，是否在关闭前保存？")
            btn_save = msg_box.addButton("保存", QMessageBox.ButtonRole.AcceptRole)
            btn_discard = msg_box.addButton("不保存退出", QMessageBox.ButtonRole.DestructiveRole)
            btn_cancel = msg_box.addButton("取消", QMessageBox.ButtonRole.RejectRole)
            msg_box.setDefaultButton(btn_save)
            msg_box.exec()
            clicked = msg_box.clickedButton()
            if clicked == btn_save:
                if not self.quick_save_proj():
                    event.ignore()
                    return
            elif clicked == btn_cancel:
                event.ignore()
                return

        if hasattr(self, 'timer'):
            self.timer.stop()
        self._save_window_state()
        event.accept()

    def create_property_panel(self):
        panel = QWidget()
        panel.setObjectName("property_panel")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        self.property_summary_label = QLabel("未选择对象")
        self.property_summary_label.setObjectName("propertySummary")
        self.property_summary_label.setWordWrap(True)
        layout.addWidget(self.property_summary_label)

        text_group = QGroupBox("文字样式")
        text_layout = QFormLayout(text_group)
        text_layout.setContentsMargins(10, 14, 10, 10)
        text_layout.setHorizontalSpacing(8)
        text_layout.setVerticalSpacing(8)

        self.font_combo = QComboBox()
        self.font_combo.setMinimumWidth(120)
        self.font_combo.setEditable(True)
        for f in QFontDatabase.families():
            self.font_combo.addItem(f)
        default_font = self.scene.config_manager.get('default_font_family', DEFAULT_FONT)
        idx = self.font_combo.findText(default_font)
        if idx >= 0:
            self.font_combo.setCurrentIndex(idx)
        else:
            self.font_combo.setCurrentText(default_font)
        self.font_combo.currentTextChanged.connect(
            lambda name: self.change_selected_font(QFont(name))
        )
        text_layout.addRow("字体", self.font_combo)

        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(8, 200)
        default_font_size = self.scene.config_manager.get('default_font_size', DEFAULT_FONT_SIZE)
        self.font_size_spin.setValue(default_font_size)
        self.font_size_spin.setSuffix("pt")
        self.font_size_spin.valueChanged.connect(self.change_selected_font_size)
        text_layout.addRow("字号", self.font_size_spin)

        self.color_button = QPushButton()
        self.color_button.setFixedHeight(30)
        self.color_button.setToolTip("选择文字颜色")
        self.color_button.clicked.connect(self.change_selected_color)
        text_layout.addRow("颜色", self.color_button)
        self._set_color_button_color(QColor("black"))
        layout.addWidget(text_group)

        layout_group = QGroupBox("竖排参数")
        layout_form = QFormLayout(layout_group)
        layout_form.setContentsMargins(10, 14, 10, 10)
        layout_form.setHorizontalSpacing(8)
        layout_form.setVerticalSpacing(8)

        self.chars_per_column_spin = QSpinBox()
        self.chars_per_column_spin.setRange(5, 50)
        self.chars_per_column_spin.setValue(15)
        self.chars_per_column_spin.setSuffix("字")
        self.chars_per_column_spin.valueChanged.connect(self.change_chars_per_column)
        layout_form.addRow("每列字数", self.chars_per_column_spin)

        self.column_spacing_spin = QSpinBox()
        self.column_spacing_spin.setRange(0, 200)
        self.column_spacing_spin.setValue(COLUMN_SPACING)
        self.column_spacing_spin.setSuffix("px")
        self.column_spacing_spin.valueChanged.connect(self.change_column_spacing)
        layout_form.addRow("列间距", self.column_spacing_spin)

        self.character_spacing_spin = QSpinBox()
        self.character_spacing_spin.setRange(-200, 200)
        self.character_spacing_spin.setValue(0)
        self.character_spacing_spin.setSuffix("px")
        self.character_spacing_spin.valueChanged.connect(self.change_character_spacing)
        layout_form.addRow("字间距", self.character_spacing_spin)

        self.manual_line_break_btn = QPushButton("手动换行")
        self.manual_line_break_btn.setCheckable(True)
        self.manual_line_break_btn.setChecked(True)
        self.manual_line_break_btn.setProperty("class", "toggle")
        self.manual_line_break_btn.toggled.connect(self.toggle_manual_line_break)
        layout_form.addRow("", self.manual_line_break_btn)
        layout.addWidget(layout_group)

        preset_group = QGroupBox("文字格式预设")
        preset_layout = QVBoxLayout(preset_group)
        preset_layout.setContentsMargins(10, 14, 10, 10)
        preset_layout.setSpacing(8)

        self.text_preset_combo = QComboBox()
        self.text_preset_combo.setToolTip("选择后应用到当前选中的文字")
        self.text_preset_combo.activated.connect(lambda _index: self.apply_selected_text_preset())
        preset_layout.addWidget(self.text_preset_combo)

        preset_actions = QHBoxLayout()
        apply_preset_btn = QPushButton("应用")
        apply_preset_btn.clicked.connect(self.apply_selected_text_preset)
        add_preset_btn = QPushButton("新增预设")
        add_preset_btn.clicked.connect(self.add_current_text_format_preset)
        preset_actions.addWidget(apply_preset_btn)
        preset_actions.addWidget(add_preset_btn)
        preset_layout.addLayout(preset_actions)

        default_actions = QHBoxLayout()
        set_default_btn = QPushButton("当前格式设为默认")
        set_default_btn.clicked.connect(self.set_current_text_format_as_default)
        delete_preset_btn = QPushButton("删除预设")
        delete_preset_btn.clicked.connect(self.delete_selected_text_preset)
        default_actions.addWidget(set_default_btn)
        default_actions.addWidget(delete_preset_btn)
        preset_layout.addLayout(default_actions)
        layout.addWidget(preset_group)
        self._refresh_text_preset_combo()

        selection_group = QGroupBox("选择模式")
        selection_form = QFormLayout(selection_group)
        selection_form.setContentsMargins(10, 14, 10, 10)
        selection_form.setHorizontalSpacing(8)
        selection_form.setVerticalSpacing(8)

        self.marquee_mode_combo = QComboBox()
        self.marquee_mode_combo.addItem("全选", "all")
        self.marquee_mode_combo.addItem("仅图片", "images")
        self.marquee_mode_combo.addItem("仅连接点元素", "connected")
        self.marquee_mode_combo.setToolTip("框选模式，Alt+M 可循环切换")
        self.marquee_mode_combo.currentIndexChanged.connect(
            lambda idx: self.set_marquee_mode(self.marquee_mode_combo.itemData(idx))
        )
        selection_form.addRow("框选", self.marquee_mode_combo)
        layout.addWidget(selection_group)

        layout.addStretch(1)
        return panel

    def _set_color_button_color(self, color):
        self.color_button.setText(color.name().upper())
        self.color_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {color.name()};
                color: {"white" if color.lightness() < 128 else "#24292f"};
                border: 1px solid #8c959f;
                border-radius: 5px;
                padding: 4px 10px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                border: 1px solid #0969da;
            }}
        """)

    def update_property_summary(self, selected, texts, images):
        if not hasattr(self, 'property_summary_label'):
            return
        if not selected:
            self.property_summary_label.setText("未选择对象")
        elif texts and not images:
            self.property_summary_label.setText(f"已选择 {len(texts)} 个文字对象")
        elif images and not texts:
            self.property_summary_label.setText(f"已选择 {len(images)} 张图片")
        else:
            self.property_summary_label.setText(
                f"已选择 {len(selected)} 个对象：{len(texts)} 个文字，{len(images)} 张图片"
            )

    @staticmethod
    def _text_format_from_item(item):
        return {
            'font_family': item.font_family,
            'font_size': item.font_size,
            'text_color': item.text_color.name(),
            'chars_per_column': item.chars_per_column,
            'column_spacing': item.column_spacing,
            'character_spacing': item.character_spacing,
            'manual_line_break': item.manual_line_break,
        }

    def _default_text_format(self):
        stored = self.scene.config_manager.get('default_text_format', {})
        if not isinstance(stored, dict):
            stored = {}
        return {
            'font_family': stored.get(
                'font_family',
                self.scene.config_manager.get('default_font_family', DEFAULT_FONT)
            ),
            'font_size': stored.get(
                'font_size',
                self.scene.config_manager.get('default_font_size', DEFAULT_FONT_SIZE)
            ),
            'text_color': stored.get('text_color', '#000000'),
            'chars_per_column': stored.get('chars_per_column', 15),
            'column_spacing': stored.get('column_spacing', COLUMN_SPACING),
            'character_spacing': stored.get('character_spacing', 0),
            'manual_line_break': stored.get('manual_line_break', True),
        }

    def _set_shared_configs(self, values):
        for document in self._documents:
            for key, value in values.items():
                document.scene.config_manager.config[key] = copy.deepcopy(value)
        self.scene.config_manager.save_config()

    def _set_shared_config(self, key, value):
        self._set_shared_configs({key: value})

    def _text_format_presets(self):
        presets = self.scene.config_manager.get('text_format_presets', [])
        if not isinstance(presets, list):
            return []
        return [
            preset for preset in presets
            if isinstance(preset, dict)
            and isinstance(preset.get('name'), str)
            and isinstance(preset.get('format'), dict)
        ]

    def _refresh_text_preset_combo(self, select_name=None):
        if not hasattr(self, 'text_preset_combo'):
            return
        if select_name is None:
            select_name = self.text_preset_combo.currentData()
        self.text_preset_combo.blockSignals(True)
        self.text_preset_combo.clear()
        self.text_preset_combo.addItem("选择预设...", None)
        for preset in self._text_format_presets():
            self.text_preset_combo.addItem(preset['name'], preset['name'])
        index = self.text_preset_combo.findData(select_name)
        self.text_preset_combo.setCurrentIndex(index if index >= 0 else 0)
        self.text_preset_combo.blockSignals(False)

    def _selected_text_format_source(self):
        selected = [
            item for item in self.scene.selectedItems()
            if isinstance(item, VTextItem)
        ]
        if not selected:
            self.status_bar.showMessage("请先选中一个文字对象", 3000)
            return None
        return selected[0]

    def _apply_text_format(self, item, text_format):
        item.font_family = str(text_format.get('font_family', item.font_family))
        item.font_size = max(1, int(text_format.get('font_size', item.font_size)))
        color = QColor(text_format.get('text_color', item.text_color.name()))
        if color.isValid():
            item.text_color = color
        item.chars_per_column = max(1, int(text_format.get('chars_per_column', item.chars_per_column)))
        item.column_spacing = int(text_format.get('column_spacing', item.column_spacing))
        item.character_spacing = int(text_format.get('character_spacing', item.character_spacing))
        item.manual_line_break = bool(text_format.get('manual_line_break', item.manual_line_break))
        item.rebuild()
        self.scene.update_connectors(item)
        self.scene.update_image_text_connectors(item)

    def set_current_text_format_as_default(self):
        item = self._selected_text_format_source()
        if item is None:
            return
        text_format = self._text_format_from_item(item)
        self._set_shared_configs({
            'default_text_format': text_format,
            'default_font_family': text_format['font_family'],
            'default_font_size': text_format['font_size'],
        })
        self.status_bar.showMessage("当前文字样式和竖排参数已设为默认值", 3000)

    def add_current_text_format_preset(self):
        item = self._selected_text_format_source()
        if item is None:
            return
        name, ok = QInputDialog.getText(self, "新增文字预设", "预设名称:")
        name = name.strip()
        if not ok or not name:
            return
        presets = self._text_format_presets()
        existing = next((preset for preset in presets if preset['name'] == name), None)
        if existing:
            reply = QMessageBox.question(
                self, "覆盖预设", f"预设“{name}”已存在，是否覆盖？"
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            existing['format'] = self._text_format_from_item(item)
        else:
            presets.append({'name': name, 'format': self._text_format_from_item(item)})
        self._set_shared_config('text_format_presets', presets)
        self._refresh_text_preset_combo(name)
        self.status_bar.showMessage(f"已保存文字预设：{name}", 3000)

    def apply_selected_text_preset(self):
        name = self.text_preset_combo.currentData() if hasattr(self, 'text_preset_combo') else None
        if not name:
            return
        preset = next((preset for preset in self._text_format_presets() if preset['name'] == name), None)
        if preset is None:
            self._refresh_text_preset_combo()
            return
        selected = [
            item for item in self.scene.selectedItems()
            if isinstance(item, VTextItem)
        ]
        if not selected:
            self.status_bar.showMessage("请先选中要应用预设的文字对象", 3000)
            return
        for item in selected:
            self._apply_text_format(item, preset['format'])
        self.update_font_controls()
        self.status_bar.showMessage(f"已将预设“{name}”应用到 {len(selected)} 个文字对象", 3000)

    def delete_selected_text_preset(self):
        name = self.text_preset_combo.currentData() if hasattr(self, 'text_preset_combo') else None
        if not name:
            return
        reply = QMessageBox.question(self, "删除预设", f"确定删除预设“{name}”吗？")
        if reply != QMessageBox.StandardButton.Yes:
            return
        presets = [preset for preset in self._text_format_presets() if preset['name'] != name]
        self._set_shared_config('text_format_presets', presets)
        self._refresh_text_preset_combo()
        self.status_bar.showMessage(f"已删除文字预设：{name}", 3000)
    
    def create_toolbars_legacy(self):
        # 主工具栏 - 编辑和格式化
        main_toolbar = QToolBar("编辑与格式")
        main_toolbar.setObjectName("toolbar_main")
        main_toolbar.setMovable(False)
        main_toolbar.setIconSize(QSize(16, 16))
        main_toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
        self.addToolBar(Qt.ToolBarArea.TopToolBarArea, main_toolbar)
        
        # === 基本编辑操作 ===
        btn_add_text = QAction("添加文本", self)
        btn_add_text.triggered.connect(self.add_text)
        main_toolbar.addAction(btn_add_text)

        btn_add_img = QAction("插入图片", self)
        btn_add_img.triggered.connect(self.add_image)
        main_toolbar.addAction(btn_add_img)

        btn_edit_text = QAction("编辑文字", self)
        btn_edit_text.triggered.connect(self.edit_selected_text)
        main_toolbar.addAction(btn_edit_text)

        self.btn_resize = QAction("调整图片大小", self)
        self.btn_resize.setCheckable(True)
        self.btn_resize.setToolTip("开启后选中图片会显示缩放控制点")
        self.btn_resize.toggled.connect(self.toggle_resize_mode)
        main_toolbar.addAction(self.btn_resize)
        
        main_toolbar.addSeparator()
        
        btn_undo = QAction("撤销", self)
        btn_undo.setShortcut("Ctrl+Z")
        btn_undo.triggered.connect(self.undo)
        main_toolbar.addAction(btn_undo)
        
        btn_align_right = QAction("右对齐", self)
        btn_align_right.triggered.connect(self.align_right)
        main_toolbar.addAction(btn_align_right)

        btn_align_top = QAction("顶部对齐", self)
        btn_align_top.triggered.connect(self.align_top)
        main_toolbar.addAction(btn_align_top)

        btn_smart_connect = QAction("智能连接", self)
        btn_smart_connect.setToolTip("按选中顺序逐对连接元素（Ctrl+点击确定顺序）")
        btn_smart_connect.triggered.connect(self.auto_connect_selected)
        main_toolbar.addAction(btn_smart_connect)
        
        main_toolbar.addSeparator()
        
        # === 字体格式 ===
        main_toolbar.addWidget(QLabel("字体:"))
        self.font_combo = QComboBox()
        self.font_combo.setMinimumWidth(120)
        self.font_combo.setMaximumWidth(170)
        self.font_combo.setEditable(True)
        # 加载系统所有字体
        for f in QFontDatabase.families():
            self.font_combo.addItem(f)
        default_font = self.scene.config_manager.get('default_font_family', DEFAULT_FONT)
        idx = self.font_combo.findText(default_font)
        if idx >= 0:
            self.font_combo.setCurrentIndex(idx)
        else:
            self.font_combo.setCurrentText(default_font)
        self.font_combo.currentTextChanged.connect(
            lambda name: self.change_selected_font(QFont(name))
        )
        main_toolbar.addWidget(self.font_combo)
        
        main_toolbar.addWidget(QLabel("大小:"))
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(8, 200)
        # 从配置加载默认字体大小
        default_font_size = self.scene.config_manager.get('default_font_size', DEFAULT_FONT_SIZE)
        self.font_size_spin.setValue(default_font_size)
        self.font_size_spin.setSuffix("pt")
        self.font_size_spin.valueChanged.connect(self.change_selected_font_size)
        main_toolbar.addWidget(self.font_size_spin)
        
        self.color_button = QPushButton()
        self.color_button.setFixedSize(28, 24)
        self.color_button.setStyleSheet("""
            QPushButton {
                background-color: black;
                border: 2px solid rgba(255, 255, 255, 0.9);
                border-radius: 4px;
                padding: 0;
            }
            QPushButton:hover {
                border: 2px solid rgba(0, 120, 215, 0.8);
            }
            QPushButton:pressed {
            }
        """)
        self.color_button.clicked.connect(self.change_selected_color)
        main_toolbar.addWidget(self.color_button)
        
        main_toolbar.addSeparator()
        
        # === 竖排布局设置 ===
        main_toolbar.addWidget(QLabel("每列字数:"))
        self.chars_per_column_spin = QSpinBox()
        self.chars_per_column_spin.setRange(5, 50)
        self.chars_per_column_spin.setValue(15)
        self.chars_per_column_spin.setSuffix("字")
        self.chars_per_column_spin.valueChanged.connect(self.change_chars_per_column)
        main_toolbar.addWidget(self.chars_per_column_spin)
        
        main_toolbar.addWidget(QLabel("列间距:"))
        self.column_spacing_spin = QSpinBox()
        self.column_spacing_spin.setRange(0, 200)
        self.column_spacing_spin.setValue(COLUMN_SPACING)
        self.column_spacing_spin.setSuffix("px")
        self.column_spacing_spin.valueChanged.connect(self.change_column_spacing)
        main_toolbar.addWidget(self.column_spacing_spin)

        main_toolbar.addWidget(QLabel("字间距:"))
        self.character_spacing_spin = QSpinBox()
        self.character_spacing_spin.setRange(-200, 200)
        self.character_spacing_spin.setValue(0)
        self.character_spacing_spin.setSuffix("px")
        self.character_spacing_spin.valueChanged.connect(self.change_character_spacing)
        main_toolbar.addWidget(self.character_spacing_spin)
        
        self.manual_line_break_btn = QPushButton("手动换行")
        self.manual_line_break_btn.setCheckable(True)
        self.manual_line_break_btn.setChecked(True)
        self.manual_line_break_btn.setMaximumHeight(28)
        self.manual_line_break_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.9),
                    stop:1 rgba(249, 249, 249, 0.9));
                border: 1px solid rgba(0, 0, 0, 0.1);
                border-radius: 4px;
                padding: 3px 8px;
                font-weight: 500;
            }
            QPushButton:checked {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(0, 120, 215, 1.0),
                    stop:1 rgba(0, 90, 158, 1.0));
                color: white;
                border: 1px solid rgba(0, 90, 158, 1.0);
                font-weight: 600;
            }
            QPushButton:hover {
                border: 1px solid rgba(0, 120, 215, 0.4);
            }
        """)
        self.manual_line_break_btn.toggled.connect(self.toggle_manual_line_break)
        main_toolbar.addWidget(self.manual_line_break_btn)

        main_toolbar.addSeparator()
        main_toolbar.addWidget(QLabel("框选:"))
        self.marquee_mode_combo = QComboBox()
        self.marquee_mode_combo.addItem("全选", "all")
        self.marquee_mode_combo.addItem("仅图片", "images")
        self.marquee_mode_combo.addItem("仅连接点元素", "connected")
        self.marquee_mode_combo.setToolTip("框选模式（Alt+M 循环切换）")
        self.marquee_mode_combo.setFixedWidth(96)
        self.marquee_mode_combo.currentIndexChanged.connect(
            lambda idx: self.set_marquee_mode(self.marquee_mode_combo.itemData(idx))
        )
        main_toolbar.addWidget(self.marquee_mode_combo)

        main_toolbar.addSeparator()
        btn_group_connect = QAction("垂直连接", self)
        btn_group_connect.setToolTip("将选中图片+子文字视为组合，按位置顺序依次连接 b点→a点")
        btn_group_connect.triggered.connect(lambda: self.scene.group_chain_connect())
        main_toolbar.addAction(btn_group_connect)

        btn_img_connect = QAction("水平连接", self)
        btn_img_connect.setToolTip("按从右到左顺序，依次连接选中图片的连接点")
        btn_img_connect.triggered.connect(lambda: self.scene.connect_image_points_right_to_left())
        main_toolbar.addAction(btn_img_connect)

    def _add_toolbar_action(self, toolbar, text, slot, tooltip=None, shortcut=None, checkable=False):
        action = QAction(text, self)
        if tooltip:
            action.setToolTip(tooltip)
            action.setStatusTip(tooltip)
        if shortcut:
            action.setShortcut(shortcut)
        action.setCheckable(checkable)
        if checkable:
            action.toggled.connect(slot)
        else:
            action.triggered.connect(lambda checked=False: slot())
        toolbar.addAction(action)
        return action

    def create_toolbars(self):
        main_toolbar = QToolBar("主要操作")
        main_toolbar.setObjectName("toolbar_main")
        main_toolbar.setMovable(False)
        main_toolbar.setIconSize(QSize(18, 18))
        main_toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
        self.addToolBar(Qt.ToolBarArea.TopToolBarArea, main_toolbar)

        # 插入下拉组：整合添加文本、插入图片、编辑文字
        insert_menu = QMenu(self)
        insert_menu.addAction("添加文本 (Ctrl+T)", self.add_text)
        insert_menu.addAction("插入图片 (Ctrl+Shift+I)", self.add_image)
        insert_menu.addAction("编辑选中文字", self.edit_selected_text)
        
        insert_btn = QToolButton()
        insert_btn.setText("插入 ▼")
        insert_btn.setMenu(insert_menu)
        insert_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        insert_btn.setToolTip("添加文本、插入图片、编辑文字")
        main_toolbar.addWidget(insert_btn)

        main_toolbar.addSeparator()
        self._add_toolbar_action(main_toolbar, "撤销", self.undo, "撤销上一步操作", "Ctrl+Z")
        
        # 水平移动开关
        self.btn_horizontal_move = self._add_toolbar_action(
            main_toolbar,
            "🔒水平移动",
            lambda checked: self.toggle_horizontal_move_only(checked),
            "锁定垂直移动，只允许水平方向移动选中对象 (Ctrl+H)",
            checkable=True
        )
        self.btn_horizontal_move.setChecked(self.scene.config_manager.get('horizontal_move_only', False))
        
        self.btn_resize = self._add_toolbar_action(
            main_toolbar,
            "调整图片",
            self.toggle_resize_mode,
            "选中图片后显示缩放控制点",
            checkable=True
        )
        self.btn_smart_brush = self._add_toolbar_action(
            main_toolbar,
            "智能笔刷",
            self.toggle_smart_brush,
            "涂抹对象加入选区；按住 Alt 涂抹从选区中减去",
            None,
            checkable=True
        )
        main_toolbar.addWidget(QLabel("笔刷:"))
        self.smart_brush_size_spin = QSpinBox()
        self.smart_brush_size_spin.setRange(4, 120)
        self.smart_brush_size_spin.setValue(self.view._smart_brush_radius)
        self.smart_brush_size_spin.setSuffix("px")
        self.smart_brush_size_spin.setFixedWidth(72)
        self.smart_brush_size_spin.setToolTip("智能笔刷大小（快捷键 [ / ]）")
        self.smart_brush_size_spin.valueChanged.connect(self.view.set_smart_brush_radius)
        main_toolbar.addWidget(self.smart_brush_size_spin)

        main_toolbar.addSeparator()
        self._add_toolbar_action(main_toolbar, "右对齐", self.align_right, "按右边缘对齐选中对象")
        self._add_toolbar_action(main_toolbar, "顶部对齐", self.align_top, "按顶部边缘对齐选中对象")
        self._add_toolbar_action(main_toolbar, "智能连接", self.auto_connect_selected, "按选中顺序连接元素")

        main_toolbar.addSeparator()
        self._add_toolbar_action(main_toolbar, "垂直连接", lambda: self.scene.group_chain_connect(), "将选中图片+子文字视为组合，按位置顺序依次连接 b点→a点")
        self._add_toolbar_action(
            main_toolbar,
            "水平连接",
            lambda: self.scene.connect_image_points_right_to_left(),
            "按从右到左顺序连接选中图片的连接点"
        )

        main_toolbar.addSeparator()
        main_toolbar.addWidget(QLabel("框选:"))
        self.marquee_mode_combo = QComboBox()
        self.marquee_mode_combo.addItem("全选", "all")
        self.marquee_mode_combo.addItem("仅图片", "images")
        self.marquee_mode_combo.addItem("仅连接点元素", "connected")
        self.marquee_mode_combo.setToolTip("框选模式（Alt+M 循环切换）")
        self.marquee_mode_combo.setFixedWidth(110)
        self.marquee_mode_combo.currentIndexChanged.connect(
            lambda idx: self.set_marquee_mode(self.marquee_mode_combo.itemData(idx))
        )
        main_toolbar.addWidget(self.marquee_mode_combo)

        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        main_toolbar.addWidget(spacer)
        hint = QLabel("文字与排版参数在右侧「属性」页")
        hint.setObjectName("toolbarHint")
        main_toolbar.addWidget(hint)

    def create_menu_bar(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu('文件')
        
        new_action = QAction('新建文档', self)
        new_action.setShortcut('Ctrl+N')
        new_action.triggered.connect(self.new_document)
        file_menu.addAction(new_action)

        new_project_action = QAction('新建工程', self)
        new_project_action.setShortcut('Ctrl+Shift+N')
        new_project_action.triggered.connect(self.new_project)
        file_menu.addAction(new_project_action)

        close_document_action = QAction('关闭当前文档', self)
        close_document_action.setShortcut('Ctrl+W')
        close_document_action.triggered.connect(self.close_document)
        file_menu.addAction(close_document_action)

        rename_document_action = QAction('重命名当前文档', self)
        rename_document_action.setShortcut('F2')
        rename_document_action.triggered.connect(self.rename_document)
        file_menu.addAction(rename_document_action)
        
        file_menu.addSeparator()
        
        open_action = QAction('打开工程...', self)
        open_action.setShortcut('Ctrl+O')
        open_action.triggered.connect(self.load_proj)
        file_menu.addAction(open_action)
        
        save_action = QAction('快速保存', self)
        save_action.setShortcut('Ctrl+S')
        save_action.triggered.connect(self.quick_save_proj)
        file_menu.addAction(save_action)

        save_as_action = QAction('另存为...', self)
        save_as_action.setShortcut('Ctrl+Shift+S')
        save_as_action.triggered.connect(self.save_proj)
        file_menu.addAction(save_as_action)

        save_and_pdf_action = QAction('保存项目并导出全部文档 PDF', self)
        save_and_pdf_action.setShortcut('Ctrl+Shift+P')
        save_and_pdf_action.triggered.connect(self.save_proj_and_pdf)
        file_menu.addAction(save_and_pdf_action)

        print_action = QAction('打印...', self)
        print_action.setShortcut('Ctrl+P')
        print_action.triggered.connect(self.print_document)
        file_menu.addAction(print_action)

        print_preview_action = QAction('打印预览...', self)
        print_preview_action.triggered.connect(self.print_preview)
        file_menu.addAction(print_preview_action)
        
        file_menu.addSeparator()
        
        export_action = QAction('导出图片...', self)
        export_action.setShortcut('Ctrl+E')
        export_action.triggered.connect(self.export_image)
        file_menu.addAction(export_action)

        export_svg_action = QAction('导出CorelDRAW SVG...', self)
        export_svg_action.triggered.connect(self.export_coreldraw_svg)
        file_menu.addAction(export_svg_action)

        export_excel_action = QAction('导出 Excel 数据表...', self)
        export_excel_action.triggered.connect(self.export_excel)
        file_menu.addAction(export_excel_action)

        export_pdf_action = QAction('导出 PDF...', self)
        export_pdf_action.setShortcut('Ctrl+Alt+P')
        export_pdf_action.triggered.connect(self.export_pdf)
        file_menu.addAction(export_pdf_action)
        
        file_menu.addSeparator()
        
        # 设置默认字体
        set_default_font_action = QAction('设置默认字体...', self)
        set_default_font_action.triggered.connect(self.set_default_font)
        file_menu.addAction(set_default_font_action)

        set_current_text_default_action = QAction('将选中文字格式设为默认', self)
        set_current_text_default_action.triggered.connect(self.set_current_text_format_as_default)
        file_menu.addAction(set_current_text_default_action)

        add_text_preset_action = QAction('将选中文字格式新增为预设...', self)
        add_text_preset_action.triggered.connect(self.add_current_text_format_preset)
        file_menu.addAction(add_text_preset_action)

        set_save_dir_action = QAction('设置默认保存目录...', self)
        set_save_dir_action.triggered.connect(self.set_default_save_dir)
        file_menu.addAction(set_save_dir_action)

        set_img_size_action = QAction('设置插入图片默认大小...', self)
        set_img_size_action.triggered.connect(self.set_insert_image_size)
        file_menu.addAction(set_img_size_action)

        file_menu.addSeparator()

        license_expiry_action = QAction('设置软件到期日期...', self)
        license_expiry_action.triggered.connect(self.set_license_expiry_date)
        file_menu.addAction(license_expiry_action)

        license_password_action = QAction('修改授权密码...', self)
        license_password_action.triggered.connect(self.change_license_password)
        file_menu.addAction(license_password_action)
        
        # 添加素材菜单
        asset_menu = menubar.addMenu('素材')
        save_group_action = QAction('保存组合素材', self)
        save_group_action.setShortcut('Ctrl+G')
        save_group_action.triggered.connect(self.save_selected_as_group)
        asset_menu.addAction(save_group_action)

        asset_menu.addSeparator()
        batch_copy_action = QAction('批量复制...', self)
        batch_copy_action.setShortcut('Ctrl+Shift+D')
        batch_copy_action.triggered.connect(self.batch_copy)
        asset_menu.addAction(batch_copy_action)

        group_connect_action = QAction('垂直连接', self)
        group_connect_action.setToolTip("将选中图片+子文字视为组合，按位置顺序依次连接 b点→a点")
        group_connect_action.triggered.connect(lambda: self.scene.group_chain_connect())
        asset_menu.addAction(group_connect_action)

        open_library_action = QAction('切换素材库面板', self)
        open_library_action.setShortcut('F9')
        open_library_action.triggered.connect(self.open_asset_library)
        asset_menu.addAction(open_library_action)

        asset_menu.addSeparator()
        import_excel_action = QAction('从 Excel 批量导入族谱...', self)
        import_excel_action.setShortcut('Ctrl+I')
        import_excel_action.triggered.connect(self.open_family_tree_import_dialog)
        asset_menu.addAction(import_excel_action)
        
        # 添加视图菜单
        view_menu = menubar.addMenu('视图')
        
        fit_view_action = QAction('合适屏幕', self)
        fit_view_action.setShortcut('Ctrl+0')
        fit_view_action.triggered.connect(self.fit_in_view)
        view_menu.addAction(fit_view_action)
        
        fill_view_action = QAction('填充屏幕', self)
        fill_view_action.setShortcut('Ctrl+Alt+0')
        fill_view_action.triggered.connect(self.fill_view)
        view_menu.addAction(fill_view_action)
        
        actual_size_action = QAction('实际大小', self)
        actual_size_action.setShortcut('Ctrl+1')
        actual_size_action.triggered.connect(self.actual_size)
        view_menu.addAction(actual_size_action)
        
        view_menu.addSeparator()
        
        zoom_in_action = QAction('放大', self)
        zoom_in_action.setShortcut('Ctrl+=')
        zoom_in_action.triggered.connect(self.zoom_in)
        view_menu.addAction(zoom_in_action)
        
        zoom_out_action = QAction('缩小', self)
        zoom_out_action.setShortcut('Ctrl+-')
        zoom_out_action.triggered.connect(self.zoom_out)
        view_menu.addAction(zoom_out_action)
        
        zoom_selection_action = QAction('缩放到选中', self)
        zoom_selection_action.setShortcut('Ctrl+2')
        zoom_selection_action.triggered.connect(self.zoom_to_selection)
        view_menu.addAction(zoom_selection_action)

        canvas_size_action = QAction('设置画布大小...', self)
        canvas_size_action.triggered.connect(self.set_canvas_size)
        view_menu.addAction(canvas_size_action)

        view_menu.addSeparator()

        self.bg_above_connectors_action = QAction('背景图片在连线之上', self)
        self.bg_above_connectors_action.setCheckable(True)
        self.bg_above_connectors_action.setChecked(self.scene.config_manager.get('bg_above_connectors', False))
        self.bg_above_connectors_action.triggered.connect(self.toggle_bg_above_connectors)
        view_menu.addAction(self.bg_above_connectors_action)
        
        view_menu.addSeparator()
        
        # 背景设置
        set_background_action = QAction('设置默认背景图片...', self)
        set_background_action.triggered.connect(self.set_background_image)
        view_menu.addAction(set_background_action)
        
        clear_background_action = QAction('清除背景图片', self)
        clear_background_action.triggered.connect(self.clear_background_image)
        view_menu.addAction(clear_background_action)

        show_all_images_action = QAction('显示所有隐藏图片', self)
        show_all_images_action.triggered.connect(self.show_all_hidden_images)
        view_menu.addAction(show_all_images_action)

        image_manage_action = QAction('图片管理模式', self)
        image_manage_action.setShortcut('Alt+,')
        image_manage_action.triggered.connect(lambda: self.scene.toggle_image_manage_mode())
        view_menu.addAction(image_manage_action)
        
        background_opacity_action = QAction('设置背景透明度...', self)
        background_opacity_action.triggered.connect(self.set_background_opacity)
        view_menu.addAction(background_opacity_action)
        
        # 背景缩放模式子菜单
        scale_mode_menu = view_menu.addMenu('背景缩放模式')
        
        scale_fit_action = QAction('适应画布', self)
        scale_fit_action.triggered.connect(lambda: self.set_background_scale_mode('fit'))
        scale_mode_menu.addAction(scale_fit_action)
        
        scale_fill_action = QAction('填充画布', self)
        scale_fill_action.triggered.connect(lambda: self.set_background_scale_mode('fill'))
        scale_mode_menu.addAction(scale_fill_action)
        
        scale_stretch_action = QAction('拉伸填充', self)
        scale_stretch_action.triggered.connect(lambda: self.set_background_scale_mode('stretch'))
        scale_mode_menu.addAction(scale_stretch_action)
        
        scale_tile_action = QAction('平铺', self)
        scale_tile_action.triggered.connect(lambda: self.set_background_scale_mode('tile'))
        scale_mode_menu.addAction(scale_tile_action)

        # 辅助线菜单
        view_menu.addSeparator()
        toggle_guides_action = QAction('显示/隐藏辅助线', self)
        toggle_guides_action.setShortcut('Ctrl+;')
        toggle_guides_action.triggered.connect(self.toggle_guides)
        view_menu.addAction(toggle_guides_action)

        clear_guides_action = QAction('清除所有辅助线', self)
        clear_guides_action.triggered.connect(self.clear_guides)
        view_menu.addAction(clear_guides_action)

        unit_menu = view_menu.addMenu('界面单位')
        self.display_unit_group = QActionGroup(self)
        self.display_unit_group.setExclusive(True)
        self.display_unit_px_action = QAction('像素 (px)', self)
        self.display_unit_px_action.setCheckable(True)
        self.display_unit_px_action.setData('px')
        self.display_unit_mm_action = QAction('毫米 (mm)', self)
        self.display_unit_mm_action.setCheckable(True)
        self.display_unit_mm_action.setData('mm')
        self.display_unit_group.addAction(self.display_unit_px_action)
        self.display_unit_group.addAction(self.display_unit_mm_action)
        unit_menu.addActions([self.display_unit_px_action, self.display_unit_mm_action])
        current_unit = self.scene.config_manager.get('display_unit', 'mm')
        (self.display_unit_px_action if current_unit == 'px' else self.display_unit_mm_action).setChecked(True)
        self.display_unit_group.triggered.connect(self.set_display_unit)

        self.startup_horizontal_guides_action = QAction('启动时添加横向辅助线', self)
        self.startup_horizontal_guides_action.setCheckable(True)
        self.startup_horizontal_guides_action.setChecked(
            self.scene.config_manager.get('startup_horizontal_guides_enabled', False)
        )
        self.startup_horizontal_guides_action.toggled.connect(self.toggle_startup_horizontal_guides)
        view_menu.addAction(self.startup_horizontal_guides_action)

        startup_horizontal_guides_settings_action = QAction('设置启动横向辅助线位置...', self)
        startup_horizontal_guides_settings_action.triggered.connect(self.set_startup_horizontal_guides)
        view_menu.addAction(startup_horizontal_guides_settings_action)

        self.hide_points_on_image_select_action = QAction('选中图片时隐藏连接点', self)
        self.hide_points_on_image_select_action.setCheckable(True)
        self.hide_points_on_image_select_action.setChecked(
            self.scene.config_manager.get('hide_connection_points_when_image_selected', False)
        )
        self.hide_points_on_image_select_action.toggled.connect(self.toggle_hide_points_on_image_select)
        view_menu.addAction(self.hide_points_on_image_select_action)

        self.text_hover_tooltip_action = QAction('鼠标悬停显示文字内容', self)
        self.text_hover_tooltip_action.setCheckable(True)
        self.text_hover_tooltip_action.setChecked(
            self.scene.config_manager.get('show_text_hover_tooltip', False)
        )
        self.text_hover_tooltip_action.toggled.connect(self.toggle_text_hover_tooltip)
        view_menu.addAction(self.text_hover_tooltip_action)

        snap_threshold_action = QAction('设置吸附距离...', self)
        snap_threshold_action.triggered.connect(self.set_snap_threshold)
        view_menu.addAction(snap_threshold_action)
        
        # 编辑行为菜单（合并原连线菜单）
        edit_menu = menubar.addMenu('编辑设置')

        # 连线设置
        set_line_width_action = QAction('设置连线粗细...', self)
        set_line_width_action.triggered.connect(lambda: self.set_line_width())
        edit_menu.addAction(set_line_width_action)
        edit_menu.addSeparator()
        self.auto_exit_paste_action = QAction('粘贴后自动退出编辑', self)
        self.auto_exit_paste_action.setCheckable(True)
        self.auto_exit_paste_action.setChecked(self.scene.config_manager.get('auto_exit_after_paste', False))
        self.auto_exit_paste_action.toggled.connect(self.toggle_auto_exit_setting)
        edit_menu.addAction(self.auto_exit_paste_action)

        # 框选模式：三选一，替代原来的两个开关
        marquee_mode_menu = edit_menu.addMenu('框选模式  [Alt+M 循环切换]')
        self._marquee_mode_actions = {}
        for label, value in [
            ('全选（默认）', 'all'),
            ('仅选图片', 'images'),
            ('仅选有连接点的元素', 'connected'),
        ]:
            a = QAction(label, self)
            a.setCheckable(True)
            a.setData(value)
            a.triggered.connect(lambda checked, v=value: self.set_marquee_mode(v))
            marquee_mode_menu.addAction(a)
            self._marquee_mode_actions[value] = a
        self._sync_marquee_mode_ui()

        cycle_marquee_action = QAction('循环切换框选模式', self)
        cycle_marquee_action.setShortcut('Alt+M')
        cycle_marquee_action.triggered.connect(self.cycle_marquee_mode)
        edit_menu.addAction(cycle_marquee_action)

        smart_brush_action = QAction('智能笔刷', self)
        smart_brush_action.setCheckable(True)
        smart_brush_action.setShortcut('B')
        smart_brush_action.toggled.connect(self.toggle_smart_brush)
        smart_brush_action.toggled.connect(
            lambda checked: self.btn_smart_brush.setChecked(checked)
            if hasattr(self, 'btn_smart_brush') and self.btn_smart_brush.isChecked() != checked else None
        )
        edit_menu.addAction(smart_brush_action)
        self.smart_brush_action = smart_brush_action

        brush_smaller_action = QAction('缩小智能笔刷', self)
        brush_smaller_action.setShortcut('[')
        brush_smaller_action.triggered.connect(lambda: self.view.adjust_smart_brush_radius(-4))
        edit_menu.addAction(brush_smaller_action)

        brush_larger_action = QAction('放大智能笔刷', self)
        brush_larger_action.setShortcut(']')
        brush_larger_action.triggered.connect(lambda: self.view.adjust_smart_brush_radius(4))
        edit_menu.addAction(brush_larger_action)

        self.insert_image_to_bottom_action = QAction('插入图片时置于底层', self)
        self.insert_image_to_bottom_action.setCheckable(True)
        self.insert_image_to_bottom_action.setChecked(self.scene.config_manager.get('insert_image_to_bottom', False))
        self.insert_image_to_bottom_action.toggled.connect(self.toggle_insert_image_to_bottom)
        edit_menu.addAction(self.insert_image_to_bottom_action)

        self.insert_image_fit_canvas_action = QAction('插入图片时自动适应画布大小', self)
        self.insert_image_fit_canvas_action.setCheckable(True)
        self.insert_image_fit_canvas_action.setChecked(self.scene.config_manager.get('insert_image_fit_canvas', False))
        self.insert_image_fit_canvas_action.toggled.connect(self.toggle_insert_image_fit_canvas)
        edit_menu.addAction(self.insert_image_fit_canvas_action)

        self.horizontal_move_only_action = QAction('水平移动选中对象（锁定Y）', self)
        self.horizontal_move_only_action.setCheckable(True)
        self.horizontal_move_only_action.setShortcut('Ctrl+H')
        self.horizontal_move_only_action.setChecked(self.scene.config_manager.get('horizontal_move_only', False))
        self.horizontal_move_only_action.toggled.connect(self.toggle_horizontal_move_only)
        edit_menu.addAction(self.horizontal_move_only_action)

        self.image_right_edge_snap_action = QAction('图片右边缘X自动吸附', self)
        self.image_right_edge_snap_action.setCheckable(True)
        self.image_right_edge_snap_action.setChecked(
            self.scene.config_manager.get('image_right_edge_snap_enabled', False)
        )
        self.image_right_edge_snap_action.toggled.connect(self.toggle_image_right_edge_snap)
        edit_menu.addAction(self.image_right_edge_snap_action)

        self.image_top_edge_snap_action = QAction('图片顶部Y自动吸附', self)
        self.image_top_edge_snap_action.setCheckable(True)
        self.image_top_edge_snap_action.setChecked(
            self.scene.config_manager.get('image_top_edge_snap_enabled', False)
        )
        self.image_top_edge_snap_action.toggled.connect(self.toggle_image_top_edge_snap)
        edit_menu.addAction(self.image_top_edge_snap_action)

        eye_role_from_selection_menu = edit_menu.addMenu('设置选中文字眼睛颜色')
        for color_key, color_label in [
            ('yellow', '设为黄色眼睛'),
            ('red', '设为红色眼睛'),
            ('green', '设为绿色眼睛'),
        ]:
            action = QAction(color_label, self)
            action.triggered.connect(
                lambda checked=False, c=color_key: self.set_layer_eye_role_from_selection(c)
            )
            eye_role_from_selection_menu.addAction(action)
        clear_eye_action = QAction('清除眼睛颜色', self)
        clear_eye_action.triggered.connect(lambda: self.set_layer_eye_role_from_selection(None))
        eye_role_from_selection_menu.addAction(clear_eye_action)

        edit_menu.addSeparator()
        nudge_step_action = QAction('设置Shift+方向键步长...', self)
        nudge_step_action.triggered.connect(self.set_nudge_large_step)
        edit_menu.addAction(nudge_step_action)

        brush_size_action = QAction('设置笔刷默认大小...', self)
        brush_size_action.triggered.connect(self.set_default_brush_size)
        edit_menu.addAction(brush_size_action)

    def toggle_auto_exit_setting(self, enabled):
        """切换粘贴后自动退出编辑的开关"""
        self.scene.config_manager.set('auto_exit_after_paste', enabled)

    def toggle_horizontal_move_only(self, enabled):
        """切换水平移动模式：拖动/方向键移动时保持Y坐标不变。"""
        # ConfigManager is instantiated once per document, so update every
        # document copy; otherwise switching tabs restores a stale value.
        for document in self._documents:
            document.scene.config_manager.set('horizontal_move_only', enabled)
        
        # 同步工具栏按钮和菜单动作的状态
        if hasattr(self, 'btn_horizontal_move') and self.btn_horizontal_move.isChecked() != enabled:
            self.btn_horizontal_move.setChecked(enabled)
        if hasattr(self, 'horizontal_move_only_action') and self.horizontal_move_only_action.isChecked() != enabled:
            self.horizontal_move_only_action.blockSignals(True)
            self.horizontal_move_only_action.setChecked(enabled)
            self.horizontal_move_only_action.blockSignals(False)
        
        # 显示状态消息（确保 status_bar 已创建）
        if hasattr(self, 'status_bar'):
            if enabled:
                self.status_bar.showMessage('已开启：水平移动模式，移动对象时Y坐标不变', 3000)
            else:
                self.status_bar.showMessage('已关闭：水平移动模式', 3000)

    def toggle_image_right_edge_snap(self, enabled):
        """切换水平移动图片时的右边缘X吸附。"""
        for document in self._documents:
            document.scene.config_manager.set('image_right_edge_snap_enabled', enabled)
        if hasattr(self, 'status_bar'):
            if enabled:
                self.status_bar.showMessage('已开启：图片右边缘X自动吸附（阈值20px）', 3000)
            else:
                self.status_bar.showMessage('已关闭：图片右边缘X自动吸附', 3000)

    def toggle_image_top_edge_snap(self, enabled):
        """切换移动图片时的顶部Y吸附。"""
        for document in self._documents:
            document.scene.config_manager.set('image_top_edge_snap_enabled', enabled)
        if hasattr(self, 'status_bar'):
            if enabled:
                self.status_bar.showMessage('已开启：图片顶部Y自动吸附（阈值20px）', 3000)
            else:
                self.status_bar.showMessage('已关闭：图片顶部Y自动吸附', 3000)

    def set_marquee_mode(self, mode):
        """设置框选模式：all / images / connected"""
        self.scene.config_manager.set('marquee_mode', mode)
        self._sync_marquee_mode_ui()
        labels = {'all': '全选', 'images': '仅选图片', 'connected': '仅选有连接点的元素'}
        self.status_bar.showMessage(f'框选模式：{labels.get(mode, mode)}  (Alt+M 循环切换)', 3000)

    def _selected_texts_for_eye_role(self):
        selected_texts = [
            item for item in self.scene.selectedItems()
            if isinstance(item, VTextItem)
        ]
        if not selected_texts:
            self.status_bar.showMessage("请先在画布上选中文字", 3000)
            return []
        return selected_texts

    def set_layer_eye_role_from_selection(self, color_key):
        text_items = self._selected_texts_for_eye_role()
        if not text_items:
            return
        if color_key not in {'yellow', 'red', 'green', None}:
            return
        for text_item in text_items:
            text_item.layer_eye_color = color_key
        self.refresh_ui()
        color_names = {'yellow': '黄色', 'red': '红色', 'green': '绿色', None: '无'}
        self.status_bar.showMessage(
            f"已将 {len(text_items)} 个文字的眼睛颜色设为：{color_names[color_key]}",
            3000
        )

    def cycle_marquee_mode(self):
        """Alt+M 循环切换框选模式"""
        modes = ['all', 'images', 'connected']
        current = self.scene.config_manager.get('marquee_mode', 'all')
        next_mode = modes[(modes.index(current) + 1) % len(modes)] if current in modes else 'all'
        self.set_marquee_mode(next_mode)

    def _sync_marquee_mode_ui(self):
        """同步菜单和工具栏下拉框的选中状态"""
        current = self.scene.config_manager.get('marquee_mode', 'all')
        if hasattr(self, '_marquee_mode_actions'):
            for value, action in self._marquee_mode_actions.items():
                action.setChecked(value == current)
        if hasattr(self, 'marquee_mode_combo'):
            idx = self.marquee_mode_combo.findData(current)
            if idx >= 0:
                self.marquee_mode_combo.blockSignals(True)
                self.marquee_mode_combo.setCurrentIndex(idx)
                self.marquee_mode_combo.blockSignals(False)

    def toggle_marquee_filter_setting(self, enabled):
        """兼容旧调用"""
        self.set_marquee_mode('images' if enabled else 'all')

    def toggle_marquee_connected_setting(self, enabled):
        """兼容旧调用"""
        self.set_marquee_mode('connected' if enabled else 'all')

    def toggle_insert_image_to_bottom(self, enabled):
        """切换插入图片时是否置于底层"""
        self.scene.config_manager.set('insert_image_to_bottom', enabled)
        if enabled:
            self.status_bar.showMessage('已开启：插入图片时自动置于文字和连线底层', 3000)
        else:
            self.status_bar.showMessage('已关闭：插入图片时保持默认层级', 3000)

    def toggle_insert_image_fit_canvas(self, enabled):
        """切换插入图片时是否自动适应画布大小"""
        self.scene.config_manager.set('insert_image_fit_canvas', enabled)
        if enabled:
            self.status_bar.showMessage('已开启：插入图片时自动拉伸为画布大小', 3000)
        else:
            self.status_bar.showMessage('已关闭：插入图片时使用原始大小', 3000)

    def _verify_license_password(self):
        """验证管理员密码后再允许修改授权设置。"""
        password, ok = QInputDialog.getText(
            self,
            "授权验证",
            "请输入授权密码：",
            QLineEdit.EchoMode.Password
        )
        if not ok:
            return False

        stored_hash = self.scene.config_manager.get(
            "license_password_hash",
            hash_license_password(DEFAULT_LICENSE_PASSWORD)
        )
        if hash_license_password(password) != stored_hash:
            QMessageBox.warning(self, "授权验证", "密码不正确。")
            return False
        return True

    def set_license_expiry_date(self):
        """设置软件可使用的最后日期。"""
        if not self._verify_license_password():
            return

        config = self.scene.config_manager
        current_text = str(config.get("license_expiry_date", "") or "").strip()
        current_date = QDate.fromString(current_text, "yyyy-MM-dd")
        if not current_date.isValid():
            current_date = QDate.currentDate()

        dialog = QDialog(self)
        dialog.setWindowTitle("设置软件到期日期")
        layout = QFormLayout(dialog)
        layout.addRow(
            QLabel("软件在所选日期当天仍可使用，次日启动后将无法进入软件。")
        )
        expiry_edit = QDateEdit(current_date, dialog)
        expiry_edit.setCalendarPopup(True)
        expiry_edit.setDisplayFormat("yyyy-MM-dd")
        expiry_edit.setMinimumDate(QDate.currentDate())
        layout.addRow("到期日期：", expiry_edit)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("保存")
        buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("取消")
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        expiry_text = expiry_edit.date().toString("yyyy-MM-dd")
        config.set("license_expiry_date", expiry_text)
        QMessageBox.information(
            self,
            "授权设置",
            f"软件到期日期已设置为：{expiry_text}\n到期日当天仍可使用。"
        )

    def change_license_password(self):
        """修改管理员授权密码。"""
        if not self._verify_license_password():
            return

        new_password, ok = QInputDialog.getText(
            self,
            "修改授权密码",
            "请输入新密码（至少 6 位）：",
            QLineEdit.EchoMode.Password
        )
        if not ok:
            return
        if len(new_password) < 6:
            QMessageBox.warning(self, "修改授权密码", "密码至少需要 6 位。")
            return

        confirm_password, ok = QInputDialog.getText(
            self,
            "确认授权密码",
            "请再次输入新密码：",
            QLineEdit.EchoMode.Password
        )
        if not ok:
            return
        if new_password != confirm_password:
            QMessageBox.warning(self, "修改授权密码", "两次输入的密码不一致。")
            return

        self.scene.config_manager.set(
            "license_password_hash",
            hash_license_password(new_password)
        )
        QMessageBox.information(self, "修改授权密码", "授权密码已修改。")

    def set_nudge_large_step(self):
        """设置Shift+方向键大步长"""
        current = self.scene.config_manager.get('nudge_large_step', 10)
        val, ok = QInputDialog.getInt(self, "Shift+方向键步长", "步长（像素）:", current, 1, 1000)
        if ok:
            self.scene.config_manager.set('nudge_large_step', val)
            self.status_bar.showMessage(f'Shift+方向键步长已设置为 {val}px', 3000)

    def set_default_brush_size(self):
        """设置智能笔刷默认大小"""
        current = self.scene.config_manager.get('smart_brush_radius', 18)
        val, ok = QInputDialog.getInt(self, "笔刷默认大小", "笔刷半径（像素）:", current, 4, 120)
        if ok:
            self.scene.config_manager.set('smart_brush_radius', val)
            self.view._smart_brush_radius = val
            if hasattr(self, 'smart_brush_size_spin'):
                self.smart_brush_size_spin.blockSignals(True)
                self.smart_brush_size_spin.setValue(val)
                self.smart_brush_size_spin.blockSignals(False)
            self.status_bar.showMessage(f'笔刷默认大小已设置为 {val}px', 3000)

    def show_canvas_context_menu(self, pos):
        """画布空白处右键菜单"""
        menu = QMenu(self)
        point_mode_menu = menu.addMenu("独立连接点模式")
        move_mode_action = point_mode_menu.addAction("移动模式（选中/拖动/框选）")
        move_mode_action.setCheckable(True)
        connect_mode_action = point_mode_menu.addAction("连接模式（左键依次点击连接点）")
        connect_mode_action.setCheckable(True)
        current_point_mode = getattr(self.scene, 'free_connection_point_mode', 'move')
        move_mode_action.setChecked(current_point_mode == 'move')
        connect_mode_action.setChecked(current_point_mode == 'connect')
        point_mode_menu.addSeparator()
        add_free_point_here_action = menu.addAction("在此处新建连接点")
        add_free_point_action = menu.addAction("在最近辅助线交叉点新建连接点")
        menu.addSeparator()
        menu.addAction("添加文本", self.add_text)
        menu.addAction("插入图片", self.add_image)
        menu.addSeparator()
        menu.addAction("批量导入族谱... (Ctrl+I)", self.open_family_tree_import_dialog)
        menu.addSeparator()
        # 编辑组合模式下显示"更新到素材库"
        editing_id = getattr(self.scene, '_editing_group_asset_id', None)
        if editing_id is not None:
            menu.addAction(f"✅ 更新到素材库（完成编辑）", self.finish_edit_group_asset)
            menu.addAction("❌ 取消编辑组合", self.cancel_edit_group_asset)
            menu.addSeparator()
        menu.addAction("显示所有隐藏图片", self.show_all_hidden_images)
        menu.addSeparator()
        menu.addAction("打开工程\tCtrl+O", self.load_proj)
        menu.addAction("保存工程\tCtrl+S", self.save_proj)
        menu.addSeparator()
        menu.addAction("导出图片\tCtrl+E", self.export_image)
        menu.addAction("导出 PDF\tCtrl+Alt+P", self.export_pdf)
        menu.addAction("导出 CorelDRAW SVG", self.export_coreldraw_svg)
        action = menu.exec(self.view.mapToGlobal(pos))
        if action == move_mode_action:
            self.scene.set_free_connection_point_mode('move')
        elif action == connect_mode_action:
            self.scene.set_free_connection_point_mode('connect')
        elif action == add_free_point_here_action:
            self.scene.add_free_connection_point(self.view.mapToScene(pos))
        elif action == add_free_point_action:
            point = self.scene.add_free_connection_point_at_nearest_guide_intersection(
                self.view.mapToScene(pos)
            )
            if point is None:
                self.status_bar.showMessage("请先添加至少一条横向和一条竖向辅助线", 3000)

    def finish_edit_group_asset(self):
        """完成编辑，把放置时记录的完整元素列表更新到素材库"""
        asset_id = getattr(self.scene, '_editing_group_asset_id', None)
        if asset_id is None:
            return
        items = getattr(self.scene, '_editing_group_items', [])
        # 过滤掉已被删除的元素
        items = [i for i in items if i.scene() == self.scene]
        items = self._collect_complete_editing_group_items(items)
        if not items:
            QMessageBox.warning(self, "更新失败", "找不到编辑中的元素，请重新操作")
            return
        result = self.scene.asset_manager.update_group_asset(asset_id, items, self.scene)
        if result:
            self._remove_editing_group_items_from_canvas(items)
            self.scene._editing_group_asset_id = None
            self.scene._editing_group_items = []
            self.asset_library_dock.refresh_assets()
            self.status_bar.showMessage(f"组合「{result['name']}」已更新到素材库", 4000)
            QMessageBox.information(self, "更新成功", f"组合「{result['name']}」已成功更新到素材库")
        else:
            QMessageBox.warning(self, "更新失败", "更新素材库失败，请重试")

    def _collect_complete_editing_group_items(self, seed_items):
        """沿父子关系和图文连接收集编辑组合里的新增元素。"""
        collected = {
            item for item in seed_items
            if item.scene() == self.scene and isinstance(item, (VImageItem, VTextItem))
        }
        changed = True
        while changed:
            changed = False
            for item in list(collected):
                parent = item.parentItem()
                if parent and parent.scene() == self.scene and isinstance(parent, (VImageItem, VTextItem)) and parent not in collected:
                    collected.add(parent)
                    changed = True
                for child in item.childItems():
                    if child.scene() == self.scene and isinstance(child, (VImageItem, VTextItem)) and child not in collected:
                        collected.add(child)
                        changed = True
            for conn in list(self.scene.connectors) + list(self.scene.image_text_connectors):
                endpoints = []
                if hasattr(conn, 'parent_element') and hasattr(conn, 'child_element'):
                    endpoints = [conn.parent_element, conn.child_element]
                elif hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                    endpoints = [conn.image_item, conn.text_item]
                elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                    endpoints = [conn.item1, conn.item2]
                if any(item in collected for item in endpoints):
                    for item in endpoints:
                        if item and item.scene() == self.scene and isinstance(item, (VImageItem, VTextItem)) and item not in collected:
                            collected.add(item)
                            changed = True
        return sorted(collected, key=lambda item: (item.scenePos().y(), item.scenePos().x()))

    def _remove_editing_group_items_from_canvas(self, items):
        """更新素材库成功后，移除画布上的临时编辑副本。"""
        item_set = set(items)
        root_items = [
            item for item in items
            if item.scene() == self.scene and item.parentItem() not in item_set
        ]
        for item in items:
            if item.scene() == self.scene:
                self.scene.remove_all_connectors_for_item(item)
                self.scene.remove_image_text_connectors(item)
            if item in self.scene.selection_order:
                self.scene.selection_order = [i for i in self.scene.selection_order if i != item]
        for item in root_items:
            if item.scene() == self.scene:
                self.scene.removeItem(item)
        self.scene.clearSelection()
        self.scene.update()
        self.refresh_ui()

    def cancel_edit_group_asset(self):
        """取消编辑，清除编辑状态"""
        self.scene._editing_group_asset_id = None
        self.scene._editing_group_items = []
        self.status_bar.showMessage("已取消编辑组合", 3000)

    def fit_view(self):
        """初始化时适应视图"""
        # 使用新的智能适应方法
        self.view.fit_in_view()
        # 稍微缩小一点，留出边距
        self.view.scale(0.95, 0.95)
        self.view.transformChanged.emit()

    def add_text(self):
        text_format = self._default_text_format()
        t = VTextItem(
            "此处输入竖排文字\n支持自动换行\n从右向左排列",
            text_format['font_size'],
            400
        )
        self._apply_text_format(t, text_format)
        center = self.view.mapToScene(self.view.viewport().rect().center())
        t.setPos(center)
        self.scene.add_item_with_undo(t)
        
    def _calc_insert_size(self, pix, scene):
        """计算插入图片的目标宽高，返回 (width, height)，height=0 表示按比例自动"""
        # 优先：适应画布模式（强制拉伸填满画布）
        if scene.config_manager.get('insert_image_fit_canvas', False):
            canvas = scene.sceneRect()
            return int(canvas.width()), int(canvas.height())
        use_custom = scene.config_manager.get('insert_image_use_custom_size', False)
        if not use_custom:
            # 关闭自定义：按原图大小插入
            if not pix.isNull():
                return pix.width(), pix.height()
            return DEFAULT_FONT_SIZE * 4, 0

        fixed_w = scene.config_manager.get('insert_image_default_width', 0)
        fixed_h = scene.config_manager.get('insert_image_default_height', 0)
        if fixed_w and fixed_w > 0:
            w = int(fixed_w)
            h = int(fixed_h) if fixed_h and fixed_h > 0 else 0
            return w, h
        # 默认：画布宽度30%，高度按比例
        ratio = scene.config_manager.get('insert_image_max_width_ratio', 0.3)
        canvas_w = scene.sceneRect().width()
        max_w = int(canvas_w * ratio)
        w = min(pix.width(), max_w) if not pix.isNull() else DEFAULT_FONT_SIZE * 4
        return w, 0

    def _calc_insert_width(self, pix, scene):
        w, _ = self._calc_insert_size(pix, scene)
        return w

    def add_image(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "Images (*.png *.jpg *.jpeg *.bmp)")
        if path:
            pix = QPixmap(path)
            w, h = self._calc_insert_size(pix, self.scene) if not pix.isNull() else (DEFAULT_FONT_SIZE * 4, 0)
            img = VImageItem(path, target_width=w, target_height=h)
            if self.scene.config_manager.get('insert_image_fit_canvas', False):
                img.setPos(self.scene.sceneRect().topLeft())
            else:
                center = self.view.mapToScene(self.view.viewport().rect().center())
                img.setPos(center)
            if self.scene.config_manager.get('insert_image_to_bottom', False):
                img.setZValue(-1)
            self.scene.add_item_with_undo(img)
    
    def edit_selected_text(self):
        """编辑选中的文字"""
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        if selected_items:
            selected_items[0].start_inline_editing()
        else:
            print("请先选择一个文字元素")

    def toggle_smart_brush(self, enabled):
        self.view.set_smart_brush_enabled(enabled)
        if hasattr(self, 'smart_brush_action') and self.smart_brush_action.isChecked() != enabled:
            self.smart_brush_action.setChecked(enabled)

    def toggle_resize_mode(self, enabled):
        """切换图片调整大小模式"""
        self.scene.set_resize_mode(enabled)
        if enabled:
            self.status_bar.showMessage("调整大小模式：拖拽控制点缩放  回车/双击空白=确认  Esc=取消还原")
        else:
            self.status_bar.showMessage("已退出调整大小模式")
    
    def toggle_bg_above_connectors(self, enabled):
        """切换背景图片是否在连线之上"""
        self.scene.config_manager.set('bg_above_connectors', enabled)
        self.scene.update()
        # 强制所有连接器更新，因为它们的 paint 方法逻辑变了
        for c in self.scene.connectors:
            c.update()
        for c in self.scene.image_text_connectors:
            c.update()
    
    def undo(self):
        self.scene.undo()
    
    def redo(self):
        self.scene.redo()
    
    def align_top(self):
        self.scene.align_top()
    
    def align_right(self):
        self.scene.align_right()
    
    def align_center_horizontal(self):
        self.scene.align_center_horizontal()
    
    def align_center_vertical(self):
        self.scene.align_center_vertical()
    
    # 视图缩放方法
    def fit_in_view(self):
        """合适屏幕"""
        self.view.fit_in_view()
    
    def fill_view(self):
        """填充屏幕"""
        self.view.fill_view()
    
    def actual_size(self):
        """实际大小"""
        self.view.actual_size()
    
    def zoom_in(self):
        """放大"""
        self.view.zoom_in()
    
    def zoom_out(self):
        """缩小"""
        self.view.zoom_out()
    
    def zoom_to_selection(self):
        """缩放到选中内容"""
        self.view.zoom_to_selection()

    def toggle_bg_above_connectors(self, enabled):
        """切换背景图片是否在连线之上"""
        self.scene.config_manager.set('bg_above_connectors', enabled)
        self.scene.update()
        # 强制所有连接器更新，因为它们的 paint 方法逻辑变了
        for c in self.scene.connectors:
            c.update()
        for c in self.scene.image_text_connectors:
            c.update()
    
    def update_zoom_display(self):
        """更新缩放显示"""
        transform = self.view.transform()
        scale_factor = transform.m11()  # 获取x轴缩放因子
        zoom_percent = int(scale_factor * 100)
        self.zoom_label.setText(f"缩放: {zoom_percent}%")
    
    def auto_connect_selected(self):
        self.scene.auto_connect_selected_items()
    
    def clear_all_connections(self):
        self.scene.remove_all_image_text_connections()
    
    def toggle_connection_points(self):
        self.scene.toggle_connection_points()
        selected = self.scene.selectedItems() if self.scene else []
        has_images = any(isinstance(item, VImageItem) for item in selected)
        self._apply_connection_point_selection_visibility(has_images)
    
    def open_asset_library(self):
        """切换素材库面板的显示/隐藏"""
        if self.asset_library_dock.isVisible():
            self.asset_library_dock.hide()
        else:
            self.asset_library_dock.show()
            # 刷新素材库内容
            self.asset_library_dock.refresh_assets()
    
    def save_selected_as_group(self):
        self.scene.save_group_as_asset()

    def batch_copy(self):
        """步长和重复复制"""
        items = [i for i in self.scene.selectedItems() if isinstance(i, BaseElement)]
        if not items:
            QMessageBox.information(self, "步长和重复", "请先选中要复制的元素")
            return
        combined = QRectF()
        for item in items:
            combined = combined.united(QRectF(item.scenePos(), item.boundingRect().size()))
        dlg = BatchCopyDialog(item_w=combined.width(), item_h=combined.height(), parent=self)
        dlg.dspin_h_offset.setValue(-600)
        dlg.dspin_v_offset.setValue(600)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.scene.batch_copy(dlg.get_params())

    def open_family_tree_import_dialog(self):
        """从 Excel 批量导入族谱成员"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(
                self, "缺少依赖",
                "需要安装 openpyxl 库。\n请在命令行执行：pip install openpyxl"
            )
            return

        dlg = FamilyTreeImportDialog(self.scene, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        params = dlg.get_params()
        excel_path = params['excel_path']
        if not excel_path:
            QMessageBox.warning(self, "导入失败", "请先选择 Excel 文件")
            return
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, "导入失败", f"Excel 文件不存在：\n{excel_path}")
            return

        # 执行实际导入逻辑
        self._perform_excel_import(params, excel_path)

    def _open_excel_import_with_path(self, excel_path):
        """打开 Excel 导入对话框并预填文件路径（用于拖拽）"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(
                self, "缺少依赖",
                "需要安装 openpyxl 库。\n请在命令行执行：pip install openpyxl"
            )
            return
        
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, "文件不存在", f"Excel 文件不存在：\n{excel_path}")
            return
        
        dlg = FamilyTreeImportDialog(self.scene, self)
        dlg.txt_excel.setText(excel_path)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        
        # 获取参数并执行导入
        params = dlg.get_params()
        self._perform_excel_import(params, excel_path)

    def _perform_excel_import(self, params, excel_path):
        """执行Excel导入的核心逻辑"""
        try:
            import openpyxl
        except ImportError:
            return


        generation_map = {
            '1': 0, '一': 0, '一代': 0, '第1代': 0, '第一代': 0,
            '2': 1, '二': 1, '二代': 1, '第2代': 1, '第二代': 1,
            '3': 2, '三': 2, '三代': 2, '第3代': 2, '第三代': 2,
            '4': 3, '四': 3, '四代': 3, '第4代': 3, '第四代': 3,
            '5': 4, '五': 4, '五代': 4, '第5代': 4, '第五代': 4,
        }
        header_map = {}

        def cell_text(value):
            if value is None:
                return ''
            return str(value).strip()

        def column_index(column_name):
            column_name = str(column_name).strip()
            if not column_name:
                return None
            header_index = header_map.get(column_name.lower())
            if header_index is not None:
                return header_index
            if column_name.isdigit():
                return max(0, int(column_name) - 1)
            index = 0
            for char in column_name.upper():
                if not ('A' <= char <= 'Z'):
                    return None
                index = index * 26 + ord(char) - ord('A') + 1
            return index - 1

        def row_value(row, column_name):
            index = column_index(column_name)
            if index is None or index >= len(row):
                return ''
            return cell_text(row[index])

        def row_first_value(row, column_names):
            for column_name in column_names:
                value = row_value(row, column_name)
                if value:
                    return value
            return ''

        def normalize_key(value):
            return cell_text(value).lower()

        def make_full_text(main_text, sub_text1, sub_text2):
            parts = [text for text in (main_text, sub_text1, sub_text2) if text]
            return '\n'.join(parts)

        def item_scene_rect(item):
            return item.mapRectToScene(item.boundingRect())

        def rebuild_template_text(text_item, keep_template_anchor=False):
            anchor_pos = None
            if keep_template_anchor:
                anchor_pos = text_item.mapRectToScene(text_item.boundingRect()).topRight()
            text_item.rebuild()
            if anchor_pos is not None:
                new_anchor_pos = text_item.mapRectToScene(text_item.boundingRect()).topRight()
                target_scene_pos = text_item.scenePos() + (anchor_pos - new_anchor_pos)
                if text_item.parentItem():
                    text_item.setPos(text_item.parentItem().mapFromScene(target_scene_pos))
                else:
                    text_item.setPos(target_scene_pos)

        def add_extra_text_items(extra_texts, base_pos, anchor_items=None, ref_text_item=None):
            created_items = []
            if not extra_texts:
                return created_items
            # 字号、字体、列间距优先从参考文字对象（副文字2/t3/B组）取，取不到再用params默认值
            if ref_text_item and isinstance(ref_text_item, VTextItem):
                font_size = ref_text_item.font_size
                font_family = ref_text_item.font_family
                column_spacing = ref_text_item.column_spacing  # 使用副文字2的列间距
            else:
                font_size = params['font_size']
                font_family = params['font_family']
                column_spacing = COLUMN_SPACING  # 没有参考时使用默认列间距
            anchor_rect = QRectF(base_pos, QSizeF(0, 0))
            for item in anchor_items or []:
                anchor_rect = anchor_rect.united(item_scene_rect(item))
            x = anchor_rect.left() - font_size * 2 - column_spacing
            y = anchor_rect.top()
            for idx, text in enumerate(extra_texts):
                if not text:
                    continue
                text_item = VTextItem(text, font_size, 400)
                text_item.font_family = font_family
                text_item.column_spacing = column_spacing  # 设置与副文字2相同的列间距
                text_item.rebuild()
                text_item.setPos(x - idx * (font_size * 2 + column_spacing), y)
                text_cmd = AddItemCommand(self.scene, text_item)
                text_cmd.execute()
                commands.append(text_cmd)
                imported_items.append(text_item)
                created_items.append(text_item)
            return created_items

        def apply_template_text(new_items, cloned_by_asset_index, main_text, sub_text1, sub_text2):
            def remove_item_and_connectors(item):
                """删除元素及其所有相关的连接线（包括父子线和图文连接线）"""
                # 删除与该元素相关的所有连接线
                connectors_to_remove = []
                
                # 检查父子连接线
                for conn in self.scene.connectors:
                    if conn.child_element == item or conn.parent_element == item:
                        connectors_to_remove.append(conn)
                
                # 检查图文连接线
                for conn in self.scene.image_text_connectors:
                    if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                        if conn.image_item == item or conn.text_item == item:
                            connectors_to_remove.append(conn)
                    elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                        if conn.item1 == item or conn.item2 == item:
                            connectors_to_remove.append(conn)
                
                # 删除找到的连接线
                for conn in connectors_to_remove:
                    self.scene.remove_connector_item(conn)
                
                # 删除元素本身
                self.scene.removeItem(item)
            
            placeholders = params['placeholders']
            slot_values = {
                't1': main_text,
                't2': sub_text1,
                't3': sub_text2,
            }
            template_text_map = params['template_text_map']
            slot_items = {}

            mapped_any = False
            for slot, value in slot_values.items():
                asset_index = template_text_map.get(slot)
                if asset_index is None or asset_index < 0 or asset_index >= len(cloned_by_asset_index):
                    continue
                text_item = cloned_by_asset_index[asset_index]
                if isinstance(text_item, VTextItem):
                    # 如果值为空，删除该文本对象及其连接线
                    if not value or not value.strip():
                        remove_item_and_connectors(text_item)
                        cloned_by_asset_index[asset_index] = None
                        if text_item in new_items:
                            new_items.remove(text_item)
                    else:
                        text_item.full_text = value
                        rebuild_template_text(text_item, keep_template_anchor=True)
                        slot_items[slot] = text_item
                    mapped_any = True

            if mapped_any:
                return slot_items

            values = {
                placeholders['t1'].strip().lower(): slot_values['t1'],
                placeholders['t2'].strip().lower(): slot_values['t2'],
                placeholders['t3'].strip().lower(): slot_values['t3'],
            }
            placeholder_to_slot = {
                placeholders['t1'].strip().lower(): 't1',
                placeholders['t2'].strip().lower(): 't2',
                placeholders['t3'].strip().lower(): 't3',
            }
            text_items = [item for item in new_items if isinstance(item, VTextItem)]
            matched = set()

            for text_item in text_items:
                key = text_item.full_text.strip().lower()
                if key in values:
                    value = values[key]
                    # 如果值为空，删除该文本对象及其连接线
                    if not value or not value.strip():
                        remove_item_and_connectors(text_item)
                        if text_item in new_items:
                            new_items.remove(text_item)
                    else:
                        text_item.full_text = value
                        rebuild_template_text(text_item, keep_template_anchor=True)
                        slot_items[placeholder_to_slot[key]] = text_item
                    matched.add(key)

            if matched:
                return slot_items

            ordered_text_items = sorted(
                text_items,
                key=lambda item: (item.scenePos().x(), item.scenePos().y()),
                reverse=True
            )
            for slot, text_item in zip(('t1', 't2', 't3'), ordered_text_items):
                value = slot_values[slot]
                # 如果值为空，删除该文本对象及其连接线
                if not value or not value.strip():
                    remove_item_and_connectors(text_item)
                    if text_item in new_items:
                        new_items.remove(text_item)
                else:
                    text_item.full_text = value
                    rebuild_template_text(text_item, keep_template_anchor=True)
                    slot_items[slot] = text_item
            return slot_items

        def clone_group_asset(asset, base_pos):
            min_x = min(d['scene_pos'][0] for d in asset['items'])
            min_y = min(d['scene_pos'][1] for d in asset['items'])
            new_items = []
            cloned_by_asset_index = [None] * len(asset['items'])
            commands = []
            for item_index, item_data in enumerate(asset['items']):
                new_item = None
                if item_data['type'] == 'VTextItem':
                    new_item = VTextItem(item_data['text'], item_data['font_size'], item_data['box_height'])
                    new_item.font_family = item_data['font_family']
                    new_item.text_color = QColor(item_data['text_color'])
                    for key in ('chars_per_column', 'column_spacing', 'character_spacing', 'auto_height', 'manual_line_break', 'layer_eye_color'):
                        if key in item_data:
                            setattr(new_item, key, item_data[key])
                    new_item.rebuild()
                elif item_data['type'] == 'VImageItem' and os.path.exists(item_data['path']):
                    new_item = VImageItem(item_data['path'], item_data['width'])
                elif item_data['type'] == 'VImageItem':
                    image_path = item_data.get('path', '')
                    if image_path and not os.path.isabs(image_path):
                        image_path = os.path.join(APP_DIR, image_path)
                    if image_path and os.path.exists(image_path):
                        new_item = VImageItem(image_path, item_data['width'])

                if new_item:
                    off_x = item_data['scene_pos'][0] - min_x
                    off_y = item_data['scene_pos'][1] - min_y
                    new_item.setPos(base_pos.x() + off_x, base_pos.y() + off_y)
                    command = AddItemCommand(self.scene, new_item)
                    command.execute()
                    commands.append(command)
                    if item_data.get('connection_point_deleted', False):
                        new_item.delete_connection_point()
                    elif 'connection_point_visible' in item_data and new_item.connection_point:
                        new_item.connection_point.setVisible(item_data['connection_point_visible'])
                    new_items.append(new_item)
                    cloned_by_asset_index[item_index] = new_item

            for idx, item_data in enumerate(asset['items']):
                if item_data['parent_index'] != -1 and item_data['parent_index'] < len(cloned_by_asset_index):
                    child = cloned_by_asset_index[idx]
                    parent = cloned_by_asset_index[item_data['parent_index']]
                    if not child or not parent:
                        continue
                    scene_pos = child.scenePos()
                    child.setParentItem(parent)
                    child.setPos(parent.mapFromScene(scene_pos))
                    self.scene.add_connector(parent, child)

            for item1_idx, item2_idx in asset['image_text_connections']:
                if item1_idx < len(cloned_by_asset_index) and item2_idx < len(cloned_by_asset_index):
                    item1 = cloned_by_asset_index[item1_idx]
                    item2 = cloned_by_asset_index[item2_idx]
                    if not item1 or not item2:
                        continue
                    conn_cmd = self.scene._make_connector_command(item1, item2)
                    if conn_cmd:
                        conn_cmd.execute()
                        commands.append(conn_cmd)

            # 最终再统一应用连接点可见性/删除状态，防止被 add_connector/setParentItem 流程覆盖
            for item_index, item_data in enumerate(asset['items']):
                cloned = cloned_by_asset_index[item_index]
                if cloned:
                    if item_data.get('connection_point_deleted', False):
                        cloned.delete_connection_point()
                    elif 'connection_point_visible' in item_data and cloned.connection_point:
                        cloned.connection_point.setVisible(item_data['connection_point_visible'])

            return new_items, cloned_by_asset_index, commands

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"读取 Excel 失败：\n{e}")
            return

        # 设置批量导入标志，禁止显示临时对齐线
        self.scene._batch_importing = True

        if rows:
            header_map.update({
                cell_text(value).lower(): idx
                for idx, value in enumerate(rows[0])
                if cell_text(value)
            })

        generation_next_x = [params['start_x']] * len(params['y_coords'])
        generation_layout_spacing = [params['spacing']] * len(params['y_coords'])
        imported_items = []
        commands = []
        skipped = 0
        count_a = 0  # A组（主记录）数量
        count_d = 0  # D组（extra_texts）文字数量
        template_asset = params['template_asset']
        base_dir = os.path.dirname(excel_path)
        columns = params['columns']
        group_match = params['group_match']
        requested_template_name = str(params.get('template_name', '')).strip()
        if requested_template_name and requested_template_name != "— 不使用模板，纯文字 —" and template_asset is None:
            QMessageBox.warning(self, "导入失败", f"没有找到组合素材模板：\n{requested_template_name}")
            return
        canvas_rect = self.scene.sceneRect()
        import_records = []

        def clamp_to_canvas(point):
            return QPointF(
                min(max(point.x(), canvas_rect.left()), canvas_rect.right()),
                min(max(point.y(), canvas_rect.top()), canvas_rect.bottom())
            )

        def keep_items_inside_canvas(items):
            if not items:
                return
            rect = QRectF()
            for item in items:
                rect = rect.united(item_scene_rect(item))
            if rect.isEmpty():
                return
            dx = 0
            dy = 0
            if rect.left() < canvas_rect.left():
                dx = canvas_rect.left() - rect.left()
            elif rect.right() > canvas_rect.right():
                dx = canvas_rect.right() - rect.right()
            if rect.top() < canvas_rect.top():
                dy = canvas_rect.top() - rect.top()
            elif rect.bottom() > canvas_rect.bottom():
                dy = canvas_rect.bottom() - rect.bottom()
            if abs(dx) > 0.1 or abs(dy) > 0.1:
                for item in items:
                    if item.parentItem():
                        continue
                    item.moveBy(dx, dy)

        def collect_import_record(record):
            nonlocal skipped
            generation = record['generation']
            main_text = record['main_text']
            sub_text1 = record['sub_text1']
            sub_text2 = record['sub_text2']
            image_path = record['image_path']
            extra_texts = record.get('extra_texts', [])

            if generation in ('', '辈分', '代数', 'generation'):
                skipped += 1
                return

            generation_idx = generation_map.get(generation)
            if generation_idx is None or generation_idx >= len(params['y_coords']):
                skipped += 1
                return
            if not any((main_text, sub_text1, sub_text2, image_path, extra_texts)):
                skipped += 1
                return
            record['generation_idx'] = generation_idx
            record['source_order'] = len(import_records)
            import_records.append(record)

        def add_import_record(record):
            nonlocal skipped, count_a, count_d
            generation_idx = record['generation_idx']
            main_text = record['main_text']
            sub_text1 = record['sub_text1']
            sub_text2 = record['sub_text2']
            image_path = record['image_path']
            extra_texts = record.get('extra_texts', [])

            x = generation_next_x[generation_idx]
            y = params['y_coords'][generation_idx]
            base_pos = clamp_to_canvas(QPointF(x, y))
            full_text = make_full_text(main_text, sub_text1, sub_text2)
            count_a += 1
            count_d += len([t for t in extra_texts if t])

            if template_asset:
                new_items, cloned_by_asset_index, new_commands = clone_group_asset(template_asset, base_pos)
                slot_items = apply_template_text(new_items, cloned_by_asset_index, main_text, sub_text1, sub_text2)
                imported_items.extend(new_items)
                commands.extend(new_commands)
                # 取B组对应的t3文字对象作为D组的字号字体参考
                t3_index = params['template_text_map'].get('t3')
                ref_item = slot_items.get('t3') or (cloned_by_asset_index[t3_index] if t3_index is not None and 0 <= t3_index < len(cloned_by_asset_index) else None) or next((i for i in new_items if isinstance(i, VTextItem)), None)
                extra_items = add_extra_text_items(extra_texts, base_pos, new_items, ref_text_item=ref_item)
                placed_items = new_items + extra_items
                keep_items_inside_canvas(placed_items)
                placed_rect = QRectF()
                for item in placed_items:
                    placed_rect = placed_rect.united(item_scene_rect(item))
                if not placed_rect.isEmpty():
                    generation_next_x[generation_idx] = placed_rect.left() - generation_layout_spacing[generation_idx]
                return

            row_items = []
            ref_item = None
            if image_path:
                if not os.path.isabs(image_path):
                    image_path = os.path.join(base_dir, image_path)
                if os.path.exists(image_path):
                    image_item = VImageItem(image_path, params['img_width'])
                    image_item.setPos(base_pos)
                    image_cmd = AddItemCommand(self.scene, image_item)
                    image_cmd.execute()
                    commands.append(image_cmd)
                    imported_items.append(image_item)
                    row_items.append(image_item)
                else:
                    skipped += 1

            if full_text:
                text_item = VTextItem(full_text, params['font_size'], 400)
                text_item.font_family = params['font_family']
                text_item.rebuild()
                if row_items:
                    text_item.setPos(base_pos.x() - params['img_width'] - 20, base_pos.y())
                else:
                    text_item.setPos(base_pos)
                text_cmd = AddItemCommand(self.scene, text_item)
                text_cmd.execute()
                commands.append(text_cmd)
                imported_items.append(text_item)
                row_items.append(text_item)
                ref_item = text_item

            extra_items = add_extra_text_items(extra_texts, base_pos, row_items, ref_text_item=ref_item)
            placed_items = row_items + extra_items
            keep_items_inside_canvas(placed_items)
            placed_rect = QRectF()
            for item in placed_items:
                placed_rect = placed_rect.united(item_scene_rect(item))
            if not placed_rect.isEmpty():
                generation_next_x[generation_idx] = placed_rect.left() - generation_layout_spacing[generation_idx]

            if params['auto_connect'] and len(row_items) >= 2:
                conn_cmd = self.scene._make_connector_command(row_items[0], row_items[1])
                if conn_cmd:
                    conn_cmd.execute()
                    commands.append(conn_cmd)

        if group_match['enabled']:
            group_to_field = {
                normalize_key(group_match['t1']): 'main_text',
                normalize_key(group_match['t2']): 'sub_text1',
                normalize_key(group_match['t3']): 'sub_text2',
                normalize_key(group_match['extra_text']): 'extra_texts',
            }
            current_record = None
            current_fields = set()

            def flush_current_record():
                nonlocal current_record, current_fields
                if current_record and any((
                        current_record['main_text'],
                        current_record['sub_text1'],
                        current_record['sub_text2'],
                        current_record['image_path'],
                        current_record.get('extra_texts'))):
                    collect_import_record(current_record)
                current_record = None
                current_fields = set()

            for row in rows:
                generation = row_value(row, columns['generation'])
                group_value = normalize_key(row_value(row, columns['group']))
                name_text = row_first_value(row, (columns['name'], '名称', '内容'))
                image_path = row_value(row, columns['image'])

                if group_value in ('', normalize_key(columns['group'])):
                    continue

                field = group_to_field.get(group_value)
                if not field or not name_text:
                    skipped += 1
                    continue

                if current_record and (
                        (field in current_fields and field != 'extra_texts') or
                        (generation and current_record['generation'] and generation != current_record['generation'])):
                    flush_current_record()

                if current_record is None:
                    current_record = {
                        'generation': generation,
                        'main_text': '',
                        'sub_text1': '',
                        'sub_text2': '',
                        'image_path': '',
                        'extra_texts': [],
                    }
                elif generation and not current_record['generation']:
                    current_record['generation'] = generation

                if field == 'extra_texts':
                    current_record[field].append(name_text)
                else:
                    current_record[field] = name_text
                current_fields.add(field)
                if image_path and not current_record['image_path']:
                    current_record['image_path'] = image_path

            flush_current_record()
        else:
            for row in rows:
                record = {
                    'generation': row_value(row, columns['generation']),
                    'main_text': row_value(row, columns['t1']),
                    'sub_text1': row_value(row, columns['t2']),
                    'sub_text2': row_value(row, columns['t3']),
                    'image_path': row_value(row, columns['image']),
                    'extra_texts': [],
                }
                if record['generation'] in ('', '辈分', '代数', 'generation'):
                    continue
                collect_import_record(record)

        records_by_generation = {}
        for record in import_records:
            records_by_generation.setdefault(record['generation_idx'], []).append(record)
        for generation_idx, records in records_by_generation.items():
            count = len(records)
            if count <= 1:
                continue
            available_width = max(0, params['start_x'] - canvas_rect.left())
            if available_width <= 0:
                continue
            max_spacing = available_width / (count - 1)
            generation_layout_spacing[generation_idx] = max(10, min(params['spacing'], max_spacing))

        for record in sorted(import_records, key=lambda r: (r['generation_idx'], r['source_order'])):
            add_import_record(record)

        if not imported_items:
            QMessageBox.information(self, "导入结果", "没有导入任何数据，请检查 Excel 内容")
            return

        self.scene.undo_stack.push(MacroCommand(self.scene, commands))
        
        # 清除批量导入标志
        self.scene._batch_importing = False
        
        # 清除辅助线选择状态和所有辅助线，避免导入后显示对齐提示线
        self.scene.clear_guide_selection()
        
        # 强制隐藏并清除临时对齐线（导入时setPos触发的残留辅助线）
        if self.scene._temp_alignment_guide is not None:
            self.scene.removeItem(self.scene._temp_alignment_guide)
            self.scene._temp_alignment_guide = None
        self.scene._image_right_edge_snap_indicator = None
        self.scene._image_top_edge_snap_indicator = None
        self.scene.update()  # 强制刷新场景
        
        # 可选：清除所有辅助线（如果不想保留导入产生的辅助线）
        # self.scene.clear_guides()
        
        # 计算导入区域用于视图缩放
        imported_rect = QRectF()
        for item in imported_items:
            imported_rect = imported_rect.united(item_scene_rect(item))
        
        # 取消选择所有元素，避免显示蓝色选择框和对齐线
        self.scene.clearSelection()
        
        # 缩放视图到导入区域
        if not imported_rect.isEmpty():
            self.view.fitInView(imported_rect.adjusted(-100, -100, 100, 100), Qt.AspectRatioMode.KeepAspectRatio)
            self.view.transformChanged.emit()
        self.scene.update()
        self.status_bar.showMessage(f"已导入 {len(imported_items)} 个元素", 4000)
        QMessageBox.information(
            self, "导入完成",
            f"已导入 {len(imported_items)} 个元素。\n"
            f"A组（主记录）：{count_a} 条\n"
            f"D组（附加文字）：{count_d} 条\n"
            f"跳过 {skipped} 行/项。"
        )
    
    def _open_font_dialog(self):
        """弹出字体选择对话框"""
        current = QFont(self.font_combo.currentText())
        current.setPointSize(self.font_size_spin.value())
        font, ok = FontPickerDialog.get_font(current, self.scene.config_manager, self, "选择字体")
        if ok:
            self.font_combo.setCurrentText(font.family())
            self.change_selected_font(font)

    def change_selected_font(self, font):
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        for item in selected_items:
            item.font_family = font.family()
            item.rebuild()
    
    def change_selected_font_size(self, size):
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        for item in selected_items:
            item.font_size = size
            item.rebuild()
    
    def change_selected_color(self):
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        if not selected_items: return
        color = QColorDialog.getColor(selected_items[0].text_color, self, "选择文字颜色")
        if color.isValid():
            for item in selected_items:
                item.text_color = color
                item.rebuild()
            self._set_color_button_color(color)
    
    def toggle_manual_line_break(self, enabled):
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        for item in selected_items:
            item.manual_line_break = enabled
            item.rebuild()
    
    def change_chars_per_column(self, chars_count):
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        for item in selected_items:
            item.chars_per_column = chars_count
            item.rebuild()
    
    def change_column_spacing(self, spacing):
        """改变选中文字的列间距"""
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        for item in selected_items:
            item.column_spacing = spacing
            item.rebuild()

    def change_character_spacing(self, spacing):
        """改变选中文字在同一竖列中的字间距。"""
        selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
        for item in selected_items:
            item.character_spacing = spacing
            item.rebuild()
    
    def update_font_controls(self):
        try:
            # 检查场景和对象是否还存在
            if not self.scene:
                return
            
            selected_items = [item for item in self.scene.selectedItems() if isinstance(item, VTextItem)]
            if selected_items:
                item = selected_items[0]
                self.font_combo.blockSignals(True)
                self.font_size_spin.blockSignals(True)
                self.chars_per_column_spin.blockSignals(True)
                self.column_spacing_spin.blockSignals(True)
                self.character_spacing_spin.blockSignals(True)
                self.manual_line_break_btn.blockSignals(True)

                self.font_combo.setCurrentText(item.font_family)
                self.font_size_spin.setValue(item.font_size)
                self.chars_per_column_spin.setValue(item.chars_per_column)
                self.column_spacing_spin.setValue(item.column_spacing)
                self.character_spacing_spin.setValue(item.character_spacing)
                self.manual_line_break_btn.setChecked(item.manual_line_break)
                self._set_color_button_color(item.text_color)
                
                self.font_combo.blockSignals(False)
                self.font_size_spin.blockSignals(False)
                self.chars_per_column_spin.blockSignals(False)
                self.column_spacing_spin.blockSignals(False)
                self.character_spacing_spin.blockSignals(False)
                self.manual_line_break_btn.blockSignals(False)
        except (RuntimeError, AttributeError):
            # 处理 C++ 对象已被删除的情况
            pass
    
    def on_selection_changed(self):
        try:
            # 检查场景是否还存在
            if not self.scene:
                return

            if hasattr(self, 'font_combo'):
                self.update_font_controls()
            # 状态栏提示
            selected = self.scene.selectedItems()
            images = [i for i in selected if isinstance(i, VImageItem)]
            texts = [i for i in selected if isinstance(i, VTextItem)]
            # 缓存选中的图片列表，供层级面板批量操作使用
            if images:
                self._last_selected_images = list(images)
            self._apply_connection_point_selection_visibility(bool(images))
            self.update_property_summary(selected, texts, images)

            # ── 单选时在状态栏显示右上角 CDR 坐标 ──────────────────────
            single = None
            if len(images) + len(texts) == 1:
                single = images[0] if images else texts[0]

            if single is not None:
                canvas_h = self.scene.sceneRect().height()
                DPI = CORELDRAW_EXPORT_DPI
                pos = single.scenePos()
                w   = single.boundingRect().width()
                # 右上角 Qt 坐标 → CDR 坐标（mm，Y轴向上）
                cdr_x = round((pos.x() + w) * 25.4 / DPI, 2)
                cdr_y = round((canvas_h - pos.y()) * 25.4 / DPI, 2)
                kind  = "图片" if isinstance(single, VImageItem) else "文字"
                self.status_bar.showMessage(
                    f"{kind} 右上角  X: {cdr_x} mm  Y: {cdr_y} mm（CDR坐标）"
                )
            elif images and not texts:
                self.status_bar.showMessage("拖拽蓝色控制点可缩放图片  角点=等比缩放  边中点=单向拉伸  右键=更多选项")
            elif texts and not images:
                self.status_bar.showMessage("双击文字可编辑  右键=字体/颜色/列间距等设置")
            elif images and texts:
                self.status_bar.showMessage("右键=批量连接/对齐  Ctrl+G=保存组合")
            else:
                connectors = [i for i in selected if isinstance(i, (VImageTextConnector, VGenericConnector))]
                mode = self.scene.config_manager.get('marquee_mode', 'all')
                if connectors and mode == 'images':
                    self.status_bar.showMessage("已选中连线 — 当前仅图片模式，Alt+M 切换为全选模式", 5000)
                elif connectors:
                    self.status_bar.showMessage("已选中连线，右键可设置粗细或删除  Delete=删除")
                else:
                    self.status_bar.showMessage("")
            self._sync_tree_selection_to_scene()
        except (RuntimeError, AttributeError):
            pass

    def toggle_hide_points_on_image_select(self, enabled):
        self.scene.config_manager.set('hide_connection_points_when_image_selected', enabled)
        selected = self.scene.selectedItems() if self.scene else []
        has_images = any(isinstance(item, VImageItem) for item in selected)
        self._apply_connection_point_selection_visibility(has_images)

    def _apply_connection_point_selection_visibility(self, has_selected_images):
        if not self.scene:
            return
        hide_when_image_selected = self.scene.config_manager.get(
            'hide_connection_points_when_image_selected', False
        )
        visible = self.scene.show_connection_points and not (hide_when_image_selected and has_selected_images)
        for item in self.scene.items():
            if isinstance(item, (VTextItem, VImageItem)):
                item.set_connection_points_visible(visible)

    def toggle_text_hover_tooltip(self, enabled):
        self.scene.config_manager.set('show_text_hover_tooltip', enabled)
        for item in self.scene.items():
            if isinstance(item, VTextItem):
                item.update_hover_tooltip()
    
    def set_line_width(self):
        """设置连线粗细，同时更新默认值和所有现有连线"""
        current = self.scene.config_manager.get('default_line_width', DEFAULT_LINE_WIDTH)
        width, ok = QInputDialog.getInt(self, "设置连线粗细", "连线粗细 (像素):", current, 1, 20)
        if not ok:
            return
        self.scene.config_manager.set('default_line_width', width)
        count = 0
        for conn in self.scene.image_text_connectors:
            if hasattr(conn, 'set_line_width'):
                conn.set_line_width(width)
                count += 1
        print(f"连线粗细已设置为 {width}px，共更新 {count} 条")
    
    def set_canvas_size(self):
        current_rect = self.scene.sceneRect()
        w, ok1 = QInputDialog.getInt(self, "画布宽度", "宽度:", int(current_rect.width()), 100, 10000)
        h, ok2 = QInputDialog.getInt(self, "画布高度", "高度:", int(current_rect.height()), 100, 10000)
        if ok1 and ok2: self.scene.setSceneRect(0, 0, w, h)

    def _make_eye_icon(self, visible, color=None):
        """创建图层列表用的眼睛图标，支持自定义颜色。"""
        pixmap = QPixmap(22, 22)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        if color is None:
            color = QColor(0, 120, 215) if visible else QColor(150, 150, 150)
        pen = QPen(color, 2)
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        eye_path = QPainterPath()
        eye_path.moveTo(3, 11)
        eye_path.cubicTo(7, 4, 15, 4, 19, 11)
        eye_path.cubicTo(15, 18, 7, 18, 3, 11)
        painter.drawPath(eye_path)

        if visible:
            painter.setBrush(QBrush(color))
            painter.drawEllipse(QPointF(11, 11), 3, 3)
        else:
            painter.drawLine(5, 17, 17, 5)

        painter.end()
        return QIcon(pixmap)

    def _tree_item_element(self, node):
        return node.data(0, Qt.ItemDataRole.UserRole)

    def _tree_key_for_item(self, item):
        return id(item)

    def _sync_tree_node_visibility(self, node):
        item = self._tree_item_element(node)
        if not item:
            return
        visible = item.isVisible()

        # 彩色眼睛由用户给文字设置的标记决定，不再按位置推断。
        eye_color = None
        if isinstance(item, VTextItem):
            color_key = self._get_child_text_eye_color(item)
            eye_color = {
                'yellow': QColor(220, 180, 0),
                'red': QColor(200, 50, 50),
                'green': QColor(40, 160, 70),
            }.get(color_key)

        if not visible:
            eye_color = None  # 隐藏时统一用灰色
        node.setIcon(0, self._make_eye_icon(visible, eye_color))
        node.setToolTip(0, "点击隐藏" if visible else "点击显示")
        text_color = QColor(0, 100, 200) if isinstance(item, VImageItem) else QColor(30, 30, 30)
        if not visible:
            text_color = QColor(150, 150, 150)
        node.setForeground(1, QBrush(text_color))
        node.setForeground(2, QBrush(text_color))

    def _set_element_visible_from_tree(self, item, visible):
        if isinstance(item, VImageItem):
            item.set_image_visible(visible)
        else:
            item.setVisible(visible)

        # 同步父子连线可见性
        for conn in self.scene.connectors:
            if conn.parent_element == item or conn.child_element == item:
                conn.setVisible(
                    self.scene.show_connectors and
                    conn.parent_element.isVisible() and
                    conn.child_element.isVisible()
                )

        # 同步图文/通用连线可见性
        for conn in self.scene.image_text_connectors:
            if hasattr(conn, 'image_item') and hasattr(conn, 'text_item'):
                i1, i2 = conn.image_item, conn.text_item
            elif hasattr(conn, 'item1') and hasattr(conn, 'item2'):
                i1, i2 = conn.item1, conn.item2
            else:
                continue
            if i1 == item or i2 == item:
                conn.setVisible(
                    self.scene.show_image_text_connectors and
                    i1.isVisible() and i2.isVisible()
                )
        self.scene.update()

    def _sync_tree_selection_to_scene(self):
        """画布选中变化时，在右侧列表中定位并高亮对应元素。"""
        if self._tree_updating or not hasattr(self, 'tree_widget'):
            return

        selected = [
            item for item in self.scene.selectedItems()
            if isinstance(item, (VImageItem, VTextItem))
        ]
        if not selected:
            self.tree_widget.clearSelection()
            return

        ordered = [
            item for item in getattr(self.scene, 'selection_order', [])
            if item in selected
        ]
        focus_item = ordered[-1] if ordered else selected[-1]

        self._tree_updating = True
        try:
            self.tree_widget.clearSelection()
            for item in selected:
                node = self._tree_nodes_by_item.get(self._tree_key_for_item(item))
                if node:
                    node.setSelected(True)

            # 滚动到最后选中的元素，但不用 setCurrentItem 避免蓝框只跟一个节点
            focus_node = self._tree_nodes_by_item.get(self._tree_key_for_item(focus_item))
            if focus_node:
                self.tree_widget.scrollToItem(focus_node, QAbstractItemView.ScrollHint.PositionAtCenter)
        finally:
            self._tree_updating = False

    def refresh_ui(self):
        try:
            # 检查场景是否还存在
            if not self.scene:
                return

            self.tree_widget.clear()
            self._tree_nodes_by_item = {}
            self._tree_item_counter = 0
            self._tree_updating = True
            def add_node(item, parent_node):
                self._tree_item_counter += 1
                n = self._tree_item_counter
                node = QTreeWidgetItem(parent_node)

                if isinstance(item, VImageItem):
                    label = f"[{n}] 图片  {os.path.basename(item.file_path)}"
                else:
                    preview = item.full_text.replace('\n', '↵')[:12]
                    if len(item.full_text) > 12:
                        preview += '…'
                    # 如果是图片的子文字，加位置标记
                    tag = ''
                    parent = item.parentItem()
                    if isinstance(parent, VImageItem):
                        siblings = [c for c in parent.childItems() if isinstance(c, VTextItem)]
                        if len(siblings) > 1:
                            xs = [c.scenePos().x() for c in siblings]
                            ys = [c.scenePos().y() for c in siblings]
                            sx, sy = item.scenePos().x(), item.scenePos().y()
                            marks = []
                            if sx == max(xs):
                                marks.append('最右')
                            if sx == min(xs):
                                marks.append('最左')
                            if sy == min(ys):
                                marks.append('最上')
                            if sy == max(ys):
                                marks.append('最下')
                            if marks:
                                tag = f" [{'/'.join(marks)}]"
                    label = f"[{n}] 文字  {preview}{tag}"

                pos = item.scenePos()
                node.setText(1, label)
                node.setText(2, f"({int(pos.x())}, {int(pos.y())})")
                node.setData(0, Qt.ItemDataRole.UserRole, item)
                node.setData(1, Qt.ItemDataRole.UserRole, item)
                node.setData(2, Qt.ItemDataRole.UserRole, item)
                node.setFlags(node.flags() & ~Qt.ItemFlag.ItemIsUserCheckable)
                self._sync_tree_node_visibility(node)
                self._tree_nodes_by_item[self._tree_key_for_item(item)] = node

                for child in item.childItems():
                    if isinstance(child, BaseElement):
                        add_node(child, node)

            for item in self.scene.items():
                if isinstance(item, BaseElement) and item.parentItem() is None:
                    add_node(item, self.tree_widget)

            self.tree_widget.expandAll()
            self._tree_updating = False
            self._sync_tree_selection_to_scene()
            self.scene.update_all_connectors()
        except (RuntimeError, AttributeError):
            pass
        except Exception:
            pass

    def _on_tree_item_clicked(self, node, column):
        """点击眼睛切换显隐；点击列表内容选中画布元素并居中。
        若点击的是用户指定过颜色的子文字眼睛，且画布上有多个选中图片，则按同色标记批量操作。
        """
        try:
            item = self._tree_item_element(node)
            if item and item.scene():
                if column == 0:
                    # 判断是否是用户指定了彩色眼睛的子文字
                    color_key = self._get_child_text_role(item)
                    if color_key and isinstance(item, VTextItem) and isinstance(item.parentItem(), VImageItem):
                        # 用缓存的选中图片列表（避免点击面板时选中状态已变）
                        selected_images = [i for i in getattr(self, '_last_selected_images', [])
                                           if isinstance(i, VImageItem) and i.scene()]
                        if len(selected_images) > 1:
                            new_visible = not item.isVisible()
                            for img in selected_images:
                                target = self._get_role_child(img, color_key)
                                if target:
                                    self._set_element_visible_from_tree(target, new_visible)
                            self.refresh_ui()
                            return

                    self._set_element_visible_from_tree(item, not item.isVisible())
                    self._sync_tree_node_visibility(node)
                    return

                self.scene.clearSelection()
                item.setSelected(True)
                self.view.centerOn(item)
        except (RuntimeError, AttributeError):
            pass

    def _get_child_text_role(self, item):
        """返回当前子文字的用户指定眼睛颜色，否则返回 None。"""
        return self._get_child_text_eye_color(item)

    def _get_child_text_eye_color(self, item):
        """返回当前子文字应显示的彩色眼睛颜色 key，否则返回 None。"""
        if not isinstance(item, VTextItem):
            return None
        color_key = getattr(item, 'layer_eye_color', None)
        return color_key if color_key in {'yellow', 'red', 'green'} else None

    def _get_role_child(self, img, color_key):
        """从图片的子文字中找到用户标记为对应颜色的那个。"""
        siblings = [c for c in img.childItems() if isinstance(c, VTextItem)]
        if not siblings:
            return None
        for c in siblings:
            if self._get_child_text_eye_color(c) == color_key:
                return c
        return None

    def _on_tree_item_changed(self, node, column):
        """保留旧信号槽，兼容旧工程/旧 UI 调用。"""
        return

    def export_image(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export Image", "", "PNG (*.png)")
        if path:
            # 保存当前设置
            original_show_grid = self.scene.show_grid
            original_show_connectors = self.scene.show_connectors
            original_show_connection_points = self.scene.show_connection_points
            original_show_guides = self.scene.show_guides

            # 导出时隐藏网格、父子连线、连接点、辅助线
            self.scene.show_grid = False
            self.scene.set_connectors_visible(False)
            self.scene.set_connection_points_visible(False)
            self.scene.set_guides_visible(False)
            
            try:
                # 临时清除选中状态，避免选中框被渲染到图片中
                selected_items = self.scene.selectedItems()
                self.scene.clearSelection()
                
                rect = self.scene.sceneRect()
                img = QImage(rect.size().toSize(), QImage.Format.Format_ARGB32)
                img.fill(Qt.GlobalColor.white)
                p = QPainter(img)
                self.scene.render(p)
                p.end()
                img.save(path)
                print(f"图片已导出到: {path}")
                
                # 恢复选中状态
                for item in selected_items:
                    item.setSelected(True)
            finally:
                # 恢复原始设置
                self.scene.show_grid = original_show_grid
                self.scene.set_connectors_visible(original_show_connectors)
                self.scene.set_connection_points_visible(original_show_connection_points)
                self.scene.set_guides_visible(original_show_guides)

    def export_excel(self):
        """导出场景数据到 Excel：坐标已转换为 CorelDRAW 坐标系（Y轴翻转，px→mm）"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(
                self, "缺少依赖",
                "需要安装 openpyxl 库。\n请在命令行执行：pip install openpyxl"
            )
            return

        default_dir = self.scene.config_manager.get('default_save_dir', '') or \
                      (os.path.dirname(self._current_project_path) if self._current_project_path else '')
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel 数据表", default_dir, "Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'

        # ── 坐标转换工具 ──────────────────────────────────────────────
        canvas_rect = self.scene.sceneRect()
        canvas_w = canvas_rect.width()
        canvas_h = canvas_rect.height()
        # CorelDRAW 导出 DPI（与 SVG 导出保持一致）
        DPI = CORELDRAW_EXPORT_DPI  # 600

        def px_to_mm(px):
            """像素 → 毫米"""
            return round(px * 25.4 / DPI, 4)

        def to_cdr(x_px, y_px):
            """Qt 坐标（左上角原点，Y向下）→ CDR 坐标（左下角原点，Y向上），单位 mm"""
            local_x = x_px - canvas_rect.left()
            local_y = y_px - canvas_rect.top()
            cdr_x = px_to_mm(local_x)
            cdr_y = px_to_mm(canvas_h - local_y)
            return cdr_x, cdr_y

        wb = openpyxl.Workbook()

        # ── 表1：画布 ──────────────────────────────────────────────────
        ws_canvas = wb.active
        ws_canvas.title = "画布"
        ws_canvas.append(["参数", "值", "单位", "说明"])
        ws_canvas.append(["画布宽度", px_to_mm(canvas_w), "mm", "CorelDRAW 页面宽度"])
        ws_canvas.append(["画布高度", px_to_mm(canvas_h), "mm", "CorelDRAW 页面高度"])
        ws_canvas.append(["导出DPI", DPI, "DPI", "像素与毫米换算基准"])
        ws_canvas.append(["坐标原点", "左下角", "", "X向右，Y向上"])

        # ── 表2：图片 ──────────────────────────────────────────────────
        ws_img = wb.create_sheet("图片")
        ws_img.append(["图片名称", "图片路径", "X(mm)", "Y(mm)", "宽(mm)", "高(mm)"])
        for item in self.scene.items():
            if isinstance(item, VImageItem):
                name = os.path.basename(item.file_path)
                scene_item_rect = item.mapRectToScene(item.boundingRect())
                cdr_x, cdr_y = to_cdr(scene_item_rect.right(), scene_item_rect.top())
                ws_img.append([name, os.path.abspath(item.file_path),
                                cdr_x, cdr_y,
                                px_to_mm(scene_item_rect.width()),
                                px_to_mm(scene_item_rect.height())])

        # ── 表3：文字 ──────────────────────────────────────────────────
        ws_txt = wb.create_sheet("文字")
        ws_txt.append([
            "文字内容", "字体", "字号(pt)", "颜色", "X(mm)", "Y(mm)",
            "宽(mm)", "高(mm)", "每列字数", "列间距(mm)", "字间距(mm)", "自动高度", "手动换行"
        ])
        text_items = [
            item for item in self.scene.items()
            if isinstance(item, VTextItem)
        ]
        for item in text_items:
            if isinstance(item, VTextItem):
                scene_item_rect = item.mapRectToScene(item.boundingRect())
                cdr_x, cdr_y = to_cdr(scene_item_rect.right(), scene_item_rect.top())
                ws_txt.append([
                    item.full_text,
                    item.font_family,
                    item.font_size,
                    item.text_color.name(),
                    cdr_x,
                    cdr_y,
                    px_to_mm(scene_item_rect.width()),
                    px_to_mm(scene_item_rect.height()),
                    item.chars_per_column,
                    px_to_mm(item.column_spacing),
                    px_to_mm(item.character_spacing),
                    "是" if item.auto_height else "否",
                    "是" if item.manual_line_break else "否",
                ])

        # ── 表4：文字字符 ──────────────────────────────────────────────
        # 每个字符使用实际 QGraphicsSimpleTextItem 的场景包围盒，
        # 供 VBA 逐字放置，避免 CorelDRAW 重新计算竖排间距。
        ws_chars = wb.create_sheet("文字字符")
        ws_chars.append([
            "文字对象编号", "原文字内容", "字符序号", "字符",
            "字体", "字号(pt)", "颜色", "旋转角度",
            "X(mm)", "Y(mm)", "宽(mm)", "高(mm)",
            "中心X(mm)", "中心Y(mm)"
        ])
        for text_id, item in enumerate(text_items, start=1):
            char_index = 0
            for child in item.childItems():
                if not isinstance(child, QGraphicsSimpleTextItem):
                    continue
                char_text = child.text()
                if not char_text:
                    continue
                char_rect = child.mapRectToScene(child.boundingRect())
                char_x, char_y = to_cdr(char_rect.right(), char_rect.top())
                center_x, center_y = to_cdr(char_rect.center().x(), char_rect.center().y())
                child_font = child.font()
                child_color = child.brush().color().name()
                ws_chars.append([
                    text_id,
                    item.full_text,
                    char_index,
                    char_text,
                    child_font.family(),
                    child_font.pointSize(),
                    child_color,
                    round(child.rotation(), 4),
                    char_x,
                    char_y,
                    px_to_mm(char_rect.width()),
                    px_to_mm(char_rect.height()),
                    center_x,
                    center_y,
                ])
                char_index += 1

        # ── 表5：连线 ──────────────────────────────────────────────────
        ws_conn = wb.create_sheet("连线")
        ws_conn.append(["线编号", "端点类型", "X(mm)", "Y(mm)", "粗细(px)", "颜色"])

        def _endpoints(conn):
            p = conn.path()
            if p.elementCount() == 0:
                return None
            e0  = p.elementAt(0)
            elt = p.elementAt(p.elementCount() - 1)
            # 与图片/文字保持一致：距页面左边 X，距页面底边 Y
            sx, sy = to_cdr(e0.x, e0.y)
            ex, ey = to_cdr(elt.x, elt.y)
            return sx, sy, ex, ey

        all_conns = [
            c for c in self.scene.image_text_connectors
            if isinstance(c, VImageTextConnector)
            or (isinstance(c, VGenericConnector) and c.connection_type == "image-image")
        ]
        for line_no, conn in enumerate(all_conns, start=1):
            pts = _endpoints(conn)
            if pts is None:
                continue
            lw    = getattr(conn, 'line_width', 3)
            color = conn.pen().color().name()
            # 起点行
            ws_conn.append([line_no, "起点", pts[0], pts[1], lw, color])
            # 终点行
            ws_conn.append([line_no, "终点", pts[2], pts[3], lw, color])

        # 统一列宽
        for ws in (ws_canvas, ws_img, ws_txt, ws_chars, ws_conn):
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        try:
            wb.save(path)
            self.scene.config_manager.set('default_save_dir', os.path.dirname(path))
            self.status_bar.showMessage(f"Excel 已导出: {path}", 4000)
            QMessageBox.information(self, "导出完成",
                f"已导出到：\n{path}\n\n"
                f"图片：{ws_img.max_row - 1} 条  "
                f"文字：{ws_txt.max_row - 1} 条  "
                f"连线：{ws_conn.max_row - 1} 条\n\n"
                f"坐标已转换为 CorelDRAW 坐标系（单位：mm，Y轴向上，DPI={DPI}）")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"写入 Excel 失败：\n{e}")

    def export_coreldraw_svg(self):
        default_dir = self.scene.config_manager.get('default_save_dir', '') or \
                      (os.path.dirname(self._current_project_path) if self._current_project_path else '')
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出CorelDRAW SVG",
            default_dir,
            "SVG (*.svg)"
        )
        if not path:
            return
        if not path.lower().endswith('.svg'):
            path += '.svg'

        try:
            CorelSvgExporter.export(self.scene, path)
            self.scene.config_manager.set('default_save_dir', os.path.dirname(path))
            self.status_bar.showMessage(f"SVG已导出: {path}", 3000)
            QMessageBox.information(
                self,
                "导出完成",
                "SVG已导出。文字、线条和图片会以可选对象保存，适合在CorelDRAW中继续编辑。"
            )
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出SVG失败:\n{e}")

    def _configure_print_page(self, printer, scene):
        """按当前画布尺寸配置完整打印页，保持画布宽高比。"""
        rect = scene.sceneRect()
        width_mm = max(1.0, rect.width() / 96.0 * 25.4)
        height_mm = max(1.0, rect.height() / 96.0 * 25.4)
        page_size = QPageSize(
            QSizeF(width_mm, height_mm),
            QPageSize.Unit.Millimeter
        )
        # 使用完整页面，避免默认打印边距导致预览页面与画布比例不一致。
        printer.setFullPage(True)
        printer.setPageSize(page_size)
        printer.setPageMargins(
            QMarginsF(0, 0, 0, 0),
            QPageLayout.Unit.Millimeter
        )

    def _render_scene_to_printer(self, printer, scene=None):
        """将当前文档渲染到打印机或打印预览。"""
        scene = scene or self.scene
        original_show_grid = scene.show_grid
        original_show_connectors = scene.show_connectors
        original_show_image_text_connectors = scene.show_image_text_connectors
        original_show_connection_points = scene.show_connection_points
        original_show_guides = scene.show_guides
        selected_items = scene.selectedItems()
        painter = QPainter()

        try:
            scene.show_grid = False
            scene.set_connectors_visible(False)
            scene.set_image_text_connectors_visible(True)
            scene.set_connection_points_visible(False)
            scene.set_guides_visible(False)
            scene.clearSelection()

            rect = scene.sceneRect()
            target_rect = printer.paperRect(QPrinter.Unit.DevicePixel)
            if rect.isEmpty() or target_rect.isEmpty():
                return

            if not painter.begin(printer):
                raise RuntimeError("无法启动打印绘制")
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setRenderHint(QPainter.RenderHint.TextAntialiasing)
            painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)

            scale = min(
                target_rect.width() / rect.width(),
                target_rect.height() / rect.height()
            )
            painter.translate(target_rect.left(), target_rect.top())
            painter.scale(scale, scale)

            # 复用 PDF 导出的白色页面和背景图片渲染规则。
            scene._rendering_pdf = True
            painter.fillRect(
                QRectF(0, 0, rect.width(), rect.height()),
                QColor(255, 255, 255)
            )
            scene.render(
                painter,
                QRectF(0, 0, rect.width(), rect.height()),
                rect
            )
            scene._rendering_pdf = False
        finally:
            if painter.isActive():
                painter.end()
            scene._rendering_pdf = False
            scene.show_grid = original_show_grid
            scene.set_connectors_visible(original_show_connectors)
            scene.set_image_text_connectors_visible(original_show_image_text_connectors)
            scene.set_connection_points_visible(original_show_connection_points)
            scene.set_guides_visible(original_show_guides)
            for item in selected_items:
                item.setSelected(True)
            scene.update()

    def print_document(self):
        """打开打印设置并打印当前文档。"""
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        self._configure_print_page(printer, self.scene)
        dialog = QPrintDialog(printer, self)
        dialog.setWindowTitle("打印")
        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                self._render_scene_to_printer(printer, self.scene)
                self.status_bar.showMessage("打印任务已发送", 3000)
            except Exception as e:
                QMessageBox.critical(self, "打印失败", f"打印失败：\n{e}")

    def print_preview(self):
        """打开当前文档的打印预览。"""
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        self._configure_print_page(printer, self.scene)
        preview = QPrintPreviewDialog(printer, self)
        preview.setWindowTitle("打印预览")
        preview.paintRequested.connect(
            lambda preview_printer: self._render_scene_to_printer(
                preview_printer,
                self.scene
            )
        )
        preview.exec()

    def export_pdf(self):
        """导出为PDF文件 - 保留文字、图片、连线和背景图片"""
        default_dir = self.scene.config_manager.get('default_save_dir', '') or \
                      (os.path.dirname(self._current_project_path) if self._current_project_path else '')
        path, _ = QFileDialog.getSaveFileName(
            self, "导出PDF", default_dir, "PDF (*.pdf)"
        )
        if not path:
            return
        if not path.lower().endswith('.pdf'):
            path += '.pdf'
        self._do_export_pdf(path, show_dialog=True)

    def _do_export_pdf(self, path, show_dialog=False, scene=None):
        """实际执行PDF导出 - 用scene.render保证文字位置正确，图片/连线独立对象"""
        scene = scene or self.scene
        original_show_grid = scene.show_grid
        original_show_connectors = scene.show_connectors
        original_show_image_text_connectors = scene.show_image_text_connectors
        original_show_connection_points = scene.show_connection_points
        original_show_guides = scene.show_guides
        selected_items = scene.selectedItems()

        try:
            # 隐藏所有辅助元素，只保留图文连线
            scene.show_grid = False
            scene.set_connectors_visible(False)
            scene.set_image_text_connectors_visible(True)
            scene.set_connection_points_visible(False)
            scene.set_guides_visible(False)
            scene.clearSelection()

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(path)

            rect = scene.sceneRect()
            # 使用与 CorelDRAW SVG 导出相同的 DPI (600)，确保尺寸一致
            width_mm = rect.width() * 25.4 / CORELDRAW_EXPORT_DPI
            height_mm = rect.height() * 25.4 / CORELDRAW_EXPORT_DPI

            page_size = QPageSize(QSizeF(width_mm, height_mm), QPageSize.Unit.Millimeter)
            printer.setPageSize(page_size)
            printer.setPageOrientation(QPageLayout.Orientation.Portrait)
            printer.setPageMargins(QMarginsF(0, 0, 0, 0), QPageLayout.Unit.Millimeter)

            painter = QPainter()
            painter.begin(printer)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setRenderHint(QPainter.RenderHint.TextAntialiasing)
            painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)

            target_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
            scale = min(target_rect.width() / rect.width(), target_rect.height() / rect.height())
            painter.scale(scale, scale)

            # 设置PDF导出标志：跳过灰色背景/阴影，保留背景图片
            scene._rendering_pdf = True
            painter.fillRect(QRectF(0, 0, rect.width(), rect.height()), QColor(255, 255, 255))
            scene.render(painter, QRectF(0, 0, rect.width(), rect.height()), rect)
            scene._rendering_pdf = False

            painter.end()

        finally:
            scene._rendering_pdf = False
            scene.show_grid = original_show_grid
            scene.set_connectors_visible(original_show_connectors)
            scene.set_image_text_connectors_visible(original_show_image_text_connectors)
            scene.set_connection_points_visible(original_show_connection_points)
            scene.set_guides_visible(original_show_guides)
            for item in selected_items:
                item.setSelected(True)
            scene.update()

        scene.config_manager.set('default_save_dir', os.path.dirname(path))
        print(f"PDF已导出: {path}")
        if show_dialog:
            QMessageBox.information(
                self, "导出完成",
                f"PDF已成功导出到：\n{path}\n\n"
                f"画布尺寸：{int(width_mm)} × {int(height_mm)} mm\n\n"
                "已包含：背景图片、文字、图片、图文连线"
            )

    def new_project(self):
        """新建工程，关闭当前工程中的全部文档。"""
        if not self._ask_save_before_action("新建工程"):
            return
        self._current_project_path = None
        self._replace_documents([{'name': '文档 1', 'data': {}}])
        self._documents[0].dirty = False
        self._update_document_tab_title(self._documents[0])
        self.status_bar.showMessage("已新建工程", 2000)

    def save_proj(self):
        """另存为：弹出对话框选择路径"""
        default_dir = self.scene.config_manager.get('default_save_dir', '') or \
                      (os.path.dirname(self._current_project_path) if self._current_project_path else '')
        path, _ = QFileDialog.getSaveFileName(self, "保存工程", default_dir, "VLayout (*.vlayout)")
        if path:
            ProjectData.save_documents(self._documents, path)
            self._current_project_path = path
            for document in self._documents:
                document.dirty = False
                self._update_document_tab_title(document)
            self.scene.config_manager.set('default_save_dir', os.path.dirname(path))
            self.status_bar.showMessage(f"已保存: {path}", 3000)
            return True
        return False

    def quick_save_proj(self):
        """快速保存：直接覆盖当前文件，没有路径时弹对话框"""
        if self._current_project_path:
            ProjectData.save_documents(self._documents, self._current_project_path)
            for document in self._documents:
                document.dirty = False
                self._update_document_tab_title(document)
            self.status_bar.showMessage(f"已保存: {self._current_project_path}", 3000)
            return True
        else:
            return self.save_proj()

    def save_proj_and_pdf(self):
        """一键保存多文档工程，并按页码为每个文档导出一个 PDF。"""
        # 第一步：确保项目已有保存路径
        if not self._current_project_path:
            if not self.save_proj():
                return

        # 第二步：保存项目
        ProjectData.save_documents(self._documents, self._current_project_path)
        for document in self._documents:
                document.dirty = False
                self._update_document_tab_title(document)
        self.status_bar.showMessage(f"项目已保存: {self._current_project_path}", 2000)

        # 第三步：指定第一个文档的页码，后续文档自动递增
        first_page, ok = QInputDialog.getInt(
            self,
            "批量导出 PDF",
            "请输入第一个文档的页码：",
            1,
            1,
            999999,
            1
        )
        if not ok:
            self.status_bar.showMessage("已保存工程，已取消 PDF 导出", 3000)
            return

        pdf_dir = os.path.dirname(self._current_project_path)
        pdf_paths = []
        for index, document in enumerate(self._documents, start=1):
            page_number = first_page + index - 1
            filename = f"{page_number}.pdf"
            pdf_path = os.path.join(pdf_dir, filename)
            self._do_export_pdf(pdf_path, scene=document.scene)
            pdf_paths.append(pdf_path)

        self.status_bar.showMessage(
            f"已保存项目并导出 {len(pdf_paths)} 个文档 PDF", 4000
        )
        QMessageBox.information(
            self,
            "导出完成",
            "工程和文档 PDF 已全部导出：\n\n" +
            "\n".join(os.path.basename(path) for path in pdf_paths)
        )

    def load_proj(self):
        default_dir = self.scene.config_manager.get('default_save_dir', '') or \
                      (os.path.dirname(self._current_project_path) if self._current_project_path else '')
        path, _ = QFileDialog.getOpenFileName(self, "打开工程", default_dir, "VLayout (*.vlayout)")
        if path:
            if not self._ask_save_before_action("打开工程"):
                return
            document_data = ProjectData.read_documents(path)
            self._replace_documents(document_data)
            self._current_project_path = path
            for document in self._documents:
                document.dirty = False
                self._update_document_tab_title(document)
            self.scene.config_manager.set('default_save_dir', os.path.dirname(path))
            self.status_bar.showMessage(
                f"已打开: {os.path.basename(path)}（{len(self._documents)} 个文档）", 3000
            )
    
    def set_background_image(self):
        """设置默认背景图片"""
        path, _ = QFileDialog.getOpenFileName(
            self, 
            "选择背景图片", 
            "", 
            "Images (*.png *.jpg *.jpeg *.bmp *.gif)"
        )
        if path:
            if self.scene.set_background_image(path):
                QMessageBox.information(self, "成功", f"背景图片已设置\n{path}")
    
    def clear_background_image(self):
        """清除背景图片"""
        reply = QMessageBox.question(
            self,
            "确认",
            "确定要清除背景图片吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.scene.set_background_image('')
            QMessageBox.information(self, "成功", "背景图片已清除")

    def show_all_hidden_images(self):
        """显示所有被隐藏的图片"""
        count = 0
        for item in self.scene.items():
            if isinstance(item, VImageItem) and not item.isVisible():
                item.set_image_visible(True)
                count += 1
        self.status_bar.showMessage(f'已显示 {count} 张隐藏图片', 3000)

    def toggle_guides(self):
        """切换辅助线显示/隐藏"""
        self.scene.set_guides_visible(not self.scene.show_guides)

    def toggle_startup_horizontal_guides(self, enabled):
        """设置是否在启动和新建画布时自动添加横向辅助线。"""
        for document in self._documents:
            document.scene.config_manager.set('startup_horizontal_guides_enabled', enabled)
        if hasattr(self, 'status_bar'):
            message = '已开启：启动时添加横向辅助线' if enabled else '已关闭：启动时添加横向辅助线'
            self.status_bar.showMessage(message, 3000)

    def set_display_unit(self, action):
        """切换全局显示/输入单位；场景内部坐标始终为像素。"""
        unit = action.data() or 'mm'
        for document in self._documents:
            document.scene.config_manager.set('display_unit', unit)
        for view in (document.view for document in self._documents):
            view.viewport().update()
        if hasattr(self, 'status_bar'):
            self.status_bar.showMessage(f'界面单位已切换为{"毫米" if unit == "mm" else "像素"}', 3000)

    def set_startup_horizontal_guides(self):
        """设置启动横向辅助线的位置和输入单位。内部始终保存为场景像素。"""
        config = self.scene.config_manager
        positions = config.get('startup_horizontal_guides', [])
        unit = config.get('display_unit', 'mm')
        dialog = QDialog(self)
        dialog.setWindowTitle('设置启动横向辅助线位置')
        layout = QFormLayout(dialog)
        unit_combo = QComboBox(dialog)
        unit_combo.addItem('像素 (px)', 'px')
        unit_combo.addItem('毫米 (mm)', 'mm')
        unit_combo.setCurrentIndex(1 if unit == 'mm' else 0)
        values = []
        for raw in positions:
            try:
                number = float(raw)
                values.append(number * 25.4 / CORELDRAW_EXPORT_DPI if unit == 'mm' else number)
            except (TypeError, ValueError):
                continue
        value_edit = QLineEdit(', '.join(f'{number:g}' for number in values), dialog)
        layout.addRow('输入单位：', unit_combo)
        layout.addRow('Y 坐标（逗号分隔）：', value_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, parent=dialog)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        value = value_edit.text()
        selected_unit = unit_combo.currentData()
        parsed = []
        for part in value.replace('，', ',').split(','):
            part = part.strip()
            if not part:
                continue
            try:
                number = float(part)
                parsed.append(number * CORELDRAW_EXPORT_DPI / 25.4 if selected_unit == 'mm' else number)
            except ValueError:
                QMessageBox.warning(self, '输入无效', f'无法识别坐标：{part}')
                return
        for document in self._documents:
            document.scene.config_manager.set('startup_horizontal_guides', parsed)
            document.scene.config_manager.set('startup_horizontal_guides_unit', selected_unit)
            document.scene.config_manager.set('display_unit', selected_unit)
        if hasattr(self, 'display_unit_px_action'):
            self.display_unit_px_action.setChecked(selected_unit == 'px')
            self.display_unit_mm_action.setChecked(selected_unit == 'mm')
        for document in self._documents:
            document.view.viewport().update()
        if hasattr(self, 'status_bar'):
            unit_label = '毫米' if selected_unit == 'mm' else '像素'
            self.status_bar.showMessage(f'已保存 {len(parsed)} 条启动横向辅助线位置（{unit_label}）', 3000)

    def clear_guides(self):
        """清除所有辅助线"""
        self.scene.clear_guides()

    def set_snap_threshold(self):
        """设置辅助线吸附距离"""
        val, ok = QInputDialog.getInt(
            self, "吸附距离", "吸附距离（场景像素）:",
            self.scene.snap_threshold, 1, 200
        )
        if ok:
            self.scene.snap_threshold = val
            # 保存到配置文件
            self.scene.config_manager.set('snap_threshold', val)
    
    def set_background_opacity(self):
        """设置背景透明度"""
        current_opacity = self.scene.config_manager.get('background_opacity', 0.3)
        opacity, ok = QInputDialog.getDouble(
            self,
            "设置背景透明度",
            "透明度 (0.0-1.0):",
            current_opacity,
            0.0,
            1.0,
            2
        )
        if ok:
            self.scene.config_manager.set('background_opacity', opacity)
            self.scene.update()
            print(f"背景透明度已设置为: {opacity}")
    
    def set_background_scale_mode(self, mode):
        """设置背景缩放模式"""
        mode_names = {
            'fit': '适应画布',
            'fill': '填充画布',
            'stretch': '拉伸填充',
            'tile': '平铺'
        }
        self.scene.config_manager.set('background_scale_mode', mode)
        self.scene.update()
        print(f"背景缩放模式已设置为: {mode_names.get(mode, mode)}")
    
    def set_default_font(self):
        """设置默认字体"""
        # 获取当前默认字体
        current_font_family = self.scene.config_manager.get('default_font_family', DEFAULT_FONT)
        current_font_size = self.scene.config_manager.get('default_font_size', DEFAULT_FONT_SIZE)
        current_font = QFont(current_font_family)
        current_font.setPointSize(current_font_size)
        
        # 打开字体选择对话框
        font, ok = FontPickerDialog.get_font(current_font, self.scene.config_manager, self, "设置默认字体")
        if ok:
            text_format = self._default_text_format()
            text_format['font_family'] = font.family()
            text_format['font_size'] = font.pointSize()
            self._set_shared_configs({
                'default_font_family': font.family(),
                'default_font_size': font.pointSize(),
                'default_text_format': text_format,
            })
            
            # 更新工具栏显示
            self.font_combo.setCurrentText(font.family())
            self.font_size_spin.setValue(font.pointSize())
            
            QMessageBox.information(
                self,
                "设置成功",
                f"默认字体已设置为:\n字体: {font.family()}\n大小: {font.pointSize()}pt\n\n新添加的文字将使用此字体。"
            )
            print(f"默认字体已设置: {font.family()}, {font.pointSize()}pt")

    def set_default_save_dir(self):
        """设置默认保存目录"""
        current = self.scene.config_manager.get('default_save_dir', '')
        path = QFileDialog.getExistingDirectory(self, "选择默认保存目录", current)
        if path:
            self.scene.config_manager.set('default_save_dir', path)
            self.status_bar.showMessage(f"默认保存目录已设置为: {path}", 3000)

    def set_insert_image_size(self):
        """设置插入图片的默认宽高（px）"""
        cur_w = self.scene.config_manager.get('insert_image_default_width', 0)
        cur_h = self.scene.config_manager.get('insert_image_default_height', 0)
        use_custom = self.scene.config_manager.get('insert_image_use_custom_size', False)
        canvas_w = int(self.scene.sceneRect().width())
        canvas_h = int(self.scene.sceneRect().height())

        dialog = QDialog(self)
        dialog.setWindowTitle("插入图片默认大小")
        layout = QFormLayout(dialog)

        chk_enable = QCheckBox("开启自定义插入尺寸")
        chk_enable.setChecked(use_custom)
        layout.addRow(chk_enable)

        spin_w = QSpinBox()
        spin_w.setRange(1, 20000)
        spin_w.setValue(cur_w if cur_w else canvas_w)
        spin_w.setSuffix(" px")
        layout.addRow(f"宽度（画布宽={canvas_w}px）:", spin_w)

        spin_h = QSpinBox()
        spin_h.setRange(0, 20000)
        spin_h.setValue(cur_h if cur_h else canvas_h)
        spin_h.setSuffix(" px")
        spin_h.setSpecialValueText("按比例自动")
        layout.addRow(f"高度（画布高={canvas_h}px）:", spin_h)

        hint = QLabel("关闭时按原图大小插入\n高度设为0则按宽度比例自动计算")
        hint.setStyleSheet("color: gray; font-size: 11px;")
        layout.addRow(hint)

        # 联动：未开启时禁用输入框
        def on_toggle(checked):
            spin_w.setEnabled(checked)
            spin_h.setEnabled(checked)
        chk_enable.toggled.connect(on_toggle)
        on_toggle(use_custom)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addRow(btns)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            enabled = chk_enable.isChecked()
            w, h = spin_w.value(), spin_h.value()
            self.scene.config_manager.set('insert_image_use_custom_size', enabled)
            self.scene.config_manager.set('insert_image_default_width', w)
            self.scene.config_manager.set('insert_image_default_height', h)
            if not enabled:
                self.status_bar.showMessage("插入图片大小：按原图大小", 3000)
            else:
                self.status_bar.showMessage(f"插入图片默认大小已设置为 {w} × {h if h else '自动'}px", 3000)
    
    def apply_fluent_design_style(self):
        """应用Fluent Design风格样式"""
        fluent_style = """
        /* 主窗口样式 */
        QMainWindow {
            background-color: #f3f3f3;
            color: #323130;
        }
        
        /* 工具栏样式 */
        QToolBar {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(255, 255, 255, 0.95),
                stop:1 rgba(245, 245, 245, 0.95));
            border: none;
            border-bottom: 1px solid rgba(0, 0, 0, 0.1);
            spacing: 4px;
            padding: 3px 4px;
        }
        
        QToolBar::separator {
            background-color: rgba(0, 0, 0, 0.1);
            width: 1px;
            margin: 3px 2px;
        }

        QToolBar#toolbar_main QToolButton {
            padding: 3px 8px;
            min-height: 20px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: 500;
        }

        QToolBar#toolbar_main QLabel {
            font-size: 12px;
            padding-left: 2px;
            padding-right: 0px;
        }

        QToolBar#toolbar_main QComboBox,
        QToolBar#toolbar_main QSpinBox {
            min-height: 22px;
            padding: 2px 6px;
            font-size: 12px;
        }
        
        /* 按钮基础样式 - Fluent Design */
        QPushButton, QToolButton {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(255, 255, 255, 0.9),
                stop:1 rgba(249, 249, 249, 0.9));
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-radius: 6px;
            padding: 8px 16px;
            font-size: 13px;
            font-weight: 500;
            color: #323130;
            min-height: 20px;
        }
        
        QPushButton:hover, QToolButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(255, 255, 255, 1.0),
                stop:1 rgba(245, 245, 245, 1.0));
            border: 1px solid rgba(0, 120, 215, 0.4);
        }
        
        QPushButton:pressed, QToolButton:pressed {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(240, 240, 240, 1.0),
                stop:1 rgba(230, 230, 230, 1.0));
            border: 1px solid rgba(0, 120, 215, 0.6);
        }
        
        /* 主要按钮样式 */
        QPushButton[class="primary"] {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(0, 120, 215, 1.0),
                stop:1 rgba(0, 90, 158, 1.0));
            color: white;
            border: 1px solid rgba(0, 90, 158, 1.0);
            font-weight: 600;
        }
        
        QPushButton[class="primary"]:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(16, 132, 208, 1.0),
                stop:1 rgba(0, 102, 180, 1.0));
        }
        
        QPushButton[class="primary"]:pressed {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(0, 90, 158, 1.0),
                stop:1 rgba(0, 78, 140, 1.0));
        }
        
        /* 危险按钮样式 */
        QPushButton[class="danger"] {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(196, 43, 28, 1.0),
                stop:1 rgba(164, 38, 25, 1.0));
            color: white;
            border: 1px solid rgba(164, 38, 25, 1.0);
        }
        
        QPushButton[class="danger"]:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(200, 56, 39, 1.0),
                stop:1 rgba(173, 45, 31, 1.0));
        }
        
        /* 停靠面板样式 */
        QDockWidget {
            background: rgba(255, 255, 255, 0.95);
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-radius: 8px;
            titlebar-close-icon: none;
            titlebar-normal-icon: none;
        }
        
        QDockWidget::title {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(255, 255, 255, 0.95),
                stop:1 rgba(248, 248, 248, 0.95));
            border-bottom: 1px solid rgba(0, 0, 0, 0.1);
            border-top-left-radius: 8px;
            border-top-right-radius: 8px;
            padding: 8px;
            font-weight: 600;
            color: #323130;
        }
        
        /* 标签页样式 */
        QTabWidget::pane {
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-radius: 6px;
            background: rgba(255, 255, 255, 0.9);
        }
        
        QTabBar::tab {
            background: rgba(245, 245, 245, 0.9);
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-bottom: none;
            border-top-left-radius: 6px;
            border-top-right-radius: 6px;
            padding: 8px 16px;
            margin-right: 2px;
            color: #605e5c;
        }
        
        QTabBar::tab:selected {
            background: rgba(255, 255, 255, 1.0);
            color: #323130;
            font-weight: 600;
        }
        
        QTabBar::tab:hover:!selected {
            background: rgba(250, 250, 250, 1.0);
            color: #323130;
        }
        
        /* 列表样式 */
        QListWidget {
            background: rgba(255, 255, 255, 0.9);
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-radius: 6px;
            padding: 4px;
            selection-background-color: rgba(0, 120, 215, 0.2);
        }
        
        QListWidget::item {
            border-radius: 4px;
            padding: 6px;
            margin: 1px;
        }
        
        QListWidget::item:hover {
            background: rgba(0, 120, 215, 0.1);
        }
        
        QListWidget::item:selected {
            background: rgba(0, 120, 215, 0.2);
            color: #323130;
        }
        
        /* 树形控件样式 */
        QTreeWidget {
            background: rgba(255, 255, 255, 0.9);
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-radius: 6px;
            selection-background-color: rgba(0, 120, 215, 0.2);
        }
        
        /* 输入框样式 */
        QSpinBox, QLineEdit, QTextEdit {
            background: rgba(255, 255, 255, 0.9);
            border: 1px solid rgba(0, 0, 0, 0.2);
            border-radius: 4px;
            padding: 6px;
            color: #323130;
        }
        
        QSpinBox:focus, QLineEdit:focus, QTextEdit:focus {
            border: 2px solid rgba(0, 120, 215, 0.8);
        }
        
        /* 组合框样式 */
        QComboBox {
            background: rgba(255, 255, 255, 0.9);
            border: 1px solid rgba(0, 0, 0, 0.2);
            border-radius: 4px;
            padding: 6px;
            color: #323130;
        }
        
        QComboBox:hover {
            border: 1px solid rgba(0, 120, 215, 0.4);
        }
        
        QComboBox::drop-down {
            border: none;
            width: 20px;
        }
        
        /* 状态栏样式 */
        QStatusBar {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(255, 255, 255, 0.95),
                stop:1 rgba(245, 245, 245, 0.95));
            border-top: 1px solid rgba(0, 0, 0, 0.1);
            color: #605e5c;
        }
        
        /* 菜单样式 */
        QMenuBar {
            background: rgba(255, 255, 255, 0.95);
            border-bottom: 1px solid rgba(0, 0, 0, 0.1);
            color: #323130;
        }
        
        QMenuBar::item {
            background: transparent;
            padding: 8px 12px;
            border-radius: 4px;
        }
        
        QMenuBar::item:selected {
            background: rgba(0, 120, 215, 0.1);
        }
        
        QMenu {
            background: rgba(255, 255, 255, 0.95);
            border: 1px solid rgba(0, 0, 0, 0.1);
            border-radius: 6px;
            padding: 4px;
        }
        
        QMenu::item {
            padding: 8px 16px;
            border-radius: 4px;
        }
        
        QMenu::item:selected {
            background: rgba(0, 120, 215, 0.1);
        }
        
        /* 滚动条样式 */
        QScrollBar:vertical {
            background: rgba(245, 245, 245, 0.8);
            width: 12px;
            border-radius: 6px;
        }
        
        QScrollBar::handle:vertical {
            background: rgba(0, 0, 0, 0.3);
            border-radius: 6px;
            min-height: 20px;
        }
        
        QScrollBar::handle:vertical:hover {
            background: rgba(0, 0, 0, 0.5);
        }
        
        QScrollBar:horizontal {
            background: rgba(245, 245, 245, 0.8);
            height: 12px;
            border-radius: 6px;
        }
        
        QScrollBar::handle:horizontal {
            background: rgba(0, 0, 0, 0.3);
            border-radius: 6px;
            min-width: 20px;
        }
        
        QScrollBar::handle:horizontal:hover {
            background: rgba(0, 0, 0, 0.5);
        }

        /* Professional editor refinements */
        QMainWindow {
            background-color: #f6f8fa;
        }

        QToolBar {
            background: #ffffff;
            border: none;
            border-bottom: 1px solid #d8dee4;
            spacing: 6px;
            padding: 6px 8px;
        }

        QToolBar::separator {
            background-color: #d0d7de;
            width: 1px;
            margin: 5px 4px;
        }

        QToolBar#toolbar_main QToolButton {
            background: #ffffff;
            border: 1px solid #d0d7de;
            border-radius: 5px;
            color: #24292f;
            padding: 5px 11px;
            min-height: 24px;
            font-size: 12px;
            font-weight: 500;
        }

        QToolBar#toolbar_main QToolButton:hover {
            background: #f6f8fa;
            border-color: #0969da;
        }

        QToolBar#toolbar_main QToolButton:pressed,
        QToolBar#toolbar_main QToolButton:checked {
            background: #0969da;
            border-color: #0969da;
            color: #ffffff;
        }

        QLabel#toolbarHint {
            color: #57606a;
            padding-right: 4px;
            font-size: 12px;
        }

        QPushButton, QToolButton {
            background: #ffffff;
            border: 1px solid #d0d7de;
            border-radius: 5px;
            color: #24292f;
            padding: 6px 12px;
            min-height: 24px;
            font-size: 13px;
            font-weight: 500;
        }

        QPushButton:hover, QToolButton:hover {
            background: #f6f8fa;
            border-color: #0969da;
        }

        QPushButton:pressed, QToolButton:pressed {
            background: #eaeef2;
        }

        QPushButton[class="primary"] {
            background: #0969da;
            border-color: #0969da;
            color: #ffffff;
            font-weight: 600;
        }

        QPushButton[class="primary"]:hover {
            background: #075fc5;
            border-color: #075fc5;
        }

        QPushButton[class="danger"] {
            background: #cf222e;
            border-color: #cf222e;
            color: #ffffff;
            font-weight: 600;
        }

        QPushButton[class="danger"]:hover {
            background: #a40e26;
            border-color: #a40e26;
        }

        QPushButton[class="toggle"]:checked {
            background: #0969da;
            border-color: #0969da;
            color: #ffffff;
        }

        QDockWidget {
            background: #ffffff;
            border: 1px solid #d8dee4;
            border-radius: 0;
        }

        QDockWidget::title {
            background: #f6f8fa;
            border-bottom: 1px solid #d8dee4;
            padding: 7px 8px;
            font-weight: 600;
            color: #24292f;
        }

        QTabWidget::pane {
            border: 1px solid #d8dee4;
            border-radius: 4px;
            background: #ffffff;
        }

        QTabBar::tab {
            background: #f6f8fa;
            border: 1px solid #d8dee4;
            border-bottom: none;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
            padding: 7px 13px;
            color: #57606a;
        }

        QTabBar::tab:selected {
            background: #ffffff;
            color: #24292f;
            font-weight: 600;
        }

        QListWidget, QTreeWidget {
            background: #ffffff;
            border: 1px solid #d8dee4;
            border-radius: 5px;
            padding: 4px;
            alternate-background-color: #f6f8fa;
            selection-background-color: #ddf4ff;
            selection-color: #24292f;
        }

        QListWidget::item {
            border-radius: 4px;
            padding: 7px 8px;
            margin: 1px;
        }

        QListWidget::item:hover,
        QTreeWidget::item:hover {
            background: #f6f8fa;
        }

        QListWidget::item:selected,
        QTreeWidget::item:selected {
            background: #ddf4ff;
            color: #24292f;
        }

        QGroupBox {
            background: #ffffff;
            border: 1px solid #d8dee4;
            border-radius: 5px;
            margin-top: 8px;
            font-weight: 600;
            color: #24292f;
        }

        QGroupBox::title {
            subcontrol-origin: margin;
            left: 8px;
            padding: 0 4px;
            background: #ffffff;
        }

        QLabel#propertySummary {
            background: #f6f8fa;
            border: 1px solid #d8dee4;
            border-radius: 5px;
            color: #57606a;
            padding: 8px;
        }

        QSpinBox, QLineEdit, QTextEdit, QComboBox {
            background: #ffffff;
            border: 1px solid #d0d7de;
            border-radius: 5px;
            padding: 5px 7px;
            color: #24292f;
            min-height: 22px;
        }

        QSpinBox:focus, QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
            border: 1px solid #0969da;
        }

        QStatusBar {
            background: #ffffff;
            border-top: 1px solid #d8dee4;
            color: #57606a;
        }
        """
        
        self.setStyleSheet(fluent_style)
    
    def set_button_style(self, button, style_class="default"):
        """为按钮设置特定的Fluent Design样式"""
        if style_class == "primary":
            button.setProperty("class", "primary")
        elif style_class == "danger":
            button.setProperty("class", "danger")
        
        # 刷新样式
        button.style().unpolish(button)
        button.style().polish(button)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # 设置应用属性以支持更好的视觉效果 (PyQt6兼容)
    try:
        app.setAttribute(Qt.ApplicationAttribute.AA_UseHighDpiPixmaps, True)
    except AttributeError:
        # 如果属性不存在，跳过设置
        pass
    
    # 使用现代样式
    app.setStyle("Fusion")
    
    # 设置应用图标和信息
    app.setApplicationName("VertiLayout Pro")
    app.setApplicationVersion("1.0")
    app.setOrganizationName("VertiLayout")

    license_config = ConfigManager()
    license_valid, license_message = license_is_valid(license_config.config)
    if not license_valid:
        if not recover_expired_license(license_config):
            QMessageBox.critical(
                None,
                "软件已到期",
                f"{license_message}。\n授权验证失败，软件无法使用。"
            )
            sys.exit(0)

    w = MainWindow()
    w.showMaximized()  # 启动时最大化窗口
    sys.exit(app.exec())
